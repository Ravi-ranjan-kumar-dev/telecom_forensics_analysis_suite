from __future__ import annotations

from pathlib import Path
from typing import Iterable, List, Dict
import sys
import re

PROJECT_ROOT = Path(__file__).resolve().parents[1]

if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import pandas as pd
from openpyxl import load_workbook

from modules.database.cgi_repository import CGI_TABLE, create_cgi_table, cgi_count, normalize_cgi
from modules.database.master_importer import _upsert_dataframe
from modules.database.cgi_master_reader import (
    CGI_OUTPUT_COLUMNS,
    prepare_standard_cgi_dataframe,
    _clean_text,
    _clean_cgi,
    _looks_like_cgi,
    _header_score,
    _dedupe_columns,
)


BATCH_SIZE = 25000


def split_caret_row(values: Iterable[object]) -> List[str]:
    cells = [_clean_text(value) for value in values if _clean_text(value)]

    if not cells:
        return []

    joined = " ".join(cells)

    if "^" not in joined:
        return []

    return [part.strip() for part in joined.split("^")]


def flush_standard_records(records: List[Dict[str, str]], source: str) -> int:
    if not records:
        return 0

    raw = pd.DataFrame(records)
    prepared = prepare_standard_cgi_dataframe(raw, source)

    if prepared is None or prepared.empty:
        return 0

    return _upsert_dataframe(CGI_TABLE, prepared[CGI_OUTPUT_COLUMNS], "cgi")


def flush_positional_records(records: List[Dict[str, object]]) -> int:
    if not records:
        return 0

    prepared = pd.DataFrame(records)

    if prepared.empty:
        return 0

    prepared = prepared[CGI_OUTPUT_COLUMNS].drop_duplicates("cgi", keep="last")

    return _upsert_dataframe(CGI_TABLE, prepared, "cgi")


def to_float(value):
    try:
        return float(str(value).strip())
    except Exception:
        return None


def detect_positional_layout(sample_rows: List[List[str]]):
    if not sample_rows:
        return {"cgi": 0, "lat": 7, "lon": 8}

    max_columns = max(len(row) for row in sample_rows)
    best_cgi_index = 0
    best_cgi_score = -1

    for index in range(max_columns):
        score = 0

        for row in sample_rows:
            value = row[index] if index < len(row) else ""
            if _looks_like_cgi(_clean_cgi(value), f"Column{index + 1}"):
                score += 1

        if score > best_cgi_score:
            best_cgi_score = score
            best_cgi_index = index

    best_lat_index = 7
    best_lon_index = 8
    best_pair_score = -1

    for index in range(max_columns - 1):
        pair_score = 0

        for row in sample_rows:
            lat = to_float(row[index] if index < len(row) else "")
            lon = to_float(row[index + 1] if index + 1 < len(row) else "")

            if lat is None or lon is None:
                continue

            if 5 <= lat <= 40 and 65 <= lon <= 100:
                pair_score += 1

        if pair_score > best_pair_score:
            best_pair_score = pair_score
            best_lat_index = index
            best_lon_index = index + 1

    return {
        "cgi": best_cgi_index,
        "lat": best_lat_index,
        "lon": best_lon_index,
    }


def make_positional_record(values: List[str], layout, source: str):
    cgi_index = layout["cgi"]
    lat_index = layout["lat"]
    lon_index = layout["lon"]

    cgi = _clean_cgi(values[cgi_index] if cgi_index < len(values) else "")

    if not _looks_like_cgi(cgi, f"Column{cgi_index + 1}"):
        return None

    latitude = pd.to_numeric(values[lat_index] if lat_index < len(values) else "", errors="coerce")
    longitude = pd.to_numeric(values[lon_index] if lon_index < len(values) else "", errors="coerce")

    record = {
        "cgi": cgi,
        "operator": values[1] if len(values) > 1 else "",
        "circle": values[2] if len(values) > 2 else "",
        "state": "",
        "district": "",
        "police_station": "",
        "address": values[5] if len(values) > 5 else "",
        "latitude": latitude,
        "longitude": longitude,
        "source_file": source,
        "site_name": values[4] if len(values) > 4 else "",
        "town": values[3] if len(values) > 3 else "",
        "landmark": values[6] if len(values) > 6 else "",
        "azimuth": values[9] if len(values) > 9 else "",
        "technology": "",
        "status": values[10] if len(values) > 10 else "",
        "status_change_date": values[11] if len(values) > 11 else "",
        "mcc_mnc": "",
        "lac": "",
        "cid": "",
        "tac_id": "",
        "site_id": "",
        "gnb_id": "",
        "cell_id": "",
    }

    return record


def import_xlsx_streaming(file_path: Path) -> int:
    create_cgi_table()

    workbook = load_workbook(file_path, read_only=True, data_only=True)

    total_imported = 0

    print(f"[+] Workbook opened: {file_path.name}", flush=True)
    print(f"[+] Sheets found: {workbook.sheetnames}", flush=True)

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]

        print(f"\n[+] Reading sheet: {sheet_name}", flush=True)

        standard_header = None
        standard_records: List[Dict[str, str]] = []
        positional_records: List[Dict[str, object]] = []
        positional_sample: List[List[str]] = []
        positional_layout = None
        mode = None
        rows_seen = 0
        rows_used = 0
        sheet_imported = 0

        for row in worksheet.iter_rows(values_only=True):
            rows_seen += 1

            values = [_clean_text(value) for value in row]
            non_empty_values = [value for value in values if value]

            if not non_empty_values:
                continue

            caret_parts = split_caret_row(values)

            if caret_parts:
                if standard_header is None and _header_score(caret_parts) >= 80:
                    standard_header = _dedupe_columns(caret_parts)
                    mode = "caret_packed"
                    print(f"[+] Detected caret-packed header at row {rows_seen}", flush=True)
                    continue

                if standard_header is not None:
                    if _header_score(caret_parts) >= 80:
                        continue

                    if len(caret_parts) < len(standard_header):
                        caret_parts = caret_parts + [""] * (len(standard_header) - len(caret_parts))

                    record = {
                        standard_header[index]: caret_parts[index] if index < len(caret_parts) else ""
                        for index in range(len(standard_header))
                    }

                    standard_records.append(record)
                    rows_used += 1

                    if len(standard_records) >= BATCH_SIZE:
                        inserted = flush_standard_records(
                            standard_records,
                            f"{file_path.name}::{sheet_name}",
                        )
                        sheet_imported += inserted
                        total_imported += inserted
                        standard_records.clear()
                        print(
                            f"[PROGRESS] {file_path.name}::{sheet_name} | rows scanned={rows_seen} | rows used={rows_used} | inserted={sheet_imported}",
                            flush=True,
                        )

                    continue

            # Positional Column1/Column2 style fallback.
            if mode is None:
                if all(value.lower().startswith("column") for value in non_empty_values[:8]):
                    mode = "positional"
                    print(f"[+] Detected positional Column layout at row {rows_seen}", flush=True)
                    continue

            if mode == "positional":
                if all(value.lower().startswith("column") for value in non_empty_values[:8]):
                    continue

                if len(positional_sample) < 300:
                    positional_sample.append(values)
                    positional_layout = detect_positional_layout(positional_sample)

                if positional_layout is None:
                    positional_layout = detect_positional_layout(positional_sample)

                record = make_positional_record(
                    values,
                    positional_layout,
                    f"{file_path.name}::{sheet_name}",
                )

                if record:
                    positional_records.append(record)
                    rows_used += 1

                if len(positional_records) >= BATCH_SIZE:
                    inserted = flush_positional_records(positional_records)
                    sheet_imported += inserted
                    total_imported += inserted
                    positional_records.clear()
                    print(
                        f"[PROGRESS] {file_path.name}::{sheet_name} | rows scanned={rows_seen} | rows used={rows_used} | inserted={sheet_imported}",
                        flush=True,
                    )

        if standard_records:
            inserted = flush_standard_records(
                standard_records,
                f"{file_path.name}::{sheet_name}",
            )
            sheet_imported += inserted
            total_imported += inserted
            standard_records.clear()

        if positional_records:
            inserted = flush_positional_records(positional_records)
            sheet_imported += inserted
            total_imported += inserted
            positional_records.clear()

        print(
            f"[DONE] Sheet={sheet_name} | rows scanned={rows_seen} | rows used={rows_used} | inserted={sheet_imported}",
            flush=True,
        )

    workbook.close()

    return total_imported


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 -u tools/import_cgi_streaming_retry.py <file1.xlsx> [file2.xlsx ...]")
        raise SystemExit(2)

    print("[+] Streaming CGI importer started", flush=True)
    print("[+] Current CGI rows:", cgi_count(), flush=True)

    grand_total = 0

    for raw_path in sys.argv[1:]:
        file_path = Path(raw_path)

        print("\n" + "=" * 80, flush=True)
        print("[FILE]", file_path, flush=True)
        print("=" * 80, flush=True)

        if not file_path.exists():
            print("[SKIP] File not found:", file_path, flush=True)
            continue

        if file_path.suffix.lower() != ".xlsx":
            print("[SKIP] Streaming retry currently supports .xlsx only:", file_path.name, flush=True)
            continue

        imported = import_xlsx_streaming(file_path)
        grand_total += imported

        print(f"[OK] Imported from file: {imported}", flush=True)
        print(f"[DB] Total CGI rows now: {cgi_count()}", flush=True)

    print("\n" + "=" * 80, flush=True)
    print("[SUMMARY] Streaming retry imported rows:", grand_total, flush=True)
    print("[SUMMARY] Final CGI rows:", cgi_count(), flush=True)


if __name__ == "__main__":
    main()
