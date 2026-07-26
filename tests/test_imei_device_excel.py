from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from modules.reporting import (
    generate_imei_device_report,
)
from modules.reporting.imei_device_excel import (
    IMEI_DEVICE_SHEETS,
)


IMEISV_16 = "8625180548786512"


def _analysis(
    *,
    status: str = "FOUND",
) -> dict:
    source_summary = pd.DataFrame(
        [
            {
                "Evidence Source": "CDR",
                "Status": "FOUND",
                "Evidence Unit": "CDR records",
                "Matched Count": 1,
                "Message": "One CDR record found.",
            },
            {
                "Evidence Source": "IPDR",
                "Status": "FOUND",
                "Evidence Unit": "IPDR records",
                "Matched Count": 1,
                "Message": "One IPDR record found.",
            },
            {
                "Evidence Source": "GPRS",
                "Status": "FOUND",
                "Evidence Unit": "GPRS sessions",
                "Matched Count": 1,
                "Message": "One GPRS session found.",
            },
        ]
    )

    cdr_timeline = pd.DataFrame(
        [
            {
                "Date-Time": pd.Timestamp(
                    "2026-01-01 10:00:00"
                ),
                "Target Number": "9000000001",
                "Call Type": "outgoing",
                "Other Party": "8000000001",
                "Contact Category": "human_mobile",
                "Duration (Sec)": 30,
                "IMSI": "405520123456789",
                "First Cell ID": (
                    "405-52-3347-232803094"
                ),
                "Last Cell ID": (
                    "405-52-3347-232803095"
                ),
                "Source File": (
                    '=HYPERLINK("bad","bad")'
                ),
                "Source Row Number": 2,
                "Raw IMEI": IMEISV_16,
                "Normalized IMEI": IMEISV_16,
                "_internal_debug": "hidden",
            }
        ]
    )

    ipdr_timeline = pd.DataFrame(
        [
            {
                "Event Time": pd.Timestamp(
                    "2026-01-01 11:00:00"
                ),
                "Allocation End": pd.Timestamp(
                    "2026-01-01 11:05:00"
                ),
                "Subscriber / User ID": (
                    "9000000001"
                ),
                "Identifier Type": "MSISDN",
                "IMSI": "405520123456789",
                "Source IP": "10.0.0.1",
                "Destination IP": "8.8.8.8",
                "Destination Port": "443",
                "Protocol": "TCP",
                "Cell ID": (
                    "405-52-3347-232803094"
                ),
                "First Cell ID": "",
                "Source File": "ipdr.csv",
                "Source Row Number": 3,
                "Raw IMEI": IMEISV_16,
                "Normalized IMEI": IMEISV_16,
                "Traceback": "hidden",
            }
        ]
    )

    gprs_timeline = pd.DataFrame(
        [
            {
                "Session Start": pd.Timestamp(
                    "2026-01-01 12:00:00"
                ),
                "Session End": pd.Timestamp(
                    "2026-01-01 12:10:00"
                ),
                "Duration (Sec)": 600,
                "Subscriber Number": (
                    "9000000001"
                ),
                "Identifier Type": "MSISDN",
                "IMSI": "405520123456789",
                "IPv4 Address": "10.0.0.2",
                "IPv6 Address": "",
                "Total Volume": 150.0,
                "Technology": "4G",
                "Cell ID": (
                    "405-52-3347-232803094"
                ),
                "Source File": "gprs.csv",
                "Source Row Number": 4,
                "Raw IMEI": IMEISV_16,
                "Normalized IMEI": IMEISV_16,
                "Error": "hidden",
            }
        ]
    )

    return {
        "requested_imei": IMEISV_16,
        "overall_status": status,
        "message": (
            "Exact IMEI/IMEISV found in "
            "CDR, IPDR, GPRS."
        ),
        "source_summary": source_summary,
        "associated_identities": pd.DataFrame(
            [
                {
                    "Evidence Source": "CDR",
                    "Identity Type": (
                        "TARGET_MSISDN"
                    ),
                    "Identity Value": (
                        "9000000001"
                    ),
                    "Related Identity": (
                        "405520123456789"
                    ),
                    "First Seen": pd.Timestamp(
                        "2026-01-01 10:00:00"
                    ),
                    "Last Seen": pd.Timestamp(
                        "2026-01-01 10:00:00"
                    ),
                    "Matched Count": 1,
                }
            ]
        ),
        "cross_source_timeline": pd.DataFrame(
            [
                {
                    "Evidence Source": "CDR",
                    "Evidence Type": "CDR Event",
                    "Start Time": pd.Timestamp(
                        "2026-01-01 10:00:00"
                    ),
                    "End Time": pd.NaT,
                    "Target / Subscriber": (
                        "9000000001"
                    ),
                    "IMSI": "405520123456789",
                    "Contact / Endpoint": (
                        "8000000001"
                    ),
                    "IP Address": "",
                    "Cell ID": (
                        "405-52-3347-232803094"
                    ),
                    "Source File": "cdr.csv",
                    "Source Row Number": 2,
                    "Source Detail": "outgoing",
                },
                {
                    "Evidence Source": "IPDR",
                    "Evidence Type": "IPDR Record",
                    "Start Time": pd.Timestamp(
                        "2026-01-01 11:00:00"
                    ),
                    "End Time": pd.Timestamp(
                        "2026-01-01 11:05:00"
                    ),
                    "Target / Subscriber": (
                        "9000000001"
                    ),
                    "IMSI": "405520123456789",
                    "Contact / Endpoint": (
                        "8.8.8.8:443"
                    ),
                    "IP Address": "10.0.0.1",
                    "Cell ID": (
                        "405-52-3347-232803094"
                    ),
                    "Source File": "ipdr.csv",
                    "Source Row Number": 3,
                    "Source Detail": "TCP",
                },
                {
                    "Evidence Source": "GPRS",
                    "Evidence Type": "GPRS Session",
                    "Start Time": pd.Timestamp(
                        "2026-01-01 12:00:00"
                    ),
                    "End Time": pd.Timestamp(
                        "2026-01-01 12:10:00"
                    ),
                    "Target / Subscriber": (
                        "9000000001"
                    ),
                    "IMSI": "405520123456789",
                    "Contact / Endpoint": "",
                    "IP Address": "10.0.0.2",
                    "Cell ID": (
                        "405-52-3347-232803094"
                    ),
                    "Source File": "gprs.csv",
                    "Source Row Number": 4,
                    "Source Detail": "4G",
                },
            ]
        ),
        "cdr": {
            "status": "FOUND",
            "timeline": cdr_timeline,
            "towers": pd.DataFrame(
                [
                    {
                        "Cell ID": (
                            "405-52-3347-232803094"
                        ),
                        "Linked Targets": (
                            "9000000001"
                        ),
                        "Target Count": 1,
                        "Total Events": 1,
                        "First Seen": pd.Timestamp(
                            "2026-01-01 10:00:00"
                        ),
                        "Last Seen": pd.Timestamp(
                            "2026-01-01 10:00:00"
                        ),
                    }
                ]
            ),
        },
        "ipdr": {
            "status": "FOUND",
            "timeline": ipdr_timeline,
            "cells": pd.DataFrame(
                [
                    {
                        "Cell ID": (
                            "405-52-3347-232803094"
                        ),
                        "Subscribers / User IDs": (
                            "9000000001"
                        ),
                        "Subscriber Count": 1,
                        "Source Files": "ipdr.csv",
                        "Total Records": 1,
                        "First Seen": pd.Timestamp(
                            "2026-01-01 11:00:00"
                        ),
                        "Last Seen": pd.Timestamp(
                            "2026-01-01 11:00:00"
                        ),
                    }
                ]
            ),
        },
        "gprs": {
            "status": "FOUND",
            "timeline": gprs_timeline,
            "cells": pd.DataFrame(
                [
                    {
                        "Cell ID": (
                            "405-52-3347-232803094"
                        ),
                        "Latitude": 24.1,
                        "Longitude": 86.1,
                        "Subscribers": (
                            "9000000001"
                        ),
                        "Subscriber Count": 1,
                        "Source Files": "gprs.csv",
                        "Spot Names": "Tower A",
                        "Total Sessions": 1,
                        "First Seen": pd.Timestamp(
                            "2026-01-01 12:00:00"
                        ),
                        "Last Seen": pd.Timestamp(
                            "2026-01-01 12:10:00"
                        ),
                        "Total Volume": 150.0,
                    }
                ]
            ),
        },
        "review_indicators": pd.DataFrame(
            [
                {
                    "Evidence Source": (
                        "CROSS-SOURCE"
                    ),
                    "Indicator": (
                        "Device identifier appears "
                        "in multiple evidence sources"
                    ),
                    "Observation": (
                        "Found in CDR, IPDR and GPRS."
                    ),
                    "Caution": (
                        "Evidence types must remain separate."
                    ),
                }
            ]
        ),
        "data_quality": pd.DataFrame(
            [
                {
                    "Evidence Source": "CDR",
                    "Check": "Matched records",
                    "Count": 1,
                    "Meaning": (
                        "Exact matching CDR records."
                    ),
                },
                {
                    "Evidence Source": "IPDR",
                    "Check": "Matched records",
                    "Count": 1,
                    "Meaning": (
                        "Exact matching IPDR records."
                    ),
                },
                {
                    "Evidence Source": "GPRS",
                    "Check": "Matched sessions",
                    "Count": 1,
                    "Meaning": (
                        "Exact matching GPRS sessions."
                    ),
                },
            ]
        ),
        "errors": {
            "private_module": (
                "Developer-only failure"
            )
        },
    }


def _all_text(
    workbook,
) -> str:
    values = []

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    values.append(
                        str(
                            cell.value
                        )
                    )

    return "\n".join(
        values
    )


def _header_index(
    worksheet,
    header: str,
) -> int:
    for cell in worksheet[
        4
    ]:
        if cell.value == header:
            return cell.column

    raise AssertionError(
        f"Header not found: {header}"
    )


def test_imei_device_report_has_exact_contract(
    tmp_path: Path,
):
    report = generate_imei_device_report(
        case={
            "case_id": "CASE-001",
            "case_name": "Device Test",
        },
        analysis=_analysis(),
        output_dir=tmp_path,
    )

    assert report is not None
    assert report.exists()

    workbook = load_workbook(
        report,
        data_only=False,
    )

    assert tuple(
        workbook.sheetnames
    ) == IMEI_DEVICE_SHEETS


def test_source_counts_remain_separate(
    tmp_path: Path,
):
    report = generate_imei_device_report(
        case={
            "case_id": "CASE-001",
        },
        analysis=_analysis(),
        output_dir=tmp_path,
    )

    workbook = load_workbook(
        report,
        data_only=False,
    )

    overview = workbook[
        "1. Device Overview"
    ]

    text = "\n".join(
        str(
            cell.value
        )
        for row in overview.iter_rows()
        for cell in row
        if cell.value is not None
    )

    assert "CDR records" in text
    assert "IPDR records" in text
    assert "GPRS sessions" in text
    assert "Counts remain separate" in text
    assert "Total Events" not in text


def test_identifiers_are_text_and_formula_is_safe(
    tmp_path: Path,
):
    report = generate_imei_device_report(
        case={
            "case_id": "CASE-001",
        },
        analysis=_analysis(),
        output_dir=tmp_path,
    )

    workbook = load_workbook(
        report,
        data_only=False,
    )

    for sheet_name in (
        "4. CDR Evidence",
        "5. IPDR Evidence",
        "6. GPRS Evidence",
    ):
        worksheet = workbook[
            sheet_name
        ]

        column = _header_index(
            worksheet,
            "Normalized IMEI",
        )

        cell = worksheet.cell(
            row=5,
            column=column,
        )

        assert cell.value == IMEISV_16
        assert cell.data_type == "s"
        assert cell.number_format == "@"

    cdr = workbook[
        "4. CDR Evidence"
    ]

    source_column = _header_index(
        cdr,
        "Source File",
    )

    source_cell = cdr.cell(
        row=5,
        column=source_column,
    )

    assert source_cell.data_type != "f"


def test_developer_diagnostics_are_not_exported(
    tmp_path: Path,
):
    analysis = _analysis(
        status="PARTIAL"
    )

    analysis[
        "source_summary"
    ].loc[
        analysis[
            "source_summary"
        ][
            "Evidence Source"
        ].eq(
            "IPDR"
        ),
        [
            "Status",
            "Matched Count",
            "Message",
        ],
    ] = [
        "ERROR",
        0,
        (
            "TypeError: secret technical "
            "stack detail"
        ),
    ]

    report = generate_imei_device_report(
        case={
            "case_id": "CASE-001",
        },
        analysis=analysis,
        output_dir=tmp_path,
    )

    workbook = load_workbook(
        report,
        data_only=False,
    )

    text = _all_text(
        workbook
    )

    assert "secret technical stack" not in text
    assert "Developer-only failure" not in text
    assert "_internal_debug" not in text
    assert "Traceback" not in text
    assert "private_module" not in text

    assert (
        "Source analysis was unavailable"
        in text
    )


def test_report_sheets_remain_compact(
    tmp_path: Path,
):
    report = generate_imei_device_report(
        case={
            "case_id": "CASE-001",
        },
        analysis=_analysis(),
        output_dir=tmp_path,
    )

    workbook = load_workbook(
        report,
        data_only=False,
    )

    for worksheet in workbook.worksheets:
        assert worksheet.max_column <= 15
        assert worksheet.freeze_panes == "A5"
        assert (
            worksheet.sheet_view.showGridLines
            is False
        )


def test_report_generation_does_not_modify_frames(
    tmp_path: Path,
):
    analysis = _analysis()

    originals = {
        "source_summary": analysis[
            "source_summary"
        ].copy(
            deep=True
        ),
        "associated_identities": analysis[
            "associated_identities"
        ].copy(
            deep=True
        ),
        "cross_source_timeline": analysis[
            "cross_source_timeline"
        ].copy(
            deep=True
        ),
        "cdr_timeline": analysis[
            "cdr"
        ][
            "timeline"
        ].copy(
            deep=True
        ),
        "ipdr_timeline": analysis[
            "ipdr"
        ][
            "timeline"
        ].copy(
            deep=True
        ),
        "gprs_timeline": analysis[
            "gprs"
        ][
            "timeline"
        ].copy(
            deep=True
        ),
    }

    generate_imei_device_report(
        case={
            "case_id": "CASE-001",
        },
        analysis=analysis,
        output_dir=tmp_path,
    )

    pd.testing.assert_frame_equal(
        analysis[
            "source_summary"
        ],
        originals[
            "source_summary"
        ],
    )

    pd.testing.assert_frame_equal(
        analysis[
            "associated_identities"
        ],
        originals[
            "associated_identities"
        ],
    )

    pd.testing.assert_frame_equal(
        analysis[
            "cross_source_timeline"
        ],
        originals[
            "cross_source_timeline"
        ],
    )

    pd.testing.assert_frame_equal(
        analysis[
            "cdr"
        ][
            "timeline"
        ],
        originals[
            "cdr_timeline"
        ],
    )

    pd.testing.assert_frame_equal(
        analysis[
            "ipdr"
        ][
            "timeline"
        ],
        originals[
            "ipdr_timeline"
        ],
    )

    pd.testing.assert_frame_equal(
        analysis[
            "gprs"
        ][
            "timeline"
        ],
        originals[
            "gprs_timeline"
        ],
    )


def test_invalid_or_not_found_does_not_create_workbook(
    tmp_path: Path,
):
    invalid = _analysis(
        status="INVALID_IMEI"
    )

    invalid[
        "requested_imei"
    ] = ""

    assert generate_imei_device_report(
        case={
            "case_id": "CASE-001",
        },
        analysis=invalid,
        output_dir=tmp_path,
    ) is None

    not_found = _analysis(
        status="NOT_FOUND"
    )

    not_found[
        "source_summary"
    ][
        "Matched Count"
    ] = 0

    not_found[
        "cdr"
    ][
        "timeline"
    ] = pd.DataFrame()

    not_found[
        "ipdr"
    ][
        "timeline"
    ] = pd.DataFrame()

    not_found[
        "gprs"
    ][
        "timeline"
    ] = pd.DataFrame()

    assert generate_imei_device_report(
        case={
            "case_id": "CASE-001",
        },
        analysis=not_found,
        output_dir=tmp_path,
    ) is None

    assert not list(
        tmp_path.glob(
            "*.xlsx"
        )
    )
