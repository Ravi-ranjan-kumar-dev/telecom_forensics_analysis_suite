from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from modules.reporting.tower_ipdr_excel import (
    COMPLETE_REPORT_SHEET_ORDER,
    generate_tower_ipdr_complete_excel_report,
)


def _sample_tables() -> dict:
    return {
        "executive_summary": pd.DataFrame(
            {
                "metric": [
                    "Total Events",
                    "Unique Subscribers",
                ],
                "value": [
                    10,
                    3,
                ],
            }
        ),
        "data_quality": pd.DataFrame(
            {
                "check": [
                    "Missing Spot ID",
                ],
                "rows": [
                    0,
                ],
                "percentage": [
                    0.0,
                ],
            }
        ),
        "spot_cell_summary": pd.DataFrame(
            {
                "section": [
                    "SPOT",
                ],
                "spot_id": [
                    "SPOT-01",
                ],
                "event_count": [
                    10,
                ],
            }
        ),
        "priority_review_queue": pd.DataFrame(
            {
                "subscriber_number": [
                    "9000000001",
                ],
                "priority": [
                    "High",
                ],
                "why_important": [
                    "multi-Spot presence",
                ],
            }
        ),
        "rare_presence": pd.DataFrame(
            {
                "subscriber_number": [
                    "9000000002",
                ],
                "event_count": [
                    1,
                ],
            }
        ),
        "multi_spot_intelligence": pd.DataFrame(
            {
                "record_type": [
                    "MULTI_SPOT_SUBSCRIBER",
                ],
                "subscriber_number": [
                    "9000000001",
                ],
                "spots": [
                    "SPOT-01, SPOT-02",
                ],
            }
        ),
        "subscriber_activity": pd.DataFrame(
            {
                "subscriber_number": [
                    "9000000001",
                ],
                "event_count": [
                    8,
                ],
            }
        ),
        "device_sim_alerts": pd.DataFrame(
            {
                "alert_type": [
                    "SHARED_IMEI",
                ],
                "identifier": [
                    "111111111111111",
                ],
            }
        ),
        "hourly_activity": pd.DataFrame(
            {
                "event_date": [
                    "2026-06-11",
                ],
                "hour": [
                    "20",
                ],
                "event_count": [
                    10,
                ],
            }
        ),
        "source_file_summary": pd.DataFrame(
            {
                "source_path": [
                    (
                        "/home/raviranjan/Desktop/"
                        "telecom_forensics_analysis_suite/"
                        "data/tower_dump/ipdr/input/"
                        "spot_1/a.csv"
                    ),
                ],
                "source_file": [
                    (
                        "/home/raviranjan/Desktop/"
                        "telecom_forensics_analysis_suite/"
                        "data/tower_dump/ipdr/input/"
                        "spot_1/a.csv"
                    ),
                ],
                "source_relative_path": [
                    "spot_1/a.csv",
                ],
                "spot_id": [
                    "SPOT-01",
                ],
                "file_name": [
                    "a.csv",
                ],
                "status": [
                    "LOADED",
                ],
            }
        ),
        "analysis_status": pd.DataFrame(
            {
                "stage": [
                    "DuckDB Analysis",
                ],
                "status": [
                    "COMPLETED",
                ],
            }
        ),
        "methodology_limits": pd.DataFrame(
            {
                "topic": [
                    "Location interpretation",
                ],
                "guidance": [
                    (
                        "Tower presence is an investigative "
                        "indicator, not proof of exact location."
                    ),
                ],
            }
        ),
    }


def test_complete_report_has_canonical_structure(
    tmp_path,
):
    report_path = (
        tmp_path
        / "tower_ipdr_complete_analysis.xlsx"
    )

    generated = (
        generate_tower_ipdr_complete_excel_report(
            case={
                "case_id": "TEST-CASE",
                "case_name": "Test Case",
            },
            report_path=report_path,
            tables=_sample_tables(),
            generated_at="2026-07-23 15:30:00",
        )
    )

    assert generated == report_path.resolve()
    assert report_path.exists()

    workbook = load_workbook(
        report_path,
        data_only=False,
    )

    expected_names = [
        sheet_name
        for (
            _key,
            sheet_name,
            _description,
        ) in COMPLETE_REPORT_SHEET_ORDER
    ]

    assert workbook.sheetnames == expected_names

    for worksheet in workbook.worksheets:
        assert worksheet.freeze_panes == "A5"
        assert worksheet.sheet_view.showGridLines is False
        assert (
            worksheet.row_dimensions[2].height
            >= 36
        )

    assert (
        workbook[
            "1. Executive Summary"
        ].row_dimensions[2].height
        == 56
    )

    source_sheet = workbook[
        "10. Source File Summary"
    ]

    headers = [
        source_sheet.cell(
            row=4,
            column=column,
        ).value
        for column in range(
            1,
            source_sheet.max_column + 1,
        )
    ]

    assert "source_relative_path" in headers
    assert "source_path" not in headers
    assert "source_file" not in headers

    all_text = []

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(
                    cell.value,
                    str,
                ):
                    all_text.append(
                        cell.value
                    )

    combined = "\n".join(
        all_text
    )

    assert "/home/raviranjan/" not in combined
    assert (
        "spot_1/a.csv"
        in combined
    )
