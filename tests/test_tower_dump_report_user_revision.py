from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def _section_headers(
    worksheet,
    title: str,
) -> list[object]:
    for row in worksheet.iter_rows():
        if row[0].value == title:
            return [
                cell.value
                for cell in worksheet[
                    row[0].row + 1
                ]
            ]

    raise AssertionError(
        f"Section not found: {title}"
    )


def _sample_result() -> dict[str, object]:
    first_cell = "405856008CA24"
    second_cell = "405856008CA25"

    normalized = pd.DataFrame(
        {
            "subscriber_number": [
                "9000000001",
                "9000000002",
            ],
            "other_party": [
                "8000000001",
                "8000000002",
            ],
            "call_type": [
                "incoming",
                "outgoing",
            ],
            "call_datetime": pd.to_datetime(
                [
                    "2026-08-20 10:00:00",
                    "2026-08-20 11:00:00",
                ]
            ),
            "call_duration": [
                60,
                30,
            ],
            "operator": [
                "airtel",
                "jio",
            ],
            "spot_id": [
                "SPOT-01",
                "SPOT-02",
            ],
            "spot_name": [
                "Spot One",
                "Spot Two",
            ],
            "searched_cell_id": [
                first_cell,
                second_cell,
            ],
            "present_at_searched_cell": [
                "Yes",
                "Yes",
            ],
            "imei": [
                "111111111111111",
                "111111111111111",
            ],
            "imsi": [
                "404000000000001",
                "404000000000002",
            ],
            "first_cell_id": [
                first_cell,
                second_cell,
            ],
            "last_cell_id": [
                first_cell,
                second_cell,
            ],
            "source_file": [
                "one.csv",
                "two.csv",
            ],
            "source_row": [
                10,
                20,
            ],
            "is_potential_duplicate": [
                False,
                False,
            ],
            "potential_duplicate_count": [
                1,
                1,
            ],
        }
    )

    lead = pd.DataFrame(
        {
            "subscriber_number": [
                "9000000001",
            ],
            "priority": [
                "High",
            ],
            "confidence": [
                "High",
            ],
            "priority_score": [
                200,
            ],
            "event_count": [
                8,
            ],
            "cells_seen": [
                2,
            ],
            "night_event_count": [
                1,
            ],
            "imei_count": [
                1,
            ],
            "imsi_count": [
                1,
            ],
            "first_seen": pd.to_datetime(
                ["2026-08-20 10:00:00"]
            ),
            "last_seen": pd.to_datetime(
                ["2026-08-20 11:00:00"]
            ),
            "why_important": [
                "High-priority test lead",
            ],
            "next_action": [
                "Verify",
            ],
        }
    )

    uncommon = pd.DataFrame(
        {
            "subscriber_number": [
                "9000000002",
            ],
            "priority": [
                "Low",
            ],
            "confidence": [
                "Low",
            ],
            "event_count": [
                1,
            ],
            "cells_seen": [
                1,
            ],
            "searched_cells_seen": [
                1,
            ],
            "searched_cells": [
                second_cell,
            ],
            "night_event_count": [
                0,
            ],
            "first_seen": pd.to_datetime(
                ["2026-08-20 11:00:00"]
            ),
            "last_seen": pd.to_datetime(
                ["2026-08-20 11:00:00"]
            ),
            "why_important": [
                "Rare presence",
            ],
            "next_action": [
                "Verify",
            ],
        }
    )

    visitors = pd.DataFrame(
        {
            "subscriber_number": [
                "9000000001",
            ],
            "total_events": [
                8,
            ],
            "first_seen": pd.to_datetime(
                ["2026-08-20 10:00:00"]
            ),
            "last_seen": pd.to_datetime(
                ["2026-08-20 11:00:00"]
            ),
            "active_days": [
                2,
            ],
            "unique_cells": [
                2,
            ],
            "unique_operators": [
                1,
            ],
            "unique_imei": [
                1,
            ],
            "unique_imsi": [
                1,
            ],
            "unique_other_parties": [
                3,
            ],
            "total_duration_seconds": [
                120,
            ],
        }
    )

    presence = pd.DataFrame(
        {
            "subscriber_number": [
                "9000000001",
            ],
            "spots_seen_count": [
                2,
            ],
            "total_spots": [
                2,
            ],
            "match_ratio": [
                "2/2",
            ],
            "spot_ids": [
                "SPOT-01, SPOT-02",
            ],
            "spot_names": [
                "Spot One, Spot Two",
            ],
            "total_events": [
                8,
            ],
            "unique_searched_cells": [
                2,
            ],
            "searched_cell_ids": [
                f"{first_cell}, {second_cell}",
            ],
            "primary_searched_cell_id": [
                first_cell,
            ],
            "operators": [
                "airtel",
            ],
            "imei_count": [
                1,
            ],
            "imsi_count": [
                1,
            ],
            "first_seen": pd.to_datetime(
                ["2026-08-20 10:00:00"]
            ),
            "last_seen": pd.to_datetime(
                ["2026-08-20 11:00:00"]
            ),
        }
    )

    continuity = pd.DataFrame(
        {
            "subscriber_number": [
                "9000000001",
            ],
            "spots_seen_count": [
                2,
            ],
            "spot_names": [
                "Spot One, Spot Two",
            ],
            "searched_cell_ids": [
                f"{first_cell}, {second_cell}",
            ],
            "primary_searched_cell_id": [
                first_cell,
            ],
            "imei_count": [
                1,
            ],
            "imei_values": [
                "111111111111111",
            ],
            "imei_continuity": [
                "SAME IMEI ACROSS SPOTS",
            ],
            "imsi_count": [
                1,
            ],
            "imsi_values": [
                "404000000000001",
            ],
            "imsi_continuity": [
                "SAME IMSI ACROSS SPOTS",
            ],
            "confidence": [
                "High",
            ],
            "why_important": [
                "Continuity test",
            ],
            "next_verification": [
                "Verify",
            ],
        }
    )

    shared_across = pd.DataFrame(
        {
            "imei": [
                "111111111111111",
            ],
            "imsi": [
                "404000000000001",
            ],
            "spots_seen_count": [
                2,
            ],
            "spot_names": [
                "Spot One, Spot Two",
            ],
            "searched_cell_ids": [
                f"{first_cell}, {second_cell}",
            ],
            "primary_searched_cell_id": [
                first_cell,
            ],
            "unique_subscribers": [
                2,
            ],
            "subscriber_numbers": [
                "9000000001, 9000000002",
            ],
            "total_events": [
                8,
            ],
            "first_seen": pd.to_datetime(
                ["2026-08-20 10:00:00"]
            ),
            "last_seen": pd.to_datetime(
                ["2026-08-20 11:00:00"]
            ),
            "why_important": [
                "Shared identifier",
            ],
            "next_verification": [
                "Verify",
            ],
        }
    )

    device_alert = lead.assign(
        other_party_count=3
    )

    shared_device = pd.DataFrame(
        {
            "imei": [
                "111111111111111",
            ],
            "imsi": [
                "404000000000001",
            ],
            "total_events": [
                8,
            ],
            "unique_subscribers": [
                2,
            ],
            "unique_cells": [
                2,
            ],
            "unique_operators": [
                1,
            ],
            "first_seen": pd.to_datetime(
                ["2026-08-20 10:00:00"]
            ),
            "last_seen": pd.to_datetime(
                ["2026-08-20 11:00:00"]
            ),
            "operators": [
                "airtel",
            ],
            "searched_cells": [
                f"{first_cell}, {second_cell}",
            ],
            "subscribers": [
                "9000000001, 9000000002",
            ],
        }
    )

    cell_summary = pd.DataFrame(
        {
            "searched_cell_id": [
                first_cell,
            ],
            "records": [
                8,
            ],
            "searched_cell_address": [
                "Existing tower address",
            ],
            "searched_cell_longitude": [
                85.1,
            ],
            "searched_cell_source_file": [
                "cgi.xlsx",
            ],
            "searched_cell_address_found": [
                "Yes",
            ],
            "searched_cell_lookup_status": [
                "FOUND",
            ],
            "searched_cell_match_confidence": [
                "DIRECT_NORMALIZED_CGI_KEY",
            ],
        }
    )

    results = {
        "tower_dump_summary": {
            "total_records": len(
                normalized
            ),
        },
        "operator_summary": pd.DataFrame(
            {
                "operator": [
                    "airtel",
                    "jio",
                ],
                "records": [
                    1,
                    1,
                ],
            }
        ),
        "cell_summary": cell_summary,
        "call_type_summary": pd.DataFrame(
            {
                "call_type": [
                    "incoming",
                    "outgoing",
                ],
                "records": [
                    1,
                    1,
                ],
            }
        ),
        "spot_summary": pd.DataFrame(
            {
                "spot_id": [
                    "SPOT-01",
                    "SPOT-02",
                ],
                "spot_name": [
                    "Spot One",
                    "Spot Two",
                ],
                "records": [
                    1,
                    1,
                ],
            }
        ),
        "tower_cdr_priority_leads": lead,
        "tower_cdr_uncommon_numbers": uncommon,
        "repeat_visitors": visitors,
        "n_of_m_spot_presence": presence,
        "all_spot_common_numbers": presence,
        "spot_exclusive_numbers": presence.assign(
            exclusive_spot_id="SPOT-01",
            exclusive_spot_name="Spot One",
        ),
        "cross_spot_device_continuity": (
            continuity
        ),
        "shared_imei_across_spots": (
            shared_across.drop(
                columns=[
                    "imsi",
                ]
            )
        ),
        "shared_imsi_across_spots": (
            shared_across.drop(
                columns=[
                    "imei",
                ]
            )
        ),
        "tower_cdr_device_consistency": (
            device_alert
        ),
        "shared_imei": shared_device.drop(
            columns=[
                "imsi",
            ]
        ),
        "shared_imsi": shared_device.drop(
            columns=[
                "imei",
            ]
        ),
    }

    return {
        "metadata": {
            "input_folder": "synthetic",
            "files_found": 2,
            "files_loaded": 2,
        },
        "df": normalized,
        "operators": [
            "airtel",
            "jio",
        ],
        "cell_ids": [
            first_cell,
            second_cell,
        ],
        "analysis": {
            "results": results,
            "status": pd.DataFrame(),
            "errors": pd.DataFrame(),
            "function_count": len(
                results
            ),
            "completed_count": len(
                results
            ),
            "failed_count": 0,
        },
    }


def test_tower_dump_user_revision_structure_and_batch_enrichment(
    tmp_path: Path,
    monkeypatch,
):
    from modules.enrichment import (
        telecom_master_enrichment,
    )
    from modules.reporting.tower_dump_excel import (
        generate_tower_dump_excel_report,
    )

    calls = {
        "sdr": 0,
        "cgi": 0,
    }

    def fake_sdr_lookup(
        numbers,
    ) -> pd.DataFrame:
        calls["sdr"] += 1

        return pd.DataFrame(
            [
                {
                    "lookup_mobile": number,
                    "subscriber_name": (
                        f"Name {number}"
                    ),
                    "father_name": (
                        f"Father {number}"
                    ),
                    "subscriber_address": (
                        f"Address {number}"
                    ),
                    "operator": "TEST",
                    "circle": "BIHAR",
                    "sdr_found": "Yes",
                }
                for number in numbers
            ]
        )

    def fake_cgi_lookup(
        cgi_values,
    ) -> pd.DataFrame:
        calls["cgi"] += 1

        return pd.DataFrame(
            [
                {
                    "cgi": cgi,
                    "operator": "TEST",
                    "circle": "BIHAR",
                    "state": "Bihar",
                    "district": "Jehanabad",
                    "police_station": "Makhdumpur",
                    "town": "Makhdumpur",
                    "site_name": "Test Site",
                    "address": (
                        f"Tower Address {cgi}"
                    ),
                    "latitude": 25.1,
                    "longitude": 85.1,
                    "source_file": "cgi.xlsx",
                }
                for cgi in cgi_values
            ]
        )

    monkeypatch.setattr(
        telecom_master_enrichment,
        "lookup_sdr_subscribers",
        fake_sdr_lookup,
    )
    monkeypatch.setattr(
        telecom_master_enrichment,
        "lookup_cgi_addresses",
        fake_cgi_lookup,
    )

    report = generate_tower_dump_excel_report(
        _sample_result(),
        output_dir=tmp_path,
        raw_row_limit=10,
        lead_row_limit=20,
    )

    workbook = load_workbook(
        report,
        data_only=False,
    )

    assert workbook.sheetnames == [
        "1. Executive Summary",
        "2. Tower Summary",
        "3. Priority Review Queue",
        "4. Rare Uncommon",
        "5. Repeat Visitors",
        "6. N-of-M Spot Presence",
        "7. All-Spot Common Numbers",
        "8. Spot-Exclusive Numbers",
        "9. Cross-Spot Device",
        "10. Shared IMEI Across Spots",
        "11. Shared IMSI Across Spots",
        "12. Device Consistency Alerts",
        "13. Shared IMEI",
        "14. Shared IMSI",
        "15. Normalized Sample",
    ]

    assert calls == {
        "sdr": 1,
        "cgi": 1,
    }

    executive_values = [
        cell.value
        for row in workbook[
            "1. Executive Summary"
        ].iter_rows()
        for cell in row
    ]

    assert (
        "Detailed Data Availability"
        not in executive_values
    )

    tower_headers = _section_headers(
        workbook[
            "2. Tower Summary"
        ],
        "SEARCHED CELL / CGI SUMMARY",
    )

    for removed in (
        "searched_cell_longitude",
        "searched_cell_source_file",
        "searched_cell_address_found",
        "searched_cell_lookup_status",
        "searched_cell_match_confidence",
    ):
        assert removed not in tower_headers

    priority_headers = _section_headers(
        workbook[
            "3. Priority Review Queue"
        ],
        "MASTER PRIORITY REVIEW QUEUE",
    )

    assert "next_action" not in (
        priority_headers
    )
    assert "sdr_subscriber_name" in (
        priority_headers
    )

    rare_headers = _section_headers(
        workbook[
            "4. Rare Uncommon"
        ],
        "RARE / UNCOMMON NUMBERS",
    )
    repeat_headers = _section_headers(
        workbook[
            "5. Repeat Visitors"
        ],
        "REPEAT VISITORS",
    )

    assert "sdr_subscriber_name" in (
        rare_headers
    )
    assert "sdr_subscriber_name" in (
        repeat_headers
    )

    for sheet_name in (
        "6. N-of-M Spot Presence",
        "7. All-Spot Common Numbers",
        "8. Spot-Exclusive Numbers",
        "9. Cross-Spot Device",
        "10. Shared IMEI Across Spots",
        "11. Shared IMSI Across Spots",
    ):
        worksheet = workbook[
            sheet_name
        ]
        headers = [
            cell.value
            for cell in worksheet[2]
        ]

        assert "subscriber_number" in headers
        assert "sdr_subscriber_name" in (
            headers
        )
        assert (
            "primary_searched_cell_id"
            in headers
        )
        assert "primary_cell_address" in (
            headers
        )

    for sheet_name in (
        "12. Device Consistency Alerts",
        "13. Shared IMEI",
        "14. Shared IMSI",
    ):
        headers = [
            cell.value
            for cell in workbook[
                sheet_name
            ][2]
        ]

        assert "subscriber_number" in headers
        assert "sdr_subscriber_name" in (
            headers
        )

    normalized_headers = _section_headers(
        workbook[
            "15. Normalized Sample"
        ],
        "NORMALIZED RECORD SAMPLE",
    )

    for removed in (
        "source_row",
        "potential_duplicate",
        "potential_duplicate_count",
    ):
        assert removed not in (
            normalized_headers
        )

    assert "sdr_subscriber_name" in (
        normalized_headers
    )
    assert "searched_cell_address" in (
        normalized_headers
    )
