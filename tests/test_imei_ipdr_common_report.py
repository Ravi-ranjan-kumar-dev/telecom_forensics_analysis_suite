
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from modules.analysis.device.imei_ipdr_common import (
    build_common_imei_ipdr_analysis,
)
from modules.reporting import (
    generate_imei_ipdr_common_report,
)
from modules.reporting.report_paths import (
    get_imei_ipdr_common_report_path,
)


FIRST_IMEI = "862261072892730"
SECOND_IMEI = "862286069717070"


EXPECTED_SHEETS = [
    "1. Device Overview",
    "2. Common Subscribers",
    "3. Common IMSIs",
    "4. Common Destination Endpoints",
    "5. Common Source IPs",
    "6. Common Cells",
    "7. Cross Device Timeline",
    "8. Acquisition Manifest",
    "9. Review Indicators",
    "10. Data Quality",
]


def _frame(
    *,
    query_identifier: str,
    observed_imei: str,
    subscriber: str = "5754021077243",
    source_file: str,
) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "event_time": "2025-10-05 08:14:24",
                "allocation_end": "2025-10-05 08:15:56",
                "subscriber_number": subscriber,
                "imsi": "405523214527244",
                "source_ip": "2401:4900:8339:2dbc::2",
                "destination_ip": "203.0.113.10",
                "destination_port": 443.0,
                "protocol": "TCP",
                "first_cell_id": "404-10-2330-158187265",
                "query_identifier_normalized": query_identifier,
                "observed_imei_normalized": observed_imei,
                "match_basis": "QUERY_SCOPE",
                "match_relation": "SAME_BASE14",
                "source_file": source_file,
                "source_row_number": 8,
            }
        ]
    )


def _manifest(
    *,
    second_status: str = "HAS_DATA",
) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Relative Path": "first.csv",
                "Source Path": "/home/example/private/first.csv",
                "SHA-256": "a" * 64,
                "Acquisition Content Role": "UNIQUE_CONTENT",
                "Analysis Content Role": "PRIMARY_CONTENT",
                "Format": "AIRTEL_DYNAMIC_IMEI_IPDR",
                "Operator": "AIRTEL",
                "Source Type": "IPDR",
                "Query Identifier": FIRST_IMEI,
                "Query Identifier Type": "IMEI15",
                "Inspection Status": "HAS_DATA",
                "Records Declared": 1,
                "Records Normalized": 1,
                "Rejected Lines": 0,
                "Message": "Loaded.",
            },
            {
                "Relative Path": "second.csv",
                "Source Path": "/home/example/private/second.csv",
                "SHA-256": "b" * 64,
                "Acquisition Content Role": "UNIQUE_CONTENT",
                "Analysis Content Role": "PRIMARY_CONTENT",
                "Format": "AIRTEL_DYNAMIC_IMEI_IPDR",
                "Operator": "AIRTEL",
                "Source Type": "IPDR",
                "Query Identifier": SECOND_IMEI,
                "Query Identifier Type": "IMEI15",
                "Inspection Status": second_status,
                "Records Declared": (
                    0
                    if second_status == "EMPTY_NO_DATA"
                    else 1
                ),
                "Records Normalized": (
                    0
                    if second_status == "EMPTY_NO_DATA"
                    else 1
                ),
                "Rejected Lines": 0,
                "Message": (
                    "No records."
                    if second_status == "EMPTY_NO_DATA"
                    else "Loaded."
                ),
            },
        ]
    )


def _analysis() -> dict:
    return build_common_imei_ipdr_analysis(
        {
            FIRST_IMEI: _frame(
                query_identifier=FIRST_IMEI,
                observed_imei="8622610728927300",
                source_file="first.csv",
            ),
            SECOND_IMEI: _frame(
                query_identifier=SECOND_IMEI,
                observed_imei="8622860697170700",
                source_file="second.csv",
            ),
        },
        _manifest(),
    )


def _analysis_with_empty_peer() -> dict:
    return build_common_imei_ipdr_analysis(
        {
            FIRST_IMEI: _frame(
                query_identifier=FIRST_IMEI,
                observed_imei="8622610728927300",
                source_file="first.csv",
            ),
            SECOND_IMEI: pd.DataFrame(),
        },
        _manifest(
            second_status="EMPTY_NO_DATA"
        ),
    )


def _all_cell_text(
    workbook,
) -> str:
    return " ".join(
        str(
            cell.value
        )
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
        if cell.value not in {
            None,
            "",
        }
    )


def test_ipdr_common_report_path_is_source_specific(
    tmp_path: Path,
):
    path = get_imei_ipdr_common_report_path(
        "CASE-001",
        output_dir=tmp_path,
    )

    assert path.parent == tmp_path.resolve()

    assert path.name.startswith(
        "CASE-001_IMEI_IPDR_Common_Analysis_"
    )

    assert path.name.endswith(
        ".xlsx"
    )

    assert "imei_ipdr_common" in path.stem


def test_ipdr_common_workbook_uses_compact_sheet_contract(
    tmp_path: Path,
):
    path = generate_imei_ipdr_common_report(
        case={
            "case_id": "CASE-001",
        },
        analysis=_analysis(),
        output_dir=tmp_path,
    )

    assert path is not None

    workbook = load_workbook(
        path,
        data_only=False,
    )

    assert workbook.sheetnames == EXPECTED_SHEETS

    for worksheet in workbook.worksheets:
        assert worksheet.max_column <= 15, (
            worksheet.title,
            worksheet.max_column,
        )

    assert workbook[
        "1. Device Overview"
    ].max_column == 15

    assert workbook[
        "7. Cross Device Timeline"
    ].max_column == 15

    assert workbook[
        "8. Acquisition Manifest"
    ].max_column == 15

    workbook.close()


def test_ipdr_common_workbook_preserves_exact_identifiers(
    tmp_path: Path,
):
    path = generate_imei_ipdr_common_report(
        case={
            "case_id": "CASE-001",
        },
        analysis=_analysis(),
        output_dir=tmp_path,
    )

    assert path is not None

    workbook = load_workbook(
        path,
        data_only=False,
    )

    exact_values = {
        FIRST_IMEI,
        SECOND_IMEI,
        "8622610728927300",
        "8622860697170700",
        "a" * 64,
        "b" * 64,
    }

    matches = {
        value: []
        for value in exact_values
    }

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value in matches:
                    matches[
                        cell.value
                    ].append(
                        cell
                    )

    for value, cells in matches.items():
        assert cells, value

        assert all(
            cell.data_type == "s"
            and cell.number_format == "@"
            for cell in cells
        ), value

    workbook.close()


def test_ipdr_common_workbook_is_shareable_and_safe(
    tmp_path: Path,
):
    path = generate_imei_ipdr_common_report(
        case={
            "case_id": "CASE-001",
        },
        analysis=_analysis(),
        output_dir=tmp_path,
    )

    assert path is not None

    workbook = load_workbook(
        path,
        data_only=False,
    )

    text = _all_cell_text(
        workbook
    )

    assert "/home/" not in text
    assert "telecom_forensics_analysis_suite" not in text

    excel_errors = [
        (
            worksheet.title,
            cell.coordinate,
            cell.value,
        )
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
        if cell.data_type == "e"
    ]

    assert not excel_errors

    workbook.close()


def test_ipdr_common_report_retains_valid_empty_peer(
    tmp_path: Path,
):
    analysis = _analysis_with_empty_peer()

    assert analysis[
        "status"
    ] == "FOUND"

    assert analysis[
        "data_bearing_device_count"
    ] == 1

    assert analysis[
        "empty_report_count"
    ] == 1

    path = generate_imei_ipdr_common_report(
        case={
            "case_id": "CASE-001",
        },
        analysis=analysis,
        output_dir=tmp_path,
    )

    assert path is not None

    workbook = load_workbook(
        path,
        data_only=False,
    )

    text = _all_cell_text(
        workbook
    )

    assert FIRST_IMEI in text
    assert SECOND_IMEI in text
    assert "FOUND" in text
    assert "EMPTY_NO_DATA" in text

    for sheet_name in (
        "2. Common Subscribers",
        "3. Common IMSIs",
        "4. Common Destination Endpoints",
        "5. Common Source IPs",
        "6. Common Cells",
    ):
        assert (
            "No records available for this section."
            in " ".join(
                str(
                    cell.value
                )
                for row in workbook[
                    sheet_name
                ].iter_rows()
                for cell in row
                if cell.value is not None
            )
        )

    workbook.close()


def test_ipdr_common_report_rejects_non_reportable_analysis(
    tmp_path: Path,
):
    assert generate_imei_ipdr_common_report(
        case={
            "case_id": "CASE-001",
        },
        analysis={
            "status": "NOT_APPLICABLE",
            "device_count": 1,
        },
        output_dir=tmp_path,
    ) is None

    assert generate_imei_ipdr_common_report(
        case={
            "case_id": "CASE-001",
        },
        analysis={
            "status": "FOUND",
            "device_count": 1,
        },
        output_dir=tmp_path,
    ) is None
