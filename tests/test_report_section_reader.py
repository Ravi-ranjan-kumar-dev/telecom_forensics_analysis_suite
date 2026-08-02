from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from modules.reporting.report_section_reader import (
    discover_report_sections,
    read_report_section_rows,
)


_SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
_HEADER_FILL = PatternFill("solid", fgColor="5B9BD5")
_NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
_TABLE_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _write_section_title(sheet: Worksheet, row: int, title: str) -> None:
    cell = sheet.cell(row=row, column=1, value=title)
    cell.fill = _SECTION_FILL
    cell.font = Font(bold=True)


def _write_header(
    sheet: Worksheet,
    row: int,
    headers: tuple[str, ...],
) -> None:
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=column, value=header)
        cell.fill = _HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF")
        cell.border = _TABLE_BORDER


def _write_record(
    sheet: Worksheet,
    row: int,
    values: tuple[object, ...],
) -> None:
    for column, value in enumerate(values, start=1):
        cell = sheet.cell(row=row, column=column, value=value)
        cell.border = _TABLE_BORDER


def _write_structured_report(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Communication Intel"

    report_title = sheet.cell(row=1, column=1, value="Communication Intelligence")
    report_title.fill = PatternFill("solid", fgColor="1F4E78")
    report_title.font = Font(bold=True, color="FFFFFF")
    sheet.cell(row=2, column=1, value="Target")
    sheet.cell(row=2, column=2, value="9000000001")

    _write_section_title(sheet, 7, "OUTGOING SMS")
    note = sheet.cell(row=8, column=1, value="Review outgoing messages.")
    note.fill = _NOTE_FILL
    _write_header(sheet, 9, ("Other Party", "Date", "Time"))
    _write_record(sheet, 10, ("9000000002", "01/08/2026", "10:00:00"))
    _write_record(sheet, 11, (None, "01/08/2026", "10:05:00"))

    _write_section_title(sheet, 14, "EMPTY REVIEW")
    sheet.cell(
        row=15,
        column=1,
        value="No records available for this section.",
    )

    _write_section_title(sheet, 17, "INCOMING SMS")
    _write_header(sheet, 18, ("Other Party", "Date", "Time"))
    _write_record(sheet, 19, ("9000000004", "01/08/2026", "11:00:00"))

    workbook.save(path)
    workbook.close()


def test_discovers_structured_sections_without_treating_report_title_as_section(
    tmp_path: Path,
):
    report = tmp_path / "structured-report.xlsx"
    _write_structured_report(report)

    workbook = load_workbook(report, read_only=True, data_only=True)
    try:
        sections = discover_report_sections(workbook["Communication Intel"])
    finally:
        workbook.close()

    assert [section.title for section in sections] == [
        "OUTGOING SMS",
        "EMPTY REVIEW",
        "INCOMING SMS",
    ]
    assert sections[0].guidance == "Review outgoing messages."
    assert sections[0].headers == ("Other Party", "Date", "Time")
    assert sections[0].record_count == 2
    assert sections[0].data_start_row == 10
    assert sections[0].data_end_row == 11

    assert sections[1].is_empty
    assert sections[1].header_row is None
    assert sections[1].headers == ()

    assert sections[2].guidance == ""
    assert sections[2].record_count == 1


def test_reads_only_selected_section_and_honours_row_limit(tmp_path: Path):
    report = tmp_path / "structured-report.xlsx"
    _write_structured_report(report)

    workbook = load_workbook(report, read_only=True, data_only=True)
    try:
        worksheet = workbook["Communication Intel"]
        outgoing, empty, incoming = discover_report_sections(worksheet)

        assert read_report_section_rows(worksheet, outgoing, limit=1) == (
            ("9000000002", "01/08/2026", "10:00:00"),
        )
        assert read_report_section_rows(worksheet, outgoing, limit=10) == (
            ("9000000002", "01/08/2026", "10:00:00"),
            (None, "01/08/2026", "10:05:00"),
        )
        assert read_report_section_rows(worksheet, empty) == ()
        assert read_report_section_rows(worksheet, incoming) == (
            ("9000000004", "01/08/2026", "11:00:00"),
        )
    finally:
        workbook.close()


def test_plain_legacy_sheet_has_no_false_structured_sections(tmp_path: Path):
    report = tmp_path / "plain-report.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Contact", "Call Type", "Duration"])
    sheet.append(["9000000001", "Outgoing", 30])
    workbook.save(report)
    workbook.close()

    loaded = load_workbook(report, read_only=True, data_only=True)
    try:
        assert discover_report_sections(loaded.active) == ()
    finally:
        loaded.close()


def test_rejects_non_positive_section_row_limit(tmp_path: Path):
    report = tmp_path / "structured-report.xlsx"
    _write_structured_report(report)

    workbook = load_workbook(report, read_only=True, data_only=True)
    try:
        worksheet = workbook["Communication Intel"]
        section = discover_report_sections(worksheet)[0]
        with pytest.raises(ValueError, match="must be positive"):
            read_report_section_rows(worksheet, section, limit=0)
    finally:
        workbook.close()
