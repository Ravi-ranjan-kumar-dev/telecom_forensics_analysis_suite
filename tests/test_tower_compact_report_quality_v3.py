from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def test_tower_compact_report_quality_v3(
    tmp_path: Path,
):
    from modules.reporting.tower_dump_excel import (
        generate_tower_dump_excel_report,
    )

    normalized = pd.DataFrame(
        {
            "subscriber_number": [
                "9000000001",
                "9000000002",
            ],
            "a_party": [
                "8000000001",
                "9000000002",
            ],
            "other_party": [
                "8000000001",
                "8000000002",
            ],
            "b_party": [
                "9000000001",
                "8000000002",
            ],
            "call_type": [
                "incoming",
                "outgoing",
            ],
            "call_datetime": pd.to_datetime(
                [
                    "2026-06-21 10:00:00",
                    "2026-06-21 11:00:00",
                ]
            ),
            "operator": [
                "airtel",
                "jio",
            ],
            "spot_id": [
                "SPOT-01",
                "SPOT-02",
            ],
            "spot_name": [
                "Spot One",
                "Spot Two",
            ],
        }
    )

    lead = pd.DataFrame(
        {
            "subscriber_number": [
                "9000000001"
            ],
            "priority": [
                "High"
            ],
            "confidence": [
                "High"
            ],
            "priority_score": [
                200
            ],
            "event_count": [
                10
            ],
            "cells_seen": [
                2
            ],
            "searched_cells": [
                "CELL-1, CELL-2"
            ],
            "first_cells": [
                "CELL-1, CELL-2"
            ],
            "why_important": [
                "Test reason"
            ],
            "next_action": [
                "Verify"
            ],
        }
    )

    uncommon = pd.DataFrame(
        {
            "subscriber_number": [
                "9000000002"
            ],
            "priority": [
                "Low"
            ],
            "confidence": [
                "Low"
            ],
            "event_count": [
                1
            ],
            "cells_seen": [
                1
            ],
            "night_event_count": [
                1
            ],
            "why_important": [
                "Rare presence"
            ],
            "next_action": [
                "Verify"
            ],
        }
    )

    result = {
        "metadata": {
            "files_found": 2,
            "files_loaded": 2,
        },
        "df": normalized,
        "file_summary": pd.DataFrame(
            {
                "file": [
                    "one.csv",
                    "two.csv",
                ],
                "spot_id": [
                    "SPOT-01",
                    "SPOT-02",
                ],
                "records": [
                    1,
                    1,
                ],
                "status": [
                    "CACHED_STAGE",
                    "CACHED_STAGE",
                ],
            }
        ),
        "analysis": {
            "results": {
                "tower_dump_summary": {
                    "total_records": 2
                },
                "spot_summary": pd.DataFrame(
                    {
                        "spot_id": [
                            "SPOT-01",
                            "SPOT-02",
                        ],
                        "spot_name": [
                            "Spot One",
                            "Spot Two",
                        ],
                    }
                ),
                "tower_cdr_priority_leads": lead,
                "tower_cdr_uncommon_numbers": uncommon,
                "tower_cdr_device_consistency": lead,
            },
            "status": pd.DataFrame(),
            "errors": pd.DataFrame(),
        },
    }

    report = generate_tower_dump_excel_report(
        result,
        output_dir=tmp_path,
        raw_row_limit=10,
        lead_row_limit=20,
    )

    workbook = load_workbook(
        report,
        data_only=False,
    )

    quality = workbook[
        "2. Data Quality"
    ]

    quality_rows = {
        row[0].value: row
        for row in quality.iter_rows()
        if row[0].value
    }

    assert (
        quality_rows[
            "Canonical Subscriber Role Mapping"
        ][2].value
        == "INFO"
    )

    priority = workbook[
        "4. Priority Review Queue"
    ]

    priority_values = [
        cell.value
        for row in priority.iter_rows()
        for cell in row
    ]

    assert (
        "RARE / UNCOMMON SHORTLIST"
        in priority_values
    )

    assert "9000000002" in priority_values
    assert "rare_uncommon" not in priority_values

    device = workbook[
        "7. Device SIM Alerts"
    ]

    device_headers = [
        cell.value
        for cell in device[2]
    ]

    assert "subscriber_number" in device_headers
    assert "searched_cells" not in device_headers
    assert "first_cells" not in device_headers

    executive_values = [
        cell.value
        for row in workbook[
            "1. Executive Summary"
        ].iter_rows()
        for cell in row
    ]

    assert not any(
        isinstance(value, str)
        and "Full analytical tables are retained"
        in value
        for value in executive_values
    )
