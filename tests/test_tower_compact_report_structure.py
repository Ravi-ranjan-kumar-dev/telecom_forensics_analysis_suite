from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def test_tower_compact_report_structure_and_canonical_sample(
    tmp_path: Path,
):
    from modules.reporting.tower_dump_excel import (
        generate_tower_dump_excel_report,
    )

    normalized = pd.DataFrame(
        {
            "subscriber_number": [
                "9000000001",
                "9000000002",
                "9000000003",
                "9000000004",
            ],
            "a_party": [
                "9000000001",
                "9000000002",
                "9000000003",
                "9000000004",
            ],
            "other_party": [
                "8000000001",
                "8000000002",
                "8000000003",
                "8000000004",
            ],
            "b_party": [
                "8000000001",
                "8000000002",
                "8000000003",
                "8000000004",
            ],
            "call_type": [
                "incoming",
                "outgoing",
                "smsin",
                "smsout",
            ],
            "call_datetime": pd.to_datetime(
                [
                    "2026-06-21 10:00:00",
                    "2026-06-21 11:00:00",
                    "2026-06-22 10:00:00",
                    "2026-06-22 11:00:00",
                ]
            ),
            "call_duration": [
                30,
                40,
                0,
                0,
            ],
            "operator": [
                "airtel",
                "jio",
                "airtel",
                "jio",
            ],
            "spot_id": [
                "SPOT-01",
                "SPOT-01",
                "SPOT-02",
                "SPOT-02",
            ],
            "spot_name": [
                "Spot One",
                "Spot One",
                "Spot Two",
                "Spot Two",
            ],
            "searched_cell_id": [
                "CELL-1",
                "CELL-2",
                "CELL-3",
                "CELL-4",
            ],
        }
    )

    duplicate_lead = pd.DataFrame(
        {
            "subscriber_number": [
                "9000000001"
            ],
            "priority": [
                "High"
            ],
            "confidence": [
                "High"
            ],
            "priority_score": [
                250
            ],
            "event_count": [
                10
            ],
            "cells_seen": [
                2
            ],
            "why_important": [
                "Test lead"
            ],
            "next_action": [
                "Verify"
            ],
        }
    )

    spot_summary = pd.DataFrame(
        {
            "spot_id": [
                "SPOT-01",
                "SPOT-02",
            ],
            "spot_name": [
                "Spot One",
                "Spot Two",
            ],
            "records": [
                2,
                2,
            ],
        }
    )

    result = {
        "metadata": {
            "input_folder": "synthetic",
            "files_found": 4,
            "files_loaded": 4,
        },
        "df": normalized,
        "operators": [
            "airtel",
            "jio",
        ],
        "cell_ids": [
            "CELL-1",
            "CELL-2",
            "CELL-3",
            "CELL-4",
        ],
        "file_summary": pd.DataFrame(
            {
                "file": [
                    "a.csv",
                    "b.csv",
                    "c.csv",
                    "d.csv",
                ],
                "spot_id": [
                    "SPOT-01",
                    "SPOT-01",
                    "SPOT-02",
                    "SPOT-02",
                ],
                "records": [
                    1,
                    1,
                    1,
                    1,
                ],
                "status": [
                    "CACHED_STAGE",
                    "CACHED_STAGE",
                    "CACHED_STAGE",
                    "CACHED_STAGE",
                ],
            }
        ),
        "analysis": {
            "results": {
                "tower_dump_summary": {
                    "total_records": 4,
                },
                "spot_summary": spot_summary,
                "tower_cdr_priority_leads": duplicate_lead,
                "tower_cdr_multi_cell_presence": duplicate_lead,
            },
            "status": pd.DataFrame(),
            "errors": pd.DataFrame(),
            "function_count": 0,
            "completed_count": 0,
            "failed_count": 0,
        },
    }

    report = generate_tower_dump_excel_report(
        result,
        output_dir=tmp_path,
        raw_row_limit=4,
        lead_row_limit=20,
    )

    workbook = load_workbook(
        report,
        data_only=False,
    )

    assert workbook.sheetnames == [
        "1. Executive Summary",
        "2. Data Quality",
        "3. Tower Summary",
        "4. Priority Review Queue",
        "5. Visitor Intelligence",
        "6. Multi-Spot Intel",
        "7. Device SIM Alerts",
        "8. Normalized Sample",
        "9. Backend Data Guide",
        "10. Analysis Status",
        "Methodology & Limits",
    ]

    sample_sheet = workbook[
        "8. Normalized Sample"
    ]

    sample_values = [
        cell.value
        for row in sample_sheet.iter_rows()
        for cell in row
    ]

    assert "subscriber_number" in sample_values
    assert "other_party" in sample_values
    assert "a_party" not in sample_values
    assert "b_party" not in sample_values
    assert "SPOT-01" in sample_values
    assert "SPOT-02" in sample_values
    assert "airtel" in sample_values
    assert "jio" in sample_values

    priority_sheet = workbook[
        "4. Priority Review Queue"
    ]

    priority_values = [
        cell.value
        for row in priority_sheet.iter_rows()
        for cell in row
    ]

    assert priority_values.count(
        "9000000001"
    ) == 1

    visitor_sheet = workbook[
        "5. Visitor Intelligence"
    ]

    visitor_values = [
        cell.value
        for row in visitor_sheet.iter_rows()
        for cell in row
    ]

    assert (
        "NOT APPLICABLE IN WHOLE-PERIOD REPORT"
        in visitor_values
    )
