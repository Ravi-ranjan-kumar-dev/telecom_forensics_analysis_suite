from __future__ import annotations

import hashlib
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from modules.analysis.cdr.location import bottom_cgi
from modules.controllers import app_controller
from modules.loader import duplicate_flags
from modules.reporting.analysis_bundle import build_single_analysis_bundle
from modules.reporting.cdr_compact_excel import (
    generate_single_cdr_compact_report,
)


def _cdr_frame(target: str = "9000000000") -> pd.DataFrame:
    return pd.DataFrame(
        {
            "a_party": [target] * 6,
            "b_party": [
                "8000000001",
                "8000000001",
                "8000000002",
                "8000000003",
                "8000000004",
                "8000000005",
            ],
            "call_type": [
                "outgoing",
                "incoming",
                "smsout",
                "smsin",
                "outgoing",
                "incoming",
            ],
            "call_direction": [
                "OUTGOING",
                "INCOMING",
                "OUTGOING",
                "INCOMING",
                "OUTGOING",
                "INCOMING",
            ],
            "call_date": ["01-08-2026"] * 6,
            "call_time": [
                "09:00:00",
                "10:00:00",
                "11:00:00",
                "12:00:00",
                "13:00:00",
                "14:00:00",
            ],
            "datetime": pd.date_range(
                "2026-08-01 09:00:00",
                periods=6,
                freq="h",
            ),
            "call_duration": [20, 25, 0, 0, 30, 35],
            "first_cell_id": [
                "404-55-113-12101",
                "404-55-113-12101",
                "404-55-113-12101",
                "404-55-113-12102",
                "404-55-113-12103",
                "404-55-113-12104",
            ],
            "last_cell_id": ["404-55-113-12101"] * 6,
            "imei": ["490154203237518"] * 6,
            "imsi": ["405001111111111"] * 6,
        }
    )


def _section_headers(worksheet, title: str) -> list[str]:
    for row_index in range(1, worksheet.max_row + 1):
        if worksheet.cell(row=row_index, column=1).value != title:
            continue

        header_row = row_index + 1

        while (
            header_row <= worksheet.max_row
            and worksheet.cell(row=header_row, column=1).value is not None
            and worksheet.cell(row=header_row, column=2).value is None
        ):
            header_row += 1

        return [
            str(
                worksheet.cell(
                    row=header_row,
                    column=column_index,
                ).value
            )
            for column_index in range(1, worksheet.max_column + 1)
            if worksheet.cell(
                row=header_row,
                column=column_index,
            ).value is not None
        ]

    raise AssertionError(
        f"Section not found: {title}"
    )


def test_bottom_cgi_returns_least_frequent_valid_towers():
    frame = _cdr_frame()
    frame.loc[
        len(frame)
    ] = frame.iloc[
        0
    ]
    frame.loc[
        len(frame) - 1,
        "first_cell_id",
    ] = "405-51--"

    result = bottom_cgi(
        frame,
        limit=3,
    )

    assert result[
        "Cell ID"
    ].tolist() == [
        "404-55-113-12102",
        "404-55-113-12103",
        "404-55-113-12104",
    ]
    assert result[
        "Total Events"
    ].astype(
        int
    ).tolist() == [
        1,
        1,
        1,
    ]
    assert "405-51--" not in set(
        result[
            "Cell ID"
        ]
    )


def test_single_bundle_batches_bottom_contact_sdr_and_bottom_cgi(
    monkeypatch,
):
    sdr_calls: list[list[str]] = []
    cgi_calls: list[list[str]] = []

    def fake_sdr(values):
        values = list(values)
        sdr_calls.append(values)
        return pd.DataFrame(
            [
                {
                    "lookup_mobile": value,
                    "subscriber_name": f"Name {value}",
                    "father_name": "Test Father",
                    "subscriber_address": f"SDR Address {value}",
                    "operator": "AIRTEL",
                    "circle": "Bihar",
                    "sdr_found": "Yes",
                }
                for value in values
            ]
        )

    def fake_cgi(values):
        values = list(values)
        cgi_calls.append(values)
        return pd.DataFrame(
            [
                {
                    "cgi": value,
                    "town": "Patna",
                    "district": "Patna",
                    "site_name": f"Site {value}",
                    "address": f"CGI Address {value}",
                    "latitude": 25.61,
                    "longitude": 85.14,
                }
                for value in values
            ]
        )

    monkeypatch.setattr(
        "modules.enrichment.telecom_master_enrichment.lookup_sdr_subscribers",
        fake_sdr,
    )
    monkeypatch.setattr(
        "modules.enrichment.telecom_master_enrichment.lookup_cgi_addresses",
        fake_cgi,
    )

    bundle = build_single_analysis_bundle(
        _cdr_frame(),
        target="9000000000",
    )

    assert len(sdr_calls) == 1
    assert len(cgi_calls) == 1

    bottom_contacts = bundle[
        "results"
    ][
        "bottom_contacts"
    ]
    bottom_towers = bundle[
        "results"
    ][
        "bottom_cgi"
    ]

    assert set(
        bottom_contacts[
            "contact_sdr_lookup_status"
        ]
    ) == {
        "FOUND"
    }
    assert bottom_contacts[
        "contact_sdr_address"
    ].str.startswith(
        "SDR Address"
    ).all()
    assert set(
        bottom_towers[
            "cell_id_cgi_lookup_status"
        ]
    ) == {
        "FOUND"
    }
    assert bottom_towers[
        "cell_id_cgi_address"
    ].str.startswith(
        "CGI Address"
    ).all()


def test_executive_summary_displays_bottom_sdr_and_cgi_address(
    tmp_path: Path,
    monkeypatch,
):
    frame = _cdr_frame()
    bundle = {
        "results": {
            "top_contacts": pd.DataFrame(),
            "bottom_contacts": pd.DataFrame(
                [
                    {
                        "Contact": "8000000002",
                        "Total Calls": 1,
                        "contact_sdr_subscriber_name": "Bottom Contact",
                        "contact_sdr_father_name": "Test Father",
                        "contact_sdr_address": "Bottom SDR Address",
                        "contact_sdr_operator": "AIRTEL",
                        "contact_sdr_circle": "Bihar",
                        "contact_sdr_lookup_status": "FOUND",
                    }
                ]
            ),
            "bottom_cgi": pd.DataFrame(
                [
                    {
                        "Cell ID": "404-55-113-12102",
                        "Total Events": 1,
                        "cell_id_cgi_town": "Patna",
                        "cell_id_cgi_district": "Patna",
                        "cell_id_cgi_site_name": "Bottom Site",
                        "cell_id_cgi_address": "Bottom CGI Address",
                        "cell_id_cgi_latitude": 25.61,
                        "cell_id_cgi_longitude": 85.14,
                        "cell_id_cgi_lookup_status": "FOUND",
                    }
                ]
            ),
        }
    }

    monkeypatch.setattr(
        "modules.reporting.single_cdr_excel._enrich_target_metadata_with_sdr",
        lambda metadata, target: metadata,
    )
    monkeypatch.setattr(
        "modules.analysis.cdr.contact_report.lookup_cgi_addresses",
        lambda values: pd.DataFrame(),
    )
    monkeypatch.setattr(
        "modules.enrichment.sdr_subscriber_enrichment.lookup_sdr_subscribers",
        lambda values: pd.DataFrame(),
    )

    report = generate_single_cdr_compact_report(
        frame,
        "9000000000",
        metadata={
            "target": "9000000000",
        },
        analysis_bundle=bundle,
        output_dir=tmp_path,
    )

    assert report is not None
    worksheet = load_workbook(
        report,
        data_only=False,
    )[
        "1. Executive Summary"
    ]
    text = "\n".join(
        str(cell.value)
        for row in worksheet.iter_rows()
        for cell in row
        if cell.value is not None
    )

    assert "Bottom Contact" in text
    assert "Bottom SDR Address" in text
    assert "Bottom CGI Address" in text
    assert "SDR Lookup Status" in _section_headers(
        worksheet,
        "BOTTOM 10 HUMAN CONTACTS",
    )
    assert "CGI Address" in _section_headers(
        worksheet,
        "BOTTOM 10 CGI / TOWERS",
    )
    assert "CGI Lookup Status" in _section_headers(
        worksheet,
        "BOTTOM 10 CGI / TOWERS",
    )


def test_duplicate_flags_hash_only_actual_duplicate_groups(
    monkeypatch,
):
    original_sha256 = hashlib.sha256
    calls = 0

    def counting_sha256(*args, **kwargs):
        nonlocal calls
        calls += 1
        return original_sha256(
            *args,
            **kwargs,
        )

    monkeypatch.setattr(
        duplicate_flags.hashlib,
        "sha256",
        counting_sha256,
    )

    frame = pd.DataFrame(
        {
            "datetime": pd.date_range(
                "2026-08-01",
                periods=1002,
                freq="s",
            ),
            "a_party": ["9000000000"] * 1002,
            "b_party": [
                str(
                    8000000000 + index
                )
                for index in range(1002)
            ],
            "call_type": ["outgoing"] * 1002,
        }
    )
    frame.loc[
        1000
    ] = frame.loc[
        10
    ]
    frame.loc[
        1001
    ] = frame.loc[
        20
    ]

    result = duplicate_flags.flag_potential_duplicates(
        frame
    )

    assert calls == 2
    assert int(
        result[
            "is_potential_duplicate"
        ].sum()
    ) == 4


def test_multiple_cdr_defaults_to_common_report_fast_mode(
    tmp_path: Path,
    monkeypatch,
):
    loaded = {
        "9000000001": {
            "df": _cdr_frame(
                "9000000001"
            ),
            "file": "one.csv",
        },
        "9000000002": {
            "df": _cdr_frame(
                "9000000002"
            ),
            "file": "two.csv",
        },
    }
    pipeline_calls: list[str] = []
    common_path = tmp_path / "common.xlsx"

    def fake_safe_import(module_name, attribute_name):
        if attribute_name == "run_multiple":
            return lambda folder=None: loaded
        if attribute_name == "build_cross_target_analysis":
            return lambda values, min_targets=2: {
                "summary": pd.DataFrame()
            }
        if attribute_name == "generate_multi_cdr_report":
            return lambda **kwargs: str(
                common_path
            )
        return None

    monkeypatch.setattr(
        app_controller,
        "safe_import",
        fake_safe_import,
    )
    monkeypatch.setattr(
        app_controller,
        "_run_target_pipeline",
        lambda **kwargs: pipeline_calls.append(
            kwargs[
                "target"
            ]
        ),
    )
    monkeypatch.setattr(
        app_controller,
        "log_case_event",
        lambda *args, **kwargs: None,
    )
    monkeypatch.setattr(
        app_controller,
        "register_target",
        lambda *args, **kwargs: None,
    )
    monkeypatch.setattr(
        app_controller,
        "register_report",
        lambda *args, **kwargs: None,
    )
    monkeypatch.setattr(
        app_controller,
        "register_analysis_run",
        lambda *args, **kwargs: None,
    )
    monkeypatch.setattr(
        app_controller,
        "case_report_dir",
        lambda *args, **kwargs: tmp_path,
    )
    monkeypatch.setattr(
        "modules.reporting.cdr_report_source.create_cdr_source_run",
        lambda **kwargs: {
            "datasets": [],
        },
    )
    monkeypatch.setattr(
        "modules.reporting.cdr_report_source.link_report_to_source",
        lambda *args, **kwargs: None,
    )

    result = app_controller.handle_multiple_cdr(
        {
            "case_id": "CASE-FAST",
        },
        input_folder=tmp_path,
    )

    assert result is not None
    assert pipeline_calls == []
    assert result[
        "individual_reports"
    ] == {}
    assert result[
        "individual_reports_generated"
    ] is False
    assert result[
        "multiple_common_report"
    ] == str(
        common_path
    )
    assert set(
        result[
            "stage_durations"
        ]
    ) == {
        "load_and_normalize",
        "individual_reports",
        "cross_target_analysis",
        "common_report",
        "source_link",
        "total",
    }

    pipeline_calls.clear()
    optional_result = app_controller.handle_multiple_cdr(
        {
            "case_id": "CASE-OPTIONAL",
        },
        input_folder=tmp_path,
        generate_individual_reports=True,
    )

    assert optional_result is not None
    assert pipeline_calls == [
        "9000000001",
        "9000000002",
    ]
    assert optional_result[
        "individual_reports_generated"
    ] is True
