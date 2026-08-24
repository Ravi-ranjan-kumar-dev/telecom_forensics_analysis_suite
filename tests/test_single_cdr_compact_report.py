
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from modules.reporting import cdr_compact_excel
from modules.reporting import single_cdr_excel


def _sample_cdr() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "a_party": [
                "9000000000",
                "9000000000",
                "9000000000",
                "9000000000",
            ],
            "b_party": [
                "8000000001",
                "8000000002",
                "8000000003",
                "8000000004",
            ],
            "call_type": [
                "outgoing",
                "incoming",
                "smsout",
                "smsin",
            ],
            "call_date": [
                "01-01-2026",
                "01-01-2026",
                "02-01-2026",
                "02-01-2026",
            ],
            "call_time": [
                "10:00:00",
                "11:00:00",
                "12:00:00",
                "13:00:00",
            ],
            "call_duration": [
                30,
                20,
                0,
                0,
            ],
            "first_cell_id": [
                "405-52-3347-232803094",
                "405-52-3347-232803094",
                "405-52-3347-232803095",
                "405-52-3347-232803095",
            ],
            "last_cell_id": [
                "405-52-3347-232803095",
                "405-52-3347-232803094",
                "405-52-3347-232803095",
                "405-52-3347-232803095",
            ],
            "imei": [
                "111111111111111",
                "111111111111111",
                "222222222222222",
                "222222222222222",
            ],
            "imsi": [
                "405001111111111",
                "405001111111111",
                "405001111111111",
                "405001111111111",
            ],
            "roaming_circle": [
                "Bihar",
                "Bihar",
                "Jharkhand",
                "Jharkhand",
            ],
        }
    )


def _analysis_bundle() -> dict:
    contact = pd.DataFrame(
        [
            {
                "Contact": "8000000001",
                "Total Calls": 2,
                "contact_sdr_subscriber_name": "Test Person",
                "contact_sdr_lookup_status": "FOUND",
                "contact_sdr_found": "Yes",
            }
        ]
    )

    return {
        "results": {
            "cdr_summary": {
                "Total Records": 4,
            },
            "top_contacts": contact,
            "bottom_contacts": contact,
            "contact_ranking": contact,
            "social_network": contact,
            "contact_category_summary": pd.DataFrame(),
            "top_service_sender_ids": pd.DataFrame(),
            "top_short_codes": pd.DataFrame(),
            "incoming_outgoing": pd.DataFrame(),
            "other_call_type_summary": pd.DataFrame(),
            "analyze_location": pd.DataFrame(),
            "frequent_locations": pd.DataFrame(),
            "tower_transition": pd.DataFrame(),
            "movement_pattern": pd.DataFrame(),
            "tower_intelligence": pd.DataFrame(),
            "home_tower": pd.DataFrame(),
            "work_tower": pd.DataFrame(),
            "imei_intelligence": pd.DataFrame(),
            "sim_change": pd.DataFrame(),
            "activity_summary": pd.DataFrame(),
            "hourly_activity": pd.DataFrame(),
            "daily_activity": pd.DataFrame(),
            "weekly_activity": pd.DataFrame(),
            "monthly_activity": pd.DataFrame(),
            "behavioral_intelligence": pd.DataFrame(),
            "suspicious_activity": pd.DataFrame(),
            "missing_cgi_lookup": pd.DataFrame(),
            "master_enrichment_summary": pd.DataFrame(),
            "top_contact_details": pd.DataFrame(),
        },
        "status": pd.DataFrame(
            [
                {
                    "Function": "internal_test",
                    "Status": "FAILED",
                    "Error": "Developer-only failure",
                }
            ]
        ),
        "errors": {
            "internal_test": "Developer-only failure",
        },
    }


def _all_cell_text(workbook) -> str:
    values = []

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    values.append(
                        str(
                            cell.value
                        )
                    )

    return "\n".join(
        values
    )


def test_single_cdr_compact_report_has_exact_contract(
    tmp_path: Path,
    monkeypatch,
):
    monkeypatch.setattr(
        single_cdr_excel,
        "_enrich_target_metadata_with_sdr",
        lambda metadata, target: metadata,
    )

    report = (
        cdr_compact_excel
        .generate_single_cdr_compact_report(
            _sample_cdr(),
            "9000000000",
            metadata={
                "target": "9000000000",
                "target_sdr_found": "No",
            },
            analysis_bundle=_analysis_bundle(),
            output_dir=tmp_path,
        )
    )

    assert report is not None

    workbook = load_workbook(
        report,
        data_only=False,
    )

    assert tuple(
        workbook.sheetnames
    ) == (
        cdr_compact_excel
        .SINGLE_CDR_COMPACT_SHEETS
    )


def test_substantive_cdr_sections_are_preserved(
    tmp_path: Path,
    monkeypatch,
):
    monkeypatch.setattr(
        single_cdr_excel,
        "_enrich_target_metadata_with_sdr",
        lambda metadata, target: metadata,
    )

    report = (
        cdr_compact_excel
        .generate_single_cdr_compact_report(
            _sample_cdr(),
            "9000000000",
            metadata={
                "target": "9000000000",
            },
            analysis_bundle=_analysis_bundle(),
            output_dir=tmp_path,
        )
    )

    workbook = load_workbook(
        report,
        data_only=False,
    )

    text = _all_cell_text(
        workbook
    )

    required_sections = (
        "CC SUMMARY",
        "FIRST AND LAST COMMUNICATION BY DAY (FCLC)",
        "FIRST/LAST LOCATION SUMMARY (FCLC SUMMARY)",
        "FIRST/LAST CONTACT SUMMARY (FCLC OP)",
        "MOVING CALLS",
        "ROAMING SUMMARY",
        "OUTGOING SMS",
        "INCOMING SMS",
    )

    for section in required_sections:
        assert section in text


def test_developer_diagnostics_are_not_exported(
    tmp_path: Path,
    monkeypatch,
):
    monkeypatch.setattr(
        single_cdr_excel,
        "_enrich_target_metadata_with_sdr",
        lambda metadata, target: metadata,
    )

    report = (
        cdr_compact_excel
        .generate_single_cdr_compact_report(
            _sample_cdr(),
            "9000000000",
            metadata={
                "target": "9000000000",
            },
            analysis_bundle=_analysis_bundle(),
            output_dir=tmp_path,
        )
    )

    workbook = load_workbook(
        report,
        data_only=False,
    )

    text = _all_cell_text(
        workbook
    )

    assert "Analysis Status" not in workbook.sheetnames
    assert "Analysis Errors" not in workbook.sheetnames
    assert "Developer-only failure" not in text
    assert "internal_test" not in text

def test_compact_clean_frame_hides_technical_columns():
    frame = pd.DataFrame(
        [
            {
                "Cell ID": "405-52-3347-232803094",
                "Level Code": "Missing",
                "cell_id_cgi_lookup_key": (
                    "405-52-3347-232803094"
                ),
                "cell_id_cgi_record_found": "Yes",
                "cell_id_cgi_operator": "Airtel",
                "cell_id_cgi_address": "Test Address",
                "cell_id_cgi_lookup_status": "FOUND",
                "cell_id_cgi_match_confidence": (
                    "DIRECT_NORMALIZED_CGI_KEY"
                ),
                "cell_id_cgi_source_file": (
                    "/home/test/master.xlsx"
                ),
                "Score_Ruleset": "TEST-RULESET",
                "Score_Formula": "Internal formula",
            }
        ]
    )

    result = (
        cdr_compact_excel
        ._clean_frame(
            frame
        )
    )

    assert "Level Code" not in result.columns
    assert "cell_id_cgi_lookup_key" not in result.columns
    assert "cell_id_cgi_record_found" not in result.columns
    assert "cell_id_cgi_source_file" not in result.columns
    assert "Score_Ruleset" not in result.columns
    assert "Score_Formula" not in result.columns

    assert "CGI Operator" in result.columns
    assert "CGI Address" in result.columns
    assert "CGI Lookup Status" in result.columns
    assert "CGI Match Confidence" in result.columns


def test_compact_report_omits_technical_quality_and_rejected_sections(
    tmp_path: Path,
    monkeypatch,
):
    monkeypatch.setattr(
        single_cdr_excel,
        "_enrich_target_metadata_with_sdr",
        lambda metadata, target: metadata,
    )

    frame = _sample_cdr()

    frame.attrs[
        "rejected_rows"
    ] = pd.DataFrame(
        [
            {
                "parse_note": (
                    "INVALID_OR_NON_DATA_CDR_TIMESTAMP"
                )
            }
        ]
    )

    report = (
        cdr_compact_excel
        .generate_single_cdr_compact_report(
            frame,
            "9000000000",
            metadata={
                "target": "9000000000",
            },
            analysis_bundle=_analysis_bundle(),
            output_dir=tmp_path,
        )
    )

    workbook = load_workbook(
        report,
        data_only=False,
    )

    assert "9. Priority Review Queue" not in workbook.sheetnames
    assert "10. Data Quality & Guide" not in workbook.sheetnames
    assert "INVALID_OR_NON_DATA_CDR_TIMESTAMP" not in _all_cell_text(
        workbook
    )


def test_executive_top_contacts_exclude_service_senders(
    tmp_path: Path,
    monkeypatch,
):
    monkeypatch.setattr(
        single_cdr_excel,
        "_enrich_target_metadata_with_sdr",
        lambda metadata, target: metadata,
    )

    frame = _sample_cdr()

    service_row = frame.iloc[
        [
            0
        ]
    ].copy()

    service_row[
        "b_party"
    ] = "AD-AIRBNK-S"

    service_row[
        "call_type"
    ] = "smsin"

    frame = pd.concat(
        [
            frame,
            *[
                service_row
                for _ in range(
                    10
                )
            ],
        ],
        ignore_index=True,
    )

    report = (
        cdr_compact_excel
        .generate_single_cdr_compact_report(
            frame,
            "9000000000",
            metadata={
                "target": "9000000000",
            },
            analysis_bundle=_analysis_bundle(),
            output_dir=tmp_path,
        )
    )

    workbook = load_workbook(
        report,
        data_only=False,
    )

    worksheet = workbook[
        "1. Executive Summary"
    ]

    values = [
        str(
            cell.value
        )
        for row in worksheet.iter_rows()
        for cell in row
        if cell.value is not None
    ]

    text = "\n".join(
        values
    )

    assert "TOP 10 HUMAN CONTACTS" in text
    assert "BOTTOM 10 HUMAN CONTACTS" in text
    assert "8000000001" in text
    assert "AD-AIRBNK-S" not in text

def test_common_lookup_indexes_enrich_legacy_tables():
    results = {
        "contact_ranking": pd.DataFrame(
            [
                {
                    "Contact": "8002310903",
                    "contact_sdr_lookup_mobile": (
                        "8002310903"
                    ),
                    "contact_sdr_subscriber_name": (
                        "BHANU DEVI"
                    ),
                    "contact_sdr_father_name": (
                        "KAILASH SAO"
                    ),
                    "contact_sdr_address": (
                        "TEST SDR ADDRESS"
                    ),
                    "contact_sdr_operator": (
                        "AIRTEL"
                    ),
                    "contact_sdr_circle": (
                        "BIHAR"
                    ),
                    "contact_sdr_found": "Yes",
                    "contact_sdr_lookup_status": (
                        "FOUND"
                    ),
                    "contact_sdr_match_confidence": (
                        "DIRECT_NORMALIZED_MSISDN"
                    ),
                }
            ]
        ),
        "tower_intelligence": pd.DataFrame(
            [
                {
                    "Cell ID": (
                        "405-52-3347-232803094"
                    ),
                    "cell_id_cgi_lookup_key": (
                        "405-52-3347-232803094"
                    ),
                    "cell_id_cgi_operator": (
                        "Airtel"
                    ),
                    "cell_id_cgi_state": (
                        "Jharkhand"
                    ),
                    "cell_id_cgi_district": (
                        "Giridih"
                    ),
                    "cell_id_cgi_town": (
                        "Sariya"
                    ),
                    "cell_id_cgi_address": (
                        "TEST CGI ADDRESS"
                    ),
                    "cell_id_cgi_lookup_status": (
                        "FOUND"
                    ),
                    "cell_id_cgi_match_confidence": (
                        "DIRECT_NORMALIZED_CGI_KEY"
                    ),
                }
            ]
        ),
    }

    sdr_index = (
        cdr_compact_excel
        ._build_sdr_contact_index(
            results
        )
    )

    cgi_index = (
        cdr_compact_excel
        ._build_cgi_location_index(
            results
        )
    )

    legacy = pd.DataFrame(
        [
            {
                "Other Party": (
                    "8002310903"
                ),
                "Name": "",
                "SDR Address": "",
                "Cell ID": (
                    "405-52-3347-232803094"
                ),
                "Address": "",
                "End Cell ID": (
                    "405-52-3347-232803094"
                ),
                "End Address": "",
            }
        ]
    )

    enriched = (
        cdr_compact_excel
        ._enrich_event_table(
            legacy,
            sdr_index=sdr_index,
            cgi_index=cgi_index,
        )
    )

    assert enriched.loc[
        0,
        "Name",
    ] == "BHANU DEVI"

    assert enriched.loc[
        0,
        "SDR Address",
    ] == "TEST SDR ADDRESS"

    assert enriched.loc[
        0,
        "SDR Lookup Status",
    ] == "FOUND"

    assert enriched.loc[
        0,
        "Address",
    ] == "TEST CGI ADDRESS"

    assert enriched.loc[
        0,
        "End Address",
    ] == "TEST CGI ADDRESS"

    assert enriched.loc[
        0,
        "Start Tower Town",
    ] == "Sariya"

    assert enriched.loc[
        0,
        "Start Tower District",
    ] == "Giridih"

    assert enriched.loc[
        0,
        "Start Tower Lookup Status",
    ] == "FOUND"


def test_lookup_indexes_prefer_found_complete_records():
    results = {
        "top_contacts": pd.DataFrame(
            [
                {
                    "Contact": "8002310903",
                    "contact_sdr_lookup_mobile": (
                        "8002310903"
                    ),
                    "contact_sdr_found": "No",
                    "contact_sdr_lookup_status": (
                        "NOT_FOUND"
                    ),
                }
            ]
        ),
        "contact_ranking": pd.DataFrame(
            [
                {
                    "Contact": "8002310903",
                    "contact_sdr_lookup_mobile": (
                        "8002310903"
                    ),
                    "contact_sdr_subscriber_name": (
                        "FOUND PERSON"
                    ),
                    "contact_sdr_found": "Yes",
                    "contact_sdr_lookup_status": (
                        "FOUND"
                    ),
                }
            ]
        ),
    }

    index = (
        cdr_compact_excel
        ._build_sdr_contact_index(
            results
        )
    )

    assert len(
        index
    ) == 1

    assert index.loc[
        0,
        "Name",
    ] == "FOUND PERSON"

    assert index.loc[
        0,
        "SDR Lookup Status",
    ] == "FOUND"


def test_lookup_index_helpers_do_not_call_database(
    monkeypatch,
):
    def fail_lookup(
        *args,
        **kwargs,
    ):
        raise AssertionError(
            "No database lookup should occur."
        )

    monkeypatch.setattr(
        single_cdr_excel,
        "_enrich_target_metadata_with_sdr",
        fail_lookup,
    )

    results = {
        "contact_ranking": pd.DataFrame(
            [
                {
                    "Contact": "8002310903",
                    "contact_sdr_lookup_mobile": (
                        "8002310903"
                    ),
                    "contact_sdr_lookup_status": (
                        "FOUND"
                    ),
                }
            ]
        ),
        "tower_intelligence": pd.DataFrame(
            [
                {
                    "Cell ID": (
                        "405-52-3347-232803094"
                    ),
                    "cell_id_cgi_lookup_key": (
                        "405-52-3347-232803094"
                    ),
                    "cell_id_cgi_lookup_status": (
                        "FOUND"
                    ),
                }
            ]
        ),
    }

    sdr_index = (
        cdr_compact_excel
        ._build_sdr_contact_index(
            results
        )
    )

    cgi_index = (
        cdr_compact_excel
        ._build_cgi_location_index(
            results
        )
    )

    assert len(
        sdr_index
    ) == 1

    assert len(
        cgi_index
    ) == 1

def test_wide_investigator_sections_have_compact_profiles():
    required_profiles = {
        "CC SUMMARY",
        "OUTGOING VOICE CALLS",
        "INCOMING VOICE CALLS",
        "OUTGOING SMS",
        "INCOMING SMS",
        "CELL ID SUMMARY",
        "TOWER INTELLIGENCE",
        "FIRST AND LAST COMMUNICATION BY DAY (FCLC)",
        "FIRST/LAST LOCATION SUMMARY (FCLC SUMMARY)",
        "FIRST/LAST CONTACT SUMMARY (FCLC OP)",
        "MOVING CALLS",
        "MOVEMENT EVENTS",
        "TOWER TRANSITIONS",
        "MOVEMENT PATTERNS",
        "DEVICE SUMMARY",
        "SIM SUMMARY",
        "DEVICE / SIM CHANGE INDICATORS",
    }

    assert required_profiles.issubset(
        cdr_compact_excel
        .SECTION_COLUMN_PROFILES
    )

    for title in required_profiles:
        columns = (
            cdr_compact_excel
            .SECTION_COLUMN_PROFILES[
                title
            ]
        )

        assert len(
            columns
        ) <= 15


def test_incoming_voice_profile_omits_tower_columns():
    columns = cdr_compact_excel.SECTION_COLUMN_PROFILES[
        "INCOMING VOICE CALLS"
    ]

    assert "Cell ID" not in columns
    assert "Address" not in columns
    assert "End Cell ID" not in columns
    assert "End Address" not in columns


def test_location_overview_profile_omits_lookup_status():
    columns = cdr_compact_excel.SECTION_COLUMN_PROFILES[
        "CELL ID SUMMARY"
    ]

    assert "Start Tower Lookup Status" not in columns


def test_identifier_cells_are_forced_to_exact_text():
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Device Key", "IMSI", "Most Used Device Key"])
    worksheet.append(
        [86811606664317, 405001111111111, 86811606664317]
    )

    cdr_compact_excel._finish_sheet(worksheet)

    for cell, expected in zip(
        worksheet[2],
        (
            "86811606664317",
            "405001111111111",
            "86811606664317",
        ),
    ):
        assert cell.value == expected
        assert cell.data_type == "s"
        assert cell.number_format == "@"
        assert cell.quotePrefix is True
        assert cell.alignment.horizontal == "left"


def test_profiled_moving_calls_section_is_not_overwide():
    from openpyxl import Workbook

    source = pd.DataFrame(
        [
            {
                "Mobile Number": "9000000000",
                "Other Party": "8002310903",
                "Name": "BHANU DEVI",
                "Father Name": "KAILASH SAO",
                "SDR Address": "TEST SDR ADDRESS",
                "Call Type": "Outgoing Call",
                "IMEI": "111111111111111",
                "Date": "01-01-2026",
                "Time": "10:00:00",
                "Duration": 30,
                "Cell ID": "405-52-3347-232803094",
                "Address": "START TOWER ADDRESS",
                "Last Cell ID": "405-52-3347-232803095",
                "Last Address": "END TOWER ADDRESS",
                "SDR Lookup Status": "FOUND",
                "Start Tower Operator": "Airtel",
                "Start Tower Circle": "119",
                "Start Tower State": "Jharkhand",
                "Start Tower District": "Giridih",
                "Start Tower Town": "Sariya",
                "Start Tower Latitude": 24.19,
                "Start Tower Longitude": 85.89,
                "Start Tower Lookup Status": "FOUND",
                "Last Tower Operator": "Airtel",
                "Last Tower District": "Giridih",
                "Last Tower Latitude": 24.20,
                "Last Tower Longitude": 85.90,
                "Last Tower Lookup Status": "FOUND",
            }
        ]
    )

    workbook = Workbook()
    worksheet = workbook.active

    cdr_compact_excel._write_section(
        worksheet,
        1,
        title="MOVING CALLS",
        frame=source,
    )

    headers = [
        worksheet.cell(
            row=2,
            column=column,
        ).value
        for column in range(
            1,
            worksheet.max_column + 1,
        )
        if worksheet.cell(
            row=2,
            column=column,
        ).value is not None
    ]

    assert headers == [
        "Mobile Number",
        "Other Party",
        "Name",
        "Call Type",
        "IMEI",
        "Date",
        "Time",
        "Duration",
        "Cell ID",
        "Address",
        "Last Cell ID",
        "Last Address",
        "SDR Lookup Status",
    ]

    assert "Start Tower Latitude" not in headers
    assert "Last Tower Longitude" not in headers


def test_section_projection_does_not_modify_source_evidence():
    from openpyxl import Workbook

    source = pd.DataFrame(
        [
            {
                "Other Party": "8002310903",
                "Name": "BHANU DEVI",
                "Total Calls": 10,
                "Internal Extra Column": "PRESERVED",
            }
        ]
    )

    original = source.copy(
        deep=True
    )

    workbook = Workbook()
    worksheet = workbook.active

    cdr_compact_excel._write_section(
        worksheet,
        1,
        title="CC SUMMARY",
        frame=source,
    )

    pd.testing.assert_frame_equal(
        source,
        original,
    )

    assert (
        source.loc[
            0,
            "Internal Extra Column",
        ]
        == "PRESERVED"
    )
