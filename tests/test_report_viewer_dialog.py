from __future__ import annotations

import os
from pathlib import Path
import pytest
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from gui.app import build_application
from gui.widgets.report_viewer_dialog import (
    ReportViewerDialog,
    canonical_phone_number,
    detect_identifier_columns,
    detect_number_columns,
    prepare_related_records,
)


_SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
_HEADER_FILL = PatternFill("solid", fgColor="5B9BD5")
_TABLE_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _write_report(path: Path) -> None:
    workbook = Workbook()

    summary = workbook.active
    summary.title = "Call Summary"
    summary.append(["Contact", "Call Type", "Duration"])
    summary.append(["919000000001", "Outgoing", 30])
    summary.append(["9000000001", "Incoming", 20])

    details = workbook.create_sheet("Common Contacts")
    details.append(["Common Number", "Source Target", "Count"])
    details.append(["9000000002", "9000000001", 3])

    workbook.save(path)


def _write_section_title(sheet: Worksheet, row: int, title: str) -> None:
    cell = sheet.cell(row=row, column=1, value=title)
    cell.fill = _SECTION_FILL
    cell.font = Font(bold=True)


def _write_section_header(
    sheet: Worksheet,
    row: int,
    headers: tuple[str, ...],
) -> None:
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=column, value=header)
        cell.fill = _HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF")
        cell.border = _TABLE_BORDER


def _write_section_record(
    sheet: Worksheet,
    row: int,
    values: tuple[object, ...],
) -> None:
    for column, value in enumerate(values, start=1):
        cell = sheet.cell(row=row, column=column, value=value)
        cell.border = _TABLE_BORDER


def _write_structured_report(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Movement & Daily Routine"

    report_title = sheet.cell(row=1, column=1, value="Movement Analysis")
    report_title.fill = PatternFill("solid", fgColor="1F4E78")
    report_title.font = Font(bold=True, color="FFFFFF")

    _write_section_title(
        sheet,
        7,
        "FIRST AND LAST COMMUNICATION BY DAY (FCLC)",
    )
    sheet.cell(row=8, column=1, value="Review the first and last daily events.")
    _write_section_header(
        sheet,
        9,
        ("Other Party", "Event Type", "Date"),
    )
    _write_section_record(
        sheet,
        10,
        ("9000000002", "Outgoing", "01/08/2026"),
    )
    _write_section_record(
        sheet,
        11,
        ("9000000003", "Incoming", "01/08/2026"),
    )

    _write_section_title(sheet, 14, "MOVING CALLS")
    _write_section_header(sheet, 15, ("Cell ID", "Event Type", "Date"))
    _write_section_record(
        sheet,
        16,
        ("405-51-834-15492631", "Outgoing", "01/08/2026"),
    )

    _write_section_title(sheet, 19, "OBSERVED IDENTIFIER VARIATIONS")
    sheet.cell(row=20, column=1, value="Review only confirmed variations.")
    sheet.cell(
        row=21,
        column=1,
        value="No records available for this section.",
    )

    workbook.save(path)
    workbook.close()


def _write_large_structured_report(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Large Section"

    _write_section_title(sheet, 3, "LARGE CONTACT SECTION")
    _write_section_header(sheet, 4, ("Record", "Event Type"))
    for index in range(501):
        _write_section_record(
            sheet,
            5 + index,
            (f"record-{index:04d}", "Outgoing"),
        )

    _write_section_title(sheet, 508, "SMALL CONTACT SECTION")
    _write_section_header(sheet, 509, ("Record", "Event Type"))
    _write_section_record(sheet, 510, ("small-record", "Incoming"))

    workbook.save(path)
    workbook.close()


def _write_large_legacy_report(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Legacy Records"
    sheet.append(["Contact", "Event Type"])
    for index in range(501):
        sheet.append((f"900000{index:04d}", "Outgoing"))

    workbook.save(path)
    workbook.close()


def test_number_column_detection_supports_report_labels():
    assert detect_number_columns(
        ["Contact", "Call Type", "Common Number", "Source Target"]
    ) == (0, 2, 3)


def test_phone_number_normalization_matches_common_indian_formats():
    assert canonical_phone_number("+91 90000-00001") == "9000000001"
    assert canonical_phone_number("09000000001") == "9000000001"


def test_identifier_detection_separates_phone_cell_imei_and_imsi():
    assert detect_identifier_columns(
        ["Other Party", "Cell ID", "Old IMEI", "New IMSI", "Total Calls"]
    ) == {
        0: "phone",
        1: "cell_id",
        2: "imei",
        3: "imsi",
    }

    assert detect_identifier_columns(
        ["CGI Lookup Status", "Tower Address", "IMEI Status", "Total Calls"]
    ) == {}


def test_viewer_lists_sheets_and_detects_number_columns(tmp_path: Path):
    build_application(["report-viewer-test"])

    report = tmp_path / "report.xlsx"
    _write_report(report)

    dialog = ReportViewerDialog(report)

    assert dialog._sheet_selector.count() == 2
    assert dialog._section_selector.isHidden()
    assert dialog._page_controls.isHidden()
    assert dialog._table.rowCount() == 2
    assert dialog.number_columns == (0,)

    dialog._sheet_selector.setCurrentText("Common Contacts")

    assert dialog._table.rowCount() == 1
    assert dialog._section_selector.isHidden()
    assert dialog.number_columns == (0, 1)

    dialog.close()


def test_viewer_navigates_structured_sections_without_cross_section_rows(
    tmp_path: Path,
):
    build_application(["report-viewer-section-test"])

    report = tmp_path / "structured-report.xlsx"
    _write_structured_report(report)

    dialog = ReportViewerDialog(report)

    try:
        assert not dialog._section_selector.isHidden()
        assert dialog._section_selector.count() == 3
        assert dialog._section_selector.currentText() == (
            "FIRST AND LAST COMMUNICATION BY DAY (FCLC)"
        )
        assert dialog._table.columnCount() == 3
        assert dialog._table.rowCount() == 2
        assert dialog._table.horizontalHeaderItem(0).text() == "Other Party"
        assert dialog._table.item(0, 0).text() == "9000000002"
        assert dialog.number_columns == (0,)
        assert dialog._page_controls.isHidden()
        assert "Records: 2." in dialog._status.text()
        assert "Review the first and last daily events." in dialog._status.text()

        dialog._section_selector.setCurrentText("MOVING CALLS")

        assert dialog._table.columnCount() == 3
        assert dialog._table.rowCount() == 1
        assert dialog._table.horizontalHeaderItem(0).text() == "Cell ID"
        assert dialog._table.item(0, 0).text() == "405-51-834-15492631"
        assert dialog.number_columns == ()
        assert dialog._page_controls.isHidden()
        assert "Section: MOVING CALLS" in dialog._status.text()
        assert "9000000002" not in dialog._status.text()

        dialog._section_selector.setCurrentText(
            "OBSERVED IDENTIFIER VARIATIONS"
        )

        assert dialog._table.columnCount() == 0
        assert dialog._table.rowCount() == 0
        assert dialog._page_controls.isHidden()
        assert dialog._page_label.text() == "Page 0 of 0"
        assert "Records: 0." in dialog._status.text()
        assert (
            "No records are available for this section."
            in dialog._status.text()
        )
        assert "Review only confirmed variations." in dialog._status.text()
    finally:
        dialog.close()


def test_viewer_filters_selected_section_and_resets_search_on_navigation(
    tmp_path: Path,
):
    build_application(["report-viewer-filter-test"])

    report = tmp_path / "structured-filter-report.xlsx"
    _write_structured_report(report)

    dialog = ReportViewerDialog(report)

    try:
        assert dialog._search_input.isEnabled()
        assert dialog._record_count_label.text() == (
            "Visible records: 2 of 2 loaded."
        )

        dialog._search_input.setText("incoming")

        assert dialog._table.rowCount() == 1
        assert dialog._table.item(0, 0).text() == "9000000003"
        assert dialog._record_count_label.text() == (
            "Visible records: 1 of 2 loaded."
        )
        assert dialog.number_columns == (0,)
        assert dialog._table.isSortingEnabled()

        dialog._search_input.setText("not-present")

        assert dialog._table.rowCount() == 0
        assert dialog._record_count_label.text() == (
            "Visible records: 0 of 2 loaded."
        )

        dialog._section_selector.setCurrentText("MOVING CALLS")

        assert dialog._search_input.text() == ""
        assert dialog._search_input.isEnabled()
        assert dialog._table.rowCount() == 1
        assert dialog._record_count_label.text() == (
            "Visible records: 1 of 1 loaded."
        )

        dialog._section_selector.setCurrentText(
            "OBSERVED IDENTIFIER VARIATIONS"
        )

        assert dialog._search_input.text() == ""
        assert not dialog._search_input.isEnabled()
        assert dialog._record_count_label.text() == (
            "Visible records: 0 of 0 loaded."
        )
    finally:
        dialog.close()


def test_viewer_filters_legacy_sheet_and_clears_query_on_sheet_change(
    tmp_path: Path,
):
    build_application(["report-viewer-legacy-filter-test"])

    report = tmp_path / "legacy-filter-report.xlsx"
    _write_report(report)

    dialog = ReportViewerDialog(report)

    try:
        dialog._search_input.setText("OUTGOING")

        assert dialog._table.rowCount() == 1
        assert dialog._table.item(0, 0).text() == "919000000001"
        assert dialog._record_count_label.text() == (
            "Visible records: 1 of 2 loaded."
        )

        dialog._sheet_selector.setCurrentText("Common Contacts")

        assert dialog._search_input.text() == ""
        assert dialog._table.rowCount() == 1
        assert dialog._record_count_label.text() == (
            "Visible records: 1 of 1 loaded."
        )
    finally:
        dialog.close()


def test_viewer_pages_through_large_structured_section_and_resets_search(
    tmp_path: Path,
):
    build_application(["report-viewer-filter-limit-test"])

    report = tmp_path / "large-structured-report.xlsx"
    _write_large_structured_report(report)

    dialog = ReportViewerDialog(report)

    try:
        assert dialog._table.rowCount() == 500
        assert dialog._table.item(0, 0).text() == "record-0000"
        assert dialog._table.item(499, 0).text() == "record-0499"
        assert dialog._record_count_label.text() == (
            "Visible records: 500 of 500 loaded. Section total: 501."
        )
        assert not dialog._page_controls.isHidden()
        assert dialog._page_label.text() == "Page 1 of 2"
        assert not dialog._previous_page_button.isEnabled()
        assert dialog._next_page_button.isEnabled()
        assert "Showing records 1-500 of 501." in dialog._status.text()

        dialog._search_input.setText("record-0499")

        assert dialog._table.rowCount() == 1
        assert dialog._table.item(0, 0).text() == "record-0499"
        assert dialog._record_count_label.text() == (
            "Visible records: 1 of 500 loaded. Section total: 501."
        )

        dialog._next_page_button.click()

        assert dialog._search_input.text() == ""
        assert dialog._table.rowCount() == 1
        assert dialog._table.item(0, 0).text() == "record-0500"
        assert dialog._record_count_label.text() == (
            "Visible records: 1 of 1 loaded. Section total: 501."
        )
        assert dialog._table.isSortingEnabled()
        assert dialog._page_label.text() == "Page 2 of 2"
        assert dialog._previous_page_button.isEnabled()
        assert not dialog._next_page_button.isEnabled()
        assert "Showing records 501-501 of 501." in dialog._status.text()

        dialog._previous_page_button.click()

        assert dialog._table.rowCount() == 500
        assert dialog._table.item(0, 0).text() == "record-0000"
        assert dialog._page_label.text() == "Page 1 of 2"

        dialog._next_page_button.click()
        dialog._section_selector.setCurrentText("SMALL CONTACT SECTION")

        assert dialog._page_controls.isHidden()
        assert dialog._page_label.text() == "Page 1 of 1"
        assert dialog._table.rowCount() == 1
        assert dialog._table.item(0, 0).text() == "small-record"

        dialog._section_selector.setCurrentText("LARGE CONTACT SECTION")

        assert dialog._page_label.text() == "Page 1 of 2"
        assert dialog._table.rowCount() == 500
        assert dialog._table.item(0, 0).text() == "record-0000"
    finally:
        dialog.close()


def test_viewer_pages_through_large_legacy_sheet(tmp_path: Path):
    build_application(["report-viewer-legacy-page-test"])

    report = tmp_path / "large-legacy-report.xlsx"
    _write_large_legacy_report(report)

    dialog = ReportViewerDialog(report)

    try:
        assert dialog._section_selector.isHidden()
        assert not dialog._page_controls.isHidden()
        assert dialog._page_label.text() == "Page 1 of 2"
        assert dialog._table.rowCount() == 500
        assert dialog._table.item(0, 0).text() == "9000000000"
        assert dialog._table.item(499, 0).text() == "9000000499"

        dialog._next_page_button.click()

        assert dialog._page_label.text() == "Page 2 of 2"
        assert dialog._table.rowCount() == 1
        assert dialog._table.item(0, 0).text() == "9000000500"
        assert dialog.number_columns == (0,)
        assert dialog._table.isSortingEnabled()
        assert "Showing records 501-501 of 501." in dialog._status.text()
    finally:
        dialog.close()


def test_viewer_searches_all_structured_pages_and_pages_matching_results(
    tmp_path: Path,
):
    build_application(["report-viewer-all-pages-search-test"])

    report = tmp_path / "large-structured-search-report.xlsx"
    _write_large_structured_report(report)

    dialog = ReportViewerDialog(report)

    try:
        dialog._search_input.setText("record-0500")

        assert dialog._table.rowCount() == 0
        assert dialog._search_all_pages_button.isEnabled()

        dialog._search_all_pages_button.click()

        assert dialog._search_input.text() == "record-0500"
        assert dialog._table.rowCount() == 1
        assert dialog._table.item(0, 0).text() == "record-0500"
        assert dialog._record_count_label.text() == (
            "Visible search results: 1 of 1 loaded. Total matches: 1. "
            "Section total: 501."
        )
        assert dialog._page_label.text() == "Search Page 1 of 1"
        assert dialog._page_controls.isHidden()
        assert 'Search all pages: "record-0500"' in dialog._status.text()

        dialog._search_input.setText("not-present")
        dialog._search_all_pages_button.click()

        assert dialog._table.rowCount() == 0
        assert dialog._search_input.isEnabled()
        assert dialog._page_label.text() == "Search Page 0 of 0"
        assert dialog._record_count_label.text() == (
            "Visible search results: 0 of 0 loaded. Total matches: 0. "
            "Section total: 501."
        )
        assert "No matching records were found across all pages." in (
            dialog._status.text()
        )

        dialog._search_input.setText("outgoing")

        assert dialog._page_label.text() == "Page 1 of 2"
        assert dialog._table.rowCount() == 500

        dialog._search_all_pages_button.click()

        assert dialog._page_label.text() == "Search Page 1 of 2"
        assert not dialog._page_controls.isHidden()
        assert dialog._record_count_label.text() == (
            "Visible search results: 500 of 500 loaded. "
            "Total matches: 501. Section total: 501."
        )
        assert "Showing matching records 1-500 of 501." in (
            dialog._status.text()
        )

        dialog._next_page_button.click()

        assert dialog._search_input.text() == "outgoing"
        assert dialog._page_label.text() == "Search Page 2 of 2"
        assert dialog._table.rowCount() == 1
        assert dialog._table.item(0, 0).text() == "record-0500"
        assert dialog._table.isSortingEnabled()

        dialog._search_input.clear()

        assert dialog._page_label.text() == "Page 1 of 2"
        assert dialog._table.rowCount() == 500
        assert dialog._record_count_label.text() == (
            "Visible records: 500 of 500 loaded. Section total: 501."
        )
    finally:
        dialog.close()


def test_viewer_searches_all_legacy_pages_and_returns_to_page_filter(
    tmp_path: Path,
):
    build_application(["report-viewer-legacy-all-pages-search-test"])

    report = tmp_path / "large-legacy-search-report.xlsx"
    _write_large_legacy_report(report)

    dialog = ReportViewerDialog(report)

    try:
        dialog._search_input.setText("9000000500")
        assert dialog._table.rowCount() == 0

        dialog._search_all_pages_button.click()

        assert dialog._table.rowCount() == 1
        assert dialog._table.item(0, 0).text() == "9000000500"
        assert dialog._page_label.text() == "Search Page 1 of 1"
        assert dialog._record_count_label.text() == (
            "Visible search results: 1 of 1 loaded. Total matches: 1. "
            "Section total: 501."
        )

        dialog._search_input.setText("9000000001")

        assert dialog._page_label.text() == "Page 1 of 2"
        assert dialog._table.rowCount() == 1
        assert dialog._table.item(0, 0).text() == "9000000001"
        assert dialog._record_count_label.text() == (
            "Visible records: 1 of 500 loaded. Section total: 501."
        )
    finally:
        dialog.close()


def test_number_double_click_opens_verified_source_records(
    tmp_path: Path,
    monkeypatch,
):
    build_application(["report-viewer-source-test"])

    report = tmp_path / "report.xlsx"
    _write_report(report)

    dialog = ReportViewerDialog(report)

    from gui.widgets import report_viewer_dialog
    from modules.reporting import cdr_report_source

    expected = pd.DataFrame(
        [
            {
                "b_party": "9000000001",
                "call_type": "Incoming",
            }
        ]
    )

    monkeypatch.setattr(
        cdr_report_source,
        "load_verified_source_link",
        lambda path: {"verified": True},
    )
    monkeypatch.setattr(
        cdr_report_source,
        "query_related_records",
        lambda link, number, identifier_type="phone": expected,
    )

    opened = {}

    class FakeRelatedDialog:
        def __init__(
            self,
            number,
            records,
            parent,
            identifier_label="Number",
        ):
            opened["number"] = number
            opened["records"] = records
            opened["identifier_label"] = identifier_label

        def exec(self):
            opened["executed"] = True

    monkeypatch.setattr(
        report_viewer_dialog,
        "RelatedRecordsDialog",
        FakeRelatedDialog,
    )

    dialog._show_number_summary(0, 0)

    assert opened["number"] == "919000000001"
    assert opened["records"].equals(expected)
    assert opened["executed"] is True

    dialog.close()


def test_cell_id_double_click_uses_typed_verified_query(
    tmp_path: Path,
    monkeypatch,
):
    build_application(["report-viewer-cell-test"])

    report = tmp_path / "cell-report.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Cell ID", "Total Calls"])
    sheet.append(["405-51-834-15492631", 12])
    workbook.save(report)

    dialog = ReportViewerDialog(report)

    from gui.widgets import report_viewer_dialog
    from modules.reporting import cdr_report_source

    monkeypatch.setattr(
        cdr_report_source,
        "load_verified_source_link",
        lambda path: {"verified": True},
    )

    queried = {}
    expected = pd.DataFrame(
        [{"first_cell_id": "405-51-834-15492631"}]
    )

    def fake_query(
        link,
        identifier,
        *,
        identifier_type="phone",
    ):
        queried.update(
            identifier=identifier,
            identifier_type=identifier_type,
        )
        return expected

    monkeypatch.setattr(
        cdr_report_source,
        "query_related_records",
        fake_query,
    )

    opened = {}

    class FakeRelatedDialog:
        def __init__(
            self,
            value,
            records,
            parent,
            identifier_label="Number",
        ):
            opened.update(
                value=value,
                records=records,
                label=identifier_label,
            )

        def exec(self):
            opened["executed"] = True

    monkeypatch.setattr(
        report_viewer_dialog,
        "RelatedRecordsDialog",
        FakeRelatedDialog,
    )

    dialog._show_identifier_records(0, 0)

    assert queried == {
        "identifier": "4055183415492631",
        "identifier_type": "cell_id",
    }
    assert opened["label"] == "Cell ID"
    assert opened["executed"] is True

    dialog.close()


def test_prepare_related_records_removes_only_exact_duplicates():
    records = pd.DataFrame(
        [
            {
                "a_party": "9000000001",
                "b_party": "9000000002",
                "call_date": "10-01-2026",
                "call_time": "10:00:00",
                "call_type": "Outgoing",
                "call_duration": 30,
                "_source_order": 1,
            },
            {
                "a_party": "9000000001",
                "b_party": "9000000002",
                "call_date": "10-01-2026",
                "call_time": "10:00:00",
                "call_type": "Outgoing",
                "call_duration": 30,
                "_source_order": 1,
            },
            {
                "a_party": "9000000001",
                "b_party": "9000000002",
                "call_date": "10-01-2026",
                "call_time": "10:05:00",
                "call_type": "Outgoing",
                "call_duration": 30,
                "_source_order": 2,
            },
        ]
    )
    original = records.copy(deep=True)

    display_records, duplicates_hidden = prepare_related_records(records)

    assert duplicates_hidden == 1
    assert len(display_records) == 2
    assert display_records["Time"].tolist() == [
        "10:00:00",
        "10:05:00",
    ]

    pd.testing.assert_frame_equal(records, original)


def test_prepare_related_records_uses_safe_columns_and_labels():
    records = pd.DataFrame(
        [
            {
                "a_party": "9000000001",
                "target_number": "919999999999",
                "b_party": "9000000002",
                "call_date": "10-01-2026",
                "call_time": "10:00:00",
                "call_type": "Incoming",
                "call_duration": 45,
                "first_cell_id": "404-55-113-12101",
                "imei": "868116066643170",
                "imsi": "405001111111111",
                "_source_order": 7,
                "_internal_match_key": "hidden-value",
            }
        ]
    )

    display_records, duplicates_hidden = prepare_related_records(records)

    assert duplicates_hidden == 0
    assert list(display_records.columns) == [
        "Target Number",
        "Other Party",
        "Date",
        "Time",
        "Event Type",
        "Duration (Seconds)",
        "First Cell ID",
        "IMEI",
        "IMSI",
    ]
    assert display_records.iloc[0]["Target Number"] == "9000000001"
    assert "_source_order" not in display_records.columns
    assert "_internal_match_key" not in display_records.columns






@pytest.mark.parametrize(
    (
        "column_name",
        "identifier_value",
        "expected_identifier_type",
        "expected_label",
    ),
    [
        (
            "IMEI",
            "868116066643170",
            "imei",
            "IMEI",
        ),
        (
            "IMSI",
            "405001111111111",
            "imsi",
            "IMSI",
        ),
    ],
)
def test_device_identifier_double_click_uses_typed_verified_query(
    tmp_path: Path,
    monkeypatch,
    column_name: str,
    identifier_value: str,
    expected_identifier_type: str,
    expected_label: str,
):
    build_application(
        [f"report-viewer-{expected_identifier_type}-test"]
    )

    report = tmp_path / f"{expected_identifier_type}-report.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.append([column_name, "Total Calls"])
    sheet.append([identifier_value, 5])
    workbook.save(report)

    dialog = ReportViewerDialog(report)

    from gui.widgets import report_viewer_dialog
    from modules.reporting import cdr_report_source

    monkeypatch.setattr(
        cdr_report_source,
        "load_verified_source_link",
        lambda path: {"verified": True},
    )

    queried = {}
    expected_records = pd.DataFrame(
        [{expected_identifier_type: identifier_value}]
    )

    def fake_query(
        link,
        identifier,
        *,
        identifier_type="phone",
    ):
        queried.update(
            identifier=identifier,
            identifier_type=identifier_type,
        )
        return expected_records

    monkeypatch.setattr(
        cdr_report_source,
        "query_related_records",
        fake_query,
    )

    opened = {}

    class FakeRelatedDialog:
        def __init__(
            self,
            value,
            records,
            parent,
            identifier_label="Number",
        ):
            opened.update(
                value=value,
                records=records,
                label=identifier_label,
            )

        def exec(self):
            opened["executed"] = True

    monkeypatch.setattr(
        report_viewer_dialog,
        "RelatedRecordsDialog",
        FakeRelatedDialog,
    )

    dialog._show_identifier_records(0, 0)

    assert queried == {
        "identifier": identifier_value,
        "identifier_type": expected_identifier_type,
    }
    assert opened["value"] == identifier_value
    assert opened["records"].equals(expected_records)
    assert opened["label"] == expected_label
    assert opened["executed"] is True

    dialog.close()


def test_related_records_dialog_shows_investigation_summary():
    from PySide6.QtWidgets import QApplication

    from gui.widgets.report_viewer_dialog import RelatedRecordsDialog

    application = QApplication.instance() or QApplication([])
    assert application is not None

    records = pd.DataFrame(
        [
            {
                "b_party": "9000000002",
                "call_date": "03/08/2026",
                "call_type": "Incoming",
                "call_duration": 30,
            },
            {
                "b_party": "9000000002",
                "call_date": "01-08-2026",
                "call_type": "outgoing",
                "call_duration": "60",
            },
            {
                "b_party": "9000000002",
                "call_date": "02/08/2026",
                "call_type": "smsin",
                "call_duration": 0,
            },
            {
                "b_party": "9000000002",
                "call_date": "02/08/2026",
                "call_type": "SMS Out",
                "call_duration": 0,
            },
            {
                "b_party": "9000000002",
                "call_date": "02/08/2026",
                "call_type": "sms",
                "call_duration": 0,
            },
            {
                "b_party": "9000000002",
                "call_date": "unavailable",
                "call_type": "unknown",
                "call_duration": 99,
            },
        ]
    )

    dialog = RelatedRecordsDialog("9000000002", records)

    try:
        assert dialog._summary.text() == (
            "Shown-record summary: Records: 6 | Incoming: 2 | Outgoing: 2 | "
            "Calls: 2 | SMS: 3 | Call duration: 1m 30s | "
            "Date range: 01 Aug 2026 to 03 Aug 2026"
        )
    finally:
        dialog.close()


def test_related_records_dialog_reports_limit_and_enables_sorting():
    import pandas as pd
    from PySide6.QtWidgets import QApplication, QLabel, QTableWidget

    from gui.widgets.report_viewer_dialog import RelatedRecordsDialog

    application = QApplication.instance() or QApplication([])
    assert application is not None

    records = pd.DataFrame(
        {
            "Source Target": ["9000000001", "9000000001"],
            "B Party Number": ["9000000002", "9000000002"],
            "Call Date": ["01/08/2026", "01/08/2026"],
            "Call Time": ["10:00:00", "10:01:00"],
            "Call Type": ["Outgoing", "Incoming"],
        }
    )
    plain_records = records.iloc[:1].copy()
    plain_records.attrs.clear()

    records.attrs["result_limit"] = 2
    records.attrs["result_limited"] = True

    limited_dialog = RelatedRecordsDialog(
        "9000000002",
        records,
        identifier_label="Number",
    )
    plain_dialog = RelatedRecordsDialog(
        "9000000002",
        plain_records,
        identifier_label="Number",
    )

    try:
        table = limited_dialog.findChild(QTableWidget)
        assert table is not None
        assert table.isSortingEnabled()

        limited_text = "\n".join(
            label.text()
            for label in limited_dialog.findChildren(QLabel)
        )
        assert (
            "Shown-record summary (loaded subset): Records: 2 | "
            "Incoming: 1 | Outgoing: 1 | Calls: 2 | SMS: 0 | "
            "Date range: 01 Aug 2026"
            in limited_text
        )
        assert "Verified records shown: 2." in limited_text
        assert (
            "Query limit reached: at least 3 matching source records were found"
            in limited_text
        )
        assert "the first 2 source records were loaded" in limited_text

        plain_text = "\n".join(
            label.text()
            for label in plain_dialog.findChildren(QLabel)
        )
        assert (
            "Shown-record summary: Records: 1 | Incoming: 0 | Outgoing: 1 | "
            "Calls: 1 | SMS: 0 | Date range: 01 Aug 2026"
            in plain_text
        )
        assert "Verified records shown: 1." in plain_text
        assert "Query limit reached" not in plain_text
    finally:
        limited_dialog.close()
        plain_dialog.close()
