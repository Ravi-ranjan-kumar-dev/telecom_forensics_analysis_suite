from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from modules.analysis.cdr.contacts import bottom_contacts
from modules.reporting import cdr_compact_excel
from modules.reporting.cdr_contact_map import contact_map_path
from modules.reporting.cdr_movement_route import movement_route_path
from modules.reporting.multi_cdr_excel import (
    SHEET_MAP,
    generate_multi_cdr_report,
)


def _single_frame() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "a_party": ["9000000000"] * 6,
            "b_party": [
                "8000000001",
                "8000000001",
                "8000000002",
                "8000000003",
                "8000000004",
                "8000000005",
            ],
            "call_type": [
                "outgoing",
                "incoming",
                "smsout",
                "smsin",
                "outgoing",
                "incoming",
            ],
            "call_date": ["01-01-2026"] * 6,
            "call_time": [
                "09:00:00",
                "10:00:00",
                "11:00:00",
                "12:00:00",
                "13:00:00",
                "14:00:00",
            ],
            "call_duration": [20, 25, 0, 0, 30, 35],
            "first_cell_id": [
                "404-55-113-12101",
                "404-55-113-12102",
                "404-55-113-12101",
                "404-55-113-12102",
                "404-55-113-12101",
                "404-55-113-12102",
            ],
            "last_cell_id": [
                "404-55-113-12102",
                "404-55-113-12102",
                "404-55-113-12101",
                "404-55-113-12102",
                "404-55-113-12102",
                "404-55-113-12102",
            ],
            "imei": ["490154203237518"] * 6,
            "imsi": ["405001111111111"] * 6,
            "roaming_circle": ["Bihar", "Jharkhand"] * 3,
        }
    )


def _single_bundle() -> dict:
    contact = pd.DataFrame(
        [
            {
                "Contact": "8000000001",
                "Total Calls": 2,
                "contact_sdr_subscriber_name": "Test Contact",
                "contact_sdr_father_name": "Test Father",
                "contact_sdr_address": "Test Address",
                "contact_sdr_operator": "AIRTEL",
                "contact_sdr_circle": "Bihar",
                "contact_sdr_lookup_status": "FOUND",
                "contact_sdr_match_confidence": "DIRECT_NORMALIZED_MSISDN",
            }
        ]
    )
    empty = pd.DataFrame()
    return {
        "results": {
            "top_contacts": contact,
            "bottom_contacts": contact,
            "contact_ranking": contact,
            "social_network": contact,
            "contact_category_summary": empty,
            "top_service_sender_ids": empty,
            "top_short_codes": empty,
            "incoming_outgoing": empty,
            "other_call_type_summary": empty,
            "analyze_location": empty,
            "tower_movement": empty,
            "tower_transition": empty,
            "movement_pattern": empty,
            "tower_intelligence": empty,
            "home_tower": empty,
            "work_tower": empty,
            "imei_intelligence": empty,
            "sim_change": empty,
            "activity_summary": empty,
            "hourly_activity": empty,
            "daily_activity": empty,
            "weekly_activity": empty,
            "monthly_activity": empty,
        }
    }


def _sheet_text(worksheet) -> str:
    return "\n".join(
        str(cell.value)
        for row in worksheet.iter_rows()
        for cell in row
        if cell.value is not None
    )


def _section_headers(worksheet, title: str) -> list[str]:
    for row_index in range(1, worksheet.max_row + 1):
        if worksheet.cell(row=row_index, column=1).value != title:
            continue
        header_row = row_index + 1
        while (
            header_row <= worksheet.max_row
            and worksheet.cell(row=header_row, column=1).value is not None
            and worksheet.cell(row=header_row, column=2).value is None
        ):
            header_row += 1
        return [
            str(worksheet.cell(row=header_row, column=column).value)
            for column in range(1, worksheet.max_column + 1)
            if worksheet.cell(row=header_row, column=column).value is not None
        ]
    raise AssertionError(f"Section not found: {title}")


def test_bottom_contacts_returns_true_least_frequent_human_numbers():
    rows = []
    for index, count in enumerate(range(1, 13), start=1):
        rows.extend(
            {"b_party": f"8{index:09d}"}
            for _ in range(count)
        )
    rows.extend(
        {"b_party": "AD-BANK-S"}
        for _ in range(30)
    )
    rows.extend(
        {"b_party": value}
        for value in (
            "01408367994",
            "01409090210",
            "01600103944",
            "01725677701",
            "123456789012345",
        )
    )
    rows.extend(
        {"b_party": value}
        for value in (
            "08000000012",
            "+91 8000000012",
        )
    )

    result = bottom_contacts(
        pd.DataFrame(rows),
        limit=10,
    )

    assert len(result) == 10
    assert result.iloc[0]["Contact"] == "8000000001"
    assert int(result.iloc[0]["Total Calls"]) == 1
    assert "AD-BANK-S" not in set(result["Contact"])
    assert all(
        len(str(contact)) == 10 and str(contact)[0] in "6789"
        for contact in result["Contact"]
    )
    assert not any(
        str(contact).startswith("01")
        for contact in result["Contact"]
    )


def test_single_report_has_requested_independent_sheet_contract(
    tmp_path: Path,
    monkeypatch,
):
    monkeypatch.setattr(
        "modules.reporting.single_cdr_excel._enrich_target_metadata_with_sdr",
        lambda metadata, target: metadata,
    )
    monkeypatch.setattr(
        "modules.enrichment.sdr_subscriber_enrichment.lookup_sdr_subscribers",
        lambda values: pd.DataFrame(),
    )
    monkeypatch.setattr(
        "modules.analysis.cdr.contact_report.lookup_cgi_addresses",
        lambda values: pd.DataFrame(),
    )

    report = cdr_compact_excel.generate_single_cdr_compact_report(
        _single_frame(),
        "9000000000",
        metadata={"target": "9000000000"},
        analysis_bundle=_single_bundle(),
        output_dir=tmp_path,
    )

    assert report is not None
    workbook = load_workbook(report, data_only=False)
    assert tuple(workbook.sheetnames) == cdr_compact_excel.SINGLE_CDR_COMPACT_SHEETS

    executive = workbook["1. Executive Summary"]
    executive_text = _sheet_text(executive)
    assert "SDR Match" not in executive_text
    assert "TOP 10 HUMAN CONTACTS" in executive_text
    assert "BOTTOM 10 HUMAN CONTACTS" in executive_text
    assert "ROAMING SUMMARY" in executive_text
    assert "Match Confidence" not in _section_headers(
        executive,
        "TOP 10 HUMAN CONTACTS",
    )

    contact_headers = _section_headers(
        workbook["2. Full Contact Summary"],
        "FULL CONTACT SUMMARY (CC SUMMARY)",
    )
    assert "SDR Match" in _sheet_text(
        workbook["2. Full Contact Summary"]
    )
    assert "SDR Lookup Status" not in contact_headers
    assert "Last Interaction CGI Lookup Status" not in contact_headers

    for sheet_name in (
        "3. Outgoing Voice Calls",
        "4. Incoming Voice Calls",
        "5. Outgoing SMS",
        "6. Incoming SMS",
        "9. Tower Intelligence",
        "10. Probable Home Tower",
        "11. Probable Work Tower",
        "12. Daily First Last FCLC",
        "13. FCLC Location Summary",
        "14. FCLC Contact Summary",
        "15. Moving Calls",
        "16. Movement Events",
        "17. Tower Transitions",
        "18. Movement Patterns",
        "19. Device & SIM Intel",
        "20. Hourly Activity",
        "21. Daily Activity",
        "22. Weekly Activity",
        "23. Monthly Activity",
    ):
        assert sheet_name in workbook.sheetnames

    assert "3. Communication Intel" not in workbook.sheetnames
    assert "21. Activity Summary" not in workbook.sheetnames

    incoming_headers = _section_headers(
        workbook["4. Incoming Voice Calls"],
        "INCOMING VOICE CALLS",
    )
    assert "Cell ID" not in incoming_headers
    assert "Address" not in incoming_headers
    assert "End Cell ID" not in incoming_headers
    assert "End Address" not in incoming_headers

    location_headers = _section_headers(
        workbook["8. Location Overview"],
        "CELL ID SUMMARY",
    )
    assert "Start Tower Lookup Status" not in location_headers

    device_sheet = workbook["19. Device & SIM Intel"]
    device_text = _sheet_text(device_sheet)
    device_headers = _section_headers(
        device_sheet,
        "DEVICE SUMMARY",
    )
    assert "SIM SUMMARY" in device_text
    assert "DEVICE / SIM CHANGE INDICATORS" in device_text
    assert "CONFIRMED DEVICE OR SIM CHANGES" not in device_text
    assert "IMEI" not in device_headers
    assert "Valid IMEI" not in device_headers
    assert "Invalid IMEI Values" not in device_headers

    assert "5. Location & Roaming" not in workbook.sheetnames
    assert "6. Movement & Daily Routine" not in workbook.sheetnames
    assert "8. Activity & Timing" not in workbook.sheetnames
    assert "9. Priority Review Queue" not in workbook.sheetnames
    assert "10. Data Quality & Guide" not in workbook.sheetnames


def _multi_frame(
    target: str,
    other_target: str,
) -> pd.DataFrame:
    return pd.DataFrame(
        {
            "a_party": [target, target, target],
            "b_party": ["8000000001", "8000000001", other_target],
            "call_type": ["outgoing", "incoming", "outgoing"],
            "call_date": ["01-01-2026", "01-01-2026", "01-01-2026"],
            "call_time": ["09:00:00", "10:00:00", "11:00:00"],
            "call_duration": [20, 25, 30],
            "first_cell_id": [
                "404-55-113-12101",
                "404-55-113-12102",
                "404-55-113-12101",
            ],
            "last_cell_id": [
                "404-55-113-12101",
                "404-55-113-12102",
                "404-55-113-12101",
            ],
            "imei": ["490154203237518"] * 3,
            "imsi": ["405001111111111"] * 3,
        }
    )


def test_common_report_has_master_enrichment_maps_and_no_technical_sheets(
    tmp_path: Path,
    monkeypatch,
):
    sdr_calls: list[list[str]] = []
    cgi_calls: list[list[str]] = []

    def fake_sdr(values):
        values = list(values)
        sdr_calls.append(values)
        names = {
            "8000000001": "Shared Contact",
            "9000000001": "First Target",
            "9000000002": "Second Target",
        }
        return pd.DataFrame(
            [
                {
                    "lookup_mobile": number,
                    "subscriber_name": names[number],
                    "father_name": "Test Father",
                    "subscriber_address": f"Address {number}",
                    "operator": "AIRTEL",
                    "circle": "Bihar",
                    "sdr_found": "Yes",
                }
                for number in values
            ]
        )

    def fake_cgi(values):
        values = list(values)
        cgi_calls.append(values)
        rows = []
        for index, cgi in enumerate(values):
            rows.append(
                {
                    "cgi": cgi,
                    "operator": "AIRTEL",
                    "circle": "Bihar",
                    "state": "Bihar",
                    "district": "Patna",
                    "police_station": "Test PS",
                    "town": "Patna",
                    "site_name": f"Site {index + 1}",
                    "address": f"Tower Address {index + 1}",
                    "latitude": 25.61 + (index * 0.01),
                    "longitude": 85.14 + (index * 0.01),
                    "source_file": "test.xlsx",
                }
            )
        return pd.DataFrame(rows)

    monkeypatch.setattr(
        "modules.enrichment.telecom_master_enrichment.lookup_sdr_subscribers",
        fake_sdr,
    )
    monkeypatch.setattr(
        "modules.enrichment.telecom_master_enrichment.lookup_cgi_addresses",
        fake_cgi,
    )

    loaded = {
        "9000000001": {
            "df": _multi_frame("9000000001", "9000000002"),
            "file": "first.csv",
        },
        "9000000002": {
            "df": _multi_frame("9000000002", "9000000001"),
            "file": "second.csv",
        },
    }
    report = generate_multi_cdr_report(
        loaded,
        metadata={"case_name": "Master Integration Test"},
        output_dir=tmp_path,
    )

    assert report is not None
    assert len(sdr_calls) == 1
    assert len(cgi_calls) == 1

    workbook = load_workbook(report, data_only=False)
    assert tuple(workbook.sheetnames) == tuple(
        sheet_name
        for sheet_name, _, _ in SHEET_MAP
    )
    assert "13. Alerts" not in workbook.sheetnames
    assert "14. Rejected Rows" not in workbook.sheetnames
    assert "Methodology & Limits" not in workbook.sheetnames

    common_headers = [
        cell.value
        for cell in workbook["3. Common Numbers"][7]
        if cell.value is not None
    ]
    assert "Name" in common_headers
    assert "SDR Address" in common_headers
    assert "SDR Lookup Status" in common_headers

    direct_headers = [
        cell.value
        for cell in workbook["4. Direct Links"][7]
        if cell.value is not None
    ]
    assert "Source Name" in direct_headers
    assert "Destination Name" in direct_headers

    contact_matrix_headers = [
        cell.value
        for cell in workbook["8. Contact Matrix"][7]
        if cell.value is not None
    ]
    assert "Name" in contact_matrix_headers
    assert "SDR Address" in contact_matrix_headers

    for sheet_name in (
        "5. Common Towers",
        "9. Tower Matrix",
    ):
        headers = [
            cell.value
            for cell in workbook[sheet_name][7]
            if cell.value is not None
        ]
        assert "CGI Address" in headers
        assert "CGI Latitude" in headers
        assert "CGI Longitude" in headers
        assert "CGI Lookup Status" in headers

    tower_matrix_headers = [
        cell.value
        for cell in workbook["9. Tower Matrix"][7]
        if cell.value is not None
    ]
    assert "Linked Target Names" in tower_matrix_headers
    assert "Linked Target Father Names" in tower_matrix_headers
    assert "Linked Target SDR Addresses" in tower_matrix_headers
    assert "Linked Target Operators" in tower_matrix_headers
    assert "Linked Target Circles" in tower_matrix_headers
    assert "Linked Target SDR Status" in tower_matrix_headers

    tower_matrix_text = _sheet_text(
        workbook["9. Tower Matrix"]
    )
    assert "9000000001: First Target" in tower_matrix_text
    assert "9000000002: Second Target" in tower_matrix_text

    common_path = Path(report)
    map_file = contact_map_path(common_path)
    route_file = movement_route_path(common_path)
    assert map_file.is_file()
    assert route_file.is_file()

    map_text = map_file.read_text(encoding="utf-8")
    route_text = route_file.read_text(encoding="utf-8")
    assert "8000000001" in map_text
    assert "9000000001" in map_text
    assert "9000000002" in map_text
    assert "Different targets are never joined" in route_text
    assert "9000000001" in route_text
    assert "9000000002" in route_text
