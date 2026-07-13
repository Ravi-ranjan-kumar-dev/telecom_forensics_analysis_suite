from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook


def test_shared_excel_sanitizer_blocks_formula_like_values():
    from modules.reporting.excel_security import excel_safe_value

    for value in ("=1+1", "+SUM(A1:A2)", "-2+3", "@SUM(A1:A2)", "  =HYPERLINK(\"x\")", "\t+CMD"):
        safe = excel_safe_value(value)
        assert isinstance(safe, str)
        assert safe.startswith("'")

    assert excel_safe_value("9876543210") == "9876543210"
    assert excel_safe_value("CELL-A") == "CELL-A"
    assert excel_safe_value(None) == ""


def test_every_excel_renderer_uses_shared_literal_boundary():
    from modules.reporting.ipdr_excel import _safe_value as ipdr_safe
    from modules.reporting.single_cdr_excel import _excel_safe_scalar as single_safe
    from modules.reporting.tower_dump_excel import _clean_value as tower_dump_safe
    from modules.reporting.tower_gprs_excel import _safe_value as gprs_safe
    from modules.reporting.tower_ipdr_excel import _safe_value as tower_ipdr_safe
    from modules.reporting.tower_partition_excel import _safe_scalar as partition_safe

    for function in (
        ipdr_safe,
        single_safe,
        tower_dump_safe,
        gprs_safe,
        tower_ipdr_safe,
        partition_safe,
    ):
        assert function("=HYPERLINK(\"https://example.invalid\")").startswith("'")


def test_methodology_sheet_explains_corroboration_and_limits():
    from modules.reporting.report_guidance import append_methodology_sheet

    workbook = Workbook()
    workbook.remove(workbook.active)
    sheet_name = append_methodology_sheet(workbook, "Test Analysis")

    assert sheet_name == "Methodology & Limits"
    worksheet = workbook[sheet_name]
    text = " ".join(
        str(cell.value or "")
        for row in worksheet.iter_rows()
        for cell in row
    ).lower()
    assert "not a finding of guilt" in text
    assert "requires corroboration" in text
    assert "not establish a handset's exact physical position" in text
    assert "derived analytical artifact" in text


def test_multi_cdr_workbook_sanitizes_untrusted_cells_and_adds_guidance(tmp_path: Path):
    from modules.reporting.multi_cdr_excel import generate_multi_cdr_report

    loaded = {
        "9000000001": {"df": pd.DataFrame({"b_party": ["8000000001"]})},
        "9000000002": {"df": pd.DataFrame({"b_party": ["8000000002"]})},
    }
    bundle = {
        "alerts": pd.DataFrame(
            [{"Alert": "=HYPERLINK(\"https://example.invalid\",\"OPEN\")"}]
        )
    }

    report = generate_multi_cdr_report(
        loaded,
        metadata={"case_name": "Phase 3 Test"},
        analysis_bundle=bundle,
        output_dir=tmp_path,
    )

    assert report is not None
    workbook = load_workbook(report, data_only=False)
    assert "Methodology & Limits" in workbook.sheetnames

    alert_sheet = workbook["13. Alerts"]
    values = [cell.value for row in alert_sheet.iter_rows() for cell in row]
    protected = [value for value in values if isinstance(value, str) and "HYPERLINK" in value]
    assert protected
    assert all(value.startswith("'") for value in protected)
    assert all(cell.data_type != "f" for row in alert_sheet.iter_rows() for cell in row)


def test_behavioral_observations_are_neutral_and_corroboration_based():
    from modules.analysis.cdr.behavioral_intelligence import behavioral_intelligence

    dataframe = pd.DataFrame(
        {
            "b_party": ["8000000001", "8000000001", "8000000002"],
            "datetime": pd.to_datetime(
                [
                    "2026-07-10 01:00:00",
                    "2026-07-10 01:30:00",
                    "2026-07-10 15:00:00",
                ]
            ),
            "first_cell_id": ["CELL-A", "CELL-A", "CELL-B"],
            "imei": ["111111111111111", "111111111111111", "222222222222222"],
            "call_duration": [30, 60, 10],
        }
    )

    result = behavioral_intelligence(dataframe)
    assert list(result.columns) == ["Indicator", "Observation", "Caution"]

    text = " ".join(result.astype(str).stack().tolist()).lower()
    for prohibited in (
        "core associate threat",
        "primary safehouse",
        "evasion alert",
        "syndicate activity",
        "verified with",
    ):
        assert prohibited not in text

    assert "does not establish relationship" in text
    assert "not proof of exact handset or person location" in text


def test_single_cdr_workbook_sanitizes_contact_and_adds_guidance(tmp_path: Path):
    from modules.reporting.single_cdr_excel import generate_single_cdr_report

    dataframe = pd.DataFrame(
        {
            "a_party": ["9000000001"],
            "b_party": ["=HYPERLINK(\"https://example.invalid\",\"OPEN\")"],
            "call_type": ["outgoing"],
            "call_direction": ["OUTGOING"],
            "call_date": ["10/07/2026"],
            "call_time": ["12:00:00"],
            "call_duration": [30],
            "imei": ["111111111111111"],
            "imsi": ["404000000000001"],
            "first_cell_id": ["CELL-A"],
            "last_cell_id": ["CELL-A"],
            "first_location": ["Location A"],
            "last_location": ["Location A"],
        }
    )

    report = generate_single_cdr_report(
        dataframe,
        "9000000001",
        analysis_bundle={"results": {}, "errors": {}, "status": pd.DataFrame()},
        output_dir=tmp_path,
    )

    assert report is not None
    workbook = load_workbook(report, data_only=False)
    assert "Methodology & Limits" in workbook.sheetnames

    protected = []
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "HYPERLINK" in cell.value:
                    protected.append(cell)
    assert protected
    assert all(cell.value.startswith("'") and cell.data_type != "f" for cell in protected)


def test_tower_dump_workbook_sanitizes_normalized_values_and_adds_guidance(tmp_path: Path):
    from modules.reporting.tower_dump_excel import generate_tower_dump_excel_report

    result = {
        "metadata": {},
        "analysis": {
            "results": {},
            "status": pd.DataFrame(),
            "errors": pd.DataFrame(),
        },
        "df": pd.DataFrame({"subscriber_number": ["=1+1"]}),
        "warnings": [],
        "errors": [],
    }

    report = generate_tower_dump_excel_report(result, output_dir=tmp_path)
    workbook = load_workbook(report, data_only=False)
    assert "Methodology & Limits" in workbook.sheetnames

    normalized = workbook["15. Normalized Dump"]
    values = [cell for row in normalized.iter_rows() for cell in row if cell.value == "'=1+1"]
    assert values
    assert all(cell.data_type != "f" for cell in values)
