
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from modules.analysis.device.imei_common import (
    build_common_imei_cdr_analysis,
)
from modules.controllers import imei_device_controller
from modules.reporting.imei_device_excel import (
    generate_imei_device_report,
)


def _cdr_rows(
    query: str,
    observed: str,
    *,
    target: str = "9000000001",
) -> pd.DataFrame:
    rows = []

    for contact, cell in (
        (
            "9111111111",
            "404-10-100-200",
        ),
        (
            "JB-JIOINF-S",
            "405856",
        ),
        (
            "50000",
            "",
        ),
    ):
        rows.append(
            {
                "query_identifier_normalized": query,
                "observed_imei_normalized": observed,
                "imei": observed,
                "target": target,
                "call_date": "01/07/2026",
                "call_time": "10:00:00",
                "call_type": "outgoing",
                "b_party": contact,
                "call_duration": 30,
                "imsi": "405520123456789",
                "first_cell_id": cell,
                "last_cell_id": "",
                "source_file": f"{query}.csv",
                "source_path": f"/evidence/{query}.csv",
                "source_row_number": 8,
                "match_relation": "SAME_BASE14",
            }
        )

    return pd.DataFrame(
        rows
    )


def _empty_analysis(
    identifier: str,
    manifest: pd.DataFrame,
) -> dict:
    message = (
        "Valid operator report contains no result records."
    )

    return {
        "requested_imei": identifier,
        "overall_status": "EMPTY_NO_DATA",
        "message": message,
        "source_summary": pd.DataFrame(
            [
                {
                    "Evidence Source": "CDR",
                    "Status": "EMPTY_NO_DATA",
                    "Evidence Unit": "CDR records",
                    "Matched Count": 0,
                    "Message": message,
                }
            ]
        ),
        "associated_identities": pd.DataFrame(),
        "cross_source_timeline": pd.DataFrame(),
        "cdr": {
            "status": "EMPTY_NO_DATA",
            "timeline": pd.DataFrame(),
            "towers": pd.DataFrame(),
        },
        "ipdr": {
            "status": "NO_INPUT",
            "timeline": pd.DataFrame(),
        },
        "gprs": {
            "status": "NO_INPUT",
            "timeline": pd.DataFrame(),
        },
        "review_indicators": pd.DataFrame(),
        "data_quality": pd.DataFrame(),
        "acquisition_manifest": manifest,
    }


def test_empty_single_report_preserves_identifiers_and_manifest(
    tmp_path: Path,
):
    identifier = "35309885264837"
    sha256 = "1" * 64

    manifest = pd.DataFrame(
        [
            {
                "Query Identifier": identifier,
                "SHA-256": sha256,
                "Source Type": "CDR",
                "Inspection Status": "EMPTY_NO_DATA",
                "Analysis Content Role": "PRIMARY_CONTENT",
            }
        ]
    )

    path = generate_imei_device_report(
        case={
            "case_id": "CASE-001",
        },
        analysis=_empty_analysis(
            identifier,
            manifest,
        ),
        output_dir=tmp_path,
    )

    assert path is not None

    workbook = load_workbook(
        path,
        data_only=False,
    )

    assert (
        "10. Acquisition Manifest"
        in workbook.sheetnames
    )

    exact_identifier_cells = []

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value == identifier:
                    exact_identifier_cells.append(
                        cell
                    )

    assert exact_identifier_cells

    assert all(
        cell.data_type == "s"
        and cell.number_format == "@"
        for cell in exact_identifier_cells
    )

    manifest_sheet = workbook[
        "10. Acquisition Manifest"
    ]

    sha_cells = [
        cell
        for row in manifest_sheet.iter_rows()
        for cell in row
        if cell.value == sha256
    ]

    assert sha_cells

    assert all(
        cell.data_type == "s"
        and cell.number_format == "@"
        for cell in sha_cells
    )

    workbook.close()


def test_common_analysis_uses_distinct_device_families():
    base14 = "86628404348207"
    imei15 = "866284043482077"
    second_family = "86265906247429"

    result = build_common_imei_cdr_analysis(
        {
            base14: _cdr_rows(
                base14,
                "866284043482070",
            ),
            imei15: _cdr_rows(
                imei15,
                imei15,
            ),
            second_family: _cdr_rows(
                second_family,
                "862659062474290",
            ),
        }
    )

    assert result[
        "query_identifier_count"
    ] == 3

    assert result[
        "device_family_count"
    ] == 2

    target_row = result[
        "common_targets"
    ].iloc[
        0
    ]

    assert target_row[
        "Query Identifier Count"
    ] == 3

    assert target_row[
        "Device Family Count"
    ] == 2

    same_family_only = build_common_imei_cdr_analysis(
        {
            base14: _cdr_rows(
                base14,
                "866284043482070",
            ),
            imei15: _cdr_rows(
                imei15,
                imei15,
            ),
        }
    )

    assert same_family_only[
        "common_targets"
    ].empty

    assert same_family_only[
        "common_imsis"
    ].empty


def test_common_contacts_and_towers_are_separated():
    first = "86265906247429"
    second = "86672907805397"

    result = build_common_imei_cdr_analysis(
        {
            first: _cdr_rows(
                first,
                "862659062474290",
            ),
            second: _cdr_rows(
                second,
                "866729078053970",
            ),
        }
    )

    assert set(
        result[
            "common_contacts"
        ][
            "Contact Number"
        ]
    ) == {
        "9111111111",
    }

    assert set(
        result[
            "shared_service_identifiers"
        ][
            "Service / Other Identifier"
        ]
    ) == {
        "50000",
        "JB-JIOINF-S",
    }

    assert set(
        result[
            "common_towers"
        ][
            "Cell ID"
        ]
    ) == {
        "404-10-100-200",
    }

    assert "405856" not in set(
        result[
            "common_towers"
        ][
            "Cell ID"
        ]
    )


def test_inventory_separates_acquisition_and_cdr_content(
    monkeypatch,
    tmp_path: Path,
):
    first = tmp_path / "first.csv"
    duplicate = tmp_path / "duplicate.csv"
    gprs = tmp_path / "gprs.csv"

    first.write_bytes(
        b"same-cdr-content"
    )

    duplicate.write_bytes(
        b"same-cdr-content"
    )

    gprs.write_bytes(
        b"separate-gprs-content"
    )

    monkeypatch.setattr(
        imei_device_controller,
        "resolve_imei_cdr_input_folder",
        lambda case_id: tmp_path,
    )

    def fake_inspection(
        path,
    ):
        if Path(
            path
        ).name == "gprs.csv":
            return {
                "ok": True,
                "status": "HAS_DATA",
                "source_type": "GPRS",
                "format_id": "VIL_IMEI_GPRS",
                "operator": "Vodafone Idea",
                "query_identifier_normalized": (
                    "861679062132757"
                ),
                "query_identifier_type": "IMEI15",
                "record_count": 1,
                "rejected_line_count": 0,
                "message": "GPRS evidence.",
            }

        return {
            "ok": True,
            "status": "HAS_DATA",
            "source_type": "CDR",
            "format_id": "VIL_IMEI_CDR",
            "operator": "Vodafone Idea",
            "query_identifier_normalized": (
                "866284043482077"
            ),
            "query_identifier_type": "IMEI15",
            "record_count": 1,
            "rejected_line_count": 0,
            "message": "CDR evidence.",
        }

    normalizer_calls = []

    monkeypatch.setattr(
        imei_device_controller,
        "inspect_imei_evidence_file",
        fake_inspection,
    )

    def fake_normalizer(
        path,
        inspection=None,
    ):
        normalizer_calls.append(
            Path(
                path
            ).name
        )

        return {
            "ok": True,
            "status": "HAS_DATA",
            "data": pd.DataFrame(
                [
                    {
                        "query_identifier_normalized": (
                            "866284043482077"
                        ),
                        "imei": "866284043482077",
                        "target": "9000000001",
                    }
                ]
            ),
            "records_normalized": 1,
            "rejected_line_count": 0,
            "warnings": [],
            "errors": [],
            "message": "Normalized.",
        }

    monkeypatch.setattr(
        imei_device_controller,
        "normalize_imei_cdr_file",
        fake_normalizer,
    )

    result = (
        imei_device_controller
        ._load_dedicated_imei_cdr_inventory(
            "CASE-001"
        )
    )

    assert result[
        "files_found"
    ] == 3

    assert result[
        "all_content_groups"
    ] == 2

    assert result[
        "supported_cdr_content_groups"
    ] == 1

    assert result[
        "non_cdr_acquisitions"
    ] == 1

    assert result[
        "duplicate_cdr_acquisitions"
    ] == 1

    assert result[
        "analytical_records"
    ] == 1

    assert len(
        normalizer_calls
    ) == 1


def test_empty_no_data_workflow_creates_single_report(
    monkeypatch,
    tmp_path: Path,
):
    identifier = "35537311480035"

    manifest = pd.DataFrame(
        [
            {
                "Query Identifier": identifier,
                "Source Type": "CDR",
                "Inspection Status": "EMPTY_NO_DATA",
                "SHA-256": "a" * 64,
                "Analysis Content Role": "PRIMARY_CONTENT",
            }
        ]
    )

    monkeypatch.setattr(
        imei_device_controller,
        "case_report_dir",
        lambda case_id, report_type: tmp_path,
    )

    monkeypatch.setattr(
        imei_device_controller,
        "register_target",
        lambda *args, **kwargs: None,
    )

    monkeypatch.setattr(
        imei_device_controller,
        "register_report",
        lambda *args, **kwargs: None,
    )

    monkeypatch.setattr(
        imei_device_controller,
        "register_analysis_run",
        lambda *args, **kwargs: None,
    )

    monkeypatch.setattr(
        imei_device_controller,
        "log_case_event",
        lambda *args, **kwargs: None,
    )

    monkeypatch.setattr(
        imei_device_controller,
        "_print_source_summary",
        lambda analysis: None,
    )

    monkeypatch.setattr(
        imei_device_controller,
        "build_unified_imei_investigation",
        lambda *args, **kwargs: (_ for _ in ()).throw(
            AssertionError(
                "Unified event analysis must not run "
                "for a valid empty report."
            )
        ),
    )

    captured = {}

    def fake_report(
        *,
        case,
        analysis,
        output_dir,
    ):
        captured[
            "analysis"
        ] = analysis

        return tmp_path / "empty.xlsx"

    monkeypatch.setattr(
        imei_device_controller,
        "generate_imei_device_report",
        fake_report,
    )

    result = (
        imei_device_controller
        ._run_auto_single_imei_cdr(
            case={
                "case_id": "CASE-001",
            },
            identifier=identifier,
            dataframe=pd.DataFrame(),
            acquisition_manifest=manifest,
        )
    )

    assert result[
        "analysis"
    ][
        "overall_status"
    ] == "EMPTY_NO_DATA"

    assert result[
        "report"
    ] == tmp_path / "empty.xlsx"

    assert not captured[
        "analysis"
    ][
        "acquisition_manifest"
    ].empty



def test_base14_wording_and_manifest_are_investigator_safe(
    tmp_path: Path,
):
    identifier = "35309885264836"

    manifest = pd.DataFrame(
        [
            {
                "Relative Path": "evidence/base14.csv",
                "Source File": "base14.csv",
                "Source Path": (
                    "/home/investigator/project/"
                    "evidence/base14.csv"
                ),
                "SHA-256": "b" * 64,
                "Acquisition Content Role": "PRIMARY_CONTENT",
                "Duplicate Of": (
                    "/home/investigator/project/"
                    "evidence/original.csv"
                ),
                "Analysis Content Role": "PRIMARY_CONTENT",
                "Analysis Duplicate Of": (
                    "/home/investigator/project/"
                    "evidence/analysis-original.csv"
                ),
                "Format": "JIO_IMEI_CDR",
                "Operator": "Reliance Jio",
                "Source Type": "CDR",
                "Query Identifier": identifier,
                "Query Identifier Type": "BASE14",
                "Inspection Status": "EMPTY_NO_DATA",
                "Records Declared": 0,
                "Records Normalized": 0,
                "Rejected Lines": 0,
                "Message": (
                    "Valid report contains no records."
                ),
            }
        ]
    )

    path = generate_imei_device_report(
        case={
            "case_id": "CASE-BASE14",
        },
        analysis=_empty_analysis(
            identifier,
            manifest,
        ),
        output_dir=tmp_path,
    )

    assert path is not None

    workbook = load_workbook(
        path,
        data_only=False,
    )

    all_text = " ".join(
        str(
            cell.value
        )
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
        if cell.value not in {
            None,
            "",
        }
    )

    assert (
        "Requested Device Query Identifier"
        in all_text
    )

    assert (
        "BASE14, IMEI15 and IMEISV16"
        in all_text
    )

    assert (
        "Exact canonical 15- or 16-digit "
        "identifier matching."
        not in all_text
    )

    assert (
        "15-digit IMEI and 16-digit IMEISV "
        "are searched as exact identifiers."
        not in all_text
    )

    worksheet = workbook[
        "10. Acquisition Manifest"
    ]

    headers = [
        worksheet.cell(
            row=4,
            column=column_number,
        ).value
        for column_number in range(
            1,
            worksheet.max_column + 1,
        )
    ]

    assert worksheet.max_column == 15

    assert "Relative Path" in headers
    assert "Duplicate Reference" in headers

    assert "Source Path" not in headers
    assert "Duplicate Of" not in headers
    assert "Analysis Duplicate Of" not in headers

    manifest_text = " ".join(
        str(
            cell.value
        )
        for row in worksheet.iter_rows(
            min_row=5,
        )
        for cell in row
        if cell.value not in {
            None,
            "",
        }
    )

    assert "/home/" not in manifest_text

    assert (
        "analysis-original.csv"
        in manifest_text
    )

    workbook.close()
