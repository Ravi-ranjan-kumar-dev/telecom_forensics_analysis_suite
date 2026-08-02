from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from modules.analysis.cdr.activity import (
    hourly_activity,
    weekly_activity,
)
from modules.analysis.cdr.contact_report import (
    build_full_contact_summary,
)
from modules.analysis.cdr.device_quality import (
    device_change_review,
    device_summary,
    imei_is_valid,
)
from modules.reporting.multi_cdr_excel import (
    SHEET_MAP,
    generate_multi_cdr_report,
)


def _device_frame() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "b_party": [
                "8000000001",
                "8000000002",
            ],
            "call_type": [
                "outgoing",
                "incoming",
            ],
            "call_date": [
                "10-01-2026",
                "10-01-2026",
            ],
            "call_time": [
                "10:00:00",
                "11:00:00",
            ],
            "call_duration": [
                30,
                40,
            ],
            "first_cell_id": [
                "404-55-113-12101",
                "404-55-113-12101",
            ],
            "imei": [
                "868116066643170",
                "868116066643177",
            ],
            "imsi": [
                "405001111111111",
                "405001111111111",
            ],
        }
    )


def test_hourly_activity_is_chronological():
    frame = pd.DataFrame(
        {
            "call_date": [
                "10-01-2026",
                "10-01-2026",
                "10-01-2026",
            ],
            "call_time": [
                "06:10:00",
                "06:20:00",
                "18:30:00",
            ],
        }
    )

    result = hourly_activity(frame)

    assert len(result) == 24
    assert result.iloc[0]["Time Window"] == "00:00-00:59"
    assert result.iloc[23]["Time Window"] == "23:00-23:59"

    peak = result.loc[result["Hour"].eq(6)].iloc[0]

    assert int(peak["Total Events"]) == 2
    assert int(peak["Activity Rank"]) == 1


def test_weekly_activity_has_one_date_range():
    frame = pd.DataFrame(
        {
            "call_date": [
                "16-03-2026",
                "17-03-2026",
                "22-03-2026",
            ],
            "call_time": [
                "10:00:00",
                "11:00:00",
                "12:00:00",
            ],
        }
    )

    result = weekly_activity(frame)

    assert list(result.columns) == [
        "ISO Year",
        "ISO Week",
        "Year-Week",
        "Date Range",
        "Total Events",
        "Active Days",
        "Average Events per Active Day",
    ]
    assert result.iloc[0]["Year-Week"] == "2026-W12"
    assert result.iloc[0]["Date Range"] == (
        "16-03-2026 to 22-03-2026"
    )
    assert int(result.iloc[0]["Active Days"]) == 3



def test_invalid_and_valid_values_share_one_device_group():
    assert not imei_is_valid("868116066643170")
    assert imei_is_valid("868116066643177")

    result = device_summary(_device_frame())

    assert len(result) == 1

    row = result.iloc[0]

    assert row["Device Key"] == "86811606664317"
    assert row["Valid IMEI"] == "868116066643177"
    assert row["Invalid IMEI Values"] == "868116066643170"
    assert int(row["Total Events"]) == 2


def test_same_device_raw_change_is_identifier_variant():
    result = device_change_review(_device_frame())

    assert len(result) == 1

    row = result.iloc[0]

    assert row["Change Type"] == "Identifier Variant"
    assert row["Old Device Key"] == row["New Device Key"]
    assert row["Old IMSI"] == ""
    assert row["New IMSI"] == ""


def test_device_only_change_hides_unchanged_imsi():
    frame = _device_frame()
    frame.loc[1, "imei"] = "490154203237518"

    result = device_change_review(frame)

    assert len(result) == 1

    row = result.iloc[0]

    assert row["Change Type"] == "Device Change"
    assert row["Old Device Key"] != row["New Device Key"]
    assert row["Old IMSI"] == ""
    assert row["New IMSI"] == ""


def test_sim_only_change_preserves_imsi_transition():
    frame = _device_frame()
    frame.loc[1, "imei"] = frame.loc[0, "imei"]
    frame.loc[1, "imsi"] = "405002222222222"

    result = device_change_review(frame)

    assert len(result) == 1

    row = result.iloc[0]

    assert row["Change Type"] == "SIM Change"
    assert row["Old IMSI"] == "405001111111111"
    assert row["New IMSI"] == "405002222222222"


def test_device_and_sim_change_preserves_imsi_transition():
    frame = _device_frame()
    frame.loc[1, "imei"] = "490154203237518"
    frame.loc[1, "imsi"] = "405002222222222"

    result = device_change_review(frame)

    assert len(result) == 1

    row = result.iloc[0]

    assert row["Change Type"] == "Device and SIM Change"
    assert row["Old Device Key"] != row["New Device Key"]
    assert row["Old IMSI"] == "405001111111111"
    assert row["New IMSI"] == "405002222222222"


def test_contact_summary_has_map_ready_fields(monkeypatch):
    lookup = pd.DataFrame(
        [
            {
                "cgi": "404-55-113-12101",
                "operator": "AIRTEL",
                "circle": "Bihar",
                "state": "Bihar",
                "district": "Patna",
                "police_station": "",
                "town": "Patna",
                "site_name": "Test Site",
                "address": "Test Tower Address",
                "latitude": 25.61,
                "longitude": 85.14,
                "source_file": "test.xlsx",
            }
        ]
    )

    monkeypatch.setattr(
        "modules.analysis.cdr.contact_report.lookup_cgi_addresses",
        lambda values: lookup,
    )

    contacts = pd.DataFrame(
        [
            {
                "Other Party": "8000000001",
                "Name": "Test Person",
                "Total Calls": 1,
            }
        ]
    )

    result = build_full_contact_summary(
        _device_frame().iloc[[0]].copy(),
        contacts,
    )
    row = result.iloc[0]

    assert row["Most Used Target CGI"] == "404-55-113-12101"
    assert row["Most Used Tower Address"] == "Test Tower Address"
    assert float(row["Most Used Latitude"]) == 25.61
    assert float(row["Most Used Longitude"]) == 85.14


def _empty_bundle() -> dict[str, pd.DataFrame]:
    return {
        result_key: pd.DataFrame()
        for _, _, result_key in SHEET_MAP
    }


def _targets() -> dict:
    return {
        "9000000001": {"df": pd.DataFrame()},
        "9000000002": {"df": pd.DataFrame()},
    }


def test_empty_multiple_error_sheet_is_omitted(tmp_path: Path):
    report = generate_multi_cdr_report(
        _targets(),
        metadata={"case_name": "Test Case"},
        analysis_bundle=_empty_bundle(),
        output_dir=tmp_path,
    )

    assert report is not None

    workbook = load_workbook(report, read_only=True)

    assert "14. Errors" not in workbook.sheetnames
    assert "14. Rejected Rows" in workbook.sheetnames


def test_nonempty_multiple_error_sheet_is_preserved(tmp_path: Path):
    bundle = _empty_bundle()
    bundle["errors"] = pd.DataFrame(
        [{"Analysis": "test", "Error": "sample"}]
    )

    report = generate_multi_cdr_report(
        _targets(),
        metadata={"case_name": "Test Case"},
        analysis_bundle=bundle,
        output_dir=tmp_path,
    )

    assert report is not None

    workbook = load_workbook(report, read_only=True)

    assert "14. Errors" in workbook.sheetnames
    assert "15. Rejected Rows" in workbook.sheetnames
