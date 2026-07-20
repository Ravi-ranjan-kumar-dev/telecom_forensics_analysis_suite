from __future__ import annotations

from pathlib import Path

import pandas as pd

from modules.reporting.tower_partition_excel import (
    generate_tower_partition_excel_report,
)


def test_consolidated_partition_workbook(tmp_path: Path):
    partitions = ["P1", "P2", "P3", "P4", "P5"]

    subscriber_presence = pd.DataFrame(
        [
            {
                "subscriber_number": "9000000001",
                "match_count": 5,
                "total_sightings": 5,
                "match_ratio": "5/5",
                "matched_sightings": ", ".join(partitions),
                "matched_locations": "",
                "total_events": 9,
                "operators": "Airtel, Jio",
                "first_seen": pd.Timestamp("2026-07-10 12:55:00"),
                "last_seen": pd.Timestamp("2026-07-10 16:02:00"),
                **{partition: 1 for partition in partitions},
            },
            {
                "subscriber_number": "9000000002",
                "match_count": 3,
                "total_sightings": 5,
                "match_ratio": "3/5",
                "matched_sightings": "P1, P2, P4",
                "matched_locations": "",
                "total_events": 4,
                "operators": "Vi",
                "first_seen": pd.Timestamp("2026-07-10 12:58:00"),
                "last_seen": pd.Timestamp("2026-07-10 15:58:00"),
                "P1": 1,
                "P2": 1,
                "P3": 0,
                "P4": 1,
                "P5": 0,
            },
        ]
    )

    result = {
        "partition_summary": pd.DataFrame(
            [
                {
                    "sighting_id": partition,
                    "cctv_timestamp": f"2026-07-10 {12 + index}:00:00",
                    "window_start": f"2026-07-10 {11 + index}:50:00",
                    "window_end": f"2026-07-10 {12 + index}:10:00",
                    "filtered_records": 100 + index,
                    "unique_subscribers": 50 + index,
                    "unique_imei": 40 + index,
                    "unique_imsi": 45 + index,
                    "unique_searched_cells": 4,
                }
                for index, partition in enumerate(partitions, start=1)
            ]
        ),
        "subscriber_presence": subscriber_presence,
        "n_of_m_candidates": subscriber_presence.copy(),
        "strict_common_candidates": subscriber_presence.iloc[[0]].copy(),
        "imei_presence": pd.DataFrame(
            [{"imei": "123456789012345", "match_ratio": "5/5", **{p: 1 for p in partitions}}]
        ),
        "imsi_presence": pd.DataFrame(
            [{"imsi": "404000000000001", "match_ratio": "5/5", **{p: 1 for p in partitions}}]
        ),
        "total_sightings": 5,
        "total_input_records": 5000,
        "warnings": [],
        "errors": [],
        "operators": ["Airtel", "Jio", "Vi"],
        "cell_ids": ["A", "B", "C", "D"],
        "load_metadata": {
            "files_found": 4,
            "files_loaded": 4,
            "files_failed": 0,
        },
    }

    sightings = [
        {
            "sighting_id": partition,
            "cctv_timestamp": f"2026-07-10 {12 + index}:00:00",
            "window_start": f"2026-07-10 {11 + index}:50:00",
            "window_end": f"2026-07-10 {12 + index}:10:00",
            "minutes_before": 10,
            "minutes_after": 10,
            "cgi_group_id": "AUTO_ALL",
        }
        for index, partition in enumerate(partitions, start=1)
    ]

    report = generate_tower_partition_excel_report(
        result,
        case={
            "case_id": "CASE-20260710-001",
            "case_name": "Test Case",
        },
        sightings=sightings,
        output_dir=tmp_path,
        input_folder=tmp_path / "input",
        saved={
            "run_id": "partition_test",
            "run_directory": str(tmp_path / "backend"),
        },
    )

    assert report.is_file()

    from openpyxl import load_workbook

    workbook = load_workbook(report, read_only=True)
    names = workbook.sheetnames

    assert "1. Executive Summary" in names
    assert "3. Partition Summary" in names
    assert "5. N-of-M Candidates" in names
    assert "6. Strict Common" in names
    assert "9. Candidate Matrix" in names
    assert "10. Visitor Intelligence" in names
    assert "11. New Visitors" in names
    assert "12. Rare Visitors" in names
    assert "13. Repeat Relevant" in names
    assert "14. Regular Local" in names
    assert "15. Multi-Cell Relevant" in names
    assert "16. Visitor Priority Leads" in names
    assert "17. Analysis Status" in names
    assert "18. Warnings" in names
    assert "19. Partition Status" in names
    assert "20. Rejected Rows" in names
    assert "Methodology & Limits" in names


# TOWER_PARTITION_EXCEL_PRODUCTION_FREEZE_REGRESSION

def test_tower_partition_excel_freeze_contract(
    tmp_path: Path,
    monkeypatch,
):
    """Freeze visitor, scope, SDR and workbook report contracts."""

    from openpyxl import load_workbook

    import modules.enrichment.sdr_subscriber_enrichment as sdr_enrichment

    visitor_rows = pd.DataFrame(
        [
            {
                "partition_id": "P1",
                "partition_location": "Part 1",
                "partition_window_start": "2026-06-11 10:00:00",
                "partition_window_end": "2026-06-11 10:30:00",
                "partition_cgi_group_id": "AUTO_ACTIVE",
                "scope_mode": "AUTO_ACTIVE_CELLS",
                "scope_confidence": "LOW",
                "location_confirmed": "NO",
                "scope_basis": "All loaded cells were active.",
                "resolved_cell_count": 2,
                "resolved_cells": "CELL-A, CELL-B",
                "loaded_cell_count": 2,
                "subscriber_number": "9000000001",
                "visitor_type": "NEW VISITOR",
                "current_seen_count": 2,
                "baseline_seen_count": 0,
                "cells_seen": 2,
                "imei_count": 1,
                "imsi_count": 1,
                "first_seen": pd.Timestamp("2026-06-11 10:00:00"),
                "last_seen": pd.Timestamp("2026-06-11 10:20:00"),
                "rarity_score": 100,
                "priority": "HIGH",
                "confidence": "HIGH",
                "multi_cell_relevant": "YES",
                "why_important": "New multi-cell visitor.",
                "next_verification": "Verify SDR and field context.",
                "rank_reason": "INTERNAL RANK REASON",
                "investigation_hint": "INTERNAL INVESTIGATION HINT",
            },
            {
                "partition_id": "P1",
                "partition_location": "Part 1",
                "partition_window_start": "2026-06-11 10:00:00",
                "partition_window_end": "2026-06-11 10:30:00",
                "partition_cgi_group_id": "AUTO_ACTIVE",
                "scope_mode": "AUTO_ACTIVE_CELLS",
                "scope_confidence": "LOW",
                "location_confirmed": "NO",
                "scope_basis": "All loaded cells were active.",
                "resolved_cell_count": 2,
                "resolved_cells": "CELL-A, CELL-B",
                "loaded_cell_count": 2,
                "subscriber_number": "9000000002",
                "visitor_type": "RARE VISITOR",
                "current_seen_count": 1,
                "baseline_seen_count": 1,
                "cells_seen": 1,
                "imei_count": 1,
                "imsi_count": 1,
                "first_seen": pd.Timestamp("2026-06-11 10:05:00"),
                "last_seen": pd.Timestamp("2026-06-11 10:05:00"),
                "rarity_score": 80,
                "priority": "MEDIUM",
                "confidence": "MEDIUM",
                "multi_cell_relevant": "NO",
                "why_important": "Rare baseline presence.",
                "next_verification": "Verify local presence.",
                "rank_reason": "INTERNAL RANK REASON",
                "investigation_hint": "INTERNAL INVESTIGATION HINT",
            },
            {
                "partition_id": "P2",
                "partition_location": "Part 2",
                "partition_window_start": "2026-06-11 11:00:00",
                "partition_window_end": "2026-06-11 11:30:00",
                "partition_cgi_group_id": "AUTO_ACTIVE",
                "scope_mode": "AUTO_ACTIVE_CELLS",
                "scope_confidence": "LOW",
                "location_confirmed": "NO",
                "scope_basis": "All loaded cells were active.",
                "resolved_cell_count": 2,
                "resolved_cells": "CELL-A, CELL-B",
                "loaded_cell_count": 2,
                "subscriber_number": "9000000003",
                "visitor_type": "REPEAT RELEVANT VISITOR",
                "current_seen_count": 3,
                "baseline_seen_count": 3,
                "cells_seen": 2,
                "imei_count": 1,
                "imsi_count": 1,
                "first_seen": pd.Timestamp("2026-06-11 11:00:00"),
                "last_seen": pd.Timestamp("2026-06-11 11:25:00"),
                "rarity_score": 60,
                "priority": "MEDIUM",
                "confidence": "HIGH",
                "multi_cell_relevant": "YES",
                "why_important": "Repeated activity in selected Part.",
                "next_verification": "Verify movement and communication.",
                "rank_reason": "INTERNAL RANK REASON",
                "investigation_hint": "INTERNAL INVESTIGATION HINT",
            },
            {
                "partition_id": "P2",
                "partition_location": "Part 2",
                "partition_window_start": "2026-06-11 11:00:00",
                "partition_window_end": "2026-06-11 11:30:00",
                "partition_cgi_group_id": "AUTO_ACTIVE",
                "scope_mode": "AUTO_ACTIVE_CELLS",
                "scope_confidence": "LOW",
                "location_confirmed": "NO",
                "scope_basis": "All loaded cells were active.",
                "resolved_cell_count": 2,
                "resolved_cells": "CELL-A, CELL-B",
                "loaded_cell_count": 2,
                "subscriber_number": "9000000004",
                "visitor_type": "REGULAR / LOCAL PRESENCE",
                "current_seen_count": 1,
                "baseline_seen_count": 20,
                "cells_seen": 1,
                "imei_count": 1,
                "imsi_count": 1,
                "first_seen": pd.Timestamp("2026-06-11 11:10:00"),
                "last_seen": pd.Timestamp("2026-06-11 11:10:00"),
                "rarity_score": 10,
                "priority": "LOW",
                "confidence": "LOW",
                "multi_cell_relevant": "NO",
                "why_important": "Regular baseline presence.",
                "next_verification": "Normally lower investigative priority.",
                "rank_reason": "INTERNAL RANK REASON",
                "investigation_hint": "INTERNAL INVESTIGATION HINT",
            },
        ]
    )

    def fake_lookup_sdr_subscribers(numbers):
        requested = {
            str(value)
            for value in numbers
        }

        rows = []

        if "9000000001" in requested:
            rows.append(
                {
                    "lookup_mobile": "9000000001",
                    "subscriber_name": (
                        '=HYPERLINK("https://example.invalid","OPEN")'
                    ),
                    "father_name": "TEST FATHER 1",
                    "subscriber_address": "TEST ADDRESS 1",
                    "operator": "AIRTEL",
                    "circle": "BIHAR",
                    "activation_date": "2024-01-01",
                    "caf_number": "CAF-1",
                    "sdr_found": "Yes",
                }
            )

        if "9000000003" in requested:
            rows.append(
                {
                    "lookup_mobile": "9000000003",
                    "subscriber_name": "TEST SUBSCRIBER 3",
                    "father_name": "TEST FATHER 3",
                    "subscriber_address": "TEST ADDRESS 3",
                    "operator": "AIRTEL",
                    "circle": "BIHAR",
                    "activation_date": "2023-03-03",
                    "caf_number": "CAF-3",
                    "sdr_found": "Yes",
                }
            )

        return pd.DataFrame(rows)

    monkeypatch.setattr(
        sdr_enrichment,
        "lookup_sdr_subscribers",
        fake_lookup_sdr_subscribers,
    )

    partition_summary = pd.DataFrame(
        [
            {
                "sighting_id": "P1",
                "location_name": "Part 1",
                "cctv_timestamp": "2026-06-11 10:00:00",
                "window_start": "2026-06-11 10:00:00",
                "window_end": "2026-06-11 10:30:00",
                "cgi_group_id": "AUTO_ACTIVE",
                "scope_mode": "AUTO_ACTIVE_CELLS",
                "scope_confidence": "LOW",
                "location_confirmed": "NO",
                "scope_basis": "All loaded cells were active.",
                "loaded_cell_count": 2,
                "cgi_count": 2,
                "resolved_cgi_values": "CELL-A, CELL-B",
                "filtered_records": 4,
                "unique_subscribers": 2,
                "unique_imei": 2,
                "unique_imsi": 2,
                "unique_searched_cells": 2,
            },
            {
                "sighting_id": "P2",
                "location_name": "Part 2",
                "cctv_timestamp": "2026-06-11 11:00:00",
                "window_start": "2026-06-11 11:00:00",
                "window_end": "2026-06-11 11:30:00",
                "cgi_group_id": "AUTO_ACTIVE",
                "scope_mode": "AUTO_ACTIVE_CELLS",
                "scope_confidence": "LOW",
                "location_confirmed": "NO",
                "scope_basis": "All loaded cells were active.",
                "loaded_cell_count": 2,
                "cgi_count": 2,
                "resolved_cgi_values": "CELL-A, CELL-B",
                "filtered_records": 4,
                "unique_subscribers": 2,
                "unique_imei": 2,
                "unique_imsi": 2,
                "unique_searched_cells": 2,
            },
        ]
    )

    subscriber_presence = pd.DataFrame(
        [
            {
                "subscriber_number": number,
                "match_count": 1,
                "total_sightings": 2,
                "match_ratio": "1/2",
                "matched_sightings": partition,
                "matched_locations": partition,
                "total_events": 1,
                "operators": "airtel",
                "first_seen": pd.Timestamp(
                    "2026-06-11 10:00:00"
                ),
                "last_seen": pd.Timestamp(
                    "2026-06-11 11:10:00"
                ),
                "P1": int(partition == "P1"),
                "P2": int(partition == "P2"),
            }
            for number, partition in (
                ("9000000001", "P1"),
                ("9000000002", "P1"),
                ("9000000003", "P2"),
                ("9000000004", "P2"),
            )
        ]
    )

    result = {
        "partition_summary": partition_summary,
        "partition_status": partition_summary[
            [
                "sighting_id",
                "scope_mode",
                "scope_confidence",
                "location_confirmed",
            ]
        ].assign(
            status="VALID_AUTO_ACTIVE_CELLS",
            included=True,
        ),
        "rejected_rows": pd.DataFrame(),
        "subscriber_presence": subscriber_presence,
        "n_of_m_candidates": subscriber_presence.head(0).copy(),
        "strict_common_candidates": subscriber_presence.head(0).copy(),
        "imei_presence": pd.DataFrame(
            columns=["imei", "match_ratio", "P1", "P2"]
        ),
        "imsi_presence": pd.DataFrame(
            columns=["imsi", "match_ratio", "P1", "P2"]
        ),
        "partition_visitor_intelligence": visitor_rows,
        "new_visitors": visitor_rows.loc[
            visitor_rows["visitor_type"].eq("NEW VISITOR")
        ].copy(),
        "rare_visitors": visitor_rows.loc[
            visitor_rows["visitor_type"].eq("RARE VISITOR")
        ].copy(),
        "repeat_relevant_visitors": visitor_rows.loc[
            visitor_rows["visitor_type"].eq(
                "REPEAT RELEVANT VISITOR"
            )
        ].copy(),
        "regular_local_presence": visitor_rows.loc[
            visitor_rows["visitor_type"].eq(
                "REGULAR / LOCAL PRESENCE"
            )
        ].copy(),
        "multi_cell_relevant": visitor_rows.loc[
            visitor_rows["multi_cell_relevant"].eq("YES")
        ].copy(),
        "partition_priority_leads": visitor_rows.loc[
            visitor_rows["priority"].isin(
                ["HIGH", "MEDIUM"]
            )
        ].copy(),
        "total_sightings": 2,
        "total_input_records": 8,
        "warnings": [
            (
                "P1: LOW scope confidence. "
                "Location independently confirmed nahi hai."
            ),
            (
                "P2: LOW scope confidence. "
                "Location independently confirmed nahi hai."
            ),
        ],
        "errors": [],
        "operators": ["airtel"],
        "cell_ids": ["CELL-A", "CELL-B"],
        "load_metadata": {
            "files_found": 2,
            "files_loaded": 2,
            "files_failed": 0,
        },
    }

    sightings = [
        {
            "sighting_id": "P1",
            "location_name": "Part 1",
            "cctv_timestamp": "2026-06-11 10:00:00",
            "window_start": "2026-06-11 10:00:00",
            "window_end": "2026-06-11 10:30:00",
            "cgi_group_id": "AUTO_ACTIVE",
        },
        {
            "sighting_id": "P2",
            "location_name": "Part 2",
            "cctv_timestamp": "2026-06-11 11:00:00",
            "window_start": "2026-06-11 11:00:00",
            "window_end": "2026-06-11 11:30:00",
            "cgi_group_id": "AUTO_ACTIVE",
        },
    ]

    report = generate_tower_partition_excel_report(
        result,
        case={
            "case_id": "CASE-FREEZE-001",
            "case_name": "Production Freeze Test",
        },
        sightings=sightings,
        output_dir=tmp_path,
        input_folder=tmp_path / "input",
        saved={
            "run_id": "production_freeze_test",
            "run_directory": str(tmp_path / "backend"),
        },
    )

    workbook = load_workbook(
        report,
        data_only=False,
    )

    required_sheets = {
        "1. Executive Summary",
        "2. Partition Windows",
        "3. Partition Summary",
        "4. Subscriber Presence",
        "5. N-of-M Candidates",
        "6. Strict Common",
        "7. IMEI Continuity",
        "8. IMSI Continuity",
        "9. Candidate Matrix",
        "10. Visitor Intelligence",
        "11. New Visitors",
        "12. Rare Visitors",
        "13. Repeat Relevant",
        "14. Regular Local",
        "15. Multi-Cell Relevant",
        "16. Visitor Priority Leads",
        "17. Analysis Status",
        "18. Warnings",
        "19. Partition Status",
        "20. Rejected Rows",
        "Methodology & Limits",
    }

    assert required_sheets.issubset(
        set(workbook.sheetnames)
    )

    assert all(
        len(name) <= 31
        for name in workbook.sheetnames
    )

    visitor_sheet_counts = {
        "10. Visitor Intelligence": 4,
        "11. New Visitors": 1,
        "12. Rare Visitors": 1,
        "13. Repeat Relevant": 1,
        "14. Regular Local": 1,
        "15. Multi-Cell Relevant": 2,
        "16. Visitor Priority Leads": 3,
    }

    required_visitor_headers = {
        "Partition",
        "CGI Group",
        "Scope Mode",
        "Scope Confidence",
        "Location Confirmed",
        "Scope Basis",
        "Resolved Cell Count",
        "Resolved Cells",
        "Loaded Cell Count",
        "Mobile Number",
        "SDR Found",
        "Subscriber Name",
        "Father / Husband Name",
        "Full Address",
        "SDR Operator",
        "Circle",
        "Activation Date",
        "CAF Number",
        "Visitor Type",
        "Priority",
        "Confidence",
        "Why It Matters",
        "Suggested Verification",
    }

    prohibited_headers = {
        "rank_reason",
        "investigation_hint",
        "Ranking Reason",
        "Investigation Hint",
        "reason",
        "source_module",
        "priority_level",
    }

    for sheet_name, expected_rows in (
        visitor_sheet_counts.items()
    ):
        worksheet = workbook[sheet_name]

        headers = [
            worksheet.cell(4, column).value
            for column in range(
                1,
                worksheet.max_column + 1,
            )
            if worksheet.cell(
                4,
                column,
            ).value is not None
        ]

        assert required_visitor_headers.issubset(
            set(headers)
        )

        assert prohibited_headers.isdisjoint(
            set(headers)
        )

        assert worksheet.max_row - 4 == expected_rows

    intelligence = workbook[
        "10. Visitor Intelligence"
    ]

    intelligence_headers = {
        intelligence.cell(4, column).value: column
        for column in range(
            1,
            intelligence.max_column + 1,
        )
    }

    visitor_keys = [
        (
            intelligence.cell(
                row,
                intelligence_headers["Partition"],
            ).value,
            intelligence.cell(
                row,
                intelligence_headers["Mobile Number"],
            ).value,
        )
        for row in range(
            5,
            intelligence.max_row + 1,
        )
    ]

    assert len(visitor_keys) == len(
        set(visitor_keys)
    )

    priority_sheet = workbook[
        "16. Visitor Priority Leads"
    ]

    priority_headers = {
        priority_sheet.cell(4, column).value: column
        for column in range(
            1,
            priority_sheet.max_column + 1,
        )
    }

    priorities = {
        priority_sheet.cell(
            row,
            priority_headers["Priority"],
        ).value
        for row in range(
            5,
            priority_sheet.max_row + 1,
        )
    }

    assert priorities.issubset(
        {"HIGH", "MEDIUM"}
    )

    # Missing SDR records must not receive a false identity.
    for row in range(
        5,
        intelligence.max_row + 1,
    ):
        sdr_found = intelligence.cell(
            row,
            intelligence_headers["SDR Found"],
        ).value

        if sdr_found == "No":
            assert intelligence.cell(
                row,
                intelligence_headers["Subscriber Name"],
            ).value in (None, "")

            assert intelligence.cell(
                row,
                intelligence_headers[
                    "Father / Husband Name"
                ],
            ).value in (None, "")

            assert intelligence.cell(
                row,
                intelligence_headers["Full Address"],
            ).value in (None, "")

    # Formula-like SDR identity values must remain literal text.
    formula_name_cell = intelligence.cell(
        5,
        intelligence_headers["Subscriber Name"],
    )

    assert formula_name_cell.value.startswith(
        "'=HYPERLINK"
    )
    assert formula_name_cell.data_type != "f"

    executive = workbook[
        "1. Executive Summary"
    ]

    executive_values = {
        executive.cell(row, 1).value: executive.cell(
            row,
            2,
        ).value
        for row in range(
            1,
            executive.max_row + 1,
        )
    }

    assert executive_values[
        "Valid Partitions"
    ] == 2
    assert executive_values[
        "Auto Active-Cell Partitions"
    ] == 2
    assert executive_values[
        "Location-Confirmed Partitions"
    ] == 0
    assert executive_values[
        "Low Scope Confidence"
    ] == 2
    assert executive_values[
        "Visitor Classification Rows"
    ] == 4
    assert executive_values[
        "Unique Visitor Mobile Numbers"
    ] == 4
    assert executive_values[
        "Visitor SDR Profiles Found"
    ] == 2
    assert executive_values[
        "Visitor SDR Profiles Not Found"
    ] == 2
    assert executive_values[
        "New Visitors"
    ] == 1
    assert executive_values[
        "Rare Visitors"
    ] == 1
    assert executive_values[
        "Repeat Relevant Visitors"
    ] == 1
    assert executive_values[
        "Regular / Local Presence"
    ] == 1
    assert executive_values[
        "Multi-Cell Relevant Visitors"
    ] == 2
    assert executive_values[
        "Priority Visitor Leads"
    ] == 3

    warning_sheet = workbook[
        "18. Warnings"
    ]

    warning_text = " ".join(
        str(cell.value or "")
        for row in warning_sheet.iter_rows()
        for cell in row
    )

    assert "P1: LOW scope confidence" in warning_text
    assert "P2: LOW scope confidence" in warning_text
