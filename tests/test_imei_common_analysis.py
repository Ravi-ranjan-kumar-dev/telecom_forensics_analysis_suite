from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from modules.analysis.device.imei_common import (
    build_common_imei_cdr_analysis,
)
from modules.reporting.imei_device_excel import (
    generate_imei_common_report,
)


def _frame(
    *,
    query: str,
    observed: str,
    target: str,
    imsi: str,
    contact: str,
    cell: str,
    source_file: str,
) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "query_identifier_normalized": query,
                "observed_imei_normalized": observed,
                "imei": observed,
                "target": target,
                "call_date": "01/07/2026",
                "call_time": "10:00:00",
                "call_type": "outgoing",
                "b_party": contact,
                "call_duration": 30,
                "imsi": imsi,
                "first_cell_id": cell,
                "last_cell_id": "",
                "source_file": source_file,
                "source_path": f"/evidence/{source_file}",
                "source_row_number": 8,
                "match_relation": "SAME_BASE14",
            }
        ]
    )


def _analysis() -> dict:
    first = "862261072892730"
    second = "866284043482077"

    manifest = pd.DataFrame(
        [
            {
                "Query Identifier": first,
                "SHA-256": "a" * 64,
                "Source File": "first.csv",
            },
            {
                "Query Identifier": second,
                "SHA-256": "b" * 64,
                "Source File": "second.csv",
            },
        ]
    )

    return build_common_imei_cdr_analysis(
        {
            first: _frame(
                query=first,
                observed="8622610728927300",
                target="9000000001",
                imsi="405520123456789",
                contact="9111111111",
                cell="404-10-100-200",
                source_file="first.csv",
            ),
            second: _frame(
                query=second,
                observed="866284043482077",
                target="9000000001",
                imsi="405520123456789",
                contact="9111111111",
                cell="404-10-100-200",
                source_file="second.csv",
            ),
        },
        manifest,
    )


def test_common_imei_analysis_finds_shared_evidence():
    result = _analysis()

    assert result[
        "status"
    ] == "FOUND"

    assert result[
        "device_count"
    ] == 2

    assert set(
        result[
            "common_targets"
        ][
            "Target Number"
        ]
    ) == {
        "9000000001"
    }

    assert set(
        result[
            "common_imsis"
        ][
            "IMSI"
        ]
    ) == {
        "405520123456789"
    }

    assert set(
        result[
            "common_contacts"
        ][
            "Contact Number"
        ]
    ) == {
        "9111111111"
    }

    assert set(
        result[
            "common_towers"
        ][
            "Cell ID"
        ]
    ) == {
        "404-10-100-200"
    }


def test_common_analysis_requires_multiple_identifiers():
    result = build_common_imei_cdr_analysis(
        {
            "862261072892730": pd.DataFrame(),
        }
    )

    assert result[
        "status"
    ] == "NOT_APPLICABLE"

    assert result[
        "device_count"
    ] == 1


def test_common_report_uses_compact_sheet_contract(
    tmp_path: Path,
):
    path = generate_imei_common_report(
        case={
            "case_id": "CASE-001",
        },
        analysis=_analysis(),
        output_dir=tmp_path,
    )

    assert path is not None

    workbook = load_workbook(
        path,
        data_only=False,
    )

    assert workbook.sheetnames == [
        "1. Device Overview",
        "2. Common Targets",
        "3. Common IMSIs",
        "4. Common Contacts",
        "5. Shared Service IDs",
        "6. Common Towers",
        "7. Cross Device Timeline",
        "8. Acquisition Manifest",
        "9. Review Indicators",
        "10. Data Quality",
    ]

    exact_values = {
        "862261072892730",
        "866284043482077",
        "a" * 64,
    }

    matched_cells = {
        value: []
        for value in exact_values
    }

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value in exact_values:
                    matched_cells[
                        cell.value
                    ].append(
                        cell
                    )

    for value, cells in matched_cells.items():
        assert cells, value

        assert all(
            cell.data_type == "s"
            and cell.number_format == "@"
            for cell in cells
        )

    workbook.close()
