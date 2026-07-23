"""DuckDB staging importer for Tower IPDR/NAT dumps.

Production goal:
- Do not concatenate all Tower IPDR files into one huge pandas DataFrame.
- Load one source file at a time.
- Normalize with the existing verified loader.
- Append normalized events into DuckDB.
- Save a resumable manifest.
- Run later analysis using SQL instead of full pandas memory.
"""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd

from modules.loader.tower_ipdr_loader import (
    NORMALIZED_COLUMNS,
    SUPPORTED_SUFFIXES,
    load_tower_ipdr_file,
)
from modules.loader.tower_spot_layout import (
    build_tower_spot_layout,
)
from modules.staging.duckdb_store import DuckDBStore
from modules.staging.manifest import calculate_sha256


TABLE_EVENTS = "tower_ipdr_events"
TABLE_FILE_SUMMARY = "tower_ipdr_file_summary"

SPOT_PROVENANCE_COLUMNS = (
    "source_relative_path",
    "spot_id",
    "spot_name",
    "spot_folder",
)


def _now_iso() -> str:
    return datetime.now().replace(microsecond=0).isoformat()


def _json_safe(value: Any) -> Any:
    """Convert pandas/path/datetime values into JSON-safe values."""

    if value is None:
        return None

    if isinstance(value, Path):
        return str(value)

    if isinstance(value, (pd.Timestamp, datetime)):
        if pd.isna(value):
            return ""
        return value.isoformat(sep=" ")

    if isinstance(value, dict):
        return {
            str(key): _json_safe(item)
            for key, item in value.items()
        }

    if isinstance(value, list):
        return [
            _json_safe(item)
            for item in value
        ]

    if isinstance(value, tuple):
        return [
            _json_safe(item)
            for item in value
        ]

    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass

    return value


def tower_ipdr_staging_root(case_id: str) -> Path:
    return (
        Path("cases")
        / "active"
        / str(case_id)
        / "staging"
        / "tower_ipdr"
    )


def tower_ipdr_database_path(case_id: str) -> Path:
    return tower_ipdr_staging_root(case_id) / "tower_ipdr.duckdb"


def tower_ipdr_manifest_path(case_id: str) -> Path:
    return tower_ipdr_staging_root(case_id) / "manifest.json"


def _read_manifest(case_id: str) -> dict[str, Any]:
    path = tower_ipdr_manifest_path(case_id)

    if not path.exists():
        return {
            "case_id": str(case_id),
            "staging_type": "TOWER_IPDR_DUCKDB",
            "created_at": _now_iso(),
            "updated_at": _now_iso(),
            "files": [],
        }

    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {
            "case_id": str(case_id),
            "staging_type": "TOWER_IPDR_DUCKDB",
            "created_at": _now_iso(),
            "updated_at": _now_iso(),
            "files": [],
            "manifest_warning": "Previous manifest could not be read.",
        }


def _write_manifest(case_id: str, manifest: dict[str, Any]) -> None:
    root = tower_ipdr_staging_root(case_id)
    root.mkdir(parents=True, exist_ok=True)

    manifest["updated_at"] = _now_iso()

    path = tower_ipdr_manifest_path(case_id)
    tmp_path = path.with_suffix(path.suffix + ".tmp")

    tmp_path.write_text(
        json.dumps(
            _json_safe(manifest),
            ensure_ascii=False,
            indent=2,
            sort_keys=True,
        ),
        encoding="utf-8",
    )

    tmp_path.replace(path)


def _manifest_file_map(manifest: dict[str, Any]) -> dict[str, dict[str, Any]]:
    records = manifest.get("files", [])

    if not isinstance(records, list):
        return {}

    result: dict[str, dict[str, Any]] = {}

    for item in records:
        if not isinstance(item, dict):
            continue

        sha256 = str(item.get("sha256", "")).strip()
        source_path = str(item.get("source_path", "")).strip()

        key = sha256 or source_path

        if key:
            result[key] = item

    return result


def _upsert_manifest_file(
    manifest: dict[str, Any],
    record: dict[str, Any],
) -> None:
    files = manifest.setdefault("files", [])

    if not isinstance(files, list):
        manifest["files"] = []
        files = manifest["files"]

    record_key = str(record.get("sha256") or record.get("source_path") or "")

    for index, existing in enumerate(files):
        if not isinstance(existing, dict):
            continue

        existing_key = str(
            existing.get("sha256")
            or existing.get("source_path")
            or ""
        )

        if existing_key == record_key:
            files[index] = record
            return

    files.append(record)


def _candidate_files(input_folder: str | Path, recursive: bool = True) -> list[Path]:
    folder = Path(input_folder).expanduser().resolve()

    if not folder.is_dir():
        raise FileNotFoundError(f"Input folder not found: {folder}")

    iterator = folder.rglob("*") if recursive else folder.glob("*")

    return sorted(
        path
        for path in iterator
        if path.is_file()
        and path.suffix.lower() in SUPPORTED_SUFFIXES
    )


def _file_summary_frame(records: list[dict[str, Any]]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []

    for record in records:
        metadata = record.get("metadata", {})

        if not isinstance(metadata, dict):
            metadata = {}

        rows.append(
            {
                "file_name": record.get("file_name", ""),
                "source_path": record.get("source_path", ""),
                "source_relative_path": record.get(
                    "source_relative_path",
                    metadata.get(
                        "source_relative_path",
                        "",
                    ),
                ),
                "spot_id": record.get(
                    "spot_id",
                    metadata.get(
                        "spot_id",
                        "",
                    ),
                ),
                "spot_name": record.get(
                    "spot_name",
                    metadata.get(
                        "spot_name",
                        "",
                    ),
                ),
                "spot_folder": record.get(
                    "spot_folder",
                    metadata.get(
                        "spot_folder",
                        "",
                    ),
                ),
                "sha256": record.get("sha256", ""),
                "status": record.get("status", ""),
                "rows_loaded": record.get("rows_loaded", 0),
                "searched_cell_id": metadata.get("searched_cell_id", ""),
                "event_time_min": metadata.get("event_time_min", ""),
                "event_time_max": metadata.get("event_time_max", ""),
                "unique_subscribers": metadata.get("unique_subscribers", 0),
                "warnings": " | ".join(record.get("warnings", []) or []),
                "errors": " | ".join(record.get("errors", []) or []),
                "loaded_at": record.get("loaded_at", ""),
            }
        )

    return pd.DataFrame(rows)


def import_tower_ipdr_folder_to_duckdb(
    case_id: str,
    input_folder: str | Path,
    *,
    recursive: bool = True,
    force_rebuild: bool = False,
    max_files: int | None = None,
) -> dict[str, Any]:
    """Import Tower IPDR files into case DuckDB staging.

    This is intentionally file-by-file, not folder-concat.
    """

    root = tower_ipdr_staging_root(case_id)
    root.mkdir(parents=True, exist_ok=True)

    database_path = tower_ipdr_database_path(case_id)
    manifest = _read_manifest(case_id)
    manifest_map = _manifest_file_map(manifest)

    store = DuckDBStore(database_path)

    if force_rebuild:
        store.drop_table(TABLE_EVENTS)
        store.drop_table(TABLE_FILE_SUMMARY)
        manifest = {
            "case_id": str(case_id),
            "staging_type": "TOWER_IPDR_DUCKDB",
            "created_at": _now_iso(),
            "updated_at": _now_iso(),
            "files": [],
        }
        manifest_map = {}

    input_root = Path(
        input_folder
    ).expanduser().resolve()

    files = _candidate_files(
        input_root,
        recursive=recursive,
    )

    if max_files is not None:
        files = files[: int(max_files)]

    spot_layout = build_tower_spot_layout(
        input_root,
        files,
    )
    spot_assignments = spot_layout.get(
        "assignments",
        {},
    )

    loaded_files = 0
    skipped_files = 0
    failed_files = 0
    total_rows = 0

    print(f"[+] Tower IPDR staging database: {database_path}")
    print(f"[+] Candidate files: {len(files)}")

    for index, path in enumerate(files, start=1):
        relative_path = str(
            path.relative_to(
                input_root
            )
        )

        assignment = dict(
            spot_assignments.get(
                str(path.resolve()),
                {
                    "spot_id": "UNASSIGNED-ROOT",
                    "spot_name": "ROOT_LEVEL_FILES",
                    "spot_folder": ".",
                    "source_relative_path": relative_path,
                    "is_root_file": True,
                },
            )
        )

        assignment_values = {
            column: str(
                assignment.get(
                    column,
                    "",
                )
                or ""
            )
            for column in SPOT_PROVENANCE_COLUMNS
        }

        sha256 = calculate_sha256(path)
        existing = manifest_map.get(sha256)

        if (
            existing
            and existing.get("status") == "LOADED"
            and not force_rebuild
        ):
            skipped_files += 1
            print(f"[SKIP] {index}/{len(files)} already loaded: {path.name}")
            continue

        started_at = _now_iso()
        print(f"[LOAD] {index}/{len(files)} {path.name}")

        record: dict[str, Any] = {
            "file_name": path.name,
            "source_path": str(path),
            "source_relative_path": assignment_values[
                "source_relative_path"
            ],
            "spot_id": assignment_values[
                "spot_id"
            ],
            "spot_name": assignment_values[
                "spot_name"
            ],
            "spot_folder": assignment_values[
                "spot_folder"
            ],
            "sha256": sha256,
            "status": "LOADING",
            "rows_loaded": 0,
            "started_at": started_at,
            "loaded_at": "",
            "metadata": {},
            "warnings": [],
            "errors": [],
        }

        _upsert_manifest_file(manifest, record)
        _write_manifest(case_id, manifest)

        try:
            result = load_tower_ipdr_file(path)

            record["warnings"] = list(
                result.get(
                    "warnings",
                    [],
                )
                or []
            )
            record["errors"] = list(
                result.get(
                    "errors",
                    [],
                )
                or []
            )

            result_metadata = dict(
                result.get(
                    "metadata",
                    {},
                )
                or {}
            )
            result_metadata.update(
                assignment_values
            )

            record["metadata"] = _json_safe(
                result_metadata
            )

            if not result.get("ok"):
                failed_files += 1
                record["status"] = "FAILED"
                record["loaded_at"] = _now_iso()
                _upsert_manifest_file(manifest, record)
                _write_manifest(case_id, manifest)
                print(f"[FAIL] {path.name}: {record['errors']}")
                continue

            dataframe = result.get("df")

            if not isinstance(dataframe, pd.DataFrame) or dataframe.empty:
                failed_files += 1
                record["status"] = "FAILED"
                record["errors"].append("Normalized DataFrame empty.")
                record["loaded_at"] = _now_iso()
                _upsert_manifest_file(manifest, record)
                _write_manifest(case_id, manifest)
                print(f"[FAIL] {path.name}: empty normalized data")
                continue

            for column in NORMALIZED_COLUMNS:
                if column not in dataframe.columns:
                    dataframe[column] = pd.NA

            dataframe = dataframe[
                NORMALIZED_COLUMNS
            ].copy()

            for column in SPOT_PROVENANCE_COLUMNS:
                dataframe[column] = pd.Series(
                    assignment_values[column],
                    index=dataframe.index,
                    dtype="string",
                )

            rows_written = store.write_dataframe(
                dataframe,
                TABLE_EVENTS,
                mode="append",
            )

            rows_loaded = int(rows_written or len(dataframe))

            record["status"] = "LOADED"
            record["rows_loaded"] = rows_loaded
            record["loaded_at"] = _now_iso()

            loaded_files += 1
            total_rows += rows_loaded

            _upsert_manifest_file(manifest, record)
            _write_manifest(case_id, manifest)

            print(f"[OK] {path.name}: {rows_loaded:,} row(s)")

            del dataframe

        except Exception as error:
            failed_files += 1
            record["status"] = "FAILED"
            record["errors"].append(
                f"{type(error).__name__}: {error}"
            )
            record["loaded_at"] = _now_iso()
            _upsert_manifest_file(manifest, record)
            _write_manifest(case_id, manifest)
            print(f"[ERROR] {path.name}: {type(error).__name__}: {error}")

    records = [
        item
        for item in manifest.get("files", [])
        if isinstance(item, dict)
    ]

    summary_frame = _file_summary_frame(records)

    if not summary_frame.empty:
        store.drop_table(TABLE_FILE_SUMMARY)
        store.write_dataframe(
            summary_frame,
            TABLE_FILE_SUMMARY,
            mode="append",
        )

    summary = {
        "case_id": str(case_id),
        "database_path": str(database_path),
        "manifest_path": str(tower_ipdr_manifest_path(case_id)),
        "candidate_files": len(files),
        "loaded_files": loaded_files,
        "skipped_files": skipped_files,
        "failed_files": failed_files,
        "rows_loaded_this_run": total_rows,
        "total_rows_in_database": count_tower_ipdr_events(case_id),
    }

    manifest["summary"] = summary
    _write_manifest(case_id, manifest)

    return summary


def count_tower_ipdr_events(case_id: str) -> int:
    database_path = tower_ipdr_database_path(case_id)
    store = DuckDBStore(database_path)

    if not store.table_exists(TABLE_EVENTS):
        return 0

    return int(store.row_count(TABLE_EVENTS))


def tower_ipdr_cell_counts(case_id: str) -> pd.DataFrame:
    database_path = tower_ipdr_database_path(case_id)
    store = DuckDBStore(database_path)

    if not store.table_exists(TABLE_EVENTS):
        return pd.DataFrame(
            columns=[
                "searched_cell_id",
                "event_count",
                "subscriber_count",
            ]
        )

    return store.query_df(
        f"""
        SELECT
            searched_cell_id,
            COUNT(*) AS event_count,
            COUNT(DISTINCT subscriber_number) AS subscriber_count
        FROM {TABLE_EVENTS}
        GROUP BY searched_cell_id
        ORDER BY event_count DESC
        """
    )


def tower_ipdr_time_count(
    case_id: str,
    partition_time: str,
) -> pd.DataFrame:
    database_path = tower_ipdr_database_path(case_id)
    store = DuckDBStore(database_path)

    if not store.table_exists(TABLE_EVENTS):
        return pd.DataFrame(
            columns=[
                "partition_time",
                "event_count",
                "subscriber_count",
            ]
        )

    return store.query_df(
        f"""
        SELECT
            CAST(? AS TIMESTAMP) AS partition_time,
            COUNT(*) AS event_count,
            COUNT(DISTINCT subscriber_number) AS subscriber_count
        FROM {TABLE_EVENTS}
        WHERE TRY_CAST(event_time AS TIMESTAMP) = CAST(? AS TIMESTAMP)
        """,
        [partition_time, partition_time],
    )


def tower_ipdr_uncommon_at_time(
    case_id: str,
    partition_time: str,
    *,
    limit: int = 50,
) -> pd.DataFrame:
    database_path = tower_ipdr_database_path(case_id)
    store = DuckDBStore(database_path)

    if not store.table_exists(TABLE_EVENTS):
        return pd.DataFrame()

    return store.query_df(
        f"""
        WITH current_part AS (
            SELECT
                subscriber_number,
                COUNT(*) AS current_seen_count,
                COUNT(DISTINCT searched_cell_id) AS cells_seen,
                MIN(TRY_CAST(event_time AS TIMESTAMP)) AS first_seen,
                MAX(TRY_CAST(event_time AS TIMESTAMP)) AS last_seen
            FROM {TABLE_EVENTS}
            WHERE TRY_CAST(event_time AS TIMESTAMP) = CAST(? AS TIMESTAMP)
              AND subscriber_number IS NOT NULL
              AND subscriber_number <> ''
            GROUP BY subscriber_number
        ),
        baseline AS (
            SELECT DISTINCT subscriber_number
            FROM {TABLE_EVENTS}
            WHERE TRY_CAST(event_time AS TIMESTAMP) <> CAST(? AS TIMESTAMP)
              AND subscriber_number IS NOT NULL
              AND subscriber_number <> ''
        )
        SELECT
            current_part.subscriber_number,
            current_part.current_seen_count,
            0 AS baseline_seen_count,
            current_part.first_seen,
            current_part.last_seen,
            current_part.cells_seen,
            100 AS rarity_score,
            CASE
                WHEN current_part.cells_seen >= 3 THEN 'HIGH'
                WHEN current_part.current_seen_count >= 100 THEN 'HIGH'
                WHEN current_part.cells_seen >= 2 THEN 'MEDIUM_HIGH'
                ELSE 'MEDIUM'
            END AS priority_level,
            CASE
                WHEN current_part.cells_seen >= 3 THEN 'Multi-cell exact date-time presence'
                WHEN current_part.current_seen_count >= 100 THEN 'High activity at exact date-time'
                WHEN current_part.cells_seen >= 2 THEN 'Seen on multiple cells at exact date-time'
                ELSE 'Exact date-time only presence'
            END AS rank_reason
        FROM current_part
        LEFT JOIN baseline
          ON current_part.subscriber_number = baseline.subscriber_number
        WHERE baseline.subscriber_number IS NULL
        ORDER BY
            CASE priority_level
                WHEN 'HIGH' THEN 1
                WHEN 'MEDIUM_HIGH' THEN 2
                WHEN 'MEDIUM' THEN 3
                ELSE 9
            END,
            current_seen_count DESC
        LIMIT ?
        """,
        [partition_time, partition_time, int(limit)],
    )


def tower_ipdr_minute_count(
    case_id: str,
    partition_time: str,
) -> pd.DataFrame:
    """Count events in the same minute as the entered partition time."""

    database_path = tower_ipdr_database_path(case_id)
    store = DuckDBStore(database_path)

    if not store.table_exists(TABLE_EVENTS):
        return pd.DataFrame(
            columns=[
                "partition_minute",
                "event_count",
                "subscriber_count",
            ]
        )

    return store.query_df(
        f"""
        SELECT
            DATE_TRUNC('minute', CAST(? AS TIMESTAMP)) AS partition_minute,
            COUNT(*) AS event_count,
            COUNT(DISTINCT subscriber_number) AS subscriber_count
        FROM {TABLE_EVENTS}
        WHERE DATE_TRUNC('minute', TRY_CAST(event_time AS TIMESTAMP))
              = DATE_TRUNC('minute', CAST(? AS TIMESTAMP))
        """,
        [partition_time, partition_time],
    )


def tower_ipdr_uncommon_in_minute(
    case_id: str,
    partition_time: str,
    *,
    limit: int = 50,
) -> pd.DataFrame:
    """Find uncommon subscribers in the same minute as the entered time."""

    database_path = tower_ipdr_database_path(case_id)
    store = DuckDBStore(database_path)

    if not store.table_exists(TABLE_EVENTS):
        return pd.DataFrame()

    return store.query_df(
        f"""
        WITH current_part AS (
            SELECT
                subscriber_number,
                COUNT(*) AS current_seen_count,
                COUNT(DISTINCT searched_cell_id) AS cells_seen,
                MIN(TRY_CAST(event_time AS TIMESTAMP)) AS first_seen,
                MAX(TRY_CAST(event_time AS TIMESTAMP)) AS last_seen
            FROM {TABLE_EVENTS}
            WHERE DATE_TRUNC('minute', TRY_CAST(event_time AS TIMESTAMP))
                  = DATE_TRUNC('minute', CAST(? AS TIMESTAMP))
              AND subscriber_number IS NOT NULL
              AND subscriber_number <> ''
            GROUP BY subscriber_number
        ),
        baseline AS (
            SELECT DISTINCT subscriber_number
            FROM {TABLE_EVENTS}
            WHERE DATE_TRUNC('minute', TRY_CAST(event_time AS TIMESTAMP))
                  <> DATE_TRUNC('minute', CAST(? AS TIMESTAMP))
              AND subscriber_number IS NOT NULL
              AND subscriber_number <> ''
        )
        SELECT
            current_part.subscriber_number,
            current_part.current_seen_count,
            0 AS baseline_seen_count,
            current_part.first_seen,
            current_part.last_seen,
            current_part.cells_seen,
            100 AS rarity_score,
            CASE
                WHEN current_part.cells_seen >= 3 THEN 'HIGH'
                WHEN current_part.current_seen_count >= 100 THEN 'HIGH'
                WHEN current_part.cells_seen >= 2 THEN 'MEDIUM_HIGH'
                ELSE 'MEDIUM'
            END AS priority_level,
            CASE
                WHEN current_part.cells_seen >= 3 THEN 'Multi-cell same-minute presence'
                WHEN current_part.current_seen_count >= 100 THEN 'High activity in same minute'
                WHEN current_part.cells_seen >= 2 THEN 'Seen on multiple cells in same minute'
                ELSE 'Same-minute only presence'
            END AS rank_reason
        FROM current_part
        LEFT JOIN baseline
          ON current_part.subscriber_number = baseline.subscriber_number
        WHERE baseline.subscriber_number IS NULL
        ORDER BY
            CASE priority_level
                WHEN 'HIGH' THEN 1
                WHEN 'MEDIUM_HIGH' THEN 2
                WHEN 'MEDIUM' THEN 3
                ELSE 9
            END,
            current_seen_count DESC
        LIMIT ?
        """,
        [partition_time, partition_time, int(limit)],
    )


def tower_ipdr_investigation_summary(
    case_id: str,
    partition_time: str,
    *,
    mode: str = "same_minute",
    lead_limit: int = 50,
) -> dict[str, pd.DataFrame]:
    """Simple investigation-friendly Tower IPDR partition summary.

    mode:
    - exact_second: only exact timestamp match
    - same_minute: all events in same minute as partition_time

    This function is designed for non-technical investigation output.
    """

    database_path = tower_ipdr_database_path(case_id)
    store = DuckDBStore(database_path)

    empty = {
        "summary": pd.DataFrame(),
        "lead_summary": pd.DataFrame(),
        "common_numbers": pd.DataFrame(),
        "uncommon_numbers": pd.DataFrame(),
        "multi_cell_presence": pd.DataFrame(),
        "repeat_presence": pd.DataFrame(),
        "device_consistency": pd.DataFrame(),
        "priority_leads": pd.DataFrame(),
    }

    if not store.table_exists(TABLE_EVENTS):
        return empty

    mode_value = str(mode).strip().lower()

    if mode_value not in {"exact_second", "same_minute"}:
        mode_value = "same_minute"

    if mode_value == "exact_second":
        current_filter = "TRY_CAST(event_time AS TIMESTAMP) = CAST(? AS TIMESTAMP)"
        baseline_filter = "TRY_CAST(event_time AS TIMESTAMP) <> CAST(? AS TIMESTAMP)"
        mode_label = "Exact second"
    else:
        current_filter = (
            "DATE_TRUNC('minute', TRY_CAST(event_time AS TIMESTAMP)) "
            "= DATE_TRUNC('minute', CAST(? AS TIMESTAMP))"
        )
        baseline_filter = (
            "DATE_TRUNC('minute', TRY_CAST(event_time AS TIMESTAMP)) "
            "<> DATE_TRUNC('minute', CAST(? AS TIMESTAMP))"
        )
        mode_label = "Same minute"

    summary = store.query_df(
        f"""
        WITH current_part AS (
            SELECT *
            FROM {TABLE_EVENTS}
            WHERE {current_filter}
        )
        SELECT
            CAST(? AS TIMESTAMP) AS partition_time,
            ? AS analysis_mode,
            COUNT(*) AS records_found,
            COUNT(DISTINCT subscriber_number) AS numbers_found,
            COUNT(DISTINCT searched_cell_id) AS cells_involved,
            COUNT(DISTINCT imei) AS imei_found,
            COUNT(DISTINCT imsi) AS imsi_found,
            MIN(TRY_CAST(event_time AS TIMESTAMP)) AS first_activity,
            MAX(TRY_CAST(event_time AS TIMESTAMP)) AS last_activity
        FROM current_part
        """,
        [partition_time, partition_time, mode_label],
    )

    common_numbers = store.query_df(
        f"""
        WITH current_part AS (
            SELECT
                subscriber_number,
                COUNT(*) AS selected_time_records,
                COUNT(DISTINCT searched_cell_id) AS selected_time_cells,
                COUNT(DISTINCT imei) AS selected_time_imei,
                COUNT(DISTINCT imsi) AS selected_time_imsi,
                MIN(TRY_CAST(event_time AS TIMESTAMP)) AS first_seen_selected,
                MAX(TRY_CAST(event_time AS TIMESTAMP)) AS last_seen_selected
            FROM {TABLE_EVENTS}
            WHERE {current_filter}
              AND subscriber_number IS NOT NULL
              AND subscriber_number <> ''
            GROUP BY subscriber_number
        ),
        baseline AS (
            SELECT
                subscriber_number,
                COUNT(*) AS other_time_records,
                COUNT(DISTINCT searched_cell_id) AS other_time_cells,
                MIN(TRY_CAST(event_time AS TIMESTAMP)) AS first_seen_other,
                MAX(TRY_CAST(event_time AS TIMESTAMP)) AS last_seen_other
            FROM {TABLE_EVENTS}
            WHERE {baseline_filter}
              AND subscriber_number IS NOT NULL
              AND subscriber_number <> ''
            GROUP BY subscriber_number
        )
        SELECT
            current_part.subscriber_number AS mobile_number,
            current_part.selected_time_records,
            baseline.other_time_records,
            current_part.selected_time_cells,
            baseline.other_time_cells,
            'Common Number' AS finding_type,
            'This number is present in selected time and also seen in other loaded data.' AS meaning,
            'May be a local/repeated number or linked person. Verify before conclusion.' AS why_it_matters,
            'Medium' AS priority,
            CASE
                WHEN current_part.selected_time_cells >= 2 OR baseline.other_time_cells >= 2 THEN 'Medium'
                ELSE 'Low'
            END AS confidence_level,
            'Verify with CDR, SDR/CAF, IMEI/IMSI and field information.' AS suggested_action
        FROM current_part
        INNER JOIN baseline
          ON current_part.subscriber_number = baseline.subscriber_number
        ORDER BY
            current_part.selected_time_cells DESC,
            current_part.selected_time_records DESC,
            baseline.other_time_records DESC
        LIMIT ?
        """,
        [partition_time, partition_time, int(lead_limit)],
    )

    uncommon_numbers = store.query_df(
        f"""
        WITH current_part AS (
            SELECT
                subscriber_number,
                COUNT(*) AS selected_time_records,
                COUNT(DISTINCT searched_cell_id) AS selected_time_cells,
                COUNT(DISTINCT imei) AS selected_time_imei,
                COUNT(DISTINCT imsi) AS selected_time_imsi,
                MIN(TRY_CAST(event_time AS TIMESTAMP)) AS first_seen_selected,
                MAX(TRY_CAST(event_time AS TIMESTAMP)) AS last_seen_selected
            FROM {TABLE_EVENTS}
            WHERE {current_filter}
              AND subscriber_number IS NOT NULL
              AND subscriber_number <> ''
            GROUP BY subscriber_number
        ),
        baseline AS (
            SELECT DISTINCT subscriber_number
            FROM {TABLE_EVENTS}
            WHERE {baseline_filter}
              AND subscriber_number IS NOT NULL
              AND subscriber_number <> ''
        )
        SELECT
            current_part.subscriber_number AS mobile_number,
            current_part.selected_time_records,
            0 AS other_time_records,
            current_part.selected_time_cells,
            current_part.selected_time_imei,
            current_part.selected_time_imsi,
            current_part.first_seen_selected,
            current_part.last_seen_selected,
            'Uncommon / New Visitor' AS finding_type,
            'This number is present in selected time but not found in other loaded data.' AS meaning,
            'May indicate a new visitor, rare presence, or incident-time lead.' AS why_it_matters,
            CASE
                WHEN current_part.selected_time_cells >= 3 THEN 'High'
                WHEN current_part.selected_time_cells >= 2 THEN 'Medium-High'
                WHEN current_part.selected_time_records >= 10 THEN 'Medium'
                ELSE 'Low'
            END AS priority,
            CASE
                WHEN current_part.selected_time_cells >= 2 AND current_part.selected_time_records >= 5 THEN 'High'
                WHEN current_part.selected_time_records >= 3 THEN 'Medium'
                ELSE 'Low'
            END AS confidence_level,
            'Verify this number first with CDR/SDR/CAF, IMEI/IMSI continuity, and field/local input.' AS suggested_action
        FROM current_part
        LEFT JOIN baseline
          ON current_part.subscriber_number = baseline.subscriber_number
        WHERE baseline.subscriber_number IS NULL
        ORDER BY
            CASE
                WHEN current_part.selected_time_cells >= 3 THEN 1
                WHEN current_part.selected_time_cells >= 2 THEN 2
                WHEN current_part.selected_time_records >= 10 THEN 3
                ELSE 4
            END,
            current_part.selected_time_records DESC
        LIMIT ?
        """,
        [partition_time, partition_time, int(lead_limit)],
    )

    multi_cell_presence = store.query_df(
        f"""
        SELECT
            subscriber_number AS mobile_number,
            COUNT(*) AS records_found,
            COUNT(DISTINCT searched_cell_id) AS cells_seen,
            COUNT(DISTINCT imei) AS imei_count,
            COUNT(DISTINCT imsi) AS imsi_count,
            MIN(TRY_CAST(event_time AS TIMESTAMP)) AS first_seen,
            MAX(TRY_CAST(event_time AS TIMESTAMP)) AS last_seen,
            'Multi-Cell Presence' AS finding_type,
            'This number appeared in more than one searched cell during selected time.' AS meaning,
            'Multi-cell presence can be a stronger area-presence or movement lead.' AS why_it_matters,
            CASE
                WHEN COUNT(DISTINCT searched_cell_id) >= 3 THEN 'High'
                ELSE 'Medium-High'
            END AS priority,
            CASE
                WHEN COUNT(*) >= 5 THEN 'High'
                ELSE 'Medium'
            END AS confidence_level,
            'Verify movement feasibility, tower locations, CDR location and field/local information.' AS suggested_action
        FROM {TABLE_EVENTS}
        WHERE {current_filter}
          AND subscriber_number IS NOT NULL
          AND subscriber_number <> ''
        GROUP BY subscriber_number
        HAVING COUNT(DISTINCT searched_cell_id) >= 2
        ORDER BY cells_seen DESC, records_found DESC
        LIMIT ?
        """,
        [partition_time, int(lead_limit)],
    )

    repeat_presence = store.query_df(
        f"""
        WITH current_part AS (
            SELECT DISTINCT subscriber_number
            FROM {TABLE_EVENTS}
            WHERE {current_filter}
              AND subscriber_number IS NOT NULL
              AND subscriber_number <> ''
        ),
        baseline AS (
            SELECT
                subscriber_number,
                COUNT(*) AS other_time_records,
                COUNT(DISTINCT DATE_TRUNC('minute', TRY_CAST(event_time AS TIMESTAMP))) AS other_minutes_seen,
                COUNT(DISTINCT searched_cell_id) AS other_cells_seen
            FROM {TABLE_EVENTS}
            WHERE {baseline_filter}
              AND subscriber_number IS NOT NULL
              AND subscriber_number <> ''
            GROUP BY subscriber_number
        )
        SELECT
            current_part.subscriber_number AS mobile_number,
            baseline.other_time_records,
            baseline.other_minutes_seen,
            baseline.other_cells_seen,
            'Repeat Presence' AS finding_type,
            'This number is seen at selected time and repeatedly in other loaded data.' AS meaning,
            'May be local, regular visitor, or associated number depending on case context.' AS why_it_matters,
            CASE
                WHEN baseline.other_minutes_seen >= 10 OR baseline.other_cells_seen >= 3 THEN 'Medium'
                ELSE 'Low'
            END AS priority,
            CASE
                WHEN baseline.other_time_records >= 20 THEN 'Medium'
                ELSE 'Low'
            END AS confidence_level,
            'Do not treat as suspect only because it is repeated. Verify role and location context.' AS suggested_action
        FROM current_part
        INNER JOIN baseline
          ON current_part.subscriber_number = baseline.subscriber_number
        ORDER BY
            baseline.other_minutes_seen DESC,
            baseline.other_time_records DESC
        LIMIT ?
        """,
        [partition_time, partition_time, int(lead_limit)],
    )

    device_consistency = store.query_df(
        f"""
        SELECT
            subscriber_number AS mobile_number,
            COUNT(*) AS records_found,
            COUNT(DISTINCT imei) AS imei_count,
            COUNT(DISTINCT imsi) AS imsi_count,
            COUNT(DISTINCT searched_cell_id) AS cells_seen,
            'IMEI/IMSI Consistency' AS finding_type,
            CASE
                WHEN COUNT(DISTINCT imei) > 1 AND COUNT(DISTINCT imsi) > 1
                    THEN 'Multiple device and SIM identifiers seen.'
                WHEN COUNT(DISTINCT imei) > 1
                    THEN 'Multiple device identifiers seen.'
                WHEN COUNT(DISTINCT imsi) > 1
                    THEN 'Multiple SIM identifiers seen.'
                ELSE 'Device and SIM identifiers appear consistent in selected time.'
            END AS meaning,
            CASE
                WHEN COUNT(DISTINCT imei) > 1 OR COUNT(DISTINCT imsi) > 1
                    THEN 'May indicate SIM/device change, shared handset, data issue, or multiple records.'
                ELSE 'Consistent identifiers increase confidence but still need verification.'
            END AS why_it_matters,
            CASE
                WHEN COUNT(DISTINCT imei) > 1 OR COUNT(DISTINCT imsi) > 1
                    THEN 'Medium'
                ELSE 'Low'
            END AS priority,
            CASE
                WHEN COUNT(*) >= 3 THEN 'Medium'
                ELSE 'Low'
            END AS confidence_level,
            'Verify IMEI/IMSI with CDR, SDR/CAF and operator records before conclusion.' AS suggested_action
        FROM {TABLE_EVENTS}
        WHERE {current_filter}
          AND subscriber_number IS NOT NULL
          AND subscriber_number <> ''
        GROUP BY subscriber_number
        ORDER BY
            imei_count DESC,
            imsi_count DESC,
            records_found DESC
        LIMIT ?
        """,
        [partition_time, int(lead_limit)],
    )

    priority_leads = store.query_df(
        f"""
        WITH current_part AS (
            SELECT
                subscriber_number,
                COUNT(*) AS selected_time_records,
                COUNT(DISTINCT searched_cell_id) AS selected_time_cells,
                COUNT(DISTINCT imei) AS imei_count,
                COUNT(DISTINCT imsi) AS imsi_count,
                MIN(TRY_CAST(event_time AS TIMESTAMP)) AS first_seen,
                MAX(TRY_CAST(event_time AS TIMESTAMP)) AS last_seen
            FROM {TABLE_EVENTS}
            WHERE {current_filter}
              AND subscriber_number IS NOT NULL
              AND subscriber_number <> ''
            GROUP BY subscriber_number
        ),
        baseline AS (
            SELECT
                subscriber_number,
                COUNT(*) AS baseline_records
            FROM {TABLE_EVENTS}
            WHERE {baseline_filter}
              AND subscriber_number IS NOT NULL
              AND subscriber_number <> ''
            GROUP BY subscriber_number
        )
        SELECT
            current_part.subscriber_number AS mobile_number,
            current_part.selected_time_records,
            COALESCE(baseline.baseline_records, 0) AS baseline_records,
            current_part.selected_time_cells,
            current_part.imei_count,
            current_part.imsi_count,
            current_part.first_seen,
            current_part.last_seen,
            CASE
                WHEN baseline.subscriber_number IS NULL AND current_part.selected_time_cells >= 2
                    THEN 'High'
                WHEN baseline.subscriber_number IS NULL
                    THEN 'Medium'
                WHEN current_part.selected_time_cells >= 2
                    THEN 'Medium-High'
                WHEN current_part.selected_time_records >= 10
                    THEN 'Medium'
                ELSE 'Low'
            END AS priority,
            CASE
                WHEN current_part.selected_time_cells >= 2 AND current_part.selected_time_records >= 5
                    THEN 'High'
                WHEN current_part.selected_time_records >= 3
                    THEN 'Medium'
                ELSE 'Low'
            END AS confidence_level,
            CASE
                WHEN baseline.subscriber_number IS NULL AND current_part.selected_time_cells >= 2
                    THEN 'New/rare number with multi-cell presence'
                WHEN baseline.subscriber_number IS NULL
                    THEN 'New/rare number in selected time'
                WHEN current_part.selected_time_cells >= 2
                    THEN 'Common number with multi-cell presence'
                WHEN current_part.selected_time_records >= 10
                    THEN 'High activity in selected time'
                ELSE 'Low-volume presence'
            END AS simple_reason,
            'Verify priority leads with CDR/SDR/CAF, IMEI/IMSI and field/local information.' AS suggested_action
        FROM current_part
        LEFT JOIN baseline
          ON current_part.subscriber_number = baseline.subscriber_number
        ORDER BY
            CASE
                WHEN baseline.subscriber_number IS NULL AND current_part.selected_time_cells >= 2 THEN 1
                WHEN baseline.subscriber_number IS NULL THEN 2
                WHEN current_part.selected_time_cells >= 2 THEN 3
                WHEN current_part.selected_time_records >= 10 THEN 4
                ELSE 9
            END,
            current_part.selected_time_records DESC
        LIMIT ?
        """,
        [partition_time, partition_time, int(lead_limit)],
    )

    total_loaded_records = int(count_tower_ipdr_events(case_id))

    selected_records_for_warning = 0
    selected_start_for_warning = start_time
    selected_end_for_warning = end_time

    if isinstance(summary, pd.DataFrame) and not summary.empty:
        try:
            selected_records_for_warning = int(summary.iloc[0].get("records_found", 0) or 0)
        except Exception:
            selected_records_for_warning = 0

        selected_start_for_warning = summary.iloc[0].get("partition_start", start_time)
        selected_end_for_warning = summary.iloc[0].get("partition_end", end_time)

    data_scope_warnings = _build_tower_ipdr_scope_warnings(
        selected_records=selected_records_for_warning,
        total_records=total_loaded_records,
        selected_start=selected_start_for_warning,
        selected_end=selected_end_for_warning,
    )

    total_loaded_records = int(count_tower_ipdr_events(case_id))

    selected_records_for_warning = 0
    selected_start_for_warning = start_time
    selected_end_for_warning = end_time

    if isinstance(summary, pd.DataFrame) and not summary.empty:
        try:
            selected_records_for_warning = int(summary.iloc[0].get("records_found", 0) or 0)
        except Exception:
            selected_records_for_warning = 0

        selected_start_for_warning = summary.iloc[0].get("partition_start", start_time)
        selected_end_for_warning = summary.iloc[0].get("partition_end", end_time)

    data_scope_warnings = _build_tower_ipdr_scope_warnings(
        selected_records=selected_records_for_warning,
        total_records=total_loaded_records,
        selected_start=selected_start_for_warning,
        selected_end=selected_end_for_warning,
    )

    lead_summary = pd.DataFrame(
        [
            {
                "finding": "Common Numbers",
                "records": len(common_numbers),
                "meaning": "Numbers seen in selected time and also elsewhere in loaded data.",
            },
            {
                "finding": "Uncommon / New Visitor",
                "records": len(uncommon_numbers),
                "meaning": "Numbers seen in selected time but not seen elsewhere in loaded data.",
            },
            {
                "finding": "Multi-Cell Presence",
                "records": len(multi_cell_presence),
                "meaning": "Numbers seen in more than one searched cell.",
            },
            {
                "finding": "Repeat Presence",
                "records": len(repeat_presence),
                "meaning": "Numbers repeatedly seen in other loaded data also.",
            },
            {
                "finding": "IMEI/IMSI Consistency",
                "records": len(device_consistency),
                "meaning": "Device/SIM consistency or possible change indicators.",
            },
            {
                "finding": "Priority Leads",
                "records": len(priority_leads),
                "meaning": "Combined ranking based on rarity, cells, activity and confidence.",
            },
        ]
    )

    return {
        "summary": summary,
        "lead_summary": lead_summary,
        "data_scope_warnings": data_scope_warnings,
        "common_numbers": common_numbers,
        "uncommon_numbers": uncommon_numbers,
        "multi_cell_presence": multi_cell_presence,
        "repeat_presence": repeat_presence,
        "device_consistency": device_consistency,
        "priority_leads": priority_leads,
    }


def _safe_first_value(dataframe: pd.DataFrame, column: str, default: Any = "") -> Any:
    if not isinstance(dataframe, pd.DataFrame) or dataframe.empty:
        return default

    if column not in dataframe.columns:
        return default

    value = dataframe.iloc[0].get(column, default)

    try:
        if pd.isna(value):
            return default
    except Exception:
        pass

    return value


def _print_simple_leads(
    dataframe: pd.DataFrame,
    *,
    number_col: str = "mobile_number",
    max_rows: int = 10,
) -> None:
    if not isinstance(dataframe, pd.DataFrame) or dataframe.empty:
        print("   No important lead found in this section.")
        return

    for index, row in dataframe.head(max_rows).reset_index(drop=True).iterrows():
        number = row.get(number_col, "")
        priority = row.get("priority", "")
        confidence = row.get("confidence_level", "")
        reason = (
            row.get("simple_reason", "")
            or row.get("meaning", "")
            or row.get("finding_type", "")
        )
        action = row.get("suggested_action", "")

        print(f"   {index + 1}. Mobile Number : {number}")
        print(f"      Priority      : {priority or 'Not specified'}")
        print(f"      Confidence    : {confidence or 'Not specified'}")
        print(f"      Why important : {reason or 'Needs verification'}")

        if action:
            print(f"      Next Action   : {action}")

        print()




def print_tower_ipdr_investigation_summary(
    result: dict[str, Any],
    *,
    max_leads: int = 10,
) -> None:
    """Print a simple English investigation summary."""

    summary = result.get(
        "summary",
        pd.DataFrame(),
    )

    lead_summary = result.get(
        "lead_summary",
        pd.DataFrame(),
    )

    print("\n" + "=" * 78)
    print(
        "TOWER IPDR INVESTIGATION SUMMARY"
    )
    print("=" * 78)

    if summary.empty:
        print(
            "No records were found "
            "in the selected Part."
        )
        print("=" * 78)
        return

    row = summary.iloc[0]

    def value(
        name: str,
        default: Any = "",
    ) -> Any:
        return row.get(
            name,
            default,
        )

    print(
        "Selected Period    : "
        f"{value('partition_time')}"
    )

    spot_id = str(
        value(
            "spot_id",
            "",
        )
        or ""
    ).strip()

    spot_name = str(
        value(
            "spot_name",
            "",
        )
        or ""
    ).strip()

    if spot_id:
        print(
            "Selected Spot      : "
            f"{spot_id}"
            + (
                f" | {spot_name}"
                if (
                    spot_name
                    and spot_name
                    != spot_id
                )
                else ""
            )
        )
    else:
        print(
            "Selected Spot      : "
            "ALL LOADED SPOTS (legacy)"
        )

    print(
        "Spot Scope         : "
        f"{value('spot_scope_mode')}"
    )
    print(
        "Analysis Mode      : "
        f"{value('analysis_mode')}"
    )
    print(
        "Records Found      : "
        f"{int(value('records_found', 0)):,}"
    )
    print(
        "Numbers Found      : "
        f"{int(value('numbers_found', 0)):,}"
    )
    print(
        "Searched Cells     : "
        f"{int(value('cells_involved', 0)):,}"
    )
    print(
        "First Activity     : "
        f"{value('first_activity')}"
    )
    print(
        "Last Activity      : "
        f"{value('last_activity')}"
    )

    finding_meanings = {
        "Common Numbers": (
            "Numbers also seen outside "
            "the selected Part."
        ),
        "Part-Uncommon": (
            "Numbers not seen in another "
            "configured Part."
        ),
        "Spot-Uncommon": (
            "Numbers not seen in the same "
            "Spot outside the selected Part."
        ),
        "Global-Uncommon": (
            "Numbers not seen anywhere outside "
            "the selected Part."
        ),
        "Multi-Cell Presence": (
            "Numbers seen on multiple searched "
            "cells inside the selected Part."
        ),
        "Repeat Presence": (
            "Numbers repeatedly seen elsewhere "
            "in the loaded data."
        ),
        "IMEI/IMSI Consistency": (
            "Device and SIM identifier "
            "continuity checks."
        ),
        "Suspicious Timing": (
            "Numbers with notable activity "
            "inside the selected Part."
        ),
        "Priority Leads": (
            "Combined ranking for investigator "
            "review."
        ),
    }

    print("\n" + "-" * 78)
    print("IMPORTANT FINDINGS")
    print("-" * 78)

    if lead_summary.empty:
        print(
            "No lead summary is available."
        )
    else:
        for item in (
            lead_summary.itertuples(
                index=False,
            )
        ):
            finding = str(
                getattr(
                    item,
                    "finding",
                    "Finding",
                )
            )

            records = int(
                getattr(
                    item,
                    "records",
                    0,
                )
                or 0
            )

            displayed = int(
                getattr(
                    item,
                    "displayed_records",
                    records,
                )
                or 0
            )

            print(
                f"- {finding}: "
                f"{records:,} total"
            )

            if displayed < records:
                print(
                    "  Displayed: top "
                    f"{displayed:,}"
                )

            meaning = finding_meanings.get(
                finding,
                "",
            )

            if meaning:
                print(
                    f"  Meaning: {meaning}"
                )

    def section(
        title: str,
        meaning: str,
        use: str,
        key: str,
    ) -> None:
        frame = result.get(
            key,
            pd.DataFrame(),
        )

        print("\n" + "-" * 78)
        print(title)
        print("-" * 78)
        print(
            f"Meaning: {meaning}"
        )
        print(
            f"Use: {use}"
        )

        _print_simple_leads(
            frame,
            max_rows=max_leads,
        )

    section(
        "TOP PRIORITY LEADS",
        (
            "Combined ranking based on Part, "
            "Spot, global rarity, activity and "
            "multi-cell presence."
        ),
        (
            "Use this list to decide the review "
            "order. It is not a final conclusion."
        ),
        "priority_leads",
    )

    section(
        "PART-UNCOMMON NUMBERS",
        (
            "Numbers found in this Part but not "
            "in another configured Part."
        ),
        (
            "Identify leads specific to the "
            "selected incident period."
        ),
        "part_uncommon_numbers",
    )

    section(
        "SPOT-UNCOMMON / NEW-IN-SPOT",
        (
            "Numbers found in this Part but not "
            "elsewhere in the same Spot."
        ),
        (
            "Identify possible new or rare "
            "visitors at the selected Spot."
        ),
        "spot_uncommon_numbers",
    )

    section(
        "GLOBAL-UNCOMMON NUMBERS",
        (
            "Numbers not found anywhere outside "
            "the selected Part."
        ),
        (
            "Review these strongest rarity leads "
            "with independent evidence."
        ),
        "global_uncommon_numbers",
    )

    section(
        "COMMON NUMBERS",
        (
            "Numbers also present in another "
            "time or Spot."
        ),
        (
            "Check whether they are local users, "
            "regular visitors or linked persons."
        ),
        "common_numbers",
    )

    section(
        "MULTI-CELL PRESENCE",
        (
            "Numbers seen on more than one "
            "searched cell in this Part."
        ),
        (
            "Review possible area presence and "
            "movement. This is not exact location proof."
        ),
        "multi_cell_presence",
    )

    section(
        "REPEAT PRESENCE",
        (
            "Numbers also seen repeatedly outside "
            "the selected Part."
        ),
        (
            "A repeated number may be local or "
            "regular. Verify its role."
        ),
        "repeat_presence",
    )

    section(
        "IMEI / IMSI CONSISTENCY",
        (
            "Checks device and SIM identifier "
            "continuity."
        ),
        (
            "Verify possible device or SIM changes "
            "with operator records."
        ),
        "device_consistency",
    )

    section(
        "SUSPICIOUS TIMING / HIGH ACTIVITY",
        (
            "Numbers with notable activity in "
            "the selected Part."
        ),
        (
            "Compare activity with incident time, "
            "route and field information."
        ),
        "suspicious_timing",
    )

    print("\n" + "-" * 78)
    print("SUGGESTED VERIFICATION")
    print("-" * 78)
    print(
        "1. Review Global-Uncommon and "
        "high-priority leads first."
    )
    print(
        "2. Compare Part-Uncommon numbers "
        "across all configured Parts."
    )
    print(
        "3. Verify Spot-Uncommon numbers "
        "against local and regular users."
    )
    print(
        "4. Verify CDR, SDR/CAF, IMEI, IMSI "
        "and operator records."
    )
    print(
        "5. Tower or IPDR presence alone is "
        "not proof of exact location or involvement; "
        "independent verification is required."
    )
    print("=" * 78)




def _safe_report_name(value: str) -> str:
    import re

    text = str(value).strip()
    text = text.replace(":", "-").replace(" ", "_")
    text = re.sub(r"[^A-Za-z0-9_.-]+", "_", text)
    return text.strip("_") or "partition"


def tower_ipdr_fast_report_root(case_id: str) -> Path:
    return (
        Path("cases")
        / "active"
        / str(case_id)
        / "reports"
        / "tower_dump"
        / "ipdr"
        / "fast_partition"
    )


def export_tower_ipdr_investigation_summary(
    case_id: str,
    partition_time: str,
    *,
    mode: str = "same_minute",
    lead_limit: int = 50,
    max_leads_in_text: int = 20,
) -> dict[str, Any]:
    """Export fast Tower IPDR partition result for case record.

    Exports:
    - simple officer-friendly TXT summary
    - detailed CSV tables
    - manifest.json
    """

    import io
    from contextlib import redirect_stdout

    result = tower_ipdr_investigation_summary(
        case_id,
        partition_time,
        mode=mode,
        lead_limit=lead_limit,
    )

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_partition = _safe_report_name(partition_time)
    run_id = f"tower_ipdr_fast_partition_{safe_partition}_{timestamp}"

    output_dir = tower_ipdr_fast_report_root(case_id) / run_id
    output_dir.mkdir(parents=True, exist_ok=True)

    saved_files: dict[str, str] = {}

    for name, dataframe in result.items():
        if not isinstance(dataframe, pd.DataFrame):
            continue

        csv_path = output_dir / f"{name}.csv"
        dataframe.to_csv(csv_path, index=False)
        saved_files[name] = str(csv_path)

    buffer = io.StringIO()

    with redirect_stdout(buffer):
        print_tower_ipdr_investigation_summary(
            result,
            max_leads=max_leads_in_text,
        )

    text_path = output_dir / "investigation_summary.txt"
    text_path.write_text(buffer.getvalue(), encoding="utf-8")
    saved_files["investigation_summary_text"] = str(text_path)

    manifest = {
        "case_id": str(case_id),
        "run_id": run_id,
        "created_at": _now_iso(),
        "partition_time": str(partition_time),
        "analysis_mode": str(mode),
        "lead_limit": int(lead_limit),
        "max_leads_in_text": int(max_leads_in_text),
        "output_dir": str(output_dir),
        "saved_files": saved_files,
        "note": (
            "Fast Tower IPDR partition report generated from DuckDB staging. "
            "Results are investigation leads and require verification."
        ),
    }

    manifest_path.write_text(
        json.dumps(
            _json_safe(manifest),
            ensure_ascii=False,
            indent=2,
            sort_keys=True,
        ),
        encoding="utf-8",
    )
    return manifest



def tower_ipdr_range_investigation_summary(
    case_id: str,
    start_time: str,
    end_time: str,
    *,
    spot_id: str = "",
    spot_name: str = "",
    comparison_parts: list[dict[str, Any]] | None = None,
    current_part_no: int | None = None,
    lead_limit: int = 50,
) -> dict[str, pd.DataFrame]:
    """Return Spot-aware Date-Time Part intelligence.

    Main selected-scope rule:
        selected Spot
        AND start_time <= event_time < end_time

    Uncommon classifications:
    - Part-Uncommon:
      not seen in another configured Date-Time Part.
    - Spot-Uncommon:
      not seen in the same Spot outside this Part.
    - Global-Uncommon:
      not seen anywhere outside this selected Part.
    """

    database_path = (
        tower_ipdr_database_path(
            case_id
        )
    )

    store = DuckDBStore(
        database_path
    )

    empty = {
        "summary": pd.DataFrame(),
        "lead_summary": pd.DataFrame(),
        "uncommon_classification": (
            pd.DataFrame()
        ),
        "part_uncommon_numbers": (
            pd.DataFrame()
        ),
        "spot_uncommon_numbers": (
            pd.DataFrame()
        ),
        "global_uncommon_numbers": (
            pd.DataFrame()
        ),
        "common_numbers": pd.DataFrame(),
        "uncommon_numbers": pd.DataFrame(),
        "multi_cell_presence": (
            pd.DataFrame()
        ),
        "repeat_presence": pd.DataFrame(),
        "device_consistency": (
            pd.DataFrame()
        ),
        "suspicious_timing": (
            pd.DataFrame()
        ),
        "priority_leads": pd.DataFrame(),
    }

    if not store.table_exists(
        TABLE_EVENTS
    ):
        return empty

    selected_spot = str(
        spot_id
        or ""
    ).strip()

    selected_spot_name = str(
        spot_name
        or selected_spot
        or "ALL LOADED SPOTS"
    ).strip()

    safe_limit = max(
        1,
        min(
            int(lead_limit),
            5000,
        ),
    )

    def sql_text(
        value: Any,
    ) -> str:
        return (
            "'"
            + str(
                value
            ).replace(
                "'",
                "''",
            )
            + "'"
        )

    event_time_sql = (
        "TRY_CAST(event_time AS TIMESTAMP)"
    )

    current_time_condition = (
        f"{event_time_sql} >= "
        f"CAST({sql_text(start_time)} AS TIMESTAMP) "
        f"AND {event_time_sql} < "
        f"CAST({sql_text(end_time)} AS TIMESTAMP)"
    )

    if selected_spot:
        selected_spot_condition = (
            "TRIM(CAST(spot_id AS VARCHAR)) "
            f"= {sql_text(selected_spot)}"
        )
        spot_scope_mode = (
            "SELECTED_SPOT_ONLY"
        )
        spot_scope_status = (
            "VALID_SELECTED_SPOT"
        )
    else:
        selected_spot_condition = "TRUE"
        spot_scope_mode = (
            "LEGACY_ALL_SPOTS"
        )
        spot_scope_status = (
            "LEGACY_NO_SPOT_MAPPING"
        )

    current_scope_condition = (
        f"({current_time_condition}) "
        f"AND ({selected_spot_condition})"
    )

    same_spot_outside_condition = (
        f"({selected_spot_condition}) "
        f"AND NOT ({current_time_condition})"
    )

    global_outside_condition = (
        f"NOT ({current_scope_condition})"
    )

    other_part_conditions: list[str] = []

    for part in (
        comparison_parts
        or []
    ):
        try:
            part_number = int(
                part.get(
                    "part_no",
                    -1,
                )
            )
        except (
            TypeError,
            ValueError,
        ):
            part_number = -1

        if (
            current_part_no is not None
            and part_number
            == int(current_part_no)
        ):
            continue

        other_start = str(
            part.get(
                "start_time",
                "",
            )
            or ""
        ).strip()

        other_end = str(
            part.get(
                "end_time",
                "",
            )
            or ""
        ).strip()

        other_spot = str(
            part.get(
                "spot_id",
                "",
            )
            or ""
        ).strip()

        if not (
            other_start
            and other_end
        ):
            continue

        other_time_condition = (
            f"{event_time_sql} >= "
            f"CAST({sql_text(other_start)} "
            "AS TIMESTAMP) "
            f"AND {event_time_sql} < "
            f"CAST({sql_text(other_end)} "
            "AS TIMESTAMP)"
        )

        if other_spot:
            other_spot_condition = (
                "TRIM(CAST(spot_id AS VARCHAR)) "
                f"= {sql_text(other_spot)}"
            )

            other_part_conditions.append(
                f"(({other_time_condition}) "
                f"AND ({other_spot_condition}))"
            )
        else:
            other_part_conditions.append(
                f"({other_time_condition})"
            )

    other_parts_condition = (
        " OR ".join(
            other_part_conditions
        )
        if other_part_conditions
        else "FALSE"
    )

    summary = store.query_df(
        f"""
        SELECT
            CAST(
                {sql_text(start_time)}
                AS TIMESTAMP
            ) AS partition_start,
            CAST(
                {sql_text(end_time)}
                AS TIMESTAMP
            ) AS partition_end,
            CONCAT(
                CAST(
                    {sql_text(start_time)}
                    AS VARCHAR
                ),
                ' to ',
                CAST(
                    {sql_text(end_time)}
                    AS VARCHAR
                )
            ) AS partition_time,
            {sql_text(selected_spot)}
                AS spot_id,
            {sql_text(selected_spot_name)}
                AS spot_name,
            {sql_text(spot_scope_mode)}
                AS spot_scope_mode,
            {sql_text(spot_scope_status)}
                AS spot_scope_status,
            'Date-Time Range + Spot'
                AS analysis_mode,
            COUNT(*) AS records_found,
            COUNT(
                DISTINCT subscriber_number
            ) AS numbers_found,
            COUNT(
                DISTINCT searched_cell_id
            ) AS cells_involved,
            COUNT(
                DISTINCT imei
            ) AS imei_found,
            COUNT(
                DISTINCT imsi
            ) AS imsi_found,
            MIN(
                {event_time_sql}
            ) AS first_activity,
            MAX(
                {event_time_sql}
            ) AS last_activity
        FROM {TABLE_EVENTS}
        WHERE {current_scope_condition}
        """
    )

    classification = store.query_df(
        f"""
        WITH subscriber_rollup AS (
            SELECT
                TRIM(
                    CAST(
                        subscriber_number
                        AS VARCHAR
                    )
                ) AS mobile_number,

                SUM(
                    CASE
                        WHEN
                            {current_scope_condition}
                        THEN 1
                        ELSE 0
                    END
                ) AS selected_period_records,

                COUNT(
                    DISTINCT CASE
                        WHEN
                            {current_scope_condition}
                        THEN searched_cell_id
                    END
                ) AS selected_period_cells,

                COUNT(
                    DISTINCT CASE
                        WHEN
                            {current_scope_condition}
                        THEN imei
                    END
                ) AS selected_period_imei,

                COUNT(
                    DISTINCT CASE
                        WHEN
                            {current_scope_condition}
                        THEN imsi
                    END
                ) AS selected_period_imsi,

                MIN(
                    CASE
                        WHEN
                            {current_scope_condition}
                        THEN {event_time_sql}
                    END
                ) AS first_seen_selected,

                MAX(
                    CASE
                        WHEN
                            {current_scope_condition}
                        THEN {event_time_sql}
                    END
                ) AS last_seen_selected,

                SUM(
                    CASE
                        WHEN
                            {other_parts_condition}
                        THEN 1
                        ELSE 0
                    END
                ) AS other_parts_records,

                SUM(
                    CASE
                        WHEN
                            {same_spot_outside_condition}
                        THEN 1
                        ELSE 0
                    END
                ) AS same_spot_other_records,

                SUM(
                    CASE
                        WHEN
                            {global_outside_condition}
                        THEN 1
                        ELSE 0
                    END
                ) AS global_other_records,

                COUNT(
                    DISTINCT CASE
                        WHEN
                            {global_outside_condition}
                        THEN spot_id
                    END
                ) AS global_other_spot_count

            FROM {TABLE_EVENTS}
            WHERE
                subscriber_number
                    IS NOT NULL
                AND TRIM(
                    CAST(
                        subscriber_number
                        AS VARCHAR
                    )
                ) <> ''
            GROUP BY
                subscriber_number
        )

        SELECT
            mobile_number,
            selected_period_records,
            selected_period_cells,
            selected_period_imei,
            selected_period_imsi,
            first_seen_selected,
            last_seen_selected,
            other_parts_records,
            same_spot_other_records,
            global_other_records,
            global_other_spot_count,

            CASE
                WHEN other_parts_records = 0
                    THEN 'PART_ONLY'
                ELSE 'SEEN_IN_OTHER_PART'
            END AS part_status,

            CASE
                WHEN same_spot_other_records = 0
                    THEN 'NEW_IN_SPOT'
                ELSE
                    'SEEN_IN_SPOT_OUTSIDE_PART'
            END AS spot_status,

            CASE
                WHEN global_other_records = 0
                    THEN 'GLOBAL_UNCOMMON'
                ELSE 'SEEN_ELSEWHERE'
            END AS global_status

        FROM subscriber_rollup
        WHERE
            selected_period_records > 0
        ORDER BY
            selected_period_cells DESC,
            selected_period_records DESC,
            mobile_number
        """
    )

    if classification.empty:
        empty["summary"] = summary
        return empty

    numeric_columns = [
        "selected_period_records",
        "selected_period_cells",
        "selected_period_imei",
        "selected_period_imsi",
        "other_parts_records",
        "same_spot_other_records",
        "global_other_records",
        "global_other_spot_count",
    ]

    for column in numeric_columns:
        classification[column] = (
            pd.to_numeric(
                classification[column],
                errors="coerce",
            )
            .fillna(0)
            .astype("int64")
        )

    classification["spot_id"] = (
        selected_spot
    )

    classification["spot_name"] = (
        selected_spot_name
    )

    classification["cells_seen"] = (
        classification[
            "selected_period_cells"
        ]
    )

    classification["imei_count"] = (
        classification[
            "selected_period_imei"
        ]
    )

    classification["imsi_count"] = (
        classification[
            "selected_period_imsi"
        ]
    )

    classification["records_found"] = (
        classification[
            "selected_period_records"
        ]
    )

    classification["baseline_records"] = (
        classification[
            "global_other_records"
        ]
    )

    classification["simple_reason"] = (
        classification["part_status"]
        + " | "
        + classification["spot_status"]
        + " | "
        + classification["global_status"]
    )

    classification[
        "suggested_action"
    ] = (
        "Verify with CDR/SDR/CAF, "
        "IMEI/IMSI, field information "
        "and incident timing."
    )

    def decorate_uncommon(
        frame: pd.DataFrame,
        *,
        scope: str,
        meaning: str,
    ) -> pd.DataFrame:
        output = frame.copy()

        if output.empty:
            return output

        output["uncommon_scope"] = scope
        output["finding_type"] = scope
        output["meaning"] = meaning
        output[
            "why_it_matters"
        ] = (
            "This is an investigation lead; "
            "it is not proof of identity, "
            "movement or involvement."
        )

        output["priority"] = "Low"

        output.loc[
            output[
                "selected_period_records"
            ].ge(3),
            "priority",
        ] = "Medium"

        output.loc[
            output[
                "selected_period_cells"
            ].ge(2),
            "priority",
        ] = "Medium-High"

        output.loc[
            (
                output[
                    "global_status"
                ]
                == "GLOBAL_UNCOMMON"
            )
            & output[
                "selected_period_cells"
            ].ge(2),
            "priority",
        ] = "High"

        output[
            "confidence_level"
        ] = "Low"

        output.loc[
            output[
                "selected_period_records"
            ].ge(3),
            "confidence_level",
        ] = "Medium"

        output.loc[
            output[
                "selected_period_cells"
            ].ge(2)
            & output[
                "selected_period_records"
            ].ge(5),
            "confidence_level",
        ] = "High"

        return output

    part_uncommon_full = (
        classification.loc[
            classification[
                "part_status"
            ]
            == "PART_ONLY"
        ]
        .copy()
    )

    part_uncommon_full = decorate_uncommon(
        part_uncommon_full,
        scope="Part-Uncommon",
        meaning=(
            "Selected Part में मिला, "
            "लेकिन किसी दूसरे configured "
            "Part में नहीं मिला।"
        ),
    )

    spot_uncommon_full = (
        classification.loc[
            classification[
                "spot_status"
            ]
            == "NEW_IN_SPOT"
        ]
        .copy()
    )

    spot_uncommon_full = decorate_uncommon(
        spot_uncommon_full,
        scope="Spot-Uncommon",
        meaning=(
            "Selected Part में मिला, "
            "लेकिन इसी Spot के बाकी समय "
            "में नहीं मिला।"
        ),
    )

    global_uncommon_full = (
        classification.loc[
            classification[
                "global_status"
            ]
            == "GLOBAL_UNCOMMON"
        ]
        .copy()
    )

    global_uncommon_full = (
        decorate_uncommon(
            global_uncommon_full,
            scope="Global-Uncommon",
            meaning=(
                "Selected Part के बाहर "
                "पूरे loaded Tower IPDR data "
                "में कहीं नहीं मिला।"
            ),
        )
    )

    common_full = (
        classification.loc[
            classification[
                "global_status"
            ]
            == "SEEN_ELSEWHERE"
        ]
        .copy()
    )

    if not common_full.empty:
        common_full[
            "other_period_records"
        ] = common_full[
            "global_other_records"
        ]

        common_full[
            "finding_type"
        ] = "Common Number"

        common_full[
            "meaning"
        ] = (
            "Selected Part में और loaded "
            "data के किसी अन्य हिस्से में "
            "भी मौजूद है।"
        )

        common_full[
            "why_it_matters"
        ] = (
            "Local, repeated visitor, "
            "associate या linked person हो "
            "सकता है; verification required."
        )

        common_full["priority"] = "Low"

        common_full.loc[
            common_full[
                "selected_period_cells"
            ].ge(2),
            "priority",
        ] = "Medium"

        common_full[
            "confidence_level"
        ] = "Medium"

    global_numbers = set(
        global_uncommon_full[
            "mobile_number"
        ].astype(str)
    )

    spot_only = spot_uncommon_full.loc[
        ~spot_uncommon_full[
            "mobile_number"
        ].astype(str).isin(
            global_numbers
        )
    ].copy()

    selected_uncommon = (
        global_uncommon_full.copy()
    )

    selected_uncommon = pd.concat(
        [
            selected_uncommon,
            spot_only,
        ],
        ignore_index=True,
        sort=False,
    )

    selected_numbers = set(
        selected_uncommon[
            "mobile_number"
        ].astype(str)
    )

    part_only = part_uncommon_full.loc[
        ~part_uncommon_full[
            "mobile_number"
        ].astype(str).isin(
            selected_numbers
        )
    ].copy()

    uncommon_full = pd.concat(
        [
            selected_uncommon,
            part_only,
        ],
        ignore_index=True,
        sort=False,
    )

    uncommon_rank = {
        "Global-Uncommon": 1,
        "Spot-Uncommon": 2,
        "Part-Uncommon": 3,
    }

    if not uncommon_full.empty:
        uncommon_full[
            "_scope_rank"
        ] = (
            uncommon_full[
                "uncommon_scope"
            ]
            .map(uncommon_rank)
            .fillna(9)
        )

        uncommon_full = (
            uncommon_full
            .sort_values(
                [
                    "_scope_rank",
                    "selected_period_cells",
                    "selected_period_records",
                ],
                ascending=[
                    True,
                    False,
                    False,
                ],
            )
            .drop(
                columns=[
                    "_scope_rank"
                ]
            )
        )

    multi_cell_full = (
        classification.loc[
            classification[
                "selected_period_cells"
            ].ge(2)
        ]
        .copy()
    )

    if not multi_cell_full.empty:
        multi_cell_full[
            "finding_type"
        ] = "Multi-Cell Presence"

        multi_cell_full[
            "meaning"
        ] = (
            "Selected Spot और Part में "
            "एक से अधिक searched Cell ID "
            "पर presence मिली।"
        )

        multi_cell_full[
            "why_it_matters"
        ] = (
            "Area presence या movement lead "
            "हो सकता है; exact location proof "
            "नहीं है।"
        )

        multi_cell_full["priority"] = (
            "Medium-High"
        )

        multi_cell_full.loc[
            multi_cell_full[
                "selected_period_cells"
            ].ge(3),
            "priority",
        ] = "High"

        multi_cell_full[
            "confidence_level"
        ] = "High"

    repeat_full = (
        classification.loc[
            classification[
                "global_other_records"
            ].gt(0)
        ]
        .copy()
    )

    if not repeat_full.empty:
        repeat_full[
            "other_period_records"
        ] = repeat_full[
            "global_other_records"
        ]

        repeat_full[
            "finding_type"
        ] = "Repeat Presence"

        repeat_full[
            "meaning"
        ] = (
            "Selected Part के अलावा loaded "
            "data में भी presence मिली।"
        )

        repeat_full[
            "why_it_matters"
        ] = (
            "Local या repeated visitor हो "
            "सकता है; suspicious मानना उचित "
            "नहीं होगा बिना verification."
        )

        repeat_full["priority"] = "Low"

        repeat_full.loc[
            repeat_full[
                "global_other_records"
            ].ge(20),
            "priority",
        ] = "Medium"

        repeat_full[
            "confidence_level"
        ] = "Medium"

    device_full = (
        classification.copy()
    )

    device_full[
        "finding_type"
    ] = "IMEI/IMSI Consistency"

    device_full[
        "meaning"
    ] = (
        "Selected Part में number के "
        "device और SIM identifiers की "
        "continuity check."
    )

    device_full[
        "why_it_matters"
    ] = (
        "Multiple identifiers device/SIM "
        "change, shared handset या data issue "
        "दिखा सकते हैं।"
    )

    device_full["priority"] = "Low"

    device_change_mask = (
        device_full[
            "selected_period_imei"
        ].gt(1)
        | device_full[
            "selected_period_imsi"
        ].gt(1)
    )

    device_full.loc[
        device_change_mask,
        "priority",
    ] = "Medium"

    device_full[
        "confidence_level"
    ] = "Medium"

    device_full[
        "simple_reason"
    ] = (
        "IMEI count="
        + device_full[
            "selected_period_imei"
        ].astype(str)
        + ", IMSI count="
        + device_full[
            "selected_period_imsi"
        ].astype(str)
    )

    suspicious_full = (
        classification.loc[
            classification[
                "selected_period_records"
            ].ge(5)
            | classification[
                "selected_period_cells"
            ].ge(2)
        ]
        .copy()
    )

    if not suspicious_full.empty:
        suspicious_full[
            "finding_type"
        ] = (
            "Suspicious Timing / "
            "High Activity"
        )

        suspicious_full[
            "meaning"
        ] = (
            "Selected Part में notable "
            "activity मिली।"
        )

        suspicious_full[
            "why_it_matters"
        ] = (
            "Incident time और route से "
            "comparison के लिए useful lead."
        )

        suspicious_full["priority"] = (
            "Medium"
        )

        suspicious_full.loc[
            suspicious_full[
                "selected_period_cells"
            ].ge(2)
            & suspicious_full[
                "selected_period_records"
            ].ge(10),
            "priority",
        ] = "High"

        suspicious_full[
            "confidence_level"
        ] = "Medium"

    priority_full = (
        classification.copy()
    )

    priority_full[
        "priority_score"
    ] = (
        priority_full[
            "selected_period_records"
        ].clip(
            upper=100
        )
        + priority_full[
            "selected_period_cells"
        ].clip(
            upper=10
        )
        * 10
        + (
            priority_full[
                "part_status"
            ]
            == "PART_ONLY"
        ).astype(int)
        * 15
        + (
            priority_full[
                "spot_status"
            ]
            == "NEW_IN_SPOT"
        ).astype(int)
        * 25
        + (
            priority_full[
                "global_status"
            ]
            == "GLOBAL_UNCOMMON"
        ).astype(int)
        * 50
        + priority_full[
            "selected_period_imei"
        ].gt(1).astype(int)
        * 20
        + priority_full[
            "selected_period_imsi"
        ].gt(1).astype(int)
        * 20
    )

    priority_full["priority"] = "Low"

    priority_full.loc[
        priority_full[
            "priority_score"
        ].ge(70),
        "priority",
    ] = "Medium"

    priority_full.loc[
        priority_full[
            "priority_score"
        ].ge(120),
        "priority",
    ] = "High"

    priority_full[
        "confidence_level"
    ] = "Low"

    priority_full.loc[
        priority_full[
            "selected_period_records"
        ].ge(3),
        "confidence_level",
    ] = "Medium"

    priority_full.loc[
        priority_full[
            "selected_period_cells"
        ].ge(2)
        & priority_full[
            "selected_period_records"
        ].ge(5),
        "confidence_level",
    ] = "High"

    priority_full[
        "finding_type"
    ] = "Priority Lead"

    priority_full[
        "meaning"
    ] = (
        "Part, Spot, global rarity, "
        "multi-cell presence और activity "
        "का combined ranking."
    )

    priority_full[
        "why_it_matters"
    ] = (
        "Investigator review order तय करने "
        "में सहायता करता है; guilt finding "
        "नहीं है।"
    )

    priority_full = (
        priority_full
        .sort_values(
            [
                "priority_score",
                "selected_period_cells",
                "selected_period_records",
            ],
            ascending=[
                False,
                False,
                False,
            ],
        )
    )

    part_uncommon_numbers = (
        part_uncommon_full.head(
            safe_limit
        ).reset_index(
            drop=True
        )
    )

    spot_uncommon_numbers = (
        spot_uncommon_full.head(
            safe_limit
        ).reset_index(
            drop=True
        )
    )

    global_uncommon_numbers = (
        global_uncommon_full.head(
            safe_limit
        ).reset_index(
            drop=True
        )
    )

    uncommon_numbers = (
        uncommon_full.head(
            safe_limit
        ).reset_index(
            drop=True
        )
    )

    common_numbers = (
        common_full.sort_values(
            [
                "selected_period_cells",
                "selected_period_records",
                "global_other_records",
            ],
            ascending=False,
        )
        .head(
            safe_limit
        )
        .reset_index(
            drop=True
        )
    )

    multi_cell_presence = (
        multi_cell_full.sort_values(
            [
                "selected_period_cells",
                "selected_period_records",
            ],
            ascending=False,
        )
        .head(
            safe_limit
        )
        .reset_index(
            drop=True
        )
    )

    repeat_presence = (
        repeat_full.sort_values(
            [
                "global_other_records",
                "selected_period_records",
            ],
            ascending=False,
        )
        .head(
            safe_limit
        )
        .reset_index(
            drop=True
        )
    )

    device_consistency = (
        device_full.sort_values(
            [
                "selected_period_imei",
                "selected_period_imsi",
                "selected_period_records",
            ],
            ascending=False,
        )
        .head(
            safe_limit
        )
        .reset_index(
            drop=True
        )
    )

    suspicious_timing = (
        suspicious_full.sort_values(
            [
                "selected_period_cells",
                "selected_period_records",
            ],
            ascending=False,
        )
        .head(
            safe_limit
        )
        .reset_index(
            drop=True
        )
    )

    priority_leads = (
        priority_full.head(
            safe_limit
        ).reset_index(
            drop=True
        )
    )

    lead_summary = pd.DataFrame(
        [
            {
                "finding": "Common Numbers",
                "records": len(common_full),
                "displayed_records": (
                    len(common_numbers)
                ),
                "meaning": (
                    "Selected Part और loaded "
                    "data के अन्य हिस्से में "
                    "भी presence."
                ),
            },
            {
                "finding": "Part-Uncommon",
                "records": (
                    len(part_uncommon_full)
                ),
                "displayed_records": (
                    len(part_uncommon_numbers)
                ),
                "meaning": (
                    "दूसरे configured Part "
                    "में नहीं मिला."
                ),
            },
            {
                "finding": "Spot-Uncommon",
                "records": (
                    len(spot_uncommon_full)
                ),
                "displayed_records": (
                    len(spot_uncommon_numbers)
                ),
                "meaning": (
                    "इसी Spot के बाकी समय "
                    "में नहीं मिला."
                ),
            },
            {
                "finding": "Global-Uncommon",
                "records": (
                    len(global_uncommon_full)
                ),
                "displayed_records": (
                    len(global_uncommon_numbers)
                ),
                "meaning": (
                    "Selected Part के बाहर "
                    "कहीं नहीं मिला."
                ),
            },
            {
                "finding": (
                    "Multi-Cell Presence"
                ),
                "records": (
                    len(multi_cell_full)
                ),
                "displayed_records": (
                    len(multi_cell_presence)
                ),
                "meaning": (
                    "Selected Part में multiple "
                    "searched cells."
                ),
            },
            {
                "finding": "Repeat Presence",
                "records": len(repeat_full),
                "displayed_records": (
                    len(repeat_presence)
                ),
                "meaning": (
                    "Selected Part के बाहर "
                    "भी presence."
                ),
            },
            {
                "finding": (
                    "IMEI/IMSI Consistency"
                ),
                "records": (
                    len(device_full)
                ),
                "displayed_records": (
                    len(device_consistency)
                ),
                "meaning": (
                    "Device/SIM continuity "
                    "और change indicators."
                ),
            },
            {
                "finding": (
                    "Suspicious Timing"
                ),
                "records": (
                    len(suspicious_full)
                ),
                "displayed_records": (
                    len(suspicious_timing)
                ),
                "meaning": (
                    "Notable activity in "
                    "selected Part."
                ),
            },
            {
                "finding": "Priority Leads",
                "records": len(priority_full),
                "displayed_records": (
                    len(priority_leads)
                ),
                "meaning": (
                    "Combined review ranking."
                ),
            },
        ]
    )

    return {
        "summary": summary,
        "lead_summary": lead_summary,
        "uncommon_classification": (
            classification.reset_index(
                drop=True
            )
        ),
        "part_uncommon_numbers": (
            part_uncommon_numbers
        ),
        "spot_uncommon_numbers": (
            spot_uncommon_numbers
        ),
        "global_uncommon_numbers": (
            global_uncommon_numbers
        ),
        "common_numbers": common_numbers,
        "uncommon_numbers": uncommon_numbers,
        "multi_cell_presence": (
            multi_cell_presence
        ),
        "repeat_presence": (
            repeat_presence
        ),
        "device_consistency": (
            device_consistency
        ),
        "suspicious_timing": (
            suspicious_timing
        ),
        "priority_leads": priority_leads,
    }



def tower_ipdr_partwise_range_report_root(case_id: str) -> Path:
    return (
        Path("cases")
        / "active"
        / str(case_id)
        / "reports"
        / "tower_dump"
        / "ipdr"
        / "partwise_range"
    )


def _report_json_safe(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(k): _report_json_safe(v) for k, v in value.items()}

    if isinstance(value, list):
        return [_report_json_safe(item) for item in value]

    if isinstance(value, tuple):
        return [_report_json_safe(item) for item in value]

    try:
        if pd.isna(value):
            return None
    except Exception:
        pass

    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            return str(value)

    return value



def export_tower_ipdr_partwise_range_report(
    case_id: str,
    parts: list[dict[str, Any]],
    *,
    comparison_parts: list[dict[str, Any]] | None = None,
    precomputed_results: dict[int, dict[str, Any]] | None = None,
    lead_limit: int = 50,
    max_leads_in_text: int = 20,
) -> dict[str, Any]:
    """Export Spot-aware Tower IPDR Part-wise reports.

    The function reuses precomputed analysis results when available.
    It creates TXT, CSV, Excel, manifest and latest-report files.
    """

    import io
    from contextlib import redirect_stdout

    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    from openpyxl.utils import (
        get_column_letter,
    )

    if not parts:
        raise ValueError(
            "No Date-Time Parts are available."
        )

    configured_parts = list(
        comparison_parts
        or parts
    )

    cached_results = dict(
        precomputed_results
        or {}
    )

    timestamp = datetime.now().strftime(
        "%Y%m%d_%H%M%S"
    )

    run_id = (
        "tower_ipdr_partwise_range_"
        f"{timestamp}"
    )

    output_dir = (
        tower_ipdr_partwise_range_report_root(
            case_id
        )
        / run_id
    )

    output_dir.mkdir(
        parents=True,
        exist_ok=True,
    )

    saved_files: dict[str, str] = {}
    part_summaries: list[dict[str, Any]] = []
    combined_tables: dict[
        str,
        list[pd.DataFrame],
    ] = {}

    text_buffer = io.StringIO()

    print("=" * 78, file=text_buffer)
    print(
        "TOWER IPDR PART-WISE INVESTIGATION REPORT",
        file=text_buffer,
    )
    print("=" * 78, file=text_buffer)
    print(
        f"Case ID        : {case_id}",
        file=text_buffer,
    )
    print(
        f"Report Run     : {run_id}",
        file=text_buffer,
    )
    print(
        f"Parts Analyzed : {len(parts)}",
        file=text_buffer,
    )
    print(
        f"Parts Compared : {len(configured_parts)}",
        file=text_buffer,
    )
    print("", file=text_buffer)
    print(
        "Important: These findings are "
        "investigation leads. Verify every lead "
        "with independent records and field evidence.",
        file=text_buffer,
    )

    table_guidance = {
        "priority_leads": (
            "Priority Lead",
            (
                "Combined ranking based on rarity, "
                "activity and multi-cell presence."
            ),
            (
                "Review with CDR, SDR/CAF, IMEI, "
                "IMSI and field information."
            ),
        ),
        "part_uncommon_numbers": (
            "Part-Uncommon",
            (
                "Found in this Part but not in "
                "another configured Part."
            ),
            (
                "Compare with incident timing and "
                "all configured Parts."
            ),
        ),
        "spot_uncommon_numbers": (
            "Spot-Uncommon",
            (
                "Found in this Part but not elsewhere "
                "in the same Spot."
            ),
            (
                "Verify against local and regular "
                "users of the selected Spot."
            ),
        ),
        "global_uncommon_numbers": (
            "Global-Uncommon",
            (
                "Not found anywhere outside the "
                "selected Part."
            ),
            (
                "Treat as a strong rarity lead, "
                "not as proof."
            ),
        ),
        "uncommon_numbers": (
            "Uncommon Number",
            (
                "Combined uncommon list across Part, "
                "Spot and global scope."
            ),
            (
                "Review the Part, Spot and global "
                "status before prioritizing the lead."
            ),
        ),
        "common_numbers": (
            "Common Number",
            (
                "Also found outside the selected Part."
            ),
            (
                "Check whether the number is local, "
                "regular or linked."
            ),
        ),
        "multi_cell_presence": (
            "Multi-Cell Presence",
            (
                "Seen on multiple searched cells "
                "inside the selected Part."
            ),
            (
                "Review possible movement and area "
                "presence."
            ),
        ),
        "repeat_presence": (
            "Repeat Presence",
            (
                "Repeatedly seen outside the "
                "selected Part."
            ),
            (
                "Verify the person's role before "
                "drawing any conclusion."
            ),
        ),
        "device_consistency": (
            "IMEI/IMSI Consistency",
            (
                "Device and SIM identifier "
                "continuity check."
            ),
            (
                "Verify possible device or SIM "
                "changes with operator records."
            ),
        ),
        "suspicious_timing": (
            "Timing / High Activity",
            (
                "Notable activity inside the "
                "selected Part."
            ),
            (
                "Compare with incident time, route "
                "and field information."
            ),
        ),
    }

    lead_meanings = {
        "Common Numbers": (
            "Also seen outside the selected Part."
        ),
        "Part-Uncommon": (
            "Not seen in another configured Part."
        ),
        "Spot-Uncommon": (
            "Not seen elsewhere in the same Spot."
        ),
        "Global-Uncommon": (
            "Not seen anywhere outside this Part."
        ),
        "Multi-Cell Presence": (
            "Seen on multiple searched cells."
        ),
        "Repeat Presence": (
            "Repeatedly seen elsewhere."
        ),
        "IMEI/IMSI Consistency": (
            "Device and SIM continuity check."
        ),
        "Suspicious Timing": (
            "Notable activity in this Part."
        ),
        "Priority Leads": (
            "Combined investigator review ranking."
        ),
    }

    def public_frame(
        frame: pd.DataFrame,
        table_name: str,
    ) -> pd.DataFrame:
        output = frame.copy()

        if output.empty:
            return output

        if table_name == "lead_summary":
            if "finding" in output.columns:
                output["meaning"] = (
                    output["finding"]
                    .map(lead_meanings)
                    .fillna(
                        "Investigation finding."
                    )
                )

            return output

        guidance = table_guidance.get(
            table_name
        )

        if guidance:
            finding_type, meaning, action = (
                guidance
            )

            output["finding_type"] = (
                finding_type
            )
            output["meaning"] = meaning
            output[
                "why_it_matters"
            ] = (
                "This is an investigation lead "
                "and requires independent verification."
            )
            output[
                "suggested_action"
            ] = action

        return output

    for part in parts:
        part_number = int(
            part.get(
                "part_no",
                0,
            )
            or 0
        )

        part_name = str(
            part.get(
                "part_name",
                f"Part {part_number}",
            )
        )

        start_time = str(
            part.get(
                "start_time",
                "",
            )
        )

        end_time = str(
            part.get(
                "end_time",
                "",
            )
        )

        spot_id = str(
            part.get(
                "spot_id",
                "",
            )
            or ""
        )

        spot_name = str(
            part.get(
                "spot_name",
                "",
            )
            or ""
        )

        result = cached_results.get(
            part_number
        )

        if result is None:
            result = (
                tower_ipdr_range_investigation_summary(
                    case_id,
                    start_time,
                    end_time,
                    spot_id=spot_id,
                    spot_name=spot_name,
                    comparison_parts=(
                        configured_parts
                    ),
                    current_part_no=(
                        part_number
                    ),
                    lead_limit=lead_limit,
                )
            )

        print(
            "\n" + "#" * 78,
            file=text_buffer,
        )
        print(
            f"{part_name.upper()} REPORT",
            file=text_buffer,
        )
        print(
            "#" * 78,
            file=text_buffer,
        )
        print(
            f"Spot  : {spot_id or 'ALL-SPOTS'}"
            + (
                f" | {spot_name}"
                if spot_name
                else ""
            ),
            file=text_buffer,
        )
        print(
            f"Period: {start_time} to {end_time}",
            file=text_buffer,
        )

        summary_frame = result.get(
            "summary",
            pd.DataFrame(),
        )

        if (
            isinstance(
                summary_frame,
                pd.DataFrame,
            )
            and not summary_frame.empty
        ):
            summary_row = (
                summary_frame.iloc[0].to_dict()
            )
        else:
            summary_row = {}

        part_summary = {
            "part_no": part_number,
            "part_name": part_name,
            "spot_id": spot_id,
            "spot_name": spot_name,
            "spot_scope_mode": (
                summary_row.get(
                    "spot_scope_mode",
                    part.get(
                        "spot_scope_mode",
                        "",
                    ),
                )
            ),
            "start_time": start_time,
            "end_time": end_time,
            "records_found": (
                summary_row.get(
                    "records_found",
                    0,
                )
            ),
            "numbers_found": (
                summary_row.get(
                    "numbers_found",
                    0,
                )
            ),
            "cells_involved": (
                summary_row.get(
                    "cells_involved",
                    0,
                )
            ),
            "first_activity": (
                summary_row.get(
                    "first_activity",
                    "",
                )
            ),
            "last_activity": (
                summary_row.get(
                    "last_activity",
                    "",
                )
            ),
        }

        part_summaries.append(
            part_summary
        )

        for table_name, frame in (
            result.items()
        ):
            if not isinstance(
                frame,
                pd.DataFrame,
            ):
                continue

            clean_frame = public_frame(
                frame,
                table_name,
            )

            csv_path = (
                output_dir
                / (
                    f"part_{part_number:02d}_"
                    f"{table_name}.csv"
                )
            )

            clean_frame.to_csv(
                csv_path,
                index=False,
            )

            saved_files[
                (
                    f"part_{part_number:02d}_"
                    f"{table_name}"
                )
            ] = str(
                csv_path
            )

            combined = clean_frame.copy()

            metadata_columns = [
                ("part_no", part_number),
                ("part_name", part_name),
                ("spot_id", spot_id),
                ("spot_name", spot_name),
                ("part_start", start_time),
                ("part_end", end_time),
            ]

            existing_metadata = [
                column_name
                for column_name, _ in metadata_columns
                if column_name in combined.columns
            ]

            if existing_metadata:
                combined = combined.drop(
                    columns=existing_metadata,
                )

            for column_index, (
                column_name,
                column_value,
            ) in enumerate(metadata_columns):
                combined.insert(
                    column_index,
                    column_name,
                    column_value,
                )

            combined_tables.setdefault(
                table_name,
                [],
            ).append(
                combined
            )

        with redirect_stdout(
            text_buffer
        ):
            print_tower_ipdr_investigation_summary(
                result,
                max_leads=max_leads_in_text,
            )

    all_parts_summary = pd.DataFrame(
        part_summaries
    )

    summary_csv = (
        output_dir
        / "all_parts_summary.csv"
    )

    all_parts_summary.to_csv(
        summary_csv,
        index=False,
    )

    saved_files[
        "all_parts_summary"
    ] = str(
        summary_csv
    )

    combined_frames: dict[
        str,
        pd.DataFrame,
    ] = {}

    for table_name, frames in (
        combined_tables.items()
    ):
        combined_frames[
            table_name
        ] = pd.concat(
            frames,
            ignore_index=True,
            sort=False,
        )

    text_path = (
        output_dir
        / "investigation_summary_all_parts.txt"
    )

    text_path.write_text(
        text_buffer.getvalue(),
        encoding="utf-8",
    )

    saved_files[
        "investigation_summary_all_parts"
    ] = str(
        text_path
    )

    excel_path = (
        output_dir
        / "tower_ipdr_partwise_analysis.xlsx"
    )

    sheet_tables = [
        (
            "Part Summary",
            all_parts_summary,
        ),
        (
            "Lead Summary",
            combined_frames.get(
                "lead_summary",
                pd.DataFrame(),
            ),
        ),
        (
            "Priority Leads",
            combined_frames.get(
                "priority_leads",
                pd.DataFrame(),
            ),
        ),
        (
            "Part Uncommon",
            combined_frames.get(
                "part_uncommon_numbers",
                pd.DataFrame(),
            ),
        ),
        (
            "Spot Uncommon",
            combined_frames.get(
                "spot_uncommon_numbers",
                pd.DataFrame(),
            ),
        ),
        (
            "Global Uncommon",
            combined_frames.get(
                "global_uncommon_numbers",
                pd.DataFrame(),
            ),
        ),
        (
            "Common Numbers",
            combined_frames.get(
                "common_numbers",
                pd.DataFrame(),
            ),
        ),
        (
            "Multi Cell",
            combined_frames.get(
                "multi_cell_presence",
                pd.DataFrame(),
            ),
        ),
        (
            "Repeat Presence",
            combined_frames.get(
                "repeat_presence",
                pd.DataFrame(),
            ),
        ),
        (
            "Device SIM",
            combined_frames.get(
                "device_consistency",
                pd.DataFrame(),
            ),
        ),
        (
            "Timing Activity",
            combined_frames.get(
                "suspicious_timing",
                pd.DataFrame(),
            ),
        ),
        (
            "Classification",
            combined_frames.get(
                "uncommon_classification",
                pd.DataFrame(),
            ),
        ),
    ]

    methodology = pd.DataFrame(
        [
            {
                "topic": "Range rule",
                "explanation": (
                    "Each Part uses "
                    "start_time <= event_time < end_time."
                ),
            },
            {
                "topic": "Spot scope",
                "explanation": (
                    "Each new Part is limited to "
                    "its selected Spot."
                ),
            },
            {
                "topic": "Part-Uncommon",
                "explanation": (
                    "The number is not seen in another "
                    "configured Part."
                ),
            },
            {
                "topic": "Spot-Uncommon",
                "explanation": (
                    "The number is not seen elsewhere "
                    "in the same Spot."
                ),
            },
            {
                "topic": "Global-Uncommon",
                "explanation": (
                    "The number is not seen anywhere "
                    "outside the selected Part."
                ),
            },
            {
                "topic": "Investigation limit",
                "explanation": (
                    "Tower and IPDR presence is not "
                    "proof of exact location, identity "
                    "or involvement."
                ),
            },
        ]
    )

    with pd.ExcelWriter(
        excel_path,
        engine="openpyxl",
    ) as writer:
        for sheet_name, frame in (
            sheet_tables
        ):
            export_frame = (
                frame
                if not frame.empty
                else pd.DataFrame(
                    {
                        "status": [
                            "No records available"
                        ]
                    }
                )
            )

            export_frame.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
            )

        methodology.to_excel(
            writer,
            sheet_name="Methodology",
            index=False,
        )

        workbook = writer.book

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78",
        )

        header_font = Font(
            bold=True,
            color="FFFFFF",
        )

        thin_side = Side(
            style="thin",
            color="D9E2F3",
        )

        thin_border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side,
        )

        for worksheet in (
            workbook.worksheets
        ):
            worksheet.freeze_panes = "A2"

            if worksheet.max_row >= 1:
                worksheet.auto_filter.ref = (
                    worksheet.dimensions
                )

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
                cell.border = thin_border

            for row in worksheet.iter_rows(
                min_row=2,
            ):
                for cell in row:
                    cell.alignment = Alignment(
                        vertical="top",
                        wrap_text=True,
                    )
                    cell.border = thin_border

            for column_cells in (
                worksheet.columns
            ):
                column_letter = (
                    get_column_letter(
                        column_cells[0].column
                    )
                )

                maximum_length = 0

                for cell in column_cells:
                    value = (
                        ""
                        if cell.value is None
                        else str(cell.value)
                    )

                    maximum_length = max(
                        maximum_length,
                        min(
                            len(value),
                            60,
                        ),
                    )

                worksheet.column_dimensions[
                    column_letter
                ].width = min(
                    max(
                        maximum_length + 2,
                        12,
                    ),
                    45,
                )

    saved_files[
        "excel_workbook"
    ] = str(
        excel_path
    )

    manifest_path = (
        output_dir
        / "manifest.json"
    )

    saved_files[
        "manifest"
    ] = str(
        manifest_path
    )

    manifest = {
        "case_id": str(case_id),
        "run_id": run_id,
        "created_at": _now_iso(),
        "analysis_type": (
            "Tower IPDR Spot-aware "
            "Part-wise Analysis"
        ),
        "range_rule": (
            "start_time <= event_time < end_time"
        ),
        "parts_analyzed": len(parts),
        "parts_compared": (
            len(configured_parts)
        ),
        "lead_limit": int(
            lead_limit
        ),
        "output_dir": str(
            output_dir
        ),
        "parts": _report_json_safe(
            part_summaries
        ),
        "saved_files": _report_json_safe(
            saved_files
        ),
        "note": (
            "All findings require verification "
            "from independent records and "
            "field evidence."
        ),
    }

    manifest_path.write_text(
        json.dumps(
            _report_json_safe(
                manifest
            ),
            ensure_ascii=False,
            indent=2,
            sort_keys=True,
        ),
        encoding="utf-8",
    )

    latest_report_path = (
        save_tower_ipdr_partwise_latest_report(
            case_id,
            manifest,
        )
    )

    saved_files[
        "latest_report"
    ] = str(
        latest_report_path
    )

    manifest[
        "saved_files"
    ] = _report_json_safe(
        saved_files
    )

    manifest_path.write_text(
        json.dumps(
            _report_json_safe(
                manifest
            ),
            ensure_ascii=False,
            indent=2,
            sort_keys=True,
        ),
        encoding="utf-8",
    )

    return manifest



def _simple_join_unique(values: Any) -> str:
    cleaned = []

    for value in values:
        if value is None:
            continue

        text = str(value).strip()

        if not text or text.lower() in {"nan", "none"}:
            continue

        cleaned.append(text)

    return ", ".join(sorted(set(cleaned)))


def _priority_sort_value(value: Any) -> int:
    order = {
        "High": 1,
        "Medium-High": 2,
        "Medium": 3,
        "Low": 4,
    }
    return order.get(str(value), 9)


def tower_ipdr_compare_date_time_parts(
    case_id: str,
    parts: list[dict[str, Any]],
    *,
    lead_limit: int = 200,
) -> dict[str, pd.DataFrame]:
    """Compare mobile numbers across saved Date-Time Parts.

    Purpose:
    - Common across parts
    - Part-only numbers
    - Combined priority leads
    """

    database_path = tower_ipdr_database_path(case_id)
    store = DuckDBStore(database_path)

    empty = {
        "comparison_summary": pd.DataFrame(),
        "part_level_counts": pd.DataFrame(),
        "common_across_parts": pd.DataFrame(),
        "part_only_numbers": pd.DataFrame(),
        "combined_priority_leads": pd.DataFrame(),
        "all_part_presence": pd.DataFrame(),
    }

    if not parts or not store.table_exists(TABLE_EVENTS):
        return empty

    presence_frames: list[pd.DataFrame] = []

    for part in parts:
        part_no = int(part.get("part_no", 0))
        part_name = str(part.get("part_name") or f"Part {part_no}")
        start_time = str(part.get("start_time"))
        end_time = str(part.get("end_time"))

        dataframe = store.query_df(
            f"""
            SELECT
                ? AS part_no,
                ? AS part_name,
                ? AS start_time,
                ? AS end_time,
                subscriber_number AS mobile_number,
                COUNT(*) AS records_found,
                COUNT(DISTINCT searched_cell_id) AS cells_seen,
                COUNT(DISTINCT imei) AS imei_count,
                COUNT(DISTINCT imsi) AS imsi_count,
                MIN(TRY_CAST(event_time AS TIMESTAMP)) AS first_seen,
                MAX(TRY_CAST(event_time AS TIMESTAMP)) AS last_seen
            FROM {TABLE_EVENTS}
            WHERE TRY_CAST(event_time AS TIMESTAMP) >= CAST(? AS TIMESTAMP)
              AND TRY_CAST(event_time AS TIMESTAMP) < CAST(? AS TIMESTAMP)
              AND subscriber_number IS NOT NULL
              AND subscriber_number <> ''
            GROUP BY subscriber_number
            """,
            [
                part_no,
                part_name,
                start_time,
                end_time,
                start_time,
                end_time,
            ],
        )

        if isinstance(dataframe, pd.DataFrame) and not dataframe.empty:
            presence_frames.append(dataframe)

    if not presence_frames:
        return empty

    all_part_presence = pd.concat(presence_frames, ignore_index=True)
    all_part_presence["mobile_number"] = all_part_presence["mobile_number"].astype(str)

    combined_rows: list[dict[str, Any]] = []

    for mobile_number, group in all_part_presence.groupby("mobile_number", dropna=True):
        total_records = int(group["records_found"].sum())
        parts_seen_count = int(group["part_no"].nunique())
        max_cells_seen = int(group["cells_seen"].max())
        total_part_entries = int(len(group))

        if parts_seen_count >= 2 and max_cells_seen >= 2:
            priority = "High"
            confidence = "High"
            reason = "Seen in multiple Date-Time Parts and more than one searched cell"
        elif parts_seen_count >= 2:
            priority = "Medium-High"
            confidence = "Medium"
            reason = "Seen in multiple Date-Time Parts"
        elif max_cells_seen >= 2:
            priority = "Medium"
            confidence = "Medium"
            reason = "Seen in one Date-Time Part but on multiple searched cells"
        elif total_records >= 10:
            priority = "Medium"
            confidence = "Medium"
            reason = "High activity in one Date-Time Part"
        else:
            priority = "Low"
            confidence = "Low"
            reason = "Seen in one Date-Time Part with limited records"

        combined_rows.append(
            {
                "mobile_number": mobile_number,
                "parts_seen_count": parts_seen_count,
                "parts_seen": _simple_join_unique(group["part_name"].tolist()),
                "part_ranges": _simple_join_unique(
                    [
                        f"{row.start_time} to {row.end_time}"
                        for row in group.itertuples(index=False)
                    ]
                ),
                "total_records": total_records,
                "max_cells_seen_in_any_part": max_cells_seen,
                "max_imei_count_in_any_part": int(group["imei_count"].max()),
                "max_imsi_count_in_any_part": int(group["imsi_count"].max()),
                "first_seen": group["first_seen"].min(),
                "last_seen": group["last_seen"].max(),
                "priority": priority,
                "confidence_level": confidence,
                "simple_reason": reason,
                "suggested_action": (
                    "Verify with CDR/SDR/CAF, IMEI/IMSI, tower location and field information."
                ),
                "part_entries": total_part_entries,
            }
        )

    combined_priority_leads = pd.DataFrame(combined_rows)

    if combined_priority_leads.empty:
        return empty

    combined_priority_leads["_priority_sort"] = combined_priority_leads["priority"].map(
        _priority_sort_value
    )
    combined_priority_leads = combined_priority_leads.sort_values(
        by=[
            "_priority_sort",
            "parts_seen_count",
            "max_cells_seen_in_any_part",
            "total_records",
        ],
        ascending=[True, False, False, False],
    ).drop(columns=["_priority_sort"])

    common_across_parts = combined_priority_leads[
        combined_priority_leads["parts_seen_count"] >= 2
    ].head(lead_limit)

    part_only_mobile_numbers = set(
        combined_priority_leads.loc[
            combined_priority_leads["parts_seen_count"] == 1,
            "mobile_number",
        ].astype(str)
    )

    part_only_numbers = all_part_presence[
        all_part_presence["mobile_number"].astype(str).isin(part_only_mobile_numbers)
    ].copy()

    if not part_only_numbers.empty:
        part_only_numbers = part_only_numbers.merge(
            combined_priority_leads[
                [
                    "mobile_number",
                    "priority",
                    "confidence_level",
                    "simple_reason",
                    "suggested_action",
                ]
            ],
            on="mobile_number",
            how="left",
        )
        part_only_numbers["_priority_sort"] = part_only_numbers["priority"].map(
            _priority_sort_value
        )
        part_only_numbers = part_only_numbers.sort_values(
            by=["_priority_sort", "cells_seen", "records_found"],
            ascending=[True, False, False],
        ).drop(columns=["_priority_sort"]).head(lead_limit)

    part_level_rows: list[dict[str, Any]] = []

    for part_key, group in all_part_presence.groupby(
        ["part_no", "part_name", "start_time", "end_time"],
        dropna=False,
    ):
        part_no, part_name, start_time, end_time = part_key

        part_level_rows.append(
            {
                "part_no": part_no,
                "part_name": part_name,
                "start_time": start_time,
                "end_time": end_time,
                "unique_numbers": int(group["mobile_number"].nunique()),
                "total_records": int(group["records_found"].sum()),
                "multi_cell_numbers": int((group["cells_seen"] >= 2).sum()),
                "high_activity_numbers": int((group["records_found"] >= 10).sum()),
            }
        )

    part_level_counts = pd.DataFrame(part_level_rows).sort_values("part_no")

    comparison_summary = pd.DataFrame(
        [
            {
                "finding": "Total Date-Time Parts",
                "count": len(parts),
                "meaning": "Total saved investigation periods.",
            },
            {
                "finding": "Unique Numbers Across All Parts",
                "count": int(combined_priority_leads["mobile_number"].nunique()),
                "meaning": "Numbers found in any selected Date-Time Part.",
            },
            {
                "finding": "Common Across Parts",
                "count": int((combined_priority_leads["parts_seen_count"] >= 2).sum()),
                "meaning": "Numbers found in two or more Date-Time Parts.",
            },
            {
                "finding": "Part-Only Numbers",
                "count": int((combined_priority_leads["parts_seen_count"] == 1).sum()),
                "meaning": "Numbers found in only one selected Date-Time Part.",
            },
            {
                "finding": "High Priority Combined Leads",
                "count": int((combined_priority_leads["priority"] == "High").sum()),
                "meaning": "Numbers with stronger combined importance across parts.",
            },
        ]
    )

    return {
        "comparison_summary": comparison_summary,
        "part_level_counts": part_level_counts,
        "common_across_parts": common_across_parts,
        "part_only_numbers": part_only_numbers,
        "combined_priority_leads": combined_priority_leads.head(lead_limit),
        "all_part_presence": all_part_presence,
    }


def print_tower_ipdr_part_comparison_summary(
    comparison: dict[str, pd.DataFrame],
    *,
    max_rows: int = 20,
) -> None:
    """Print part comparison in simple investigation language."""

    comparison_summary = comparison.get("comparison_summary", pd.DataFrame())
    part_level_counts = comparison.get("part_level_counts", pd.DataFrame())
    common_across_parts = comparison.get("common_across_parts", pd.DataFrame())
    part_only_numbers = comparison.get("part_only_numbers", pd.DataFrame())
    combined_priority_leads = comparison.get("combined_priority_leads", pd.DataFrame())

    print("\n" + "=" * 78)
    print("DATE-TIME PART COMPARISON SUMMARY")
    print("=" * 78)

    if isinstance(comparison_summary, pd.DataFrame) and not comparison_summary.empty:
        for _, row in comparison_summary.iterrows():
            print(f"- {row.get('finding')}: {row.get('count')}")
            print(f"  Meaning: {row.get('meaning')}")
    else:
        print("No comparison summary available.")

    print("\n" + "-" * 78)
    print("PART-WISE BASIC COUNTS")
    print("-" * 78)

    if isinstance(part_level_counts, pd.DataFrame) and not part_level_counts.empty:
        for _, row in part_level_counts.iterrows():
            print(f"{row.get('part_name')} | {row.get('start_time')} to {row.get('end_time')}")
            print(f"  Unique Numbers      : {row.get('unique_numbers')}")
            print(f"  Total Records       : {row.get('total_records')}")
            print(f"  Multi-Cell Numbers  : {row.get('multi_cell_numbers')}")
            print(f"  High Activity Nos.  : {row.get('high_activity_numbers')}")
    else:
        print("No part-wise count available.")

    print("\n" + "-" * 78)
    print("TOP COMBINED PRIORITY LEADS")
    print("-" * 78)
    print("Meaning: Numbers ranked by presence across parts, multi-cell presence and activity.")
    _print_simple_leads(combined_priority_leads, max_rows=max_rows)

    print("\n" + "-" * 78)
    print("COMMON ACROSS DATE-TIME PARTS")
    print("-" * 78)
    print("Meaning: Numbers found in two or more selected Date-Time Parts.")
    print("Use: These may show repeated presence across investigation periods.")
    _print_simple_leads(common_across_parts, max_rows=max_rows)

    print("\n" + "-" * 78)
    print("PART-ONLY NUMBERS")
    print("-" * 78)
    print("Meaning: Numbers found in only one selected Date-Time Part.")
    print("Use: These may be period-specific visitors or one-time presence leads.")
    _print_simple_leads(part_only_numbers, max_rows=max_rows)


def _excel_safe_sheet_name(name: Any, used_names: set[str]) -> str:
    """Create Excel-safe unique sheet name."""

    raw = str(name or "Sheet").strip()
    cleaned = ""

    for char in raw:
        if char in r'[]:*?/\\':
            cleaned += "_"
        else:
            cleaned += char

    cleaned = cleaned.strip() or "Sheet"
    cleaned = cleaned[:31]

    candidate = cleaned
    counter = 2

    while candidate in used_names:
        suffix = f"_{counter}"
        candidate = cleaned[: 31 - len(suffix)] + suffix
        counter += 1

    used_names.add(candidate)
    return candidate


def _excel_column_letter(index: int) -> str:
    """1-based column number to Excel letter."""

    letters = ""

    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters

    return letters or "A"


def _write_dataframe_to_excel_sheet(
    workbook: Any,
    sheet_name: str,
    dataframe: pd.DataFrame,
    *,
    used_names: set[str],
    max_rows: int = 50000,
) -> None:
    """Write a DataFrame to an Excel worksheet with safe formatting."""

    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows

    safe_name = _excel_safe_sheet_name(sheet_name, used_names)
    worksheet = workbook.create_sheet(title=safe_name)

    if dataframe is None or not isinstance(dataframe, pd.DataFrame) or dataframe.empty:
        worksheet["A1"] = "No data available in this section."
        worksheet["A1"].font = Font(bold=True)
        return

    original_rows = len(dataframe)

    export_df = dataframe.copy()

    if len(export_df) > max_rows:
        export_df = export_df.head(max_rows)

    for row in dataframe_to_rows(export_df, index=False, header=True):
        worksheet.append(row)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_alignment

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap_alignment

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_index, column_cells in enumerate(worksheet.columns, start=1):
        max_length = 10

        for cell in column_cells[:200]:
            value = cell.value
            if value is None:
                continue
            max_length = max(max_length, min(len(str(value)), 50))

        worksheet.column_dimensions[_excel_column_letter(column_index)].width = min(
            max_length + 2,
            55,
        )

    if original_rows > max_rows:
        note_row = worksheet.max_row + 2
        worksheet.cell(
            row=note_row,
            column=1,
            value=(
                f"Note: Sheet truncated for Excel performance. "
                f"Showing first {max_rows:,} of {original_rows:,} rows."
            ),
        )
        worksheet.cell(row=note_row, column=1).font = Font(bold=True)


def export_tower_ipdr_excel_workbook_from_manifest(
    manifest: dict[str, Any],
    *,
    max_rows_per_sheet: int = 50000,
) -> str:
    """Create consolidated Excel workbook from an exported part-wise report manifest.

    This function uses already exported CSV files, so it does not rerun heavy analysis.
    """

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    output_dir = Path(str(manifest.get("output_dir", "")))

    if not output_dir.exists():
        raise ValueError("Report output folder nahi mila.")

    saved_files = dict(manifest.get("saved_files", {}))

    workbook_path = output_dir / "tower_ipdr_partwise_investigation_report.xlsx"

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    used_names: set[str] = set()

    # ------------------------------------------------------------------
    # Report Index Sheet
    # ------------------------------------------------------------------
    index_sheet = workbook.create_sheet(
        title=_excel_safe_sheet_name("Report Index", used_names)
    )

    index_rows = [
        ("Report Type", manifest.get("analysis_type", "Tower IPDR Part-wise Analysis")),
        ("Case ID", manifest.get("case_id", "")),
        ("Run ID", manifest.get("run_id", "")),
        ("Created At", manifest.get("created_at", "")),
        ("Parts Count", manifest.get("parts_count", "")),
        ("Display Rule", manifest.get("display_rule", "")),
        ("Internal Range Rule", manifest.get("range_rule", "")),
        (
            "Important Note",
            "This report provides investigation leads. Final conclusion requires verification.",
        ),
    ]

    for row_no, (key, value) in enumerate(index_rows, start=1):
        index_sheet.cell(row=row_no, column=1, value=key)
        index_sheet.cell(row=row_no, column=2, value=value)

    index_sheet["A1"].font = Font(bold=True)
    index_sheet["B1"].font = Font(bold=True)

    for cell in index_sheet["A"]:
        cell.font = Font(bold=True)

    index_sheet.column_dimensions["A"].width = 24
    index_sheet.column_dimensions["B"].width = 90

    start_row = len(index_rows) + 3
    index_sheet.cell(row=start_row, column=1, value="Saved Excel Sections")
    index_sheet.cell(row=start_row, column=1).font = Font(bold=True)
    index_sheet.cell(row=start_row, column=1).fill = PatternFill("solid", fgColor="D9EAF7")

    row_no = start_row + 1

    for key, path_value in sorted(saved_files.items()):
        if str(path_value).lower().endswith(".csv"):
            index_sheet.cell(row=row_no, column=1, value=key)
            index_sheet.cell(row=row_no, column=2, value=str(path_value))
            row_no += 1

    for row in index_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # ------------------------------------------------------------------
    # Preferred sheet order
    # ------------------------------------------------------------------
    preferred_order = [
        ("all_parts_summary", "All Parts Summary"),
        ("comparison_comparison_summary", "Comparison Summary"),
        ("comparison_part_level_counts", "Part Counts"),
        ("comparison_common_across_parts", "Common Across Parts"),
        ("comparison_part_only_numbers", "Part Only Numbers"),
        ("comparison_combined_priority_leads", "Combined Priority"),
        ("comparison_all_part_presence", "All Part Presence"),
    ]

    written_keys: set[str] = set()

    for key, sheet_name in preferred_order:
        csv_path = saved_files.get(key)

        if not csv_path:
            continue

        path = Path(str(csv_path))

        if not path.exists():
            continue

        dataframe = _read_csv_for_excel_safe(path)
        _write_dataframe_to_excel_sheet(
            workbook,
            sheet_name,
            dataframe,
            used_names=used_names,
            max_rows=max_rows_per_sheet,
        )
        written_keys.add(key)

    # ------------------------------------------------------------------
    # Part-specific sections
    # ------------------------------------------------------------------
    part_section_order = [
        "summary",
        "lead_summary",
        "priority_leads",
        "common_numbers",
        "uncommon_numbers",
        "multi_cell_presence",
        "repeat_presence",
        "device_consistency",
        "suspicious_timing",
    ]

    for part_no in range(1, int(manifest.get("parts_count", 0)) + 1):
        prefix = f"part_{part_no:02d}"

        for section in part_section_order:
            key = f"{prefix}_{section}"
            csv_path = saved_files.get(key)

            if not csv_path:
                continue

            path = Path(str(csv_path))

            if not path.exists():
                continue

            sheet_name = f"P{part_no} {section.replace('_', ' ').title()}"
            dataframe = _read_csv_for_excel_safe(path)

            _write_dataframe_to_excel_sheet(
                workbook,
                sheet_name,
                dataframe,
                used_names=used_names,
                max_rows=max_rows_per_sheet,
            )
            written_keys.add(key)

    # ------------------------------------------------------------------
    # Any remaining CSVs
    # ------------------------------------------------------------------
    for key, csv_path in sorted(saved_files.items()):
        if key in written_keys:
            continue

        path = Path(str(csv_path))

        if not str(path).lower().endswith(".csv"):
            continue

        if not path.exists():
            continue

        dataframe = _read_csv_for_excel_safe(path)
        sheet_name = key.replace("_", " ").title()

        _write_dataframe_to_excel_sheet(
            workbook,
            sheet_name,
            dataframe,
            used_names=used_names,
            max_rows=max_rows_per_sheet,
        )

    workbook.save(workbook_path)

    return str(workbook_path)


def _build_tower_ipdr_scope_warnings(
    *,
    selected_records: int,
    total_records: int,
    selected_start: object,
    selected_end: object,
) -> pd.DataFrame:
    """Build warnings when selected Date-Time Part covers too much or no data."""

    columns = [
        "warning_type",
        "severity",
        "selected_period",
        "selected_records",
        "total_records",
        "coverage_percent",
        "baseline_records",
        "simple_warning",
        "why_it_matters",
        "suggested_action",
    ]

    warnings: list[dict[str, object]] = []

    selected_records = int(selected_records or 0)
    total_records = int(total_records or 0)

    coverage_percent = 0.0
    baseline_records = 0

    if total_records > 0:
        coverage_percent = round((selected_records / total_records) * 100, 2)
        baseline_records = max(total_records - selected_records, 0)

    if total_records <= 0:
        warnings.append(
            {
                "warning_type": "No Loaded Data",
                "severity": "High",
                "selected_period": f"{selected_start} to {selected_end}",
                "selected_records": selected_records,
                "total_records": total_records,
                "coverage_percent": coverage_percent,
                "baseline_records": baseline_records,
                "simple_warning": "Loaded Tower IPDR data available nahi hai.",
                "why_it_matters": "Analysis result reliable nahi hoga jab dump data loaded na ho.",
                "suggested_action": "Pehle option 1: Load Dump Data chalayein.",
            }
        )

    elif selected_records <= 0:
        warnings.append(
            {
                "warning_type": "No Records In Selected Period",
                "severity": "Medium",
                "selected_period": f"{selected_start} to {selected_end}",
                "selected_records": selected_records,
                "total_records": total_records,
                "coverage_percent": coverage_percent,
                "baseline_records": baseline_records,
                "simple_warning": "Selected Date-Time Part me koi record nahi mila.",
                "why_it_matters": "Is part se investigation lead nahi niklega.",
                "suggested_action": "Date-Time range, dump period aur operator data verify karein.",
            }
        )

    elif coverage_percent >= 95:
        warnings.append(
            {
                "warning_type": "Selected Period Covers Almost Full Dump",
                "severity": "High",
                "selected_period": f"{selected_start} to {selected_end}",
                "selected_records": selected_records,
                "total_records": total_records,
                "coverage_percent": coverage_percent,
                "baseline_records": baseline_records,
                "simple_warning": "Selected Date-Time Part lagbhag poora loaded dump cover kar raha hai.",
                "why_it_matters": (
                    "Common/Uncommon comparison weak ho sakta hai, kyunki selected period ke bahar "
                    "comparison ke liye data bahut kam ya zero hai."
                ),
                "suggested_action": (
                    "Chhota incident-specific Date-Time Part banayein, ya incident se pehle/baad ka "
                    "comparison dump load karein."
                ),
            }
        )

    elif coverage_percent >= 75:
        warnings.append(
            {
                "warning_type": "Selected Period Covers Large Dump Portion",
                "severity": "Medium",
                "selected_period": f"{selected_start} to {selected_end}",
                "selected_records": selected_records,
                "total_records": total_records,
                "coverage_percent": coverage_percent,
                "baseline_records": baseline_records,
                "simple_warning": "Selected Date-Time Part loaded dump ka bada hissa cover kar raha hai.",
                "why_it_matters": "Common/Uncommon result useful hai, lekin comparison limited ho sakta hai.",
                "suggested_action": "Report interpret karte waqt baseline data size check karein.",
            }
        )

    return pd.DataFrame(warnings, columns=columns)


def tower_ipdr_partwise_latest_report_path(case_id: str) -> Path:
    return tower_ipdr_partwise_range_report_root(case_id) / "latest_report.json"


def save_tower_ipdr_partwise_latest_report(
    case_id: str,
    manifest: dict[str, Any],
) -> Path:
    """Save latest Tower IPDR part-wise report pointer.

    This is useful for:
    - View Case Reports
    - GUI report opening
    - quick latest report lookup
    """

    path = tower_ipdr_partwise_latest_report_path(case_id)
    path.parent.mkdir(parents=True, exist_ok=True)

    saved_files = dict(manifest.get("saved_files", {}))

    payload = {
        "case_id": str(case_id),
        "updated_at": _now_iso(),
        "report_type": "Tower IPDR Part-wise Date-Time Range Report",
        "run_id": manifest.get("run_id"),
        "output_dir": manifest.get("output_dir"),
        "main_report": saved_files.get("investigation_summary_all_parts"),
        "summary_csv": saved_files.get("all_parts_summary"),
        "excel_workbook": saved_files.get("excel_workbook"),
        "manifest": saved_files.get("manifest"),
        "parts_count": manifest.get("parts_count"),
        "note": (
            "Latest Tower IPDR part-wise report pointer. "
            "Generated files are investigation reports and should be verified before conclusion."
        ),
    }

    path.write_text(
        json.dumps(
            _report_json_safe(payload),
            ensure_ascii=False,
            indent=2,
            sort_keys=True,
        ),
        encoding="utf-8",
    )

    return path


def load_tower_ipdr_partwise_latest_report(case_id: str) -> dict[str, Any]:
    path = tower_ipdr_partwise_latest_report_path(case_id)

    if not path.exists():
        return {}

    return json.loads(path.read_text(encoding="utf-8"))


def _read_csv_for_excel_safe(path: Path) -> pd.DataFrame:
    """Read CSV safely for Excel export.

    Empty CSV sections should not stop the full Excel workbook export.
    """

    try:
        dataframe = pd.read_csv(path)
    except pd.errors.EmptyDataError:
        return pd.DataFrame(
            [
                {
                    "message": "No data available in this section.",
                    "source_file": str(path),
                }
            ]
        )

    if dataframe is None or (dataframe.empty and len(dataframe.columns) == 0):
        return pd.DataFrame(
            [
                {
                    "message": "No data available in this section.",
                    "source_file": str(path),
                }
            ]
        )

    return dataframe
