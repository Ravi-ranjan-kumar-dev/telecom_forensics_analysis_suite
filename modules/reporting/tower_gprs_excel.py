"""One consolidated Excel workbook for Tower GPRS Dump analysis."""

from __future__ import annotations

from modules.core.time_utils import utc_now, utc_now_iso

import math
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .excel_security import excel_safe_value
from .report_guidance import append_methodology_sheet


MAX_DATA_ROWS = 1_000_000
TITLE_FILL = PatternFill("solid", fgColor="17365D")
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
SUCCESS_FILL = PatternFill("solid", fgColor="E2F0D9")
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
ERROR_FILL = PatternFill("solid", fgColor="FCE4D6")
WHITE_BOLD = Font(color="FFFFFF", bold=True)
HEADER_FONT = Font(bold=True)
THIN = Side(style="thin", color="B7C9DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _safe_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_-]+", "_", str(value).strip())
    return cleaned.strip("_") or "CASE"


def _safe_value(value: Any) -> Any:
    """Compatibility wrapper around the shared Excel security boundary."""
    return excel_safe_value(value)


def _frame(value: Any) -> pd.DataFrame:
    if isinstance(value, pd.DataFrame):
        return value.copy()
    if isinstance(value, list):
        return pd.DataFrame(value)
    if isinstance(value, dict):
        return pd.DataFrame([value])
    return pd.DataFrame()


def _sheet_name(workbook: Workbook, requested: str) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", " ", requested)
    cleaned = re.sub(r"\s+", " ", cleaned).strip() or "Sheet"
    cleaned = cleaned[:31]
    existing = {sheet.title.lower() for sheet in workbook.worksheets}

    if cleaned.lower() not in existing:
        return cleaned

    number = 2

    while True:
        suffix = f" {number}"
        candidate = cleaned[: 31 - len(suffix)] + suffix

        if candidate.lower() not in existing:
            return candidate

        number += 1


def _title(worksheet, text: str, subtitle: str, columns: int) -> int:
    last_column = max(2, columns)
    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=last_column,
    )
    cell = worksheet.cell(1, 1, text)
    cell.fill = TITLE_FILL
    cell.font = Font(color="FFFFFF", bold=True, size=16)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 28

    worksheet.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=last_column,
    )
    sub = worksheet.cell(2, 1, subtitle)
    sub.font = Font(italic=True, color="44546A")
    sub.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )
    worksheet.row_dimensions[2].height = 24
    return 4


def _widths(worksheet) -> None:
    for cells in worksheet.columns:
        letter = get_column_letter(cells[0].column)
        maximum = 0

        for cell in cells[:5000]:
            if cell.value is not None:
                maximum = max(maximum, len(str(cell.value)))

        worksheet.column_dimensions[letter].width = min(
            max(maximum + 2, 10),
            44,
        )


def _write_page(
    worksheet,
    dataframe: pd.DataFrame,
    *,
    title: str,
    subtitle: str,
) -> None:
    header_row = _title(
        worksheet,
        title,
        subtitle,
        max(2, len(dataframe.columns)),
    )

    if dataframe.empty:
        cell = worksheet.cell(header_row, 1, "No records found.")
        cell.fill = WARNING_FILL
        cell.font = HEADER_FONT
        worksheet.column_dimensions["A"].width = 34
        worksheet.freeze_panes = f"A{header_row + 1}"
        worksheet.sheet_view.showGridLines = False
        return

    for column_index, column in enumerate(dataframe.columns, start=1):
        cell = worksheet.cell(header_row, column_index, _safe_value(str(column)))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for row_index, record in enumerate(
        dataframe.itertuples(index=False, name=None),
        start=header_row + 1,
    ):
        for column_index, value in enumerate(record, start=1):
            cell = worksheet.cell(
                row_index,
                column_index,
                _safe_value(value),
            )
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top")

            if isinstance(cell.value, (datetime, date)):
                cell.number_format = "yyyy-mm-dd hh:mm:ss"

    last_row = header_row + len(dataframe)
    last_column = get_column_letter(len(dataframe.columns))
    worksheet.auto_filter.ref = (
        f"A{header_row}:{last_column}{last_row}"
    )
    worksheet.freeze_panes = f"A{header_row + 1}"
    worksheet.sheet_view.showGridLines = False
    _widths(worksheet)


def _write_table(
    workbook: Workbook,
    base_name: str,
    dataframe: pd.DataFrame,
    *,
    subtitle: str,
) -> list[str]:
    dataframe = _frame(dataframe)

    if dataframe.empty:
        name = _sheet_name(workbook, base_name)
        worksheet = workbook.create_sheet(name)
        _write_page(
            worksheet,
            dataframe,
            title=base_name,
            subtitle=subtitle,
        )
        return [name]

    pages = max(1, math.ceil(len(dataframe) / MAX_DATA_ROWS))
    names: list[str] = []

    for page_index in range(pages):
        start = page_index * MAX_DATA_ROWS
        end = min(start + MAX_DATA_ROWS, len(dataframe))
        page = dataframe.iloc[start:end].reset_index(drop=True)
        requested = (
            base_name
            if pages == 1
            else f"{base_name} {page_index + 1}"
        )
        name = _sheet_name(workbook, requested)
        worksheet = workbook.create_sheet(name)
        page_subtitle = subtitle

        if pages > 1:
            page_subtitle = (
                f"{subtitle} | Part {page_index + 1}/{pages} | "
                f"Rows {start + 1:,}-{end:,}"
            )

        _write_page(
            worksheet,
            page,
            title=base_name,
            subtitle=page_subtitle,
        )
        names.append(name)

    return names


def _key_values(worksheet, start_row: int, rows: list[tuple[str, Any]]) -> None:
    for offset, (key, value) in enumerate(rows):
        row = start_row + offset
        key_cell = worksheet.cell(row, 1, key)
        value_cell = worksheet.cell(row, 2, _safe_value(value))
        key_cell.fill = HEADER_FILL
        key_cell.font = HEADER_FONT
        key_cell.border = BORDER
        value_cell.border = BORDER
        value_cell.alignment = Alignment(wrap_text=True, vertical="top")

    worksheet.column_dimensions["A"].width = 32
    worksheet.column_dimensions["B"].width = 58


def _warnings(
    load_result: dict[str, Any],
) -> pd.DataFrame:
    rows: list[dict[str, str]] = []

    for message in load_result.get("warnings", []) or []:
        rows.append({"Level": "WARNING", "Message": str(message)})

    for message in load_result.get("errors", []) or []:
        rows.append({"Level": "ERROR", "Message": str(message)})

    if not rows:
        rows.append(
            {
                "Level": "INFO",
                "Message": "No loader or analysis warning was reported.",
            }
        )

    return pd.DataFrame(rows)


def _portable_report_path(
    value: Any,
) -> str:
    """Return a portable report path without workstation details."""

    text = str(
        value
        or ""
    ).strip()

    if not text:
        return ""

    normalized = text.replace(
        "\\",
        "/",
    )

    for marker in (
        "data/",
        "cases/",
        "database/",
        "config/",
        "logs/",
    ):
        position = normalized.find(
            marker
        )

        if position >= 0:
            return normalized[
                position:
            ]

    path = Path(
        normalized
    )

    if path.is_absolute():
        return path.name

    return normalized


def _sanitize_gprs_report_frame(
    value: Any,
) -> pd.DataFrame:
    """Remove private paths and preserve identifiers as text."""

    frame = _frame(
        value
    ).copy()

    if frame.empty:
        return frame

    if (
        "source_relative_path"
        in frame.columns
        and "source_file"
        in frame.columns
    ):
        relative = (
            frame[
                "source_relative_path"
            ]
            .fillna("")
            .astype(str)
            .str.strip()
        )

        fallback = frame[
            "source_file"
        ].map(
            _portable_report_path
        )

        frame["source_file"] = (
            relative.where(
                relative.ne(""),
                fallback,
            )
        )

    path_columns = {
        "source_file",
        "input_folder",
        "run_directory",
        "report_path",
        "pipeline_state_path",
        "backend_run_directory",
    }

    for column in frame.columns:
        normalized = str(
            column
        ).strip().lower()

        if (
            normalized in path_columns
            or normalized.endswith(
                "_path"
            )
        ):
            frame[column] = frame[
                column
            ].map(
                _portable_report_path
            )

    identifier_columns = {
        "subscriber_number",
        "subscriber_number_raw",
        "subscribers",
        "imei",
        "imei_raw",
        "imsi",
        "imsi_raw",
        "searched_cell_id",
        "cgi",
        "cell_id",
        "identifier",
    }

    for column in frame.columns:
        normalized = str(
            column
        ).strip().lower()

        if normalized not in identifier_columns:
            continue

        frame[column] = frame[
            column
        ].map(
            lambda item: (
                ""
                if pd.isna(item)
                else str(item)
            )
        )

    # This is a legacy internal field from the old
    # CCTV-window workflow. Start and End are sufficient.
    if (
        "cctv_timestamp"
        in frame.columns
        and "window_start"
        in frame.columns
        and "window_end"
        in frame.columns
    ):
        frame = frame.drop(
            columns=[
                "cctv_timestamp",
            ]
        )

    return frame



def _build_gprs_spot_summary(
    load_result: dict[str, Any],
) -> pd.DataFrame:
    """Build accurate Spot-level file and record counts."""

    file_summary = _frame(
        load_result.get(
            "file_summary"
        )
    )

    if (
        file_summary.empty
        or "spot_id"
        not in file_summary.columns
    ):
        return _sanitize_gprs_report_frame(
            load_result.get(
                "spot_summary"
            )
        )

    rows: list[
        dict[str, Any]
    ] = []

    group_columns = [
        column
        for column in (
            "spot_id",
            "spot_name",
            "spot_folder",
        )
        if column in file_summary.columns
    ]

    for values, group in file_summary.groupby(
        group_columns,
        dropna=False,
        sort=True,
    ):
        if not isinstance(
            values,
            tuple,
        ):
            values = (
                values,
            )

        identity = dict(
            zip(
                group_columns,
                values,
            )
        )

        status = (
            group.get(
                "status",
                pd.Series(
                    "",
                    index=group.index,
                ),
            )
            .fillna("")
            .astype(str)
            .str.upper()
        )

        records = pd.to_numeric(
            group.get(
                "records",
                pd.Series(
                    0,
                    index=group.index,
                ),
            ),
            errors="coerce",
        ).fillna(0)

        searched_cells = (
            group.get(
                "searched_cell_id",
                pd.Series(
                    "",
                    index=group.index,
                ),
            )
            .fillna("")
            .astype(str)
            .str.strip()
        )

        rows.append(
            {
                **identity,
                "files_found": int(
                    len(group)
                ),
                "files_loaded": int(
                    status.eq(
                        "LOADED"
                    ).sum()
                ),
                "files_empty_no_data": int(
                    status.eq(
                        "EMPTY_NO_DATA"
                    ).sum()
                ),
                "files_failed": int(
                    status.eq(
                        "FAILED"
                    ).sum()
                ),
                "records": int(
                    records.sum()
                ),
                "searched_cells": int(
                    searched_cells[
                        searched_cells.ne("")
                    ].nunique()
                ),
            }
        )

    return pd.DataFrame(
        rows
    )


def _non_standard_identifier_row_count(
    analysis: dict[str, Any],
    metadata: dict[str, Any],
) -> int:
    """Return the verified non-standard identifier row count."""

    frame = _frame(
        analysis.get(
            "non_standard_identifiers"
        )
    )

    if not frame.empty:
        return int(
            len(frame)
        )

    raw_count = metadata.get(
        "non_standard_identifier_rows",
        0,
    )

    try:
        return max(
            int(
                raw_count
                or 0
            ),
            0,
        )
    except (
        TypeError,
        ValueError,
    ):
        return 0


def generate_tower_gprs_excel_report(
    *,
    case: dict[str, Any],
    load_result: dict[str, Any],
    analysis: dict[str, Any],
    partition: dict[str, Any] | None,
    output_dir: str | Path,
    saved: dict[str, Any] | None = None,
) -> Path:
    """Generate one consolidated and portable Tower GPRS workbook."""

    case_id = (
        str(
            case.get(
                "case_id",
                "",
            )
        ).strip()
        or "CASE"
    )
    case_name = str(
        case.get(
            "case_name",
            "",
        )
    ).strip()

    output = Path(
        output_dir
    ).expanduser().resolve()

    output.mkdir(
        parents=True,
        exist_ok=True,
    )

    timestamp = utc_now().strftime(
        "%Y%m%d_%H%M%S_%f"
    )

    report_path = output / (
        "Tower_GPRS_Dump_Analysis_"
        f"{_safe_filename(case_id)}_"
        f"{timestamp}.xlsx"
    )

    workbook = Workbook()
    workbook.remove(
        workbook.active
    )

    overview = workbook.create_sheet(
        _sheet_name(
            workbook,
            "1. Executive Summary",
        )
    )

    start = _title(
        overview,
        "Tower GPRS Dump Analysis",
        (
            f"Case: {case_id}"
            + (
                f" | {case_name}"
                if case_name
                else ""
            )
        ),
        6,
    )

    metadata = (
        load_result.get(
            "metadata",
            {},
        )
        or {}
    )

    subscriber_summary = (
        _sanitize_gprs_report_frame(
            analysis.get(
                "subscriber_summary"
            )
        )
    )

    n_of_m = _sanitize_gprs_report_frame(
        partition.get(
            "n_of_m_candidates"
        )
        if isinstance(
            partition,
            dict,
        )
        else None
    )

    strict = _sanitize_gprs_report_frame(
        partition.get(
            "strict_common_candidates"
        )
        if isinstance(
            partition,
            dict,
        )
        else None
    )

    non_standard_part = (
        _sanitize_gprs_report_frame(
            partition.get(
                "non_standard_subscriber_presence"
            )
        )
        if isinstance(
            partition,
            dict,
        )
        else pd.DataFrame()
    )

    _key_values(
        overview,
        start,
        [
            (
                "Case ID",
                case_id,
            ),
            (
                "Case Name",
                case_name,
            ),
            (
                "Source Section",
                "Tower GPRS Dump",
            ),
            (
                "Currently Supported Parser",
                "Airtel GPRS Session Dump",
            ),
            (
                "Generated At",
                utc_now_iso(),
            ),
            (
                "Input Folder",
                _portable_report_path(
                    metadata.get(
                        "input_folder",
                        "",
                    )
                ),
            ),
            (
                "Files Found",
                metadata.get(
                    "files_found",
                    0,
                ),
            ),
            (
                "Files Loaded",
                metadata.get(
                    "files_loaded",
                    0,
                ),
            ),
            (
                "Files Empty/No Data",
                metadata.get(
                    "files_empty_no_data",
                    0,
                ),
            ),
            (
                "Files Failed",
                metadata.get(
                    "files_failed",
                    0,
                ),
            ),
            (
                "Spots Found",
                metadata.get(
                    "spot_count",
                    0,
                ),
            ),
            (
                "Normalized Sessions",
                metadata.get(
                    "records",
                    0,
                ),
            ),
            (
                "Operators",
                ", ".join(
                    load_result.get(
                        "operators",
                        [],
                    )
                    or []
                ),
            ),
            (
                "Searched CGI/Cells",
                len(
                    load_result.get(
                        "cell_ids",
                        [],
                    )
                    or []
                ),
            ),
            (
                "Unique Subscriber Identifiers",
                len(
                    subscriber_summary
                ),
            ),
            (
                "Non-standard Identifier Rows",
                _non_standard_identifier_row_count(
                    analysis,
                    metadata,
                ),
            ),
            (
                "Dynamic Parts",
                (
                    partition.get(
                        "total_partitions",
                        0,
                    )
                    if isinstance(
                        partition,
                        dict,
                    )
                    else 0
                ),
            ),
            (
                "Mobile Numbers in 2+ Parts",
                len(
                    n_of_m
                ),
            ),
            (
                "Strict Common Mobile Numbers",
                len(
                    strict
                ),
            ),
            (
                "Part Non-standard Identifiers",
                len(
                    non_standard_part
                ),
            ),
            (
                "Session-overlap Rule",
                (
                    partition.get(
                        "overlap_rule",
                        "",
                    )
                    if isinstance(
                        partition,
                        dict,
                    )
                    else (
                        "session_start < part_end AND "
                        "session_end > part_start"
                    )
                ),
            ),
            (
                "Spot Rule",
                (
                    partition.get(
                        "spot_rule",
                        "",
                    )
                    if isinstance(
                        partition,
                        dict,
                    )
                    else (
                        "Complete analysis covers "
                        "all loaded Spots"
                    )
                ),
            ),
            (
                "Identifier Handling",
                (
                    "Valid 10-digit Indian MSISDN "
                    "leads are separated from "
                    "non-standard identifiers"
                ),
            ),
            (
                "Backend Run Directory",
                _portable_report_path(
                    saved.get(
                        "run_directory",
                        "",
                    )
                    if isinstance(
                        saved,
                        dict,
                    )
                    else ""
                ),
            ),
        ],
    )

    overview.freeze_panes = "A4"
    overview.sheet_view.showGridLines = False

    core_tables = [
        (
            "2. Source Files",
            load_result.get(
                "file_summary"
            ),
            (
                "Source-file status with "
                "portable relative paths."
            ),
        ),
        (
            "3. Spot Summary",
            _build_gprs_spot_summary(
                load_result
            ),
            (
                "Spot-wise files, loaded records "
                "and searched-cell coverage."
            ),
        ),
        (
            "4. Session Summary",
            analysis.get(
                "summary"
            ),
            "Core GPRS session metrics.",
        ),
        (
            "5. Technology",
            analysis.get(
                "technology_summary"
            ),
            "Technology distribution.",
        ),
        (
            "6. Connection Type",
            analysis.get(
                "pre_post_summary"
            ),
            "Prepaid and postpaid distribution.",
        ),
        (
            "7. Roaming",
            analysis.get(
                "roaming_summary"
            ),
            "Roaming-circle distribution.",
        ),
        (
            "8. Subscribers",
            analysis.get(
                "subscriber_summary"
            ),
            "Subscriber-identifier session intelligence.",
        ),
        (
            "9. Repeat Subscribers",
            analysis.get(
                "repeat_subscribers"
            ),
            "Subscriber identifiers with multiple sessions.",
        ),
        (
            "10. Common Mobile",
            analysis.get(
                "gprs_common_numbers"
            ),
            "Valid mobile numbers with repeat GPRS presence.",
        ),
        (
            "11. Rare Mobile",
            analysis.get(
                "gprs_uncommon_numbers"
            ),
            "Valid mobile numbers with rare or single-session presence.",
        ),
        (
            "12. Multi Cell Mobile",
            analysis.get(
                "gprs_multi_cell_presence"
            ),
            "Valid mobile numbers seen across multiple cells.",
        ),
        (
            "13. Device Check",
            analysis.get(
                "gprs_device_consistency"
            ),
            "IMEI, IMSI and IP consistency review.",
        ),
        (
            "14. Timing Leads",
            analysis.get(
                "gprs_suspicious_timing"
            ),
            "Timing-based GPRS leads.",
        ),
        (
            "15. Priority Mobile",
            analysis.get(
                "gprs_priority_leads"
            ),
            "Ranked valid mobile-number leads.",
        ),
        (
            "16. Nonstandard Leads",
            analysis.get(
                "gprs_non_standard_leads"
            ),
            (
                "Non-standard subscriber identifiers "
                "preserved separately for verification."
            ),
        ),
        (
            "17. IMEI Summary",
            analysis.get(
                "imei_summary"
            ),
            "Device identity summary.",
        ),
        (
            "18. Shared IMEI",
            analysis.get(
                "shared_imei"
            ),
            "IMEI associated with multiple subscriber identifiers.",
        ),
        (
            "19. IMSI Summary",
            analysis.get(
                "imsi_summary"
            ),
            "SIM/subscriber identity summary.",
        ),
        (
            "20. Shared IMSI",
            analysis.get(
                "shared_imsi"
            ),
            "IMSI associated with multiple subscriber identifiers.",
        ),
        (
            "21. IP Analysis",
            analysis.get(
                "ip_summary"
            ),
            "IPv4 and IPv6 usage summary.",
        ),
        (
            "22. Duration Buckets",
            analysis.get(
                "duration_buckets"
            ),
            "Session-duration distribution.",
        ),
        (
            "23. Hourly Activity",
            analysis.get(
                "hourly_activity"
            ),
            "Session-start activity by hour.",
        ),
        (
            "24. Long Sessions",
            analysis.get(
                "long_sessions"
            ),
            "Longest sessions for verification.",
        ),
        (
            "25. Zero Volume",
            analysis.get(
                "zero_volume_sessions"
            ),
            "Sessions with zero total volume.",
        ),
        (
            "26. Nonstandard Raw",
            analysis.get(
                "non_standard_identifiers"
            ),
            (
                "Raw non-standard identifier rows "
                "with evidence provenance."
            ),
        ),
        (
            "27. Data Quality",
            analysis.get(
                "data_quality"
            ),
            "Validation and quality checks.",
        ),
        (
            "28. Rejected Rows",
            analysis.get(
                "rejected_rows"
            ),
            "Malformed or non-data rows kept in quarantine.",
        ),
    ]

    for name, dataframe, subtitle in core_tables:
        _write_table(
            workbook,
            name,
            _sanitize_gprs_report_frame(
                dataframe
            ),
            subtitle=subtitle,
        )

    partition_tables: list[
        tuple[str, str, str]
    ] = []

    if isinstance(
        partition,
        dict,
    ):
        partition_tables = [
            (
                "29. Part Windows",
                "partition_windows",
                (
                    "Exact Start and End Date-Time "
                    "with selected Spot scope."
                ),
            ),
            (
                "30. Part Summary",
                "partition_summary",
                (
                    "Spot-wise session-overlap "
                    "and subscriber counts."
                ),
            ),
            (
                "31. Part Status",
                "partition_status",
                "Part validation and scope status.",
            ),
            (
                "32. Location Exclusions",
                "time_only_excluded_by_location",
                (
                    "Time-overlapping sessions excluded "
                    "because the selected scope did not match."
                ),
            ),
            (
                "33. Mobile Presence",
                "subscriber_presence",
                (
                    "Valid mobile-number presence "
                    "across configured Parts."
                ),
            ),
            (
                "34. Mobile in 2+ Parts",
                "n_of_m_candidates",
                (
                    "Valid mobile numbers present "
                    "in two or more Parts."
                ),
            ),
            (
                "35. Strict Common Mobile",
                "strict_common_candidates",
                (
                    "Valid mobile numbers present "
                    "in every configured Part."
                ),
            ),
            (
                "36. Part Nonstandard IDs",
                "non_standard_subscriber_presence",
                (
                    "Non-standard identifiers preserved "
                    "with Part presence classification."
                ),
            ),
            (
                "37. IMEI Continuity",
                "imei_presence",
                "IMEI continuity across Parts.",
            ),
            (
                "38. IMSI Continuity",
                "imsi_presence",
                "IMSI continuity across Parts.",
            ),
            (
                "39. IPv4 Continuity",
                "ipv4_presence",
                "IPv4 continuity across Parts.",
            ),
            (
                "40. IPv6 Continuity",
                "ipv6_presence",
                "IPv6 continuity across Parts.",
            ),
        ]

        for name, key, subtitle in partition_tables:
            _write_table(
                workbook,
                name,
                _sanitize_gprs_report_frame(
                    partition.get(
                        key
                    )
                ),
                subtitle=subtitle,
            )

    status_number = (
        41
        if partition_tables
        else 29
    )
    warning_number = (
        42
        if partition_tables
        else 30
    )

    status = pd.DataFrame(
        [
            {
                "Stage": "Tower GPRS Loading",
                "Status": "COMPLETED",
                "Details": (
                    f"{metadata.get('records', 0):,} "
                    "normalized sessions"
                ),
            },
            {
                "Stage": "Core Analysis",
                "Status": "COMPLETED",
                "Details": (
                    f"{len(core_tables)} "
                    "analytical table groups"
                ),
            },
            {
                "Stage": "Spot-based Part Analysis",
                "Status": (
                    "COMPLETED"
                    if partition_tables
                    else "NOT REQUESTED"
                ),
                "Details": (
                    f"{partition.get('total_partitions', 0)} Parts"
                    if partition_tables
                    else ""
                ),
            },
            {
                "Stage": "Consolidated Excel",
                "Status": "COMPLETED",
                "Details": _portable_report_path(
                    report_path
                ),
            },
        ]
    )

    _write_table(
        workbook,
        f"{status_number}. Analysis Status",
        status,
        subtitle="Execution status for this report.",
    )

    warning_names = _write_table(
        workbook,
        f"{warning_number}. Warnings",
        _sanitize_gprs_report_frame(
            _warnings(
                load_result
            )
        ),
        subtitle="Loader and normalization warnings.",
    )

    for sheet_name in warning_names:
        worksheet = workbook[
            sheet_name
        ]

        for row in range(
            5,
            worksheet.max_row + 1,
        ):
            level = str(
                worksheet.cell(
                    row,
                    1,
                ).value
                or ""
            ).upper()

            worksheet.cell(
                row,
                1,
            ).fill = (
                ERROR_FILL
                if level == "ERROR"
                else WARNING_FILL
                if level == "WARNING"
                else SUCCESS_FILL
            )

    workbook.properties.title = (
        f"Tower GPRS Dump Analysis - {case_id}"
    )
    workbook.properties.subject = (
        "GPRS session, Spot and Date-Time Part analysis"
    )
    workbook.properties.creator = (
        "Telecom Forensics Analysis Suite"
    )

    append_methodology_sheet(
        workbook,
        "Tower GPRS Dump Analysis",
    )

    workbook.save(
        report_path
    )

    return report_path
