"""Consolidated Excel reporting for target/reverse multi-operator IPDR."""

from __future__ import annotations

from modules.core.time_utils import utc_now, utc_now_iso

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .excel_security import excel_safe_value
from .report_guidance import append_methodology_sheet


REPORT_TABLE_ROW_LIMIT = 5_000
EVENT_PREVIEW_ROWS = 5_000

TITLE_FILL = PatternFill("solid", fgColor="17365D")
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
INFO_FILL = PatternFill("solid", fgColor="E2F0D9")
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
ERROR_FILL = PatternFill("solid", fgColor="FCE4D6")
THIN = Side(style="thin", color="B7C9DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _safe_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_-]+", "_", str(value).strip())
    return cleaned.strip("_") or "CASE"


def _safe_value(value: Any) -> Any:
    """Compatibility wrapper around the shared Excel security boundary."""
    return excel_safe_value(value)


def _frame(value: Any) -> pd.DataFrame:
    if isinstance(value, pd.DataFrame):
        return value.copy()

    if isinstance(value, list):
        return pd.DataFrame(value)

    if isinstance(value, dict):
        return pd.DataFrame([value])

    return pd.DataFrame()


def _sheet_name(workbook: Workbook, requested: str) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", " ", requested)
    cleaned = re.sub(r"\s+", " ", cleaned).strip() or "Sheet"
    cleaned = cleaned[:31]
    existing = {sheet.title.lower() for sheet in workbook.worksheets}

    if cleaned.lower() not in existing:
        return cleaned

    number = 2

    while True:
        suffix = f" {number}"
        candidate = cleaned[: 31 - len(suffix)] + suffix

        if candidate.lower() not in existing:
            return candidate

        number += 1


def _write_title(
    worksheet,
    title: str,
    subtitle: str,
    columns: int,
) -> int:
    last_column = max(2, columns)
    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=last_column,
    )
    title_cell = worksheet.cell(1, 1, title)
    title_cell.fill = TITLE_FILL
    title_cell.font = Font(
        color="FFFFFF",
        bold=True,
        size=16,
    )
    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )
    worksheet.row_dimensions[1].height = 28

    worksheet.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=last_column,
    )
    subtitle_cell = worksheet.cell(2, 1, subtitle)
    subtitle_cell.font = Font(
        italic=True,
        color="44546A",
    )
    subtitle_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )
    worksheet.row_dimensions[2].height = 26
    return 4


def _set_widths(worksheet) -> None:
    for cells in worksheet.columns:
        letter = get_column_letter(cells[0].column)
        maximum = 0

        for cell in cells[:500]:
            if cell.value is not None:
                maximum = max(maximum, len(str(cell.value)))

        worksheet.column_dimensions[letter].width = min(
            max(maximum + 2, 10),
            42,
        )


def _write_page(
    worksheet,
    dataframe: pd.DataFrame,
    *,
    title: str,
    subtitle: str,
) -> None:
    header_row = _write_title(
        worksheet,
        title,
        subtitle,
        max(2, len(dataframe.columns)),
    )

    if dataframe.empty:
        cell = worksheet.cell(
            header_row,
            1,
            "No records found.",
        )
        cell.fill = WARNING_FILL
        cell.font = Font(bold=True)
        worksheet.column_dimensions["A"].width = 36
        worksheet.freeze_panes = f"A{header_row + 1}"
        worksheet.sheet_view.showGridLines = False
        return

    for column_index, column in enumerate(
        dataframe.columns,
        start=1,
    ):
        cell = worksheet.cell(
            header_row,
            column_index,
            _safe_value(str(column)),
        )
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
        cell.border = BORDER
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for row_index, record in enumerate(
        dataframe.itertuples(index=False, name=None),
        start=header_row + 1,
    ):
        for column_index, value in enumerate(record, start=1):
            cell = worksheet.cell(
                row_index,
                column_index,
                _safe_value(value),
            )
            cell.border = BORDER
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=False,
            )

            if isinstance(cell.value, (datetime, date)):
                cell.number_format = "yyyy-mm-dd hh:mm:ss"

    last_row = header_row + len(dataframe)
    last_column = get_column_letter(len(dataframe.columns))
    worksheet.auto_filter.ref = (
        f"A{header_row}:{last_column}{last_row}"
    )
    worksheet.freeze_panes = f"A{header_row + 1}"
    worksheet.sheet_view.showGridLines = False
    _set_widths(worksheet)


def _write_table(
    workbook: Workbook,
    base_name: str,
    dataframe: pd.DataFrame,
    *,
    subtitle: str,
) -> list[str]:
    dataframe = _frame(dataframe)
    total_rows = len(dataframe)

    if total_rows > REPORT_TABLE_ROW_LIMIT:
        dataframe = dataframe.head(
            REPORT_TABLE_ROW_LIMIT
        ).reset_index(drop=True)
        subtitle = (
            f"{subtitle} | Showing first "
            f"{REPORT_TABLE_ROW_LIMIT:,} of {total_rows:,} rows. "
            "Complete table is preserved in the backend CSV."
        )

    name = _sheet_name(workbook, base_name)
    worksheet = workbook.create_sheet(name)
    _write_page(
        worksheet,
        dataframe,
        title=base_name,
        subtitle=subtitle,
    )
    return [name]


def _key_values(
    worksheet,
    start_row: int,
    rows: list[tuple[str, Any]],
) -> None:
    for offset, (key, value) in enumerate(rows):
        row = start_row + offset
        key_cell = worksheet.cell(row, 1, key)
        value_cell = worksheet.cell(row, 2, _safe_value(value))
        key_cell.fill = HEADER_FILL
        key_cell.font = Font(bold=True)
        key_cell.border = BORDER
        value_cell.border = BORDER
        value_cell.alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )

        if isinstance(value_cell.value, (datetime, date)):
            value_cell.number_format = "yyyy-mm-dd hh:mm:ss"

    worksheet.column_dimensions["A"].width = 34
    worksheet.column_dimensions["B"].width = 72


def _messages(load_result: dict[str, Any]) -> pd.DataFrame:
    rows: list[dict[str, str]] = []

    for message in load_result.get("warnings", []) or []:
        rows.append(
            {"Level": "WARNING", "Message": str(message)}
        )

    for message in load_result.get("errors", []) or []:
        rows.append(
            {"Level": "ERROR", "Message": str(message)}
        )

    if not rows:
        rows.append(
            {
                "Level": "INFO",
                "Message": "No loader warning or error was reported.",
            }
        )

    return pd.DataFrame(rows)


def generate_ipdr_excel_report(
    *,
    case: dict[str, Any],
    mode: str,
    load_result: dict[str, Any],
    analysis: dict[str, Any],
    output_dir: str | Path,
    saved: dict[str, Any] | None = None,
) -> Path:
    """Generate one consolidated IPDR workbook."""

    case_id = str(case.get("case_id", "")).strip() or "CASE"
    case_name = str(case.get("case_name", "")).strip()
    mode = str(mode).strip().lower()
    output = Path(output_dir).expanduser().resolve()
    output.mkdir(parents=True, exist_ok=True)
    timestamp = utc_now().strftime("%Y%m%d_%H%M%S_%f")
    report_path = output / (
        f"IPDR_Analysis_{mode.title()}_"
        f"{_safe_filename(case_id)}_{timestamp}.xlsx"
    )

    workbook = Workbook()
    workbook.remove(workbook.active)

    overview = workbook.create_sheet(
        _sheet_name(workbook, "1. Executive Summary")
    )
    start = _write_title(
        overview,
        f"{mode.title()} IPDR Analysis",
        f"Case: {case_id}"
        + (f" | {case_name}" if case_name else ""),
        6,
    )

    metadata = load_result.get("metadata", {}) or {}
    subscriber_summary = _frame(
        analysis.get("subscriber_summary")
    )
    data_quality = _frame(
        analysis.get("data_quality")
    )
    error_count = 0
    warning_count = 0

    if not data_quality.empty and "Severity" in data_quality:
        error_count = int(
            data_quality.loc[
                data_quality["Severity"].eq("ERROR"),
                "Rows",
            ].sum()
        )
        warning_count = int(
            data_quality.loc[
                data_quality["Severity"].eq("WARNING"),
                "Rows",
            ].sum()
        )

    _key_values(
        overview,
        start,
        [
            ("Case ID", case_id),
            ("Case Name", case_name),
            ("Analysis Mode", mode.upper()),
            ("Generated At", utc_now_iso()),
            ("Input Folder", load_result.get("folder", "")),
            ("Files Found", metadata.get("files_found", 0)),
            ("Files Loaded", metadata.get("files_loaded", 0)),
            ("Files Failed", metadata.get("files_failed", 0)),
            ("Normalized Records", metadata.get("total_records", 0)),
            ("IPDR Event Records", metadata.get("event_records", 0)),
            (
                "Allocation-only Records",
                metadata.get("allocation_records", 0),
            ),
            ("Search Requests", metadata.get("search_requests", 0)),
            (
                "Operators",
                ", ".join(metadata.get("operators", []) or []),
            ),
            (
                "Report Scopes",
                ", ".join(metadata.get("scopes", []) or []),
            ),
            ("Unique Subscribers/User IDs", len(subscriber_summary)),
            ("Data-quality Error Rows", error_count),
            ("Data-quality Warning Rows", warning_count),
            (
                "Complete Backend Tables",
                saved.get("run_directory", "")
                if isinstance(saved, dict)
                else "",
            ),
            (
                "Raw/Normalized Event Policy",
                "Complete normalized events are preserved in backend CSV. "
                f"Excel includes first {EVENT_PREVIEW_ROWS:,} normalized rows and first {REPORT_TABLE_ROW_LIMIT:,} rows of each analytical table.",
            ),
        ],
    )
    overview.freeze_panes = "A4"
    overview.sheet_view.showGridLines = False

    tables = [
        ("2. Source Files", analysis.get("file_summary"), "Source classification, operator and load status."),
        ("3. Query Summary", analysis.get("query_summary"), "Target/reverse query and source-file metrics."),
        ("4. Subscribers", analysis.get("subscriber_summary"), "Subscriber/user-ID level activity."),
        ("5. Multi-File Subs", analysis.get("multi_file_subscribers"), "Subscribers observed in more than one source file."),
        ("6. Subscriber Matrix", analysis.get("subscriber_file_presence"), "Dynamic source-file presence matrix."),
        ("7. IMEI Summary", analysis.get("imei_summary"), "Device identity activity."),
        ("8. Shared IMEI", analysis.get("shared_imei"), "IMEI linked with multiple subscriber identities."),
        ("9. IMEI Matrix", analysis.get("imei_file_presence"), "IMEI presence across source files."),
        ("10. IMSI Summary", analysis.get("imsi_summary"), "IMSI activity summary."),
        ("11. Shared IMSI", analysis.get("shared_imsi"), "IMSI linked with multiple subscriber identities."),
        ("12. IMSI Matrix", analysis.get("imsi_file_presence"), "IMSI presence across source files."),
        ("13. Source IP", analysis.get("source_ip_summary"), "Source public/private IP activity."),
        ("14. NAT IP", analysis.get("translated_ip_summary"), "Translated/NAT IP activity."),
        ("15. Destination IP", analysis.get("destination_ip_summary"), "Destination IP activity."),
        ("16. Destination Ports", analysis.get("destination_port_summary"), "Destination-port frequency and subscriber coverage."),
        ("17. Endpoints", analysis.get("destination_endpoint_summary"), "Destination IP and port combinations."),
        ("18. Allocations", analysis.get("allocation_records"), "Deduplicated session/allocation records."),
        ("19. APN", analysis.get("apn_summary"), "Access Point Name activity."),
        ("20. Technology", analysis.get("technology_summary"), "2G/4G/5G/RAT summary."),
        ("21. CGI Cells", analysis.get("cgi_summary"), "CGI/cell activity."),
        ("22. Cell Movement", analysis.get("cell_movement"), "First/last cell continuity and changes."),
        ("23. Hourly Activity", analysis.get("hourly_activity"), "Date and hour-wise activity."),
        ("24. Reverse Query Check", analysis.get("reverse_query_validation"), "Validation of queried destination IP against result rows."),
        ("25. Search Requests", analysis.get("search_requests"), "Reverse-IP search input workbook records."),
        ("26. Data Quality", analysis.get("data_quality"), "Validation checks; raw values remain unchanged."),
        ("27. Rejected Rows", analysis.get("rejected_rows"), "Malformed/non-data rows quarantined with physical source-line provenance."),
    ]

    for name, dataframe, subtitle in tables:
        _write_table(
            workbook,
            name,
            _frame(dataframe),
            subtitle=subtitle,
        )

    normalized = _frame(analysis.get("normalized_events"))
    preview = normalized.head(EVENT_PREVIEW_ROWS).copy()
    _write_table(
        workbook,
        "28. Event Preview",
        preview,
        subtitle=(
            f"First {min(len(normalized), EVENT_PREVIEW_ROWS):,} "
            f"of {len(normalized):,} normalized records. "
            "Complete data is saved in backend normalized_events.csv."
        ),
    )

    status = pd.DataFrame(
        [
            {
                "Stage": "Multi-operator Loading",
                "Status": "COMPLETED",
                "Details": (
                    f"{metadata.get('total_records', 0):,} "
                    "normalized result records"
                ),
            },
            {
                "Stage": "Search Request Intake",
                "Status": "COMPLETED",
                "Details": (
                    f"{metadata.get('search_requests', 0):,} "
                    "request rows"
                ),
            },
            {
                "Stage": "Core Analysis",
                "Status": "COMPLETED",
                "Details": f"{len(tables)} analysis groups",
            },
            {
                "Stage": "Backend Persistence",
                "Status": "COMPLETED",
                "Details": (
                    saved.get("run_directory", "")
                    if isinstance(saved, dict)
                    else ""
                ),
            },
            {
                "Stage": "Consolidated Excel",
                "Status": "COMPLETED",
                "Details": str(report_path),
            },
        ]
    )
    _write_table(
        workbook,
        "28. Analysis Status",
        status,
        subtitle="Execution status for this analysis run.",
    )

    message_sheets = _write_table(
        workbook,
        "29. Warnings",
        _messages(load_result),
        subtitle="Loader warnings, duplicate-file notices and errors.",
    )

    for sheet_name in message_sheets:
        worksheet = workbook[sheet_name]

        for row in range(5, worksheet.max_row + 1):
            level = str(
                worksheet.cell(row, 1).value or ""
            ).upper()
            worksheet.cell(row, 1).fill = (
                ERROR_FILL
                if level == "ERROR"
                else WARNING_FILL
                if level == "WARNING"
                else INFO_FILL
            )

    workbook.properties.title = (
        f"{mode.title()} IPDR Analysis - {case_id}"
    )
    workbook.properties.subject = (
        "Target, reverse-IP and broadband allocation analysis"
    )
    workbook.properties.creator = (
        "Telecom Forensics Analysis Suite"
    )
    append_methodology_sheet(workbook, "Target / Reverse IPDR Analysis")
    workbook.save(report_path)
    return report_path
