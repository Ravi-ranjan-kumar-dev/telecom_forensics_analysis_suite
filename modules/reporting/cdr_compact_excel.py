
"""Compact investigator-facing CDR Excel reports.

This renderer consumes already-computed analysis results. It does not rerun
registered analysis functions and does not expose developer diagnostics.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import single_cdr_excel as detailed
from .excel_security import excel_safe_value
from .report_paths import get_single_report_path


SINGLE_CDR_COMPACT_SHEETS = (
    "1. Executive Summary",
    "2. Priority Contacts",
    "3. Communication Intel",
    "4. Network Intelligence",
    "5. Location & Roaming",
    "6. Movement & Daily Routine",
    "7. Device & SIM Intel",
    "8. Activity & Timing",
    "9. Priority Review Queue",
    "10. Data Quality & Guide",
)

TECHNICAL_COLUMNS = {
    "Group",
    "Result Key",
    "Module",
    "Function",
    "Duration (sec)",
    "Error",
    "Traceback",
    "Score_Ruleset",
    "Score_Formula",
    "Strength_Ruleset",
    "Strength_Formula",
    "Ruleset",
    "source_table",
}

SECTION_COLUMN_PROFILES = {
    "CC SUMMARY": (
        "Other Party",
        "Name",
        "Father Name",
        "SDR Address",
        "Total Calls",
        "Total Duration",
        "Avg. Call Duration",
        "Out Count",
        "IN Count",
        "Out SMS Count",
        "In SMS Count",
        "First Call Time",
        "Last Call Time",
        "SDR Lookup Status",
    ),
    "TOP HUMAN CONTACTS": (
        "Contact",
        "Name",
        "Father Name",
        "Address",
        "Operator",
        "Total Calls",
        "SDR Found",
        "SDR Lookup Status",
        "Match Confidence",
    ),
    "CONTACT RANKING": (
        "Contact",
        "Name",
        "Father Name",
        "Address",
        "Total_Events",
        "Incoming",
        "Outgoing",
        "SMS",
        "Total_Duration",
        "Unique_Towers",
        "Score",
        "SDR Lookup Status",
    ),
    "OUTGOING VOICE CALLS": (
        "Sub Call Type",
        "Other Party",
        "Name",
        "SDR Address",
        "Date",
        "Time",
        "Duration",
        "Cell ID",
        "Address",
        "End Cell ID",
        "End Address",
        "SDR Lookup Status",
    ),
    "INCOMING VOICE CALLS": (
        "Sub Call Type",
        "Other Party",
        "Name",
        "SDR Address",
        "Date",
        "Time",
        "Duration",
        "Cell ID",
        "Address",
        "End Cell ID",
        "End Address",
        "SDR Lookup Status",
    ),
    "OUTGOING SMS": (
        "Sub Call Type",
        "Other Party",
        "Name",
        "SDR Address",
        "Date",
        "Time",
        "Cell ID",
        "Address",
        "End Cell ID",
        "End Address",
        "SDR Lookup Status",
    ),
    "INCOMING SMS": (
        "Sub Call Type",
        "Other Party",
        "Name",
        "SDR Address",
        "Date",
        "Time",
        "Cell ID",
        "Address",
        "End Cell ID",
        "End Address",
        "SDR Lookup Status",
    ),
    "STRONGEST COMMUNICATION LINKS": (
        "Contact",
        "Name",
        "Father Name",
        "Address",
        "Total_Events",
        "Incoming",
        "Outgoing",
        "SMS",
        "Total_Duration",
        "Unique_Towers",
        "Strength",
        "SDR Lookup Status",
    ),
    "CELL ID SUMMARY": (
        "Cell ID",
        "Total Calls",
        "Address",
        "Start Tower Town",
        "Start Tower District",
        "Start Tower Lookup Status",
    ),
    "FREQUENT TOWERS": (
        "Cell ID",
        "Total Events",
        "CGI Town",
        "CGI District",
        "CGI Address",
        "CGI Lookup Status",
    ),
    "TOWER INTELLIGENCE": (
        "Cell ID",
        "First Seen",
        "Last Seen",
        "Total Events",
        "Unique IMEIs",
        "Unique Human Contacts",
        "Most Human Contacted",
        "CGI Town",
        "CGI District",
        "CGI Address",
        "CGI Lookup Status",
    ),
    "PROBABLE HOME TOWER INDICATORS": (
        "Cell ID",
        "Night Events",
        "Unique Days",
        "Unique Human Contacts",
        "Window",
        "CGI Town",
        "CGI District",
        "CGI Address",
        "CGI Lookup Status",
    ),
    "PROBABLE WORK TOWER INDICATORS": (
        "Cell ID",
        "Office Events",
        "Working Days",
        "Unique Human Contacts",
        "CGI Town",
        "CGI District",
        "CGI Address",
        "CGI Lookup Status",
    ),
    "FIRST AND LAST COMMUNICATION BY DAY (FCLC)": (
        "Date",
        "Time",
        "Description",
        "Call Type",
        "Other Party",
        "Name",
        "Duration",
        "Cell ID",
        "Address",
        "SDR Lookup Status",
        "Start Tower Town",
        "Start Tower District",
        "Start Tower Lookup Status",
    ),
    "FIRST/LAST LOCATION SUMMARY (FCLC SUMMARY)": (
        "Cell ID",
        "Total Calls",
        "Address",
        "Start Tower Town",
        "Start Tower District",
        "Start Tower Lookup Status",
    ),
    "FIRST/LAST CONTACT SUMMARY (FCLC OP)": (
        "Other Party",
        "Name",
        "Father Name",
        "SDR Address",
        "Total Calls",
        "First Call Time",
        "Last Call Time",
        "SDR Lookup Status",
    ),
    "MOVING CALLS": (
        "Mobile Number",
        "Other Party",
        "Name",
        "Call Type",
        "IMEI",
        "Date",
        "Time",
        "Duration",
        "Cell ID",
        "Address",
        "Last Cell ID",
        "Last Address",
        "SDR Lookup Status",
    ),
    "MOVEMENT EVENTS": (
        "Other Party",
        "Name",
        "Call Type",
        "Sub Call Type",
        "IMEI",
        "Date",
        "Time",
        "Duration",
        "Roaming Circle",
        "Cell ID",
        "Address",
        "SDR Lookup Status",
        "Start Tower Town",
        "Start Tower District",
        "Start Tower Lookup Status",
    ),
    "TOWER TRANSITIONS": (
        "call_date",
        "call_time",
        "From Tower",
        "From Tower Town",
        "From Tower Address",
        "To Tower",
        "To Tower Town",
        "To Tower Address",
        "b_party",
        "Name",
        "call_type",
        "SDR Lookup Status",
    ),
    "MOVEMENT PATTERNS": (
        "From Tower",
        "From Tower Town",
        "From Tower Address",
        "To Tower",
        "To Tower Town",
        "To Tower Address",
        "Occurrences",
    ),
    "IMEI SUMMARY": (
        "imei",
        "IMEI",
        "First Seen",
        "Last Seen",
        "Total Events",
        "Unique Human Contacts",
        "Unique Valid Towers",
        "Total Duration (Sec)",
    ),
    "IMEI INTELLIGENCE": (
        "IMEI",
        "First Seen",
        "Last Seen",
        "Total Events",
        "Unique Human Contacts",
        "Unique Valid Towers",
        "Most Used Valid Tower",
        "Most Human Contacted",
    ),
    "SIM OR DEVICE CHANGES": (
        "Date",
        "Time",
        "Old IMEI",
        "New IMEI",
        "Tower",
        "Contact",
        "Event",
    ),
    "BEHAVIORAL OBSERVATIONS": (
        "Indicator",
        "Observation",
        "Caution",
    ),
    "REVIEW INDICATORS": (
        "Type",
        "Count",
        "Remark",
    ),
}

THIN_BORDER = Border(
    left=Side(style="thin", color="D9E1F2"),
    right=Side(style="thin", color="D9E1F2"),
    top=Side(style="thin", color="D9E1F2"),
    bottom=Side(style="thin", color="D9E1F2"),
)

TITLE_FILL = PatternFill(
    "solid",
    fgColor="1F4E78",
)

SECTION_FILL = PatternFill(
    "solid",
    fgColor="D9EAF7",
)

HEADER_FILL = PatternFill(
    "solid",
    fgColor="5B9BD5",
)

NOTE_FILL = PatternFill(
    "solid",
    fgColor="FFF2CC",
)


def _bundle_results(
    bundle: dict[str, Any] | None,
) -> dict[str, Any]:
    """Return the canonical analysis result dictionary."""

    if not isinstance(
        bundle,
        dict,
    ):
        return {}

    results = bundle.get(
        "results"
    )

    if isinstance(
        results,
        dict,
    ):
        return results

    return bundle


def _as_frame(
    value: Any,
) -> pd.DataFrame:
    """Convert one analysis result to a DataFrame."""

    if value is None:
        return pd.DataFrame()

    if isinstance(
        value,
        pd.DataFrame,
    ):
        return value.copy()

    return detailed._result_as_dataframe(
        value
    )


def _friendly_enrichment_name(
    column: Any,
) -> str | None:
    """Return a readable name for one CGI enrichment column."""

    name = str(
        column
    )

    prefixes = (
        (
            "cell_id_cgi_",
            "CGI",
        ),
        (
            "first_cell_",
            "First Cell",
        ),
        (
            "last_cell_",
            "Last Cell",
        ),
    )

    suffix_names = {
        "operator": "Operator",
        "circle": "Circle",
        "state": "State",
        "district": "District",
        "police_station": "Police Station",
        "town": "Town",
        "site_name": "Site Name",
        "address": "Address",
        "latitude": "Latitude",
        "longitude": "Longitude",
        "address_found": "Address Found",
        "lookup_status": "Lookup Status",
        "match_confidence": "Match Confidence",
    }

    for prefix, label in prefixes:
        if not name.startswith(
            prefix
        ):
            continue

        suffix = name[
            len(
                prefix
            ):
        ]

        readable_suffix = suffix_names.get(
            suffix
        )

        if readable_suffix is None:
            return None

        return (
            f"{label} "
            f"{readable_suffix}"
        )

    return None

def _clean_frame(
    frame: pd.DataFrame,
    *,
    display_context: str = "",
) -> pd.DataFrame:
    """Return an investigator-facing table without internal columns."""

    if not isinstance(
        frame,
        pd.DataFrame,
    ):
        return pd.DataFrame()

    output = frame.copy()

    if display_context:
        output = detailed._enrich_contact_report_dataframe(
            display_context,
            output,
        )

    sanitizer = getattr(
        detailed,
        "_sanitize_report_paths",
        None,
    )

    if callable(
        sanitizer
    ):
        output = sanitizer(
            output
        )

    drop_columns = []

    for column in output.columns:
        name = str(
            column
        )

        normalized = name.lower()

        if (
            name in TECHNICAL_COLUMNS
            or name.startswith(
                "_"
            )
            or name.startswith(
                "contact_sdr_"
            )
            or normalized.endswith(
                "_lookup_key"
            )
            or normalized.endswith(
                "_record_found"
            )
            or normalized.endswith(
                "_source_file"
            )
            or normalized.endswith(
                "_ruleset"
            )
            or normalized.endswith(
                "_formula"
            )
        ):
            drop_columns.append(
                column
            )

    output = output.drop(
        columns=drop_columns,
        errors="ignore",
    )

    rename_columns = {}

    for column in output.columns:
        readable_name = (
            _friendly_enrichment_name(
                column
            )
        )

        if (
            readable_name
            and readable_name
            not in output.columns
        ):
            rename_columns[
                column
            ] = readable_name

    output = output.rename(
        columns=rename_columns
    )

    if "Level Code" in output.columns:
        level_values = (
            output[
                "Level Code"
            ]
            .fillna(
                ""
            )
            .astype(
                str
            )
            .str.strip()
            .str.lower()
        )

        blank_like = {
            "",
            "missing",
            "unknown",
            "n/a",
            "na",
            "none",
            "not available",
        }

        if level_values.isin(
            blank_like
        ).all():
            output = output.drop(
                columns=[
                    "Level Code",
                ]
            )

    empty_columns = []

    for column in output.columns:
        values = output[
            column
        ]

        blank_mask = (
            values.isna()
            | values.astype(
                str
            )
            .str.strip()
            .str.lower()
            .isin(
                {
                    "",
                    "nan",
                    "none",
                }
            )
        )

        if blank_mask.all():
            empty_columns.append(
                column
            )

    output = output.drop(
        columns=empty_columns,
        errors="ignore",
    )

    return output



def _preferred_columns(
    frame: pd.DataFrame,
    columns: list[str] | tuple[str, ...] | None,
) -> pd.DataFrame:
    """Keep preferred columns when they are available."""

    if (
        not isinstance(
            frame,
            pd.DataFrame,
        )
        or not columns
    ):
        return frame

    available = [
        column
        for column in columns
        if column in frame.columns
    ]

    if not available:
        return frame

    return frame.loc[
        :,
        available,
    ].copy()


def _blank_mask(
    values: pd.Series,
) -> pd.Series:
    """Return True for empty investigator-facing values."""

    return (
        values.isna()
        | values.astype(
            str
        )
        .str.strip()
        .str.lower()
        .isin(
            {
                "",
                "nan",
                "none",
                "missing",
                "not available",
                "n/a",
                "na",
            }
        )
    )


def _coalesce_series(
    frame: pd.DataFrame,
    candidates: tuple[str, ...],
) -> pd.Series:
    """Return the first available non-blank source column."""

    result = pd.Series(
        pd.NA,
        index=frame.index,
        dtype="object",
    )

    for column in candidates:
        if column not in frame.columns:
            continue

        source = frame[
            column
        ]

        result = result.where(
            ~_blank_mask(
                result
            ),
            source,
        )

    return result


def _normalize_mobile_key(
    value: Any,
) -> str:
    """Return a normalized 10-digit Indian mobile key."""

    digits = "".join(
        character
        for character in str(
            value
            if value is not None
            else ""
        )
        if character.isdigit()
    )

    if (
        len(
            digits
        )
        == 11
        and digits.startswith(
            "0"
        )
    ):
        digits = digits[
            -10:
        ]

    if (
        len(
            digits
        )
        > 10
        and digits.endswith(
            digits[
                -10:
            ]
        )
    ):
        digits = digits[
            -10:
        ]

    if (
        len(
            digits
        )
        != 10
        or digits[
            0
        ]
        not in {
            "6",
            "7",
            "8",
            "9",
        }
    ):
        return ""

    return digits


def _normalize_cgi_key(
    value: Any,
) -> str:
    """Return a stable CGI lookup key."""

    if value is None:
        return ""

    text = str(
        value
    ).strip().upper()

    if text.lower() in {
        "",
        "nan",
        "none",
        "missing",
        "invalid",
        "n/a",
        "na",
    }:
        return ""

    return text


def _build_sdr_contact_index(
    results: dict[str, Any],
) -> pd.DataFrame:
    """Build one reusable SDR index from enriched analysis results."""

    sources = []

    for key in (
        "contact_ranking",
        "top_contacts",
        "social_network",
    ):
        frame = _as_frame(
            results.get(
                key
            )
        )

        if frame.empty:
            continue

        source = pd.DataFrame(
            index=frame.index
        )

        source[
            "_contact_key"
        ] = _coalesce_series(
            frame,
            (
                "contact_sdr_lookup_mobile",
                "Contact",
                "Other Party",
                "b_party",
            ),
        ).map(
            _normalize_mobile_key
        )

        field_sources = {
            "Name": (
                "contact_sdr_subscriber_name",
                "Name",
                "Subscriber Name",
            ),
            "Father Name": (
                "contact_sdr_father_name",
                "Father Name",
            ),
            "SDR Address": (
                "contact_sdr_address",
                "SDR Address",
            ),
            "SDR Operator": (
                "contact_sdr_operator",
                "SDR Operator",
                "Operator",
            ),
            "SDR Circle": (
                "contact_sdr_circle",
                "SDR Circle",
                "Circle",
            ),
            "Activation Date": (
                "contact_sdr_activation_date",
                "Activation Date",
            ),
            "CAF Number": (
                "contact_sdr_caf_number",
                "CAF Number",
            ),
            "SDR Found": (
                "contact_sdr_found",
                "SDR Found",
            ),
            "SDR Lookup Status": (
                "contact_sdr_lookup_status",
                "SDR Lookup Status",
            ),
            "SDR Match Confidence": (
                "contact_sdr_match_confidence",
                "SDR Match Confidence",
                "Match Confidence",
            ),
        }

        for output_column, candidates in field_sources.items():
            source[
                output_column
            ] = _coalesce_series(
                frame,
                candidates,
            )

        sources.append(
            source
        )

    if not sources:
        return pd.DataFrame(
            columns=[
                "_contact_key",
                "Name",
                "Father Name",
                "SDR Address",
                "SDR Operator",
                "SDR Circle",
                "Activation Date",
                "CAF Number",
                "SDR Found",
                "SDR Lookup Status",
                "SDR Match Confidence",
            ]
        )

    index = pd.concat(
        sources,
        ignore_index=True,
    )

    index = index.loc[
        index[
            "_contact_key"
        ].ne(
            ""
        )
    ].copy()

    found_rank = (
        index[
            "SDR Lookup Status"
        ]
        .fillna(
            ""
        )
        .astype(
            str
        )
        .str.upper()
        .eq(
            "FOUND"
        )
        .astype(
            int
        )
    )

    completeness = (
        ~index[
            [
                "Name",
                "Father Name",
                "SDR Address",
                "SDR Operator",
                "SDR Circle",
            ]
        ].apply(
            _blank_mask
        )
    ).sum(
        axis=1
    )

    index[
        "_found_rank"
    ] = found_rank

    index[
        "_completeness"
    ] = completeness

    index = (
        index.sort_values(
            [
                "_contact_key",
                "_found_rank",
                "_completeness",
            ],
            ascending=[
                True,
                False,
                False,
            ],
            kind="stable",
        )
        .drop_duplicates(
            subset=[
                "_contact_key",
            ],
            keep="first",
        )
        .drop(
            columns=[
                "_found_rank",
                "_completeness",
            ]
        )
        .reset_index(
            drop=True
        )
    )

    return index


def _build_cgi_location_index(
    results: dict[str, Any],
) -> pd.DataFrame:
    """Build one reusable CGI index from tower analysis results."""

    sources = []

    for key in (
        "tower_intelligence",
        "frequent_locations",
        "home_tower",
        "work_tower",
    ):
        frame = _as_frame(
            results.get(
                key
            )
        )

        if frame.empty:
            continue

        source = pd.DataFrame(
            index=frame.index
        )

        source[
            "_cgi_key"
        ] = _coalesce_series(
            frame,
            (
                "cell_id_cgi_lookup_key",
                "Cell ID",
                "first_cell_lookup_key",
                "first_cell_id",
            ),
        ).map(
            _normalize_cgi_key
        )

        field_sources = {
            "Tower Operator": (
                "cell_id_cgi_operator",
                "first_cell_operator",
            ),
            "Tower Circle": (
                "cell_id_cgi_circle",
                "first_cell_circle",
            ),
            "Tower State": (
                "cell_id_cgi_state",
                "first_cell_state",
            ),
            "Tower District": (
                "cell_id_cgi_district",
                "first_cell_district",
            ),
            "Tower Police Station": (
                "cell_id_cgi_police_station",
                "first_cell_police_station",
            ),
            "Tower Town": (
                "cell_id_cgi_town",
                "first_cell_town",
            ),
            "Tower Site Name": (
                "cell_id_cgi_site_name",
                "first_cell_site_name",
            ),
            "Tower Address": (
                "cell_id_cgi_address",
                "first_cell_address",
            ),
            "Tower Latitude": (
                "cell_id_cgi_latitude",
                "first_cell_latitude",
            ),
            "Tower Longitude": (
                "cell_id_cgi_longitude",
                "first_cell_longitude",
            ),
            "CGI Address Found": (
                "cell_id_cgi_address_found",
                "first_cell_address_found",
            ),
            "CGI Lookup Status": (
                "cell_id_cgi_lookup_status",
                "first_cell_lookup_status",
            ),
            "CGI Match Confidence": (
                "cell_id_cgi_match_confidence",
                "first_cell_match_confidence",
            ),
        }

        for output_column, candidates in field_sources.items():
            source[
                output_column
            ] = _coalesce_series(
                frame,
                candidates,
            )

        sources.append(
            source
        )

    if not sources:
        return pd.DataFrame(
            columns=[
                "_cgi_key",
                "Tower Operator",
                "Tower Circle",
                "Tower State",
                "Tower District",
                "Tower Police Station",
                "Tower Town",
                "Tower Site Name",
                "Tower Address",
                "Tower Latitude",
                "Tower Longitude",
                "CGI Address Found",
                "CGI Lookup Status",
                "CGI Match Confidence",
            ]
        )

    index = pd.concat(
        sources,
        ignore_index=True,
    )

    index = index.loc[
        index[
            "_cgi_key"
        ].ne(
            ""
        )
    ].copy()

    found_rank = (
        index[
            "CGI Lookup Status"
        ]
        .fillna(
            ""
        )
        .astype(
            str
        )
        .str.upper()
        .eq(
            "FOUND"
        )
        .astype(
            int
        )
    )

    completeness = (
        ~index[
            [
                "Tower Address",
                "Tower Town",
                "Tower District",
                "Tower Latitude",
                "Tower Longitude",
            ]
        ].apply(
            _blank_mask
        )
    ).sum(
        axis=1
    )

    index[
        "_found_rank"
    ] = found_rank

    index[
        "_completeness"
    ] = completeness

    index = (
        index.sort_values(
            [
                "_cgi_key",
                "_found_rank",
                "_completeness",
            ],
            ascending=[
                True,
                False,
                False,
            ],
            kind="stable",
        )
        .drop_duplicates(
            subset=[
                "_cgi_key",
            ],
            keep="first",
        )
        .drop(
            columns=[
                "_found_rank",
                "_completeness",
            ]
        )
        .reset_index(
            drop=True
        )
    )

    return index


def _merge_lookup_fields(
    frame: pd.DataFrame,
    *,
    lookup: pd.DataFrame,
    lookup_key: str,
    row_keys: pd.Series,
    field_map: dict[str, str],
) -> pd.DataFrame:
    """Merge reusable lookup fields without overwriting evidence."""

    if (
        frame.empty
        or lookup.empty
    ):
        return frame.copy()

    output = frame.copy()

    temporary_key = "__lookup_join_key"

    output[
        temporary_key
    ] = row_keys

    selected_columns = [
        lookup_key,
        *field_map.keys(),
    ]

    selected_columns = [
        column
        for column in selected_columns
        if column in lookup.columns
    ]

    lookup_frame = (
        lookup[
            selected_columns
        ]
        .rename(
            columns={
                lookup_key: temporary_key,
                **{
                    source: target
                    for source, target in field_map.items()
                    if source in lookup.columns
                },
            }
        )
        .copy()
    )

    existing_columns = set(
        output.columns
    )

    output = output.merge(
        lookup_frame,
        on=temporary_key,
        how="left",
        suffixes=(
            "",
            "__lookup",
        ),
        sort=False,
        validate="many_to_one",
    )

    for target_column in field_map.values():
        lookup_column = (
            f"{target_column}__lookup"
            if target_column in existing_columns
            else target_column
        )

        if lookup_column not in output.columns:
            continue

        if target_column not in existing_columns:
            continue

        mask = _blank_mask(
            output[
                target_column
            ]
        )

        output.loc[
            mask,
            target_column,
        ] = output.loc[
            mask,
            lookup_column,
        ]

        output = output.drop(
            columns=[
                lookup_column,
            ],
            errors="ignore",
        )

    output = output.drop(
        columns=[
            temporary_key,
        ],
        errors="ignore",
    )

    return output


def _enrich_contact_table(
    frame: pd.DataFrame,
    sdr_index: pd.DataFrame,
) -> pd.DataFrame:
    """Apply the reusable SDR index to one report table."""

    if (
        not isinstance(
            frame,
            pd.DataFrame,
        )
        or frame.empty
        or sdr_index.empty
    ):
        return (
            frame.copy()
            if isinstance(
                frame,
                pd.DataFrame,
            )
            else pd.DataFrame()
        )

    contact_column = next(
        (
            column
            for column in (
                "Contact",
                "Other Party",
                "b_party",
                "B Party Number",
                "Contact Number",
                "Counterparty",
            )
            if column in frame.columns
        ),
        None,
    )

    if contact_column is None:
        return frame.copy()

    return _merge_lookup_fields(
        frame,
        lookup=sdr_index,
        lookup_key="_contact_key",
        row_keys=frame[
            contact_column
        ].map(
            _normalize_mobile_key
        ),
        field_map={
            "Name": "Name",
            "Father Name": "Father Name",
            "SDR Address": "SDR Address",
            "SDR Operator": "SDR Operator",
            "SDR Circle": "SDR Circle",
            "Activation Date": "Activation Date",
            "CAF Number": "CAF Number",
            "SDR Found": "SDR Found",
            "SDR Lookup Status": "SDR Lookup Status",
            "SDR Match Confidence": "SDR Match Confidence",
        },
    )


def _enrich_one_cgi_column(
    frame: pd.DataFrame,
    cgi_index: pd.DataFrame,
    *,
    key_column: str,
    label: str,
    address_column: str | None = None,
) -> pd.DataFrame:
    """Apply CGI fields for one Cell ID column."""

    if (
        frame.empty
        or cgi_index.empty
        or key_column not in frame.columns
    ):
        return frame.copy()

    field_map = {
        "Tower Operator": (
            f"{label} Operator"
        ),
        "Tower Circle": (
            f"{label} Circle"
        ),
        "Tower State": (
            f"{label} State"
        ),
        "Tower District": (
            f"{label} District"
        ),
        "Tower Police Station": (
            f"{label} Police Station"
        ),
        "Tower Town": (
            f"{label} Town"
        ),
        "Tower Site Name": (
            f"{label} Site Name"
        ),
        "Tower Address": (
            address_column
            or f"{label} Address"
        ),
        "Tower Latitude": (
            f"{label} Latitude"
        ),
        "Tower Longitude": (
            f"{label} Longitude"
        ),
        "CGI Address Found": (
            f"{label} Address Found"
        ),
        "CGI Lookup Status": (
            f"{label} Lookup Status"
        ),
        "CGI Match Confidence": (
            f"{label} Match Confidence"
        ),
    }

    return _merge_lookup_fields(
        frame,
        lookup=cgi_index,
        lookup_key="_cgi_key",
        row_keys=frame[
            key_column
        ].map(
            _normalize_cgi_key
        ),
        field_map=field_map,
    )


def _enrich_event_table(
    frame: pd.DataFrame,
    *,
    sdr_index: pd.DataFrame,
    cgi_index: pd.DataFrame,
) -> pd.DataFrame:
    """Apply common SDR and CGI maps to a legacy event table."""

    output = _enrich_contact_table(
        frame,
        sdr_index,
    )

    tower_columns = (
        (
            "Cell ID",
            "Start Tower",
            "Address",
        ),
        (
            "End Cell ID",
            "End Tower",
            "End Address",
        ),
        (
            "First Cell ID",
            "First Tower",
            "First Address",
        ),
        (
            "Last Cell ID",
            "Last Tower",
            "Last Address",
        ),
        (
            "first_cell_id",
            "First Tower",
            None,
        ),
        (
            "last_cell_id",
            "Last Tower",
            None,
        ),
        (
            "From Tower",
            "From Tower",
            None,
        ),
        (
            "To Tower",
            "To Tower",
            None,
        ),
    )

    for key_column, label, address_column in tower_columns:
        if key_column not in output.columns:
            continue

        output = _enrich_one_cgi_column(
            output,
            cgi_index,
            key_column=key_column,
            label=label,
            address_column=address_column,
        )

    return output

def _write_sheet_title(
    worksheet,
    *,
    title: str,
    metadata: dict[str, Any],
) -> int:
    """Write the report title and compact target metadata."""

    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=8,
    )

    title_cell = worksheet.cell(
        row=1,
        column=1,
        value=title,
    )

    title_cell.fill = TITLE_FILL
    title_cell.font = Font(
        color="FFFFFF",
        bold=True,
        size=15,
    )
    title_cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
    )

    worksheet.row_dimensions[
        1
    ].height = 26

    metadata_rows = (
        (
            "Target",
            metadata.get(
                "target",
                "",
            ),
        ),
        (
            "Subscriber",
            metadata.get(
                "subscriber_name",
                "",
            ),
        ),
        (
            "Observation Period",
            (
                f"{metadata.get('from_date', '')} "
                f"to {metadata.get('to_date', '')}"
            ).strip(),
        ),
        (
            "SDR Match",
            metadata.get(
                "target_sdr_found",
                "No",
            ),
        ),
    )

    row = 2

    for label, value in metadata_rows:
        worksheet.cell(
            row=row,
            column=1,
            value=label,
        ).font = Font(
            bold=True
        )

        worksheet.cell(
            row=row,
            column=2,
            value=excel_safe_value(
                detailed._excel_safe_scalar(
                    value
                )
            ),
        )

        row += 1

    return row + 1


def _write_section(
    worksheet,
    start_row: int,
    *,
    title: str,
    frame: pd.DataFrame,
    guidance: str = "",
    max_rows: int | None = None,
    preferred_columns: list[str] | tuple[str, ...] | None = None,
    display_context: str = "",
) -> int:
    """Write one titled table section and return the next free row."""

    frame = _clean_frame(
        frame,
        display_context=display_context,
    )

    selected_columns = preferred_columns

    if selected_columns is None:
        selected_columns = (
            SECTION_COLUMN_PROFILES.get(
                title
            )
        )

    frame = _preferred_columns(
        frame,
        selected_columns,
    )

    total_rows = len(
        frame
    )

    if (
        max_rows is not None
        and total_rows > max_rows
    ):
        shown = frame.head(
            max_rows
        ).copy()
    else:
        shown = frame.copy()

    column_count = max(
        2,
        len(
            shown.columns
        ),
    )

    worksheet.merge_cells(
        start_row=start_row,
        start_column=1,
        end_row=start_row,
        end_column=column_count,
    )

    section_cell = worksheet.cell(
        row=start_row,
        column=1,
        value=title,
    )

    section_cell.fill = SECTION_FILL
    section_cell.font = Font(
        bold=True,
        color="1F1F1F",
        size=11,
    )

    row = start_row + 1

    note_parts = []

    if guidance:
        note_parts.append(
            guidance
        )

    if (
        max_rows is not None
        and total_rows > max_rows
    ):
        note_parts.append(
            f"Showing first {max_rows:,} of "
            f"{total_rows:,} rows. Full detail remains "
            "available in the detailed report or source evidence."
        )

    if note_parts:
        worksheet.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=column_count,
        )

        note_cell = worksheet.cell(
            row=row,
            column=1,
            value=" ".join(
                note_parts
            ),
        )

        note_cell.fill = NOTE_FILL
        note_cell.alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )

        row += 1

    if shown.empty:
        worksheet.cell(
            row=row,
            column=1,
            value="No records available for this section.",
        )

        return row + 2

    headers = list(
        shown.columns
    )

    for column_index, header in enumerate(
        headers,
        start=1,
    ):
        cell = worksheet.cell(
            row=row,
            column=column_index,
            value=excel_safe_value(
                str(
                    header
                )
            ),
        )

        cell.fill = HEADER_FILL
        cell.font = Font(
            color="FFFFFF",
            bold=True,
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = THIN_BORDER

    row += 1

    for values in shown.itertuples(
        index=False,
        name=None,
    ):
        for column_index, value in enumerate(
            values,
            start=1,
        ):
            cell = worksheet.cell(
                row=row,
                column=column_index,
                value=excel_safe_value(
                    detailed._excel_safe_scalar(
                        value
                    )
                ),
            )

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )
            cell.border = THIN_BORDER

        row += 1

    return row + 2


def _finish_sheet(
    worksheet,
) -> None:
    """Apply readable widths and page settings."""

    for column_index in range(
        1,
        worksheet.max_column + 1,
    ):
        maximum_length = 0

        for row_index in range(
            1,
            min(
                worksheet.max_row,
                250,
            )
            + 1,
        ):
            value = worksheet.cell(
                row=row_index,
                column=column_index,
            ).value

            if value is None:
                continue

            maximum_length = max(
                maximum_length,
                len(
                    str(
                        value
                    )
                ),
            )

        worksheet.column_dimensions[
            get_column_letter(
                column_index
            )
        ].width = min(
            max(
                maximum_length + 2,
                12,
            ),
            42,
        )

    worksheet.freeze_panes = "A7"
    worksheet.sheet_view.showGridLines = False
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0


def _rejected_row_count(
    original_data: pd.DataFrame | None,
    metadata: dict[str, Any],
) -> int:
    """Return one consistent rejected-row count."""

    frame_candidates = [
        metadata.get(
            "rejected_rows"
        ),
        metadata.get(
            "quarantined_rows"
        ),
    ]

    if isinstance(
        original_data,
        pd.DataFrame,
    ):
        frame_candidates.extend(
            [
                original_data.attrs.get(
                    "rejected_rows"
                ),
                original_data.attrs.get(
                    "quarantined_rows"
                ),
            ]
        )

    for candidate in frame_candidates:
        if (
            isinstance(
                candidate,
                pd.DataFrame,
            )
            and not candidate.empty
        ):
            return int(
                len(
                    candidate
                )
            )

    for key in (
        "rejected_rows_count",
        "quarantined_rows_count",
        "invalid_rows_count",
    ):
        value = metadata.get(
            key
        )

        if value is None:
            continue

        try:
            return int(
                value
            )
        except (
            TypeError,
            ValueError,
        ):
            continue

    return 0


def _compact_extract_summary(
    data: pd.DataFrame,
) -> pd.DataFrame:
    """Return summary metrics without mixed contact lead rows."""

    frame = detailed._extract_table(
        data,
        {},
    )

    if "Header" not in frame.columns:
        return frame

    header_values = (
        frame[
            "Header"
        ]
        .fillna(
            ""
        )
        .astype(
            str
        )
        .str.strip()
    )

    return frame.loc[
        ~header_values.eq(
            "High-Frequency Contact"
        )
    ].reset_index(
        drop=True
    )

def _quality_summary(
    data: pd.DataFrame,
    metadata: dict[str, Any],
    *,
    original_data: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """Build a simple data-quality summary for investigators."""

    valid_cells = detailed._normalized_valid_cell_ids(
        data
    )

    rejected_count = _rejected_row_count(
        original_data,
        metadata,
    )

    rows = [
        (
            "Records analyzed",
            len(
                data
            ),
            "Rows accepted for CDR analysis.",
        ),
        (
            "Rejected or quarantined rows",
            rejected_count,
            "Rows excluded because they were footer, metadata or invalid event rows.",
        ),
        (
            "Unique valid Cell IDs",
            int(
                valid_cells.dropna().nunique()
            ),
            "Normalized Cell IDs eligible for location analysis.",
        ),
        (
            "Rows with missing or invalid Cell ID",
            int(
                valid_cells.isna().sum()
            ),
            "These rows cannot support reliable tower-location interpretation.",
        ),
        (
            "Unique IMEIs",
            int(
                data[
                    "imei"
                ]
                .replace(
                    "",
                    pd.NA,
                )
                .dropna()
                .nunique()
            ),
            "Distinct device identifiers recorded in the source CDR.",
        ),
        (
            "Unique IMSIs",
            int(
                data[
                    "imsi"
                ]
                .replace(
                    "",
                    pd.NA,
                )
                .dropna()
                .nunique()
            ),
            "Distinct SIM identities recorded in the source CDR.",
        ),
    ]

    return pd.DataFrame(
        rows,
        columns=[
            "Quality Check",
            "Value",
            "Meaning",
        ],
    )



def _rejected_rows_summary(
    original_data: pd.DataFrame,
    metadata: dict[str, Any],
) -> pd.DataFrame:
    """Return a reason-level rejected-row summary when available."""

    candidates = [
        metadata.get(
            "rejected_rows"
        ),
        metadata.get(
            "quarantined_rows"
        ),
        original_data.attrs.get(
            "rejected_rows"
        ),
        original_data.attrs.get(
            "quarantined_rows"
        ),
    ]

    rejected = next(
        (
            candidate
            for candidate in candidates
            if (
                isinstance(
                    candidate,
                    pd.DataFrame,
                )
                and not candidate.empty
            )
        ),
        None,
    )

    if rejected is None:
        return pd.DataFrame(
            [
                {
                    "Rejected Rows": (
                        metadata.get(
                            "rejected_rows_count",
                            metadata.get(
                                "quarantined_rows_count",
                                0,
                            ),
                        )
                        or 0
                    ),
                    "Meaning": (
                        "Detailed rejected-row diagnostics remain "
                        "in loader evidence and developer logs."
                    ),
                }
            ]
        )

    reason_column = next(
        (
            column
            for column in (
                "parse_note",
                "reason",
                "rejection_reason",
                "status",
            )
            if column in rejected.columns
        ),
        None,
    )

    if reason_column is None:
        return pd.DataFrame(
            [
                {
                    "Reason": "Rejected or quarantined source row",
                    "Rows": len(
                        rejected
                    ),
                }
            ]
        )

    return (
        rejected[
            reason_column
        ]
        .fillna(
            "Unspecified reason"
        )
        .astype(
            str
        )
        .value_counts(
            dropna=False
        )
        .rename_axis(
            "Reason"
        )
        .reset_index(
            name="Rows"
        )
    )


def _interpretation_guide() -> pd.DataFrame:
    """Return plain-language evidentiary limitations."""

    return pd.DataFrame(
        [
            {
                "Point": "Cell-site evidence",
                "Meaning": (
                    "A CDR Cell ID shows network association. "
                    "It is not proof of the exact position of a person or handset."
                ),
            },
            {
                "Point": "High-frequency contact",
                "Meaning": (
                    "Frequent communication does not by itself establish "
                    "relationship, intent or participation."
                ),
            },
            {
                "Point": "Home and work indicators",
                "Meaning": (
                    "These are time-pattern indicators and must be verified "
                    "with field evidence and other records."
                ),
            },
            {
                "Point": "SDR and CGI matches",
                "Meaning": (
                    "Results depend on the coverage, age and correctness "
                    "of the available master databases."
                ),
            },
            {
                "Point": "Investigator verification",
                "Meaning": (
                    "Important findings should be verified against source CDR, "
                    "CAF/SDR, tower records and independent case evidence."
                ),
            },
        ]
    )


def _review_checklist() -> pd.DataFrame:
    """Return the standard investigator follow-up checklist."""

    return pd.DataFrame(
        [
            {
                "Review Step": "Verify priority contacts",
                "Suggested Action": (
                    "Confirm identity, ownership and case relevance "
                    "using SDR/CAF and field verification."
                ),
            },
            {
                "Review Step": "Verify important towers",
                "Suggested Action": (
                    "Check CGI address, coverage area and event timing "
                    "against independent location evidence."
                ),
            },
            {
                "Review Step": "Review device changes",
                "Suggested Action": (
                    "Verify IMEI changes against seizure records, "
                    "subscriber statements and handset evidence."
                ),
            },
            {
                "Review Step": "Review unusual timing",
                "Suggested Action": (
                    "Compare late-night or high-volume activity "
                    "with the incident timeline."
                ),
            },
        ]
    )


def generate_single_cdr_compact_report(
    df: pd.DataFrame,
    target: str,
    metadata: dict[str, Any] | None = None,
    analysis_bundle: dict[str, Any] | None = None,
    output_dir: str | Path | None = None,
) -> str | None:
    """Generate the canonical 10-sheet Single CDR investigator report."""

    if (
        df is None
        or not isinstance(
            df,
            pd.DataFrame,
        )
        or df.empty
    ):
        print(
            "[-] Compact Single CDR report aborted: "
            "DataFrame is empty or invalid."
        )
        return None

    try:
        data = detailed._prepare_dataframe(
            df,
            str(
                target
            ),
        )

        report_metadata = detailed._metadata(
            str(
                target
            ),
            data,
            metadata,
        )

        if isinstance(
            analysis_bundle,
            dict,
        ):
            bundle = analysis_bundle
        else:
            from .analysis_bundle import (
                build_single_analysis_bundle,
            )

            bundle = build_single_analysis_bundle(
                data,
                target=str(
                    target
                ),
            )

        results = _bundle_results(
            bundle
        )

        cc_summary = detailed._contact_summary(
            data
        )
        cell_summary = detailed._cell_summary(
            data
        )
        imei_summary = detailed._imei_summary(
            data
        )
        movements = detailed._movements(
            data
        )
        fclc = detailed._fclc(
            data
        )
        fclc_summary = detailed._fclc_summary(
            fclc
        )
        fclc_op = detailed._fclc_op(
            data
        )
        moving_calls = detailed._moving_calls(
            data
        )
        roaming_summary = detailed._roaming_summary(
            data
        )

        outgoing_voice = detailed._direction_sheet(
            data,
            detailed._voice_out_mask(
                data
            ),
        )
        incoming_voice = detailed._direction_sheet(
            data,
            detailed._voice_in_mask(
                data
            ),
        )
        outgoing_sms = detailed._direction_sheet(
            data,
            detailed._sms_out_mask(
                data
            ),
        )
        incoming_sms = detailed._direction_sheet(
            data,
            detailed._sms_in_mask(
                data
            ),
        )

        # Reuse already-enriched analysis results.
        # No additional SDR or CGI database query is performed here.
        sdr_contact_index = _build_sdr_contact_index(
            results
        )

        cgi_location_index = _build_cgi_location_index(
            results
        )

        cc_summary = _enrich_contact_table(
            cc_summary,
            sdr_contact_index,
        )

        cell_summary = _enrich_event_table(
            cell_summary,
            sdr_index=sdr_contact_index,
            cgi_index=cgi_location_index,
        )

        movements = _enrich_event_table(
            movements,
            sdr_index=sdr_contact_index,
            cgi_index=cgi_location_index,
        )

        fclc = _enrich_event_table(
            fclc,
            sdr_index=sdr_contact_index,
            cgi_index=cgi_location_index,
        )

        fclc_summary = _enrich_event_table(
            fclc_summary,
            sdr_index=sdr_contact_index,
            cgi_index=cgi_location_index,
        )

        fclc_op = _enrich_contact_table(
            fclc_op,
            sdr_contact_index,
        )

        moving_calls = _enrich_event_table(
            moving_calls,
            sdr_index=sdr_contact_index,
            cgi_index=cgi_location_index,
        )

        outgoing_voice = _enrich_event_table(
            outgoing_voice,
            sdr_index=sdr_contact_index,
            cgi_index=cgi_location_index,
        )

        incoming_voice = _enrich_event_table(
            incoming_voice,
            sdr_index=sdr_contact_index,
            cgi_index=cgi_location_index,
        )

        outgoing_sms = _enrich_event_table(
            outgoing_sms,
            sdr_index=sdr_contact_index,
            cgi_index=cgi_location_index,
        )

        incoming_sms = _enrich_event_table(
            incoming_sms,
            sdr_index=sdr_contact_index,
            cgi_index=cgi_location_index,
        )

        tower_transition = _enrich_event_table(
            _as_frame(
                results.get(
                    "tower_transition"
                )
            ),
            sdr_index=sdr_contact_index,
            cgi_index=cgi_location_index,
        )

        movement_pattern = _enrich_event_table(
            _as_frame(
                results.get(
                    "movement_pattern"
                )
            ),
            sdr_index=sdr_contact_index,
            cgi_index=cgi_location_index,
        )

        workbook = Workbook()
        workbook.remove(
            workbook.active
        )

        # 1. Executive Summary
        worksheet = workbook.create_sheet(
            SINGLE_CDR_COMPACT_SHEETS[
                0
            ]
        )
        row = _write_sheet_title(
            worksheet,
            title="Single CDR Investigation Summary",
            metadata=report_metadata,
        )
        row = _write_section(
            worksheet,
            row,
            title="CDR SUMMARY",
            frame=_compact_extract_summary(
                data
            ),
            guidance=(
                "Start here for traffic volume, devices, SIM identities "
                "and valid tower counts."
            ),
        )
        row = _write_section(
            worksheet,
            row,
            title="TOP FIVE HUMAN CONTACTS",
            frame=_as_frame(
                results.get(
                    "top_contacts"
                )
            ),
            max_rows=5,
            preferred_columns=(
                "Contact",
                "Name",
                "Father Name",
                "Address",
                "Operator",
                "Circle",
                "SDR Found",
                "SDR Lookup Status",
                "Match Confidence",
                "Total Calls",
            ),
            display_context="16. Top Human Contacts",
            guidance=(
                "Automated sender IDs and short codes are excluded "
                "from this human-contact list."
            ),
        )
        row = _write_section(
            worksheet,
            row,
            title="TOP FIVE TOWERS",
            frame=cell_summary.head(
                5
            ),
            max_rows=5,
        )
        _finish_sheet(
            worksheet
        )

        # 2. Priority Contacts
        worksheet = workbook.create_sheet(
            SINGLE_CDR_COMPACT_SHEETS[
                1
            ]
        )
        row = _write_sheet_title(
            worksheet,
            title="Priority Contact Review",
            metadata=report_metadata,
        )
        row = _write_section(
            worksheet,
            row,
            title="CC SUMMARY",
            frame=cc_summary,
            max_rows=50,
            guidance=(
                "Review frequent contacts together with duration, "
                "direction and SDR identity. Frequency alone is not a conclusion."
            ),
        )
        row = _write_section(
            worksheet,
            row,
            title="TOP HUMAN CONTACTS",
            frame=_as_frame(
                results.get(
                    "top_contacts"
                )
            ),
            max_rows=30,
            display_context="16. Top Human Contacts",
        )
        row = _write_section(
            worksheet,
            row,
            title="CONTACT RANKING",
            frame=_as_frame(
                results.get(
                    "contact_ranking"
                )
            ),
            max_rows=30,
            display_context="20. Contact Ranking",
        )
        row = _write_section(
            worksheet,
            row,
            title="TOP CONTACT DETAILS",
            frame=_as_frame(
                results.get(
                    "top_contact_details"
                )
            ),
            max_rows=30,
        )
        row = _write_section(
            worksheet,
            row,
            title="OUTGOING VOICE CALLS",
            frame=outgoing_voice,
            max_rows=50,
        )
        row = _write_section(
            worksheet,
            row,
            title="INCOMING VOICE CALLS",
            frame=incoming_voice,
            max_rows=50,
        )
        _finish_sheet(
            worksheet
        )

        # 3. Communication Intelligence
        worksheet = workbook.create_sheet(
            SINGLE_CDR_COMPACT_SHEETS[
                2
            ]
        )
        row = _write_sheet_title(
            worksheet,
            title="Communication Intelligence",
            metadata=report_metadata,
        )
        for title, key, limit in (
            (
                "CONTACT CATEGORY SUMMARY",
                "contact_category_summary",
                30,
            ),
            (
                "SERVICE SENDER IDS",
                "top_service_sender_ids",
                30,
            ),
            (
                "SHORT CODES",
                "top_short_codes",
                30,
            ),
            (
                "CALL TYPE SUMMARY",
                "incoming_outgoing",
                30,
            ),
            (
                "OTHER OR UNKNOWN CALL TYPES",
                "other_call_type_summary",
                30,
            ),
        ):
            row = _write_section(
                worksheet,
                row,
                title=title,
                frame=_as_frame(
                    results.get(
                        key
                    )
                ),
                max_rows=limit,
            )

        row = _write_section(
            worksheet,
            row,
            title="OUTGOING SMS",
            frame=outgoing_sms,
            max_rows=100,
            guidance=(
                "This section is retained as substantive communication evidence. "
                "The compact report shows a review preview."
            ),
        )
        row = _write_section(
            worksheet,
            row,
            title="INCOMING SMS",
            frame=incoming_sms,
            max_rows=100,
            guidance=(
                "Review automated sender IDs separately from normal human contacts."
            ),
        )
        _finish_sheet(
            worksheet
        )

        # 4. Network Intelligence
        worksheet = workbook.create_sheet(
            SINGLE_CDR_COMPACT_SHEETS[
                3
            ]
        )
        row = _write_sheet_title(
            worksheet,
            title="Communication Network Intelligence",
            metadata=report_metadata,
        )
        row = _write_section(
            worksheet,
            row,
            title="STRONGEST COMMUNICATION LINKS",
            frame=_as_frame(
                results.get(
                    "social_network"
                )
            ),
            max_rows=50,
            display_context="23. Social Network",
            guidance=(
                "Network strength is a descriptive communication measure. "
                "It does not prove relationship or intent."
            ),
        )
        _finish_sheet(
            worksheet
        )

        # 5. Location and Roaming
        worksheet = workbook.create_sheet(
            SINGLE_CDR_COMPACT_SHEETS[
                4
            ]
        )
        row = _write_sheet_title(
            worksheet,
            title="Location and Roaming Intelligence",
            metadata=report_metadata,
        )
        for title, frame, limit in (
            (
                "LOCATION OVERVIEW",
                _as_frame(
                    results.get(
                        "analyze_location"
                    )
                ),
                30,
            ),
            (
                "CELL ID SUMMARY",
                cell_summary,
                50,
            ),
            (
                "FREQUENT TOWERS",
                _as_frame(
                    results.get(
                        "frequent_locations"
                    )
                ),
                30,
            ),
            (
                "TOWER INTELLIGENCE",
                _as_frame(
                    results.get(
                        "tower_intelligence"
                    )
                ),
                30,
            ),
            (
                "PROBABLE HOME TOWER INDICATORS",
                _as_frame(
                    results.get(
                        "home_tower"
                    )
                ),
                20,
            ),
            (
                "PROBABLE WORK TOWER INDICATORS",
                _as_frame(
                    results.get(
                        "work_tower"
                    )
                ),
                20,
            ),
            (
                "ROAMING SUMMARY",
                roaming_summary,
                50,
            ),
        ):
            row = _write_section(
                worksheet,
                row,
                title=title,
                frame=frame,
                max_rows=limit,
            )

        _finish_sheet(
            worksheet
        )

        # 6. Movement and Daily Routine
        worksheet = workbook.create_sheet(
            SINGLE_CDR_COMPACT_SHEETS[
                5
            ]
        )
        row = _write_sheet_title(
            worksheet,
            title="Movement and Daily Routine",
            metadata=report_metadata,
        )
        row = _write_section(
            worksheet,
            row,
            title="FIRST AND LAST COMMUNICATION BY DAY (FCLC)",
            frame=fclc,
            max_rows=100,
        )
        row = _write_section(
            worksheet,
            row,
            title="FIRST/LAST LOCATION SUMMARY (FCLC SUMMARY)",
            frame=fclc_summary,
            max_rows=50,
        )
        row = _write_section(
            worksheet,
            row,
            title="FIRST/LAST CONTACT SUMMARY (FCLC OP)",
            frame=fclc_op,
            max_rows=50,
        )
        row = _write_section(
            worksheet,
            row,
            title="MOVING CALLS",
            frame=moving_calls,
            max_rows=100,
            guidance=(
                "A changing first and last Cell ID is a network movement indicator. "
                "Verify important events against source records."
            ),
        )
        row = _write_section(
            worksheet,
            row,
            title="MOVEMENT EVENTS",
            frame=movements,
            max_rows=100,
        )
        row = _write_section(
            worksheet,
            row,
            title="TOWER TRANSITIONS",
            frame=tower_transition,
            max_rows=50,
        )
        row = _write_section(
            worksheet,
            row,
            title="MOVEMENT PATTERNS",
            frame=movement_pattern,
            max_rows=30,
        )
        _finish_sheet(
            worksheet
        )

        # 7. Device and SIM Intelligence
        worksheet = workbook.create_sheet(
            SINGLE_CDR_COMPACT_SHEETS[
                6
            ]
        )
        row = _write_sheet_title(
            worksheet,
            title="Device and SIM Intelligence",
            metadata=report_metadata,
        )
        row = _write_section(
            worksheet,
            row,
            title="IMEI SUMMARY",
            frame=imei_summary,
            max_rows=30,
        )
        row = _write_section(
            worksheet,
            row,
            title="IMEI INTELLIGENCE",
            frame=_as_frame(
                results.get(
                    "imei_intelligence"
                )
            ),
            max_rows=30,
        )
        row = _write_section(
            worksheet,
            row,
            title="SIM OR DEVICE CHANGES",
            frame=_as_frame(
                results.get(
                    "sim_change"
                )
            ),
            max_rows=50,
        )
        _finish_sheet(
            worksheet
        )

        # 8. Activity and Timing
        worksheet = workbook.create_sheet(
            SINGLE_CDR_COMPACT_SHEETS[
                7
            ]
        )
        row = _write_sheet_title(
            worksheet,
            title="Activity and Timing Intelligence",
            metadata=report_metadata,
        )
        for title, key, limit in (
            (
                "ACTIVITY SUMMARY",
                "activity_summary",
                30,
            ),
            (
                "HOURLY ACTIVITY",
                "hourly_activity",
                30,
            ),
            (
                "DAILY ACTIVITY",
                "daily_activity",
                120,
            ),
            (
                "WEEKLY ACTIVITY",
                "weekly_activity",
                60,
            ),
            (
                "MONTHLY ACTIVITY",
                "monthly_activity",
                36,
            ),
        ):
            row = _write_section(
                worksheet,
                row,
                title=title,
                frame=_as_frame(
                    results.get(
                        key
                    )
                ),
                max_rows=limit,
            )

        _finish_sheet(
            worksheet
        )

        # 9. Priority Review Queue
        worksheet = workbook.create_sheet(
            SINGLE_CDR_COMPACT_SHEETS[
                8
            ]
        )
        row = _write_sheet_title(
            worksheet,
            title="Priority Review Queue",
            metadata=report_metadata,
        )
        row = _write_section(
            worksheet,
            row,
            title="BEHAVIORAL OBSERVATIONS",
            frame=_as_frame(
                results.get(
                    "behavioral_intelligence"
                )
            ),
            max_rows=30,
        )
        row = _write_section(
            worksheet,
            row,
            title="REVIEW INDICATORS",
            frame=_as_frame(
                results.get(
                    "suspicious_activity"
                )
            ),
            max_rows=50,
            guidance=(
                "These are review indicators, not automated findings "
                "of guilt, identity or intent."
            ),
        )
        row = _write_section(
            worksheet,
            row,
            title="INVESTIGATOR VERIFICATION CHECKLIST",
            frame=_review_checklist(),
        )
        _finish_sheet(
            worksheet
        )

        # 10. Data Quality and Interpretation Guide
        worksheet = workbook.create_sheet(
            SINGLE_CDR_COMPACT_SHEETS[
                9
            ]
        )
        row = _write_sheet_title(
            worksheet,
            title="Data Quality and Interpretation Guide",
            metadata=report_metadata,
        )
        row = _write_section(
            worksheet,
            row,
            title="DATA QUALITY SUMMARY",
            frame=_quality_summary(
                data,
                report_metadata,
                original_data=df,
            ),
        )
        row = _write_section(
            worksheet,
            row,
            title="MISSING CGI LOOKUP",
            frame=_as_frame(
                results.get(
                    "missing_cgi_lookup"
                )
            ),
            max_rows=30,
        )
        row = _write_section(
            worksheet,
            row,
            title="MASTER DATA ENRICHMENT",
            frame=_as_frame(
                results.get(
                    "master_enrichment_summary"
                )
            ),
            max_rows=50,
        )
        row = _write_section(
            worksheet,
            row,
            title="REJECTED ROWS SUMMARY",
            frame=_rejected_rows_summary(
                df,
                report_metadata,
            ),
            max_rows=30,
        )
        row = _write_section(
            worksheet,
            row,
            title="INTERPRETATION AND LIMITATIONS",
            frame=_interpretation_guide(),
        )

        _finish_sheet(
            worksheet
        )

        report_path = Path(
            get_single_report_path(
                target,
                output_dir,
            )
        )

        report_path.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        workbook.save(
            report_path
        )

        print(
            "[+] Compact Single CDR report generated:",
            report_path,
        )
        print(
            "[+] Investigator workbook sheets:",
            len(
                workbook.sheetnames
            ),
        )

        return str(
            report_path
        )

    except Exception as error:
        print(
            "[-] Compact Single CDR report generation failed:",
            type(
                error
            ).__name__,
            "|",
            str(
                error
            ),
        )
        return None
