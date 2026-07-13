"""Reusable OpenPyXL styles for forensic Excel reports."""

from __future__ import annotations

from typing import Iterable

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FILL = PatternFill("solid", fgColor="5B9BD5")
SUBHEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
WHITE_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(color="FFFFFF", bold=True, size=14)
HEADER_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
THIN_SIDE = Side(style="thin", color="B7B7B7")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


def style_metadata_block(ws, metadata_rows: list[tuple[str, object]], max_column: int) -> int:
    """Write the metadata/title area and return the next available row."""
    row = 1
    for label, value in metadata_rows:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_column)
        cell = ws.cell(row=row, column=1, value=f"{label} - {'' if value is None else value}")
        cell.font = BOLD_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.fill = SECTION_FILL
        row += 1

    row += 1
    return row


def style_table_header(ws, row: int, column_count: int) -> None:
    for column in range(1, column_count + 1):
        cell = ws.cell(row=row, column=column)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 30


def style_data_area(ws, start_row: int, end_row: int, column_count: int) -> None:
    if end_row < start_row:
        return
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=column_count):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def set_sensible_widths(ws, headers: Iterable[str], min_width: int = 10, max_width: int = 40) -> None:
    for index, header in enumerate(headers, start=1):
        header_text = str(header or "")
        width = max(min_width, min(max_width, len(header_text) + 4))
        if any(token in header_text.lower() for token in ("address", "observation", "details")):
            width = max_width
        elif any(token in header_text.lower() for token in ("date", "time", "imei", "cell", "party", "number")):
            width = max(width, 18)
        ws.column_dimensions[get_column_letter(index)].width = width


def finish_sheet(ws, header_row: int, last_row: int, last_column: int) -> None:
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(last_column)}{max(header_row, last_row)}"
    ws.sheet_view.showGridLines = False
