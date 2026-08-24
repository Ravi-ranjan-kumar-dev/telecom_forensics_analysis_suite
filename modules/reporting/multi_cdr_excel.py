"""Generate a separate cross-target Excel workbook for multiple CDR analysis."""

from __future__ import annotations

from modules.core.time_utils import utc_now, utc_now_iso

from datetime import datetime
from pathlib import Path
from typing import Any
import traceback

import pandas as pd
from openpyxl import Workbook

from modules.analysis.cdr.cross_target import build_cross_target_analysis
from modules.enrichment.telecom_master_enrichment import (
    MULTI_CDR_TABLE_SPECS,
    enrich_analysis_bundle,
)
from .cdr_contact_map import generate_cdr_contact_map
from .cdr_movement_route import generate_multi_cdr_movement_route
from .excel_styles import (
    finish_sheet,
    set_sensible_widths,
    style_data_area,
    style_metadata_block,
    style_table_header,
)
from .excel_security import excel_safe_value
from .report_paths import get_multi_report_path


SHEET_MAP = [
    ("1. Cross Summary", "Multiple CDR Cross Summary", "summary"),
    ("2. Target Overview", "Target Overview", "target_overview"),
    ("3. Common Numbers", "Common Contact Numbers", "common_numbers"),
    ("4. Direct Links", "Direct Target-to-Target Links", "direct_target_links"),
    ("5. Common Towers", "Common Tower IDs", "common_towers"),
    ("6. Common IMEI", "Common IMEI / Shared Device", "common_imeis"),
    ("7. Common IMSI", "Common IMSI", "common_imsis"),
    ("8. Contact Matrix", "Common Number vs Target Matrix", "contact_matrix"),
    ("9. Tower Matrix", "Common Tower vs Target Matrix", "tower_matrix"),
    ("10. IMEI Matrix", "Common IMEI vs Target Matrix", "imei_matrix"),
    ("11. IMSI Matrix", "Common IMSI vs Target Matrix", "imsi_matrix"),
    ("12. Source Files", "Source Files", "source_files"),
]


_COMMON_NUMBER_SDR_RENAME = {
    "common_number_sdr_subscriber_name": "Name",
    "common_number_sdr_father_name": "Father Name",
    "common_number_sdr_address": "SDR Address",
    "common_number_sdr_operator": "SDR Operator",
    "common_number_sdr_circle": "SDR Circle",
    "common_number_sdr_lookup_status": "SDR Lookup Status",
}

_COMMON_TOWER_CGI_RENAME = {
    "common_tower_id_cgi_operator": "CGI Operator",
    "common_tower_id_cgi_circle": "CGI Circle",
    "common_tower_id_cgi_district": "CGI District",
    "common_tower_id_cgi_town": "CGI Town",
    "common_tower_id_cgi_site_name": "CGI Site Name",
    "common_tower_id_cgi_address": "CGI Address",
    "common_tower_id_cgi_latitude": "CGI Latitude",
    "common_tower_id_cgi_longitude": "CGI Longitude",
    "common_tower_id_cgi_lookup_status": "CGI Lookup Status",
}

_TOWER_MATRIX_TARGET_SDR_FIELDS = (
    (
        "target_sdr_subscriber_name",
        "Linked Target Names",
    ),
    (
        "target_sdr_father_name",
        "Linked Target Father Names",
    ),
    (
        "target_sdr_address",
        "Linked Target SDR Addresses",
    ),
    (
        "target_sdr_operator",
        "Linked Target Operators",
    ),
    (
        "target_sdr_circle",
        "Linked Target Circles",
    ),
    (
        "target_sdr_lookup_status",
        "Linked Target SDR Status",
    ),
)


def _ordered_columns(
    frame: pd.DataFrame,
    preferred: list[str],
) -> pd.DataFrame:
    """Keep preferred investigator fields first without losing metrics."""

    first = [
        column
        for column in preferred
        if column in frame.columns
    ]
    remaining = [
        column
        for column in frame.columns
        if column not in first
    ]
    return frame.loc[
        :,
        [*first, *remaining],
    ].copy()


def _display_text(
    value: Any,
) -> str:
    """Return one clean value for compact multi-target profile fields."""

    if value is None:
        return ""

    try:
        if pd.isna(
            value
        ):
            return ""
    except (
        TypeError,
        ValueError,
    ):
        pass

    return str(
        value
    ).strip()


def _has_events(
    value: Any,
) -> bool:
    """Return whether one dynamic target matrix cell has event evidence."""

    try:
        return float(
            value
        ) > 0
    except (
        TypeError,
        ValueError,
    ):
        return False


def _add_tower_matrix_target_sdr(
    frame: pd.DataFrame,
    target_profiles: pd.DataFrame | None,
) -> pd.DataFrame:
    """Attach linked-target SDR profiles without treating CGI as a number."""

    output = frame.copy()

    if (
        not isinstance(
            target_profiles,
            pd.DataFrame,
        )
        or target_profiles.empty
        or "Target" not in target_profiles.columns
    ):
        return output

    profiles: dict[str, dict[str, str]] = {}

    for record in target_profiles.to_dict(
        orient="records"
    ):
        target = _display_text(
            record.get(
                "Target"
            )
        )

        if not target:
            continue

        profiles[
            target
        ] = {
            source: _display_text(
                record.get(
                    source
                )
            )
            for source, _ in _TOWER_MATRIX_TARGET_SDR_FIELDS
        }

    target_columns = [
        target
        for target in profiles
        if target in output.columns
    ]

    if not target_columns:
        return output

    for source, heading in _TOWER_MATRIX_TARGET_SDR_FIELDS:
        values: list[str] = []

        for _, row in output.iterrows():
            linked = []

            for target in target_columns:
                if not _has_events(
                    row.get(
                        target
                    )
                ):
                    continue

                profile_value = profiles[
                    target
                ].get(
                    source,
                    "",
                )
                linked.append(
                    f"{target}: {profile_value or 'Not found'}"
                )

            values.append(
                "; ".join(
                    linked
                )
            )

        output[
            heading
        ] = values

    return output


def _present_multi_frame(
    result_key: str,
    frame: pd.DataFrame | None,
    *,
    target_profiles: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """Convert internal batch-enrichment fields to compact report columns."""

    output = (
        frame.copy()
        if isinstance(frame, pd.DataFrame)
        else pd.DataFrame()
    )

    rename_map: dict[str, str] = {}
    preferred: list[str] = []

    if result_key == "target_overview":
        rename_map.update(
            {
                "target_sdr_subscriber_name": "Subscriber Name",
                "target_sdr_father_name": "Father Name",
                "target_sdr_address": "SDR Address",
                "target_sdr_operator": "SDR Operator",
                "target_sdr_circle": "SDR Circle",
                "target_sdr_lookup_status": "SDR Lookup Status",
            }
        )
        preferred = [
            "Target",
            "Subscriber Name",
            "Father Name",
            "SDR Address",
            "SDR Operator",
            "SDR Circle",
            "SDR Lookup Status",
        ]

    elif result_key in {
        "common_numbers",
        "contact_matrix",
    }:
        rename_map.update(
            _COMMON_NUMBER_SDR_RENAME
        )
        preferred = [
            "Common Number",
            "Name",
            "Father Name",
            "SDR Address",
            "SDR Operator",
            "SDR Circle",
            "SDR Lookup Status",
        ]

    elif result_key == "direct_target_links":
        for prefix, label in (
            ("source_target_sdr_", "Source"),
            ("destination_target_sdr_", "Destination"),
        ):
            rename_map.update(
                {
                    f"{prefix}subscriber_name": f"{label} Name",
                    f"{prefix}father_name": f"{label} Father Name",
                    f"{prefix}address": f"{label} SDR Address",
                    f"{prefix}operator": f"{label} Operator",
                    f"{prefix}circle": f"{label} Circle",
                    f"{prefix}lookup_status": f"{label} SDR Lookup Status",
                }
            )
        preferred = [
            "Source Target",
            "Source Name",
            "Source Father Name",
            "Source SDR Address",
            "Source Operator",
            "Source Circle",
            "Source SDR Lookup Status",
            "Destination Target",
            "Destination Name",
            "Destination Father Name",
            "Destination SDR Address",
            "Destination Operator",
            "Destination Circle",
            "Destination SDR Lookup Status",
        ]

    elif result_key in {
        "common_towers",
        "tower_matrix",
    }:
        if result_key == "tower_matrix":
            output = _add_tower_matrix_target_sdr(
                output,
                target_profiles,
            )

        rename_map.update(
            _COMMON_TOWER_CGI_RENAME
        )
        preferred = [
            "Common Tower ID",
            "CGI Operator",
            "CGI Circle",
            "CGI District",
            "CGI Town",
            "CGI Site Name",
            "CGI Address",
            "CGI Latitude",
            "CGI Longitude",
            "CGI Lookup Status",
        ]

        if result_key == "tower_matrix":
            preferred.extend(
                heading
                for _, heading in _TOWER_MATRIX_TARGET_SDR_FIELDS
            )

    output = output.rename(
        columns=rename_map
    )

    internal_columns = [
        column
        for column in output.columns
        if (
            "_sdr_" in str(column)
            or "_cgi_" in str(column)
        )
    ]
    output = output.drop(
        columns=internal_columns,
        errors="ignore",
    )

    return _ordered_columns(
        output,
        preferred,
    )


def _contact_map_frame(
    frame: pd.DataFrame | None,
) -> pd.DataFrame:
    """Return the canonical fields consumed by the contact map renderer."""

    if not isinstance(frame, pd.DataFrame):
        return pd.DataFrame()

    rename_map = {
        "other_party_sdr_subscriber_name": "Name",
        "other_party_sdr_lookup_status": "SDR Lookup Status",
        "most_used_target_cgi_cgi_lookup_status": "Most Used CGI Lookup Status",
        "most_used_target_cgi_cgi_site_name": "Most Used Site Name",
        "most_used_target_cgi_cgi_address": "Most Used Tower Address",
        "most_used_target_cgi_cgi_latitude": "Most Used Latitude",
        "most_used_target_cgi_cgi_longitude": "Most Used Longitude",
        "last_interaction_cgi_cgi_lookup_status": "Last Interaction CGI Lookup Status",
        "last_interaction_cgi_cgi_site_name": "Last Interaction Site Name",
        "last_interaction_cgi_cgi_address": "Last Interaction Tower Address",
        "last_interaction_cgi_cgi_latitude": "Last Interaction Latitude",
        "last_interaction_cgi_cgi_longitude": "Last Interaction Longitude",
    }
    return frame.rename(
        columns=rename_map
    ).copy()


def _metadata_rows(
    metadata: dict[str, Any],
    report_name: str,
    target_count: int,
) -> list[tuple[str, Any]]:
    return [
        ("Case", metadata.get("case_name", "")),
        ("Report", report_name),
        ("Targets Analyzed", target_count),
        ("Generated On", utc_now_iso()),
        ("Minimum Common Threshold", metadata.get("min_targets", 2)),
    ]


def _write_dataframe_sheet(
    workbook: Workbook,
    sheet_name: str,
    report_name: str,
    metadata: dict[str, Any],
    target_count: int,
    frame: pd.DataFrame | None,
) -> None:
    worksheet = workbook.create_sheet(title=sheet_name)
    data = frame.copy() if isinstance(frame, pd.DataFrame) else pd.DataFrame()

    if data.shape[1] == 0:
        data = pd.DataFrame(columns=["Result"])

    headers = list(data.columns)
    header_row = style_metadata_block(
        worksheet,
        _metadata_rows(metadata, report_name, target_count),
        max(1, len(headers)),
    )

    for column_index, header in enumerate(headers, start=1):
        worksheet.cell(row=header_row, column=column_index, value=excel_safe_value(str(header)))
    style_table_header(worksheet, header_row, len(headers))

    for row_index, row in enumerate(data.itertuples(index=False, name=None), start=header_row + 1):
        for column_index, value in enumerate(row, start=1):
            worksheet.cell(
                row=row_index,
                column=column_index,
                value=excel_safe_value(value),
            )

    last_row = header_row + len(data)
    style_data_area(worksheet, header_row + 1, last_row, len(headers))
    set_sensible_widths(worksheet, headers)
    finish_sheet(worksheet, header_row, last_row, len(headers))

    for column_index, header in enumerate(headers, start=1):
        header_lower = str(header).lower()
        if "first seen" in header_lower or "last seen" in header_lower or header_lower in {"from date", "to date"}:
            for row_index in range(header_row + 1, last_row + 1):
                cell = worksheet.cell(row=row_index, column=column_index)
                if hasattr(cell.value, "year"):
                    cell.number_format = "dd-mm-yyyy hh:mm:ss"


def generate_multi_cdr_report(
    loaded_cdrs: dict[str, dict[str, Any]],
    metadata: dict[str, Any] | None = None,
    analysis_bundle: dict[str, Any] | None = None,
    output_dir: str | Path | None = None,
    min_targets: int = 2,
) -> str | None:
    """Generate the separate multiple-CDR common-analysis workbook."""
    if not isinstance(loaded_cdrs, dict) or len(loaded_cdrs) < 2:
        print("[-] Multi-CDR Excel report requires at least two loaded targets.")
        return None

    try:
        report_metadata = dict(metadata or {})
        report_metadata["min_targets"] = max(2, int(min_targets))
        bundle = analysis_bundle or build_cross_target_analysis(
            loaded_cdrs,
            min_targets=report_metadata["min_targets"],
        )

        enrichment = enrich_analysis_bundle(
            bundle,
            table_specs=MULTI_CDR_TABLE_SPECS,
        )
        bundle = enrichment[
            "bundle"
        ]

        for warning in enrichment.get(
            "warnings",
            [],
        ):
            print(
                "[!] Multiple CDR master-data enrichment:",
                warning,
            )

        workbook = Workbook()
        workbook.remove(workbook.active)

        for sheet_name, report_name, result_key in SHEET_MAP:
            frame = _present_multi_frame(
                result_key,
                bundle.get(
                    result_key
                ),
                target_profiles=bundle.get(
                    "target_overview"
                ),
            )

            _write_dataframe_sheet(
                workbook=workbook,
                sheet_name=sheet_name,
                report_name=report_name,
                metadata=report_metadata,
                target_count=len(loaded_cdrs),
                frame=frame,
            )

        path = get_multi_report_path(
            case_name=report_metadata.get("case_name"),
            output_dir=output_dir,
        )
        workbook.save(path)

        map_path = None
        route_path = None

        try:
            map_path = generate_cdr_contact_map(
                _contact_map_frame(
                    bundle.get(
                        "common_contact_map"
                    )
                ),
                target=(
                    "Multiple CDR Common Contacts "
                    f"({len(loaded_cdrs)} targets)"
                ),
                report_path=path,
            )
        except Exception as map_error:
            print(
                "[!] Multiple CDR contact map generation skipped:",
                type(map_error).__name__,
                "|",
                str(map_error),
            )

        try:
            route_path = generate_multi_cdr_movement_route(
                bundle.get(
                    "movement_route_events",
                    pd.DataFrame(),
                ),
                report_path=path,
            )
        except Exception as route_error:
            print(
                "[!] Multiple CDR movement route generation skipped:",
                type(route_error).__name__,
                "|",
                str(route_error),
            )

        print(f"[+] Multiple CDR common-analysis Excel report generated: {path}")

        if map_path is not None:
            print(f"[+] Multiple CDR contact map generated: {map_path}")

        if route_path is not None:
            print(f"[+] Multiple CDR movement route generated: {route_path}")

        return str(path)

    except Exception as error:
        print("[-] Multiple CDR common-analysis Excel report generation failed.")
        print(f"    Error Type : {type(error).__name__}")
        print(f"    Message    : {error}")
        print(traceback.format_exc(limit=4).rstrip())
        return None
