from __future__ import annotations

from modules.core.time_utils import utc_now, utc_now_iso

from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .excel_security import excel_safe_value
from .report_guidance import append_methodology_sheet


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "output" / "reports" / "tower_dump"

TITLE_FILL = "1F4E78"
HEADER_FILL = "BDD7EE"
WHITE = "FFFFFF"


def _safe_sheet_name(name: str) -> str:
    invalid = ["\\", "/", "?", "*", "[", "]", ":"]
    for char in invalid:
        name = name.replace(char, " ")
    return name[:31].strip() or "Sheet"


def _clean_value(value: Any) -> Any:
    """Compatibility wrapper around the shared Excel security boundary."""
    return excel_safe_value(value)


def _as_dataframe(value: Any) -> pd.DataFrame:
    if isinstance(value, pd.DataFrame):
        return value.copy()
    if isinstance(value, pd.Series):
        return value.reset_index()
    if isinstance(value, dict):
        return pd.DataFrame(
            [{"Field": key, "Value": _clean_value(item)} for key, item in value.items()]
        )
    if isinstance(value, list):
        if not value:
            return pd.DataFrame()
        if all(isinstance(item, dict) for item in value):
            return pd.DataFrame(value)
        return pd.DataFrame({"Value": value})
    if value is None:
        return pd.DataFrame()
    return pd.DataFrame({"Value": [value]})


def _style_worksheet(ws) -> None:
    ws.freeze_panes = "A2"
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.font = Font(bold=True, color="000000")
            cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for column_cells in ws.columns:
        length = 0
        letter = get_column_letter(column_cells[0].column)
        for cell in column_cells[:300]:
            text = "" if cell.value is None else str(cell.value)
            length = max(length, min(len(text), 70))
        ws.column_dimensions[letter].width = max(12, min(length + 2, 45))


def _write_dataframe(
    wb: Workbook,
    sheet_name: str,
    value: Any,
    *,
    max_rows: int | None = None,
) -> None:
    ws = wb.create_sheet(_safe_sheet_name(sheet_name))
    df = _as_dataframe(value)

    if max_rows is not None and len(df) > max_rows:
        df = df.head(max_rows).copy()
        marker = {column: "" for column in df.columns}
        if len(df.columns) > 0:
            marker[df.columns[0]] = f"... trimmed to first {max_rows} rows"
            df.loc[len(df)] = marker

    if df.empty:
        ws.append(["Message"])
        ws.append(["No records found."])
        _style_worksheet(ws)
        return

    ws.append([_clean_value(str(column)) for column in df.columns])

    for _, row in df.iterrows():
        ws.append([_clean_value(row.get(column)) for column in df.columns])

    _style_worksheet(ws)


def _write_section(
    ws,
    title: str,
    row_index: int,
    value: Any,
    *,
    max_rows: int | None = None,
) -> int:
    ws.cell(row_index, 1, title)
    ws.cell(row_index, 1).font = Font(bold=True, color=WHITE)
    ws.cell(row_index, 1).fill = PatternFill("solid", fgColor=TITLE_FILL)
    ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=8)
    row_index += 1

    df = _as_dataframe(value)

    if max_rows is not None and len(df) > max_rows:
        df = df.head(max_rows).copy()

    if df.empty:
        ws.cell(row_index, 1, "No records found.")
        return row_index + 3

    for col_index, column in enumerate(df.columns, start=1):
        cell = ws.cell(row_index, col_index, _clean_value(str(column)))
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_index += 1

    for _, data_row in df.iterrows():
        for col_index, column in enumerate(df.columns, start=1):
            ws.cell(row_index, col_index, _clean_value(data_row.get(column)))
        row_index += 1

    return row_index + 2


def _style_section_sheet(ws) -> None:
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    for column_cells in ws.columns:
        length = 0
        letter = get_column_letter(column_cells[0].column)
        for cell in column_cells[:300]:
            text = "" if cell.value is None else str(cell.value)
            length = max(length, min(len(text), 70))
        ws.column_dimensions[letter].width = max(12, min(length + 2, 45))


# TOWER_COMPACT_REPORT_QUALITY_V2
# TOWER_COMPACT_REPORT_QUALITY_V3


def _compact_int(
    value: Any,
    default: int = 0,
) -> int:
    """Convert one scalar value to int safely."""

    try:
        if value is None:
            return default

        if isinstance(
            value,
            (
                list,
                tuple,
                dict,
                set,
                pd.Series,
                pd.DataFrame,
            ),
        ):
            return default

        if pd.isna(value):
            return default

        return int(value)

    except (
        TypeError,
        ValueError,
        OverflowError,
    ):
        return default


def _compact_first_positive_int(
    *values: Any,
    default: int = 0,
) -> int:
    """Return the first positive integer value."""

    for value in values:
        converted = _compact_int(
            value,
            default=0,
        )

        if converted > 0:
            return converted

    return default


def _compact_summary_mapping(
    value: Any,
) -> dict[str, Any]:
    """Convert a summary dict/DataFrame into a lookup mapping."""

    if isinstance(
        value,
        dict,
    ):
        return dict(value)

    dataframe = _as_dataframe(
        value
    )

    if dataframe.empty:
        return {}

    for key_column, value_column in (
        (
            "Field",
            "Value",
        ),
        (
            "field",
            "value",
        ),
        (
            "metric",
            "value",
        ),
        (
            "Metric",
            "Value",
        ),
    ):
        if {
            key_column,
            value_column,
        }.issubset(
            dataframe.columns
        ):
            return dict(
                zip(
                    dataframe[
                        key_column
                    ].astype(str),
                    dataframe[
                        value_column
                    ],
                )
            )

    if len(dataframe) == 1:
        return dataframe.iloc[
            0
        ].to_dict()

    return {}


def _compact_alias_mismatch_count(
    dataframe: pd.DataFrame,
    canonical_column: str,
    alias_column: str,
) -> int:
    """Count non-equivalent canonical/legacy identity values."""

    if (
        canonical_column
        not in dataframe.columns
        or alias_column
        not in dataframe.columns
    ):
        return 0

    canonical = (
        dataframe[
            canonical_column
        ]
        .astype("string")
        .fillna("")
        .str.strip()
    )

    alias = (
        dataframe[
            alias_column
        ]
        .astype("string")
        .fillna("")
        .str.strip()
    )

    relevant = (
        canonical.ne("")
        | alias.ne("")
    )

    return int(
        (
            relevant
            & canonical.ne(alias)
        ).sum()
    )


def _compact_report_context(
    result: dict[str, Any],
) -> dict[str, Any]:
    """Build one reusable and evidence-aware report context."""

    metadata = (
        result.get(
            "metadata",
            {},
        )
        or {}
    )

    analysis = (
        result.get(
            "analysis",
            {},
        )
        or {}
    )

    results = (
        analysis.get(
            "results",
            {},
        )
        or {}
    )

    dataframe = result.get(
        "df"
    )

    if not isinstance(
        dataframe,
        pd.DataFrame,
    ):
        dataframe = pd.DataFrame()

    file_summary = _as_dataframe(
        result.get(
            "file_summary",
            pd.DataFrame(),
        )
    )

    spot_summary = _as_dataframe(
        results.get(
            "spot_summary",
            result.get(
                "spot_summary",
                pd.DataFrame(),
            ),
        )
    )

    tower_summary = (
        _compact_summary_mapping(
            results.get(
                "tower_dump_summary",
                {},
            )
        )
    )

    scalable_pipeline = (
        result.get(
            "scalable_pipeline",
            {},
        )
        or {}
    )

    stage = (
        scalable_pipeline.get(
            "stage",
            {},
        )
        or {}
    )

    if not isinstance(
        stage,
        dict,
    ):
        stage = {}

    total_records = (
        len(dataframe)
        if not dataframe.empty
        else _compact_first_positive_int(
            tower_summary.get(
                "total_records"
            ),
            metadata.get(
                "records_after_deduplication"
            ),
            metadata.get(
                "total_records"
            ),
        )
    )

    duplicates_removed = _compact_int(
        metadata.get(
            "duplicates_removed"
        ),
        default=0,
    )

    records_before = _compact_first_positive_int(
        metadata.get(
            "records_before_deduplication"
        ),
        total_records
        + duplicates_removed,
        default=total_records,
    )

    records_after = _compact_first_positive_int(
        metadata.get(
            "records_after_deduplication"
        ),
        total_records,
        default=total_records,
    )

    if "status" in file_summary.columns:
        status_series = (
            file_summary[
                "status"
            ]
            .astype("string")
            .fillna("")
            .str.upper()
        )
    else:
        status_series = pd.Series(
            "",
            index=file_summary.index,
            dtype="string",
        )

    if "spot_id" in file_summary.columns:
        root_mask = (
            file_summary[
                "spot_id"
            ]
            .astype("string")
            .fillna("")
            .str.upper()
            .eq(
                "UNASSIGNED-ROOT"
            )
        )
    else:
        root_mask = pd.Series(
            False,
            index=file_summary.index,
        )

    root_file_count = int(
        root_mask.sum()
    )

    if "records" in file_summary.columns:
        file_records = (
            pd.to_numeric(
                file_summary[
                    "records"
                ],
                errors="coerce",
            )
            .fillna(0)
        )

        root_record_count = int(
            file_records.loc[
                root_mask
            ].sum()
        )

        logical_loaded_entries = int(
            file_records.gt(0).sum()
        )
    else:
        root_record_count = 0
        logical_loaded_entries = 0

    physical_files = _compact_first_positive_int(
        metadata.get(
            "files_found"
        ),
        stage.get(
            "input_files"
        ),
        stage.get(
            "files_found"
        ),
        default=len(
            file_summary
        ),
    )

    physical_loaded_files = _compact_first_positive_int(
        metadata.get(
            "files_loaded"
        ),
        stage.get(
            "files_loaded"
        ),
        default=logical_loaded_entries,
    )

    files_empty = _compact_int(
        metadata.get(
            "files_empty_no_data"
        ),
        default=int(
            status_series.str.contains(
                "EMPTY",
                regex=False,
            ).sum()
        ),
    )

    files_failed = _compact_int(
        metadata.get(
            "files_failed"
        ),
        default=int(
            status_series.str.contains(
                "FAIL",
                regex=False,
            ).sum()
        ),
    )

    if (
        not spot_summary.empty
        and "spot_id"
        in spot_summary.columns
    ):
        valid_spot_mask = (
            spot_summary[
                "spot_id"
            ]
            .astype("string")
            .fillna("")
            .str.upper()
            .ne("")
            & spot_summary[
                "spot_id"
            ]
            .astype("string")
            .fillna("")
            .str.upper()
            .ne(
                "UNASSIGNED-ROOT"
            )
        )

        valid_spots = spot_summary.loc[
            valid_spot_mask
        ].copy()

    elif (
        not dataframe.empty
        and "spot_id"
        in dataframe.columns
    ):
        valid_spots = (
            dataframe.loc[
                dataframe[
                    "spot_id"
                ]
                .astype("string")
                .fillna("")
                .str.upper()
                .ne("")
                & dataframe[
                    "spot_id"
                ]
                .astype("string")
                .fillna("")
                .str.upper()
                .ne(
                    "UNASSIGNED-ROOT"
                ),
                [
                    column
                    for column in (
                        "spot_id",
                        "spot_name",
                    )
                    if column
                    in dataframe.columns
                ],
            ]
            .drop_duplicates()
        )

    else:
        valid_spots = pd.DataFrame()

    if (
        not valid_spots.empty
        and "spot_id"
        in valid_spots.columns
    ):
        valid_spot_count = int(
            valid_spots[
                "spot_id"
            ].nunique()
        )
    else:
        valid_spot_count = 0

    if (
        not valid_spots.empty
        and "spot_name"
        in valid_spots.columns
    ):
        spot_names = sorted(
            value
            for value in (
                valid_spots[
                    "spot_name"
                ]
                .astype("string")
                .fillna("")
                .str.strip()
                .unique()
                .tolist()
            )
            if value
        )
    else:
        spot_names = []

    if (
        not dataframe.empty
        and "operator"
        in dataframe.columns
    ):
        operators = sorted(
            value
            for value in (
                dataframe[
                    "operator"
                ]
                .astype("string")
                .fillna("")
                .str.strip()
                .str.lower()
                .unique()
                .tolist()
            )
            if value
        )
    else:
        operators = sorted(
            str(value)
            for value in (
                result.get(
                    "operators",
                    [],
                )
                or []
            )
            if str(value).strip()
        )

    if (
        not dataframe.empty
        and "searched_cell_id"
        in dataframe.columns
    ):
        searched_cell_count = int(
            dataframe[
                "searched_cell_id"
            ]
            .astype("string")
            .fillna("")
            .str.strip()
            .replace(
                "",
                pd.NA,
            )
            .nunique(
                dropna=True
            )
        )
    else:
        searched_cell_count = len(
            result.get(
                "cell_ids",
                [],
            )
            or []
        )

    date_from = metadata.get(
        "date_from",
        "",
    )

    date_to = metadata.get(
        "date_to",
        "",
    )

    if (
        (
            not date_from
            or not date_to
        )
        and not dataframe.empty
        and "call_datetime"
        in dataframe.columns
    ):
        datetimes = pd.to_datetime(
            dataframe[
                "call_datetime"
            ],
            errors="coerce",
        ).dropna()

        if not datetimes.empty:
            if not date_from:
                date_from = str(
                    datetimes.min()
                )

            if not date_to:
                date_to = str(
                    datetimes.max()
                )

    input_mode = str(
        metadata.get(
            "input_mode",
            "",
        )
        or ""
    ).strip()

    if not input_mode:
        cached_status_count = int(
            status_series.eq(
                "CACHED_STAGE"
            ).sum()
        )

        if (
            len(file_summary) > 0
            and cached_status_count
            == len(file_summary)
        ):
            input_mode = (
                "Cached normalized Parquet stage "
                "(raw parsing skipped)"
            )
        else:
            input_mode = (
                "Normalized Tower CDR processing"
            )

    summary_total_records = _compact_first_positive_int(
        tower_summary.get(
            "total_records"
        ),
        default=0,
    )

    if summary_total_records <= 0:
        record_count_status = "INFO"
        record_count_note = (
            "No separate analytical summary count "
            "was available for comparison."
        )

    elif summary_total_records == total_records:
        record_count_status = "PASS"
        record_count_note = (
            "Normalized DataFrame and analysis summary "
            "record counts match."
        )

    else:
        record_count_status = "REVIEW"
        record_count_note = (
            f"Normalized records={total_records:,}; "
            f"analysis summary={summary_total_records:,}."
        )

    logical_source_entries = len(
        file_summary
    )

    if (
        physical_files > 0
        and logical_source_entries > 0
        and physical_files
        != logical_source_entries
    ):
        file_inventory_status = "INFO"
        file_inventory_note = (
            f"{physical_files:,} physical input files and "
            f"{logical_source_entries:,} indexed logical "
            "source entries were recorded. These are "
            "different inventory layers."
        )
    else:
        file_inventory_status = "PASS"
        file_inventory_note = (
            "Physical and indexed source inventory "
            "counts are consistent."
        )

    subscriber_alias_mismatches = (
        _compact_alias_mismatch_count(
            dataframe,
            "subscriber_number",
            "a_party",
        )
    )

    other_party_alias_mismatches = (
        _compact_alias_mismatch_count(
            dataframe,
            "other_party",
            "b_party",
        )
    )

    alias_status = (
        "PASS"
        if (
            subscriber_alias_mismatches == 0
            and other_party_alias_mismatches == 0
        )
        else "REVIEW"
    )

    return {
        "metadata": metadata,
        "analysis": analysis,
        "results": results,
        "dataframe": dataframe,
        "file_summary": file_summary,
        "spot_summary": spot_summary,
        "total_records": total_records,
        "records_before": records_before,
        "records_after": records_after,
        "duplicates_removed": duplicates_removed,
        "physical_files": physical_files,
        "physical_loaded_files": physical_loaded_files,
        "logical_source_entries": logical_source_entries,
        "logical_loaded_entries": logical_loaded_entries,
        "files_empty": files_empty,
        "files_failed": files_failed,
        "valid_spot_count": valid_spot_count,
        "spot_names": spot_names,
        "root_file_count": root_file_count,
        "root_record_count": root_record_count,
        "operators": operators,
        "searched_cell_count": searched_cell_count,
        "date_from": date_from,
        "date_to": date_to,
        "input_mode": input_mode,
        "summary_total_records": summary_total_records,
        "record_count_status": record_count_status,
        "record_count_note": record_count_note,
        "file_inventory_status": file_inventory_status,
        "file_inventory_note": file_inventory_note,
        "subscriber_alias_mismatches": (
            subscriber_alias_mismatches
        ),
        "other_party_alias_mismatches": (
            other_party_alias_mismatches
        ),
        "alias_status": alias_status,
    }


def _compact_data_quality_overview(
    context: dict[str, Any],
) -> pd.DataFrame:
    """Build concise report-level data-quality checks."""

    rows = [
        {
            "Check": "Normalized Record Count",
            "Value": context[
                "total_records"
            ],
            "Status": context[
                "record_count_status"
            ],
            "Meaning": context[
                "record_count_note"
            ],
        },
        {
            "Check": "Physical Input Files",
            "Value": context[
                "physical_files"
            ],
            "Status": "INFO",
            "Meaning": (
                "Files physically detected in the "
                "Tower CDR input inventory."
            ),
        },
        {
            "Check": "Indexed Logical Sources",
            "Value": context[
                "logical_source_entries"
            ],
            "Status": context[
                "file_inventory_status"
            ],
            "Meaning": context[
                "file_inventory_note"
            ],
        },
        {
            "Check": "Loaded Logical Sources",
            "Value": context[
                "logical_loaded_entries"
            ],
            "Status": (
                "PASS"
                if context[
                    "logical_loaded_entries"
                ] > 0
                else "REVIEW"
            ),
            "Meaning": (
                "Indexed source entries containing "
                "usable normalized records."
            ),
        },
        {
            "Check": "Failed Files",
            "Value": context[
                "files_failed"
            ],
            "Status": (
                "PASS"
                if context[
                    "files_failed"
                ] == 0
                else "REVIEW"
            ),
            "Meaning": (
                "Failed source files require review "
                "before relying on complete coverage."
            ),
        },
        {
            "Check": "Empty / No-Data Files",
            "Value": context[
                "files_empty"
            ],
            "Status": "INFO",
            "Meaning": (
                "Valid reports that contained no "
                "telecom records."
            ),
        },
        {
            "Check": "Valid Investigation Spots",
            "Value": context[
                "valid_spot_count"
            ],
            "Status": (
                "PASS"
                if context[
                    "valid_spot_count"
                ] > 0
                else "REVIEW"
            ),
            "Meaning": (
                "UNASSIGNED-ROOT is not counted as "
                "an investigation Spot."
            ),
        },
        {
            "Check": "Unassigned Root Files",
            "Value": context[
                "root_file_count"
            ],
            "Status": (
                "REVIEW"
                if context[
                    "root_file_count"
                ] > 0
                else "PASS"
            ),
            "Meaning": (
                f"These files contain "
                f"{context['root_record_count']:,} records "
                "but are not assigned to a Spot."
            ),
        },
        {
            "Check": "Canonical Subscriber Role Mapping",
            "Value": context[
                "subscriber_alias_mismatches"
            ],
            "Status": "INFO",
            "Meaning": (
                "Rows where subscriber_number differs "
                "from original a_party because the "
                "canonical subscriber identifies the "
                "party associated with the searched "
                "tower record. This can be expected for "
                "incoming or reversed-direction records."
            ),
        },
        {
            "Check": "Canonical Other-Party Role Mapping",
            "Value": context[
                "other_party_alias_mismatches"
            ],
            "Status": "INFO",
            "Meaning": (
                "Rows where other_party differs from "
                "original b_party after directional "
                "role normalization. This is not a data "
                "error by itself."
            ),
        },
    ]

    return pd.DataFrame(
        rows
    )


def _compact_report_view(
    value: Any,
    preferred_columns: list[str],
) -> pd.DataFrame:
    """Return only investigator-friendly available columns."""

    dataframe = _as_dataframe(
        value
    )

    if dataframe.empty:
        return dataframe

    available = [
        column
        for column in preferred_columns
        if column
        in dataframe.columns
    ]

    enrichment_columns = [
        "sdr_lookup_status",
        "sdr_subscriber_name",
        "sdr_father_name",
        "sdr_address",
        "sdr_operator",
        "sdr_circle",
        "sdr_activation_date",
        "sdr_caf_number",
        "searched_cell_lookup_status",
        "searched_cell_address",
        "searched_cell_district",
        "searched_cell_police_station",
        "first_cell_lookup_status",
        "first_cell_address",
        "last_cell_lookup_status",
        "last_cell_address",
        "cgi_lookup_status",
        "cgi_address",
    ]

    for column in enrichment_columns:
        if (
            column in dataframe.columns
            and column not in available
        ):
            available.append(
                column
            )

    if not available:
        return dataframe

    return dataframe.loc[
        :,
        available,
    ].copy()


def _compact_priority_review_queue(
    results: dict[str, Any],
    *,
    limit: int,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Build one deduplicated subscriber-level review queue."""

    definitions = (
        (
            "Priority Lead",
            "tower_cdr_priority_leads",
        ),
        (
            "Rare / Uncommon",
            "tower_cdr_uncommon_numbers",
        ),
        (
            "Multi-Cell",
            "tower_cdr_multi_cell_presence",
        ),
        (
            "Suspicious Time",
            "tower_cdr_suspicious_timing",
        ),
        (
            "High Activity",
            "tower_cdr_common_numbers",
        ),
        (
            "Device / SIM Alert",
            "tower_cdr_device_consistency",
        ),
    )

    frames: list[pd.DataFrame] = []
    coverage_rows: list[dict[str, Any]] = []

    for category, result_name in definitions:
        dataframe = _as_dataframe(
            results.get(
                result_name
            )
        )

        if (
            dataframe.empty
            or "subscriber_number"
            not in dataframe.columns
        ):
            coverage_rows.append(
                {
                    "Lead Category": category,
                    "Distinct Subscribers": 0,
                }
            )
            continue

        frame = dataframe.copy()

        frame[
            "subscriber_number"
        ] = (
            frame[
                "subscriber_number"
            ]
            .astype("string")
            .fillna("")
            .str.strip()
        )

        frame = frame.loc[
            frame[
                "subscriber_number"
            ].ne("")
        ].copy()

        frame[
            "_lead_category"
        ] = category

        coverage_rows.append(
            {
                "Lead Category": category,
                "Distinct Subscribers": int(
                    frame[
                        "subscriber_number"
                    ].nunique()
                ),
            }
        )

        frames.append(
            frame
        )

    coverage = pd.DataFrame(
        coverage_rows
    )

    output_columns = [
        "subscriber_number",
        "priority",
        "confidence",
        "priority_score",
        "lead_categories",
        "rare_uncommon",
        "multi_cell",
        "suspicious_time",
        "high_activity",
        "device_sim_alert",
        "event_count",
        "cells_seen",
        "night_event_count",
        "imei_count",
        "imsi_count",
        "first_seen",
        "last_seen",
        "why_important",
        "next_action",
    ]

    if not frames:
        return (
            pd.DataFrame(
                columns=output_columns
            ),
            coverage,
        )

    combined = pd.concat(
        frames,
        ignore_index=True,
        sort=False,
    )

    defaults: dict[str, Any] = {
        "priority": "",
        "confidence": "",
        "priority_score": 0,
        "event_count": 0,
        "cells_seen": 0,
        "night_event_count": 0,
        "imei_count": 0,
        "imsi_count": 0,
        "first_seen": pd.NaT,
        "last_seen": pd.NaT,
        "why_important": "",
        "next_action": "",
    }

    for column, default in defaults.items():
        if column not in combined.columns:
            combined[
                column
            ] = default

    for column in (
        "priority_score",
        "event_count",
        "cells_seen",
        "night_event_count",
        "imei_count",
        "imsi_count",
    ):
        combined[
            column
        ] = pd.to_numeric(
            combined[
                column
            ],
            errors="coerce",
        ).fillna(0)

    combined[
        "_priority_rank"
    ] = (
        combined[
            "priority"
        ]
        .astype("string")
        .str.upper()
        .map(
            {
                "HIGH": 3,
                "MEDIUM": 2,
                "LOW": 1,
            }
        )
        .fillna(0)
    )

    combined[
        "_confidence_rank"
    ] = (
        combined[
            "confidence"
        ]
        .astype("string")
        .str.upper()
        .map(
            {
                "HIGH": 3,
                "MEDIUM": 2,
                "LOW": 1,
            }
        )
        .fillna(0)
    )

    category_map = (
        combined.groupby(
            "subscriber_number",
            sort=False,
            observed=True,
        )[
            "_lead_category"
        ]
        .agg(
            lambda values: ", ".join(
                dict.fromkeys(
                    str(value)
                    for value in values
                )
            )
        )
        .reset_index(
            name="lead_categories"
        )
    )

    best = (
        combined.sort_values(
            [
                "_priority_rank",
                "_confidence_rank",
                "priority_score",
                "event_count",
                "cells_seen",
                "subscriber_number",
            ],
            ascending=[
                False,
                False,
                False,
                False,
                False,
                True,
            ],
            kind="stable",
        )
        .drop_duplicates(
            subset=[
                "subscriber_number"
            ],
            keep="first",
        )
        .merge(
            category_map,
            on="subscriber_number",
            how="left",
            validate="one_to_one",
        )
    )

    flag_map = {
        "rare_uncommon": "Rare / Uncommon",
        "multi_cell": "Multi-Cell",
        "suspicious_time": "Suspicious Time",
        "high_activity": "High Activity",
        "device_sim_alert": "Device / SIM Alert",
    }

    for column, category in flag_map.items():
        best[
            column
        ] = (
            best[
                "lead_categories"
            ]
            .astype("string")
            .str.contains(
                category,
                regex=False,
            )
            .map(
                {
                    True: "Yes",
                    False: "",
                }
            )
        )

    for column in (
        "why_important",
        "next_action",
        "priority",
        "confidence",
        "lead_categories",
    ):
        best[
            column
        ] = best[
            column
        ].fillna("")

    return (
        best.loc[
            :,
            output_columns,
        ]
        .head(
            max(
                1,
                int(limit),
            )
        )
        .reset_index(
            drop=True
        ),
        coverage,
    )


def _compact_balanced_normalized_sample(
    dataframe: pd.DataFrame,
    *,
    max_rows: int,
) -> pd.DataFrame:
    """Build a balanced Spot/operator sample with canonical columns."""

    report_columns = [
        "subscriber_number",
        "other_party",
        "call_type",
        "call_datetime",
        "call_duration",
        "operator",
        "spot_id",
        "spot_name",
        "searched_cell_id",
        "present_at_searched_cell",
        "imei",
        "imsi",
        "first_cell_id",
        "last_cell_id",
        "searched_tower_address",
        "searched_tower_district",
        "searched_tower_state",
        "searched_tower_latitude",
        "searched_tower_longitude",
        "source_file",
        "source_row",
        "is_potential_duplicate",
        "potential_duplicate_count",
    ]

    renamed_columns = {
        "call_duration": "duration_seconds",
        "present_at_searched_cell": (
            "searched_cell_presence"
        ),
        "is_potential_duplicate": (
            "potential_duplicate"
        ),
    }

    if not isinstance(
        dataframe,
        pd.DataFrame,
    ) or dataframe.empty:
        return pd.DataFrame(
            columns=[
                renamed_columns.get(
                    column,
                    column,
                )
                for column in report_columns
            ]
        )

    source_for_column: dict[str, str] = {}

    for canonical, alias in (
        (
            "subscriber_number",
            "a_party",
        ),
        (
            "other_party",
            "b_party",
        ),
    ):
        if canonical in dataframe.columns:
            source_for_column[
                canonical
            ] = canonical

        elif alias in dataframe.columns:
            source_for_column[
                canonical
            ] = alias

    for column in report_columns:
        if column in source_for_column:
            continue

        if column in dataframe.columns:
            source_for_column[
                column
            ] = column

    source_columns = list(
        dict.fromkeys(
            source_for_column.values()
        )
    )

    if not source_columns:
        return pd.DataFrame()

    work = dataframe.loc[
        :,
        source_columns,
    ].copy()

    for canonical, source_column in (
        source_for_column.items()
    ):
        if (
            source_column != canonical
            and source_column
            in work.columns
        ):
            work = work.rename(
                columns={
                    source_column: canonical
                }
            )

    work = work.loc[
        :,
        [
            column
            for column in report_columns
            if column
            in work.columns
        ],
    ]

    limit = min(
        max(
            1,
            int(max_rows),
        ),
        len(work),
    )

    group_columns = [
        column
        for column in (
            "spot_id",
            "operator",
        )
        if column
        in work.columns
    ]

    if not group_columns:
        sample = work.head(
            limit
        ).copy()

    else:
        group_catalog = (
            work[
                group_columns
            ]
            .fillna("")
            .drop_duplicates()
        )

        group_count = max(
            1,
            len(group_catalog),
        )

        quota = max(
            1,
            limit
            // group_count,
        )

        sample = (
            work.groupby(
                group_columns,
                sort=True,
                dropna=False,
                observed=True,
                group_keys=False,
            )
            .head(
                quota
            )
        )

        if len(sample) < limit:
            remaining = work.loc[
                ~work.index.isin(
                    sample.index
                )
            ].head(
                limit
                - len(sample)
            )

            sample = pd.concat(
                [
                    sample,
                    remaining,
                ],
                axis=0,
            )

        sample = sample.head(
            limit
        ).copy()

    sort_columns = [
        column
        for column in (
            "spot_id",
            "operator",
            "call_datetime",
            "source_file",
            "source_row",
        )
        if column
        in sample.columns
    ]

    if sort_columns:
        sample = sample.sort_values(
            sort_columns,
            kind="stable",
            ignore_index=True,
        )
    else:
        sample = sample.reset_index(
            drop=True
        )

    return sample.rename(
        columns=renamed_columns
    )


def _compact_normalized_sample_info(
    context: dict[str, Any],
    sample: pd.DataFrame,
) -> pd.DataFrame:
    """Describe the bounded normalized sample."""

    if (
        not sample.empty
        and {
            "spot_id",
            "operator",
        }.issubset(
            sample.columns
        )
    ):
        coverage_pairs = (
            sample[
                [
                    "spot_id",
                    "operator",
                ]
            ]
            .fillna("")
            .drop_duplicates()
        )

        coverage = "; ".join(
            (
                f"{row.spot_id or 'NO-SPOT'}"
                f" / "
                f"{row.operator or 'NO-OPERATOR'}"
            )
            for row in coverage_pairs.itertuples(
                index=False
            )
        )
    else:
        coverage = (
            "Spot/operator grouping not available."
        )

    rows = [
        (
            "Purpose",
            (
                "Bounded investigation sample; "
                "not the complete Tower Dump."
            ),
        ),
        (
            "Sample Rows",
            len(sample),
        ),
        (
            "Complete Backend Records",
            context[
                "total_records"
            ],
        ),
        (
            "Sampling Method",
            (
                "Balanced deterministic sample by "
                "Spot and operator."
            ),
        ),
        (
            "Canonical Subscriber Column",
            "subscriber_number",
        ),
        (
            "Canonical Other-Party Column",
            "other_party",
        ),
        (
            "Coverage",
            coverage,
        ),
        (
            "Full Data Location",
            (
                "DuckDB / partitioned Parquet "
                "analytical backend."
            ),
        ),
    ]

    return pd.DataFrame(
        rows,
        columns=[
            "Field",
            "Value",
        ],
    )


def _compact_visitor_scope_guidance() -> pd.DataFrame:
    """Explain visitor intelligence scope honestly."""

    return pd.DataFrame(
        [
            {
                "Status": (
                    "NOT APPLICABLE IN "
                    "WHOLE-PERIOD REPORT"
                ),
                "Meaning": (
                    "New, Rare and Repeat-Relevant "
                    "visitor classification requires "
                    "a selected Date-Time Part and the "
                    "same Spot baseline."
                ),
                "Action": (
                    "Run New Date-Time Partition Analysis "
                    "for scope-based visitor intelligence."
                ),
                "Caution": (
                    "The software must not claim an "
                    "absolute first visit outside the "
                    "loaded data."
                ),
            }
        ]
    )


def _compact_multi_spot_scope_guidance(
    total_spots: int,
) -> pd.DataFrame:
    """Explain which cross-Spot sections are meaningful."""

    if total_spots >= 3:
        message = (
            "N-of-M analysis is shown because three "
            "or more valid Spots are available."
        )

    elif total_spots == 2:
        message = (
            "N-of-M is not repeated separately because "
            "with two valid Spots, 2-of-2 presence is "
            "equivalent to All-Spot Common Numbers."
        )

    else:
        message = (
            "Cross-Spot comparison requires at least "
            "two valid Spot folders."
        )

    return pd.DataFrame(
        [
            {
                "Valid Spots": total_spots,
                "Interpretation": message,
                "Caution": (
                    "Telecom-record presence is not "
                    "independent proof of a person's "
                    "physical movement."
                ),
            }
        ]
    )


def _overview(
    result: dict[str, Any],
    *,
    context: dict[str, Any] | None = None,
) -> pd.DataFrame:
    """Build an accurate compact executive summary."""

    report_context = (
        context
        if isinstance(
            context,
            dict,
        )
        else _compact_report_context(
            result
        )
    )

    analysis = (
        report_context.get(
            "analysis",
            {},
        )
        or {}
    )

    metadata = (
        report_context.get(
            "metadata",
            {},
        )
        or {}
    )

    rows = [
        (
            "Report Generated At",
            utc_now_iso(),
        ),
        (
            "Report Profile",
            "Compact Investigation Report",
        ),
        (
            "Input Folder",
            metadata.get(
                "input_folder",
                "",
            ),
        ),
        (
            "Input Mode",
            report_context[
                "input_mode"
            ],
        ),
        (
            "Total Normalized Records",
            report_context[
                "total_records"
            ],
        ),
        (
            "Records Before Deduplication",
            report_context[
                "records_before"
            ],
        ),
        (
            "Records After Deduplication",
            report_context[
                "records_after"
            ],
        ),
        (
            "Duplicates Removed",
            report_context[
                "duplicates_removed"
            ],
        ),
        (
            "Valid Investigation Spots",
            report_context[
                "valid_spot_count"
            ],
        ),
        (
            "Spot Names",
            ", ".join(
                report_context[
                    "spot_names"
                ]
            ),
        ),
        (
            "Unassigned Root Files",
            report_context[
                "root_file_count"
            ],
        ),
        (
            "Unassigned Root Records",
            report_context[
                "root_record_count"
            ],
        ),
        (
            "Physical Input Files Detected",
            report_context[
                "physical_files"
            ],
        ),
        (
            "Indexed Logical Source Entries",
            report_context[
                "logical_source_entries"
            ],
        ),
        (
            "Files Loaded With Records",
            report_context[
                "physical_loaded_files"
            ],
        ),
        (
            "Files Empty / No Data",
            report_context[
                "files_empty"
            ],
        ),
        (
            "Files Failed",
            report_context[
                "files_failed"
            ],
        ),
        (
            "Operators",
            ", ".join(
                report_context[
                    "operators"
                ]
            ),
        ),
        (
            "Searched Cell IDs",
            report_context[
                "searched_cell_count"
            ],
        ),
        (
            "Date From",
            report_context[
                "date_from"
            ],
        ),
        (
            "Date To",
            report_context[
                "date_to"
            ],
        ),
        (
            "Analysis Functions",
            analysis.get(
                "function_count",
                0,
            ),
        ),
        (
            "Completed Analyses",
            analysis.get(
                "completed_count",
                0,
            ),
        ),
        (
            "Failed Analyses",
            analysis.get(
                "failed_count",
                0,
            ),
        ),
        (
            "Detailed Data Availability",
            (
                "Indexed normalized evidence is retained "
                "in DuckDB / Parquet. Detailed "
                "analyses can be regenerated from "
                "that indexed evidence."
            ),
        ),
    ]

    return pd.DataFrame(
        rows,
        columns=[
            "Field",
            "Value",
        ],
    )



def _report_name(case_name: str | None = None) -> str:
    stamp = utc_now().strftime("%Y%m%d_%H%M%S_%f")
    if case_name:
        clean = "".join(ch if ch.isalnum() or ch in (" ", "_", "-") else "_" for ch in case_name)
        clean = "_".join(clean.split())
        return f"Tower_Dump_{clean}_{stamp}.xlsx"
    return f"Tower_Dump_Analysis_{stamp}.xlsx"


def generate_tower_dump_excel_report(
    result: dict[str, Any],
    *,
    output_dir: str | Path | None = None,
    case_name: str | None = None,
    raw_row_limit: int = 200000,
    report_profile: str = "COMPACT",
    lead_row_limit: int = 200,
) -> Path:
    """Generate a compact investigator-facing Excel report.

    Large analytical tables remain preserved in the
    DuckDB/Parquet backend instead of being copied into Excel.
    """

    from time import perf_counter

    # COMPACT_NORMALIZED_SAMPLE_V1
    # Legacy parameter is retained for compatibility, but
    # compact reports never copy huge raw tables into Excel.
    normalized_row_limit = min(
        max(
            1,
            int(raw_row_limit),
        ),
        1000,
    )

    profile = str(
        report_profile
        or "COMPACT"
    ).strip().upper()

    if profile != "COMPACT":
        raise ValueError(
            "Supported Tower CDR report profile: COMPACT"
        )

    lead_limit = max(
        1,
        int(lead_row_limit),
    )

    target_dir = (
        Path(output_dir)
        .expanduser()
        .resolve()
        if output_dir
        else DEFAULT_OUTPUT_DIR
    )

    target_dir.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook_path = (
        target_dir
        / _report_name(case_name)
    )

    temporary_path = (
        workbook_path.with_name(
            f".{workbook_path.stem}.temporary.xlsx"
        )
    )

    report_started = perf_counter()

    analysis = (
        result.get(
            "analysis",
            {},
        )
        or {}
    )

    results = (
        analysis.get(
            "results",
            {},
        )
        or {}
    )

    report_context = _compact_report_context(
        result
    )

    wb = Workbook()
    wb.remove(
        wb.active
    )

    total_sheets = 11
    current_sheet = 0

    def show_progress(
        sheet_name: str,
    ) -> None:
        nonlocal current_sheet

        current_sheet += 1

        print(
            "[+] Writing compact report "
            f"sheet {current_sheet}/{total_sheets}: "
            f"{sheet_name}"
        )

    print()
    print(
        "[+] Analysis completed."
    )
    print(
        "[+] Preparing compact investigation report..."
    )

    # ----------------------------------------------------------
    # 1. Executive Summary
    # ----------------------------------------------------------

    show_progress(
        "Executive Summary"
    )

    _write_dataframe(
        wb,
        "1. Executive Summary",
        _overview(
            result,
            context=report_context,
        ),
    )


    # ----------------------------------------------------------
    # 2. Data Quality
    # ----------------------------------------------------------

    show_progress(
        "Data Quality"
    )

    ws = wb.create_sheet(
        "2. Data Quality"
    )

    row = 1

    row = _write_section(
        ws,
        "DATA QUALITY OVERVIEW",
        row,
        _compact_data_quality_overview(
            report_context
        ),
        max_rows=100,
    )

    row = _write_section(
        ws,
        "MASTER DATA ENRICHMENT",
        row,
        analysis.get(
            "master_enrichment_summary",
            results.get(
                "master_enrichment_summary",
                pd.DataFrame(),
            ),
        ),
        max_rows=100,
    )

    row = _write_section(
        ws,
        "FILE SUMMARY",
        row,
        result.get(
            "file_summary",
            pd.DataFrame(),
        ),
        max_rows=1000,
    )

    row = _write_section(
        ws,
        "SPOT INGESTION SUMMARY",
        row,
        result.get(
            "spot_summary",
            pd.DataFrame(),
        ),
        max_rows=500,
    )

    row = _write_section(
        ws,
        "LOADER WARNINGS",
        row,
        result.get(
            "warnings",
            [],
        ),
        max_rows=200,
    )

    row = _write_section(
        ws,
        "LOADER ERRORS",
        row,
        result.get(
            "errors",
            [],
        ),
        max_rows=200,
    )

    _style_section_sheet(
        ws
    )

    # ----------------------------------------------------------
    # 3. Tower Summary
    # ----------------------------------------------------------

    show_progress(
        "Tower Summary"
    )

    ws = wb.create_sheet(
        "3. Tower Summary"
    )

    row = 1

    row = _write_section(
        ws,
        "OPERATOR SUMMARY",
        row,
        results.get(
            "operator_summary"
        ),
    )

    row = _write_section(
        ws,
        "SEARCHED CELL / CGI SUMMARY",
        row,
        results.get(
            "cell_summary"
        ),
        max_rows=2000,
    )

    row = _write_section(
        ws,
        "CALL TYPE SUMMARY",
        row,
        results.get(
            "call_type_summary"
        ),
    )

    row = _write_section(
        ws,
        "SPOT SUMMARY",
        row,
        results.get(
            "spot_summary"
        ),
        max_rows=500,
    )

    _style_section_sheet(
        ws
    )

    # ----------------------------------------------------------
# ----------------------------------------------------------
    # 4. Priority Review Queue
    # ----------------------------------------------------------

    show_progress(
        "Priority Review Queue"
    )

    ws = wb.create_sheet(
        "4. Priority Review Queue"
    )

    row = 1

    (
        priority_queue,
        category_coverage,
    ) = _compact_priority_review_queue(
        results,
        limit=lead_limit,
    )

    priority_columns = [
        "subscriber_number",
        "priority",
        "confidence",
        "priority_score",
        "lead_categories",
        "event_count",
        "cells_seen",
        "night_event_count",
        "imei_count",
        "imsi_count",
        "first_seen",
        "last_seen",
        "why_important",
        "next_action",
    ]

    uncommon_columns = [
        "subscriber_number",
        "priority",
        "confidence",
        "event_count",
        "cells_seen",
        "night_event_count",
        "first_seen",
        "last_seen",
        "why_important",
        "next_action",
    ]

    row = _write_section(
        ws,
        "MASTER PRIORITY REVIEW QUEUE",
        row,
        _compact_report_view(
            priority_queue,
            priority_columns,
        ),
        max_rows=lead_limit,
    )

    row = _write_section(
        ws,
        "RARE / UNCOMMON SHORTLIST",
        row,
        _compact_report_view(
            results.get(
                "tower_cdr_uncommon_numbers"
            ),
            uncommon_columns,
        ),
        max_rows=min(
            50,
            lead_limit,
        ),
    )

    row = _write_section(
        ws,
        "LEAD CATEGORY COVERAGE",
        row,
        category_coverage,
        max_rows=100,
    )

    _style_section_sheet(
        ws
    )

# ----------------------------------------------------------
    # 5. Visitor Intelligence
    # ----------------------------------------------------------

    show_progress(
        "Visitor Intelligence"
    )

    ws = wb.create_sheet(
        "5. Visitor Intelligence"
    )

    row = 1

    row = _write_section(
        ws,
        "SCOPE GUIDANCE",
        row,
        _compact_visitor_scope_guidance(),
        max_rows=10,
    )

    visitor_columns = [
        "subscriber_number",
        "total_events",
        "first_seen",
        "last_seen",
        "active_days",
        "unique_cells",
        "unique_operators",
        "unique_imei",
        "unique_imsi",
        "unique_other_parties",
        "total_duration_seconds",
    ]

    row = _write_section(
        ws,
        "FREQUENT VISITORS",
        row,
        _compact_report_view(
            results.get(
                "frequent_visitors"
            ),
            visitor_columns,
        ),
        max_rows=lead_limit,
    )

    row = _write_section(
        ws,
        "REPEAT VISITORS",
        row,
        _compact_report_view(
            results.get(
                "repeat_visitors"
            ),
            visitor_columns,
        ),
        max_rows=lead_limit,
    )

    _style_section_sheet(
        ws
    )

# ----------------------------------------------------------
    # 6. Multi-Spot Intelligence
    # ----------------------------------------------------------

    show_progress(
        "Multi-Spot Intelligence"
    )

    ws = wb.create_sheet(
        "6. Multi-Spot Intel"
    )

    row = 1

    total_spots = report_context[
        "valid_spot_count"
    ]

    row = _write_section(
        ws,
        "SCOPE GUIDANCE",
        row,
        _compact_multi_spot_scope_guidance(
            total_spots
        ),
        max_rows=10,
    )

    presence_columns = [
        "subscriber_number",
        "spots_seen_count",
        "total_spots",
        "match_ratio",
        "spot_ids",
        "spot_names",
        "total_events",
        "unique_searched_cells",
        "operators",
        "imei_count",
        "imsi_count",
        "first_seen",
        "last_seen",
    ]

    multi_spot_sections: list[
        tuple[
            str,
            str,
            list[str],
        ]
    ] = []

    if total_spots >= 3:
        multi_spot_sections.append(
            (
                "N-OF-M SPOT PRESENCE",
                "n_of_m_spot_presence",
                presence_columns,
            )
        )

    multi_spot_sections.extend(
        [
            (
                "ALL-SPOT COMMON NUMBERS",
                "all_spot_common_numbers",
                presence_columns,
            ),
            (
                "SPOT-EXCLUSIVE NUMBERS",
                "spot_exclusive_numbers",
                [
                    *presence_columns,
                    "exclusive_spot_id",
                    "exclusive_spot_name",
                ],
            ),
            (
                "DEVICE CONTINUITY ACROSS SPOTS",
                "cross_spot_device_continuity",
                [
                    "subscriber_number",
                    "spots_seen_count",
                    "spot_names",
                    "imei_count",
                    "imei_values",
                    "imei_continuity",
                    "imsi_count",
                    "imsi_values",
                    "imsi_continuity",
                    "confidence",
                    "why_important",
                    "next_verification",
                ],
            ),
            (
                "SHARED IMEI ACROSS SPOTS",
                "shared_imei_across_spots",
                [
                    "imei",
                    "spots_seen_count",
                    "spot_names",
                    "unique_subscribers",
                    "subscriber_numbers",
                    "total_events",
                    "first_seen",
                    "last_seen",
                    "why_important",
                    "next_verification",
                ],
            ),
            (
                "SHARED IMSI ACROSS SPOTS",
                "shared_imsi_across_spots",
                [
                    "imsi",
                    "spots_seen_count",
                    "spot_names",
                    "unique_subscribers",
                    "subscriber_numbers",
                    "total_events",
                    "first_seen",
                    "last_seen",
                    "why_important",
                    "next_verification",
                ],
            ),
        ]
    )

    for (
        title,
        result_name,
        preferred_columns,
    ) in multi_spot_sections:
        row = _write_section(
            ws,
            title,
            row,
            _compact_report_view(
                results.get(
                    result_name
                ),
                preferred_columns,
            ),
            max_rows=lead_limit,
        )

    _style_section_sheet(
        ws
    )

    # 7. Device / SIM Alerts
    # ----------------------------------------------------------

    show_progress(
        "Device and SIM Alerts"
    )

    ws = wb.create_sheet(
        "7. Device SIM Alerts"
    )

    row = 1

    device_columns = [
        "subscriber_number",
        "priority",
        "confidence",
        "priority_score",
        "event_count",
        "cells_seen",
        "night_event_count",
        "imei_count",
        "imsi_count",
        "other_party_count",
        "first_seen",
        "last_seen",
        "why_important",
        "next_action",
    ]

    shared_identifier_columns = [
        "imei",
        "imsi",
        "total_events",
        "unique_subscribers",
        "unique_cells",
        "unique_operators",
        "first_seen",
        "last_seen",
        "operators",
        "subscribers",
    ]

    row = _write_section(
        ws,
        "DEVICE CONSISTENCY ALERTS",
        row,
        _compact_report_view(
            results.get(
                "tower_cdr_device_consistency"
            ),
            device_columns,
        ),
        max_rows=lead_limit,
    )

    row = _write_section(
        ws,
        "SHARED IMEI",
        row,
        _compact_report_view(
            results.get(
                "shared_imei"
            ),
            shared_identifier_columns,
        ),
        max_rows=lead_limit,
    )

    row = _write_section(
        ws,
        "SHARED IMSI",
        row,
        _compact_report_view(
            results.get(
                "shared_imsi"
            ),
            shared_identifier_columns,
        ),
        max_rows=lead_limit,
    )

    _style_section_sheet(
        ws
    )

    # ----------------------------------------------------------
    # 8. Normalized Sample
    # ----------------------------------------------------------

    show_progress(
        "Normalized Sample"
    )

    normalized_sample = (
        _compact_balanced_normalized_sample(
            report_context[
                "dataframe"
            ],
            max_rows=normalized_row_limit,
        )
    )

    ws = wb.create_sheet(
        "8. Normalized Sample"
    )

    row = 1

    row = _write_section(
        ws,
        "SAMPLE INFORMATION",
        row,
        _compact_normalized_sample_info(
            report_context,
            normalized_sample,
        ),
        max_rows=20,
    )

    row = _write_section(
        ws,
        "NORMALIZED RECORD SAMPLE",
        row,
        normalized_sample,
        max_rows=normalized_row_limit,
    )

    _style_section_sheet(
        ws
    )

    # 9. Backend Data Guide
    # ----------------------------------------------------------

    show_progress(
        "Backend Data Guide"
    )

    backend_tables = (
        "subscriber_summary",
        "imei_summary",
        "imsi_summary",
        "common_subscriber_matrix",
        "subscriber_spot_detail",
        "subscriber_spot_presence",
        "cross_spot_sequence",
        "subscriber_movements",
        "investigative_indicators",
    )

    backend_rows = []

    for table_name in backend_tables:
        value = results.get(
            table_name
        )

        row_count = (
            len(value)
            if hasattr(
                value,
                "__len__",
            )
            else 0
        )

        backend_rows.append(
            {
                "backend_table": table_name,
                "rows_available": int(
                    row_count
                ),
                "included_in_compact_excel": "No",
                "storage": (
                    "Indexed evidence: DuckDB / Parquet"
                ),
                "access": (
                    "Regenerate analysis or create a "
                    "selected detailed export"
                ),
                "reason": (
                    "Derived table excluded from compact "
                    "Excel. Row count reflects the "
                    "current analysis run."
                ),
            }
        )

    _write_dataframe(
        wb,
        "9. Backend Data Guide",
        pd.DataFrame(
            backend_rows
        ),
    )

    # ----------------------------------------------------------
    # 10. Analysis Status
    # ----------------------------------------------------------

    show_progress(
        "Analysis Status"
    )

    ws = wb.create_sheet(
        "10. Analysis Status"
    )

    row = 1

    row = _write_section(
        ws,
        "ANALYSIS STATUS",
        row,
        analysis.get(
            "status",
            pd.DataFrame(),
        ),
        max_rows=1000,
    )

    row = _write_section(
        ws,
        "ANALYSIS ERRORS",
        row,
        analysis.get(
            "errors",
            pd.DataFrame(),
        ),
        max_rows=500,
    )

    _style_section_sheet(
        ws
    )


    # ----------------------------------------------------------
    # 11. Methodology and Limits
    # ----------------------------------------------------------

    show_progress(
        "Methodology & Limits"
    )

    append_methodology_sheet(
        wb,
        "Tower CDR Dump Analysis",
    )

    print(
        "[+] Saving compact Excel report..."
    )

    try:
        wb.save(
            temporary_path
        )

        temporary_path.replace(
            workbook_path
        )

    finally:
        if temporary_path.exists():
            temporary_path.unlink(
                missing_ok=True
            )

    report_seconds = (
        perf_counter()
        - report_started
    )

    print(
        "[+] Compact investigation report generated."
    )
    print(
        "[+] Report generation time: "
        f"{report_seconds:.2f} seconds"
    )
    print(
        f"[+] Report path: {workbook_path}"
    )

    return workbook_path
