"""Standard methodology and limitation sheet for analytical workbooks."""

from __future__ import annotations

from openpyxl.styles import Alignment, Font, PatternFill

from modules.core.time_utils import utc_now_iso

from .excel_security import excel_safe_value
from .excel_styles import THIN_BORDER


_GUIDANCE_ROWS = (
    ("Purpose", "This workbook supports investigative lead generation from supplied telecom records; it is not a finding of guilt or identity."),
    ("Corroboration", "Every indicator, ranking, association, movement inference and location inference requires corroboration with independent evidence."),
    ("Association limitation", "Communication frequency or shared identifiers show data association only; they do not by themselves establish relationship, intent, conspiracy or participation."),
    ("Location limitation", "Cell-site records indicate network handling or sector association and do not establish a handset's exact physical position."),
    ("Identity limitation", "MSISDN, IMSI, IMEI, IP address and subscriber details may be shared, reassigned, translated, proxied, spoofed or affected by operator processing."),
    ("Time limitation", "Timestamps depend on operator format, timezone, clock quality and source normalization. Material events should be checked against the original record."),
    ("Data quality", "Missing, malformed, rejected, duplicated or incomplete records can affect totals and rankings. Review diagnostics and rejected-row sheets."),
    ("Scoring limitation", "Scores, thresholds and ranked results are prioritization aids, not probabilities, validated risk levels or legal conclusions."),
    ("Evidence handling", "Retain the original source files, acquisition details and cryptographic hashes. This workbook is a derived analytical artifact."),
    ("Excel safety", "Text beginning with formula-like characters is stored as literal text to prevent spreadsheet formula execution."),
)


def append_methodology_sheet(workbook, report_type: str) -> str:
    """Append one consistently styled methodology sheet and return its name."""
    base_name = "Methodology & Limits"
    name = base_name
    suffix = 2
    existing = {sheet.title.lower() for sheet in workbook.worksheets}
    while name.lower() in existing:
        tail = f" {suffix}"
        name = base_name[: 31 - len(tail)] + tail
        suffix += 1

    worksheet = workbook.create_sheet(name)
    worksheet.merge_cells("A1:B1")
    title = worksheet["A1"]
    title.value = "REPORT METHODOLOGY, INTERPRETATION AND LIMITATIONS"
    title.font = Font(color="FFFFFF", bold=True, size=14)
    title.fill = PatternFill("solid", fgColor="1F4E78")
    title.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.append(["Report Type", excel_safe_value(report_type)])
    worksheet.append(["Generated At", utc_now_iso()])
    worksheet.append([])
    worksheet.append(["Section", "Guidance"])

    for section, guidance in _GUIDANCE_ROWS:
        worksheet.append([excel_safe_value(section), excel_safe_value(guidance)])

    for cell in worksheet[5]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    worksheet.column_dimensions["A"].width = 24
    worksheet.column_dimensions["B"].width = 105
    worksheet.freeze_panes = "A5"
    worksheet.sheet_view.showGridLines = False
    return name
