"""Consolidated Excel report for Tower IPDR/NAT analysis.

This renderer consumes an already-computed analysis bundle. It does not rerun
loading or analysis logic.
"""

from __future__ import annotations

from modules.core.time_utils import utc_now, utc_now_iso

import math
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd

TOWER_IPDR_EXCEL_PREVIEW_ROWS = 5000


def _excel_preview(dataframe, max_rows: int = TOWER_IPDR_EXCEL_PREVIEW_ROWS):
    """Return Excel-safe preview data.

    Full forensic evidence remains preserved in backend CSV files.
    Excel should contain summaries and limited previews only.
    """

    if isinstance(dataframe, pd.DataFrame) and len(dataframe) > max_rows:
        return dataframe.head(max_rows).copy()

    return dataframe


DETAIL_SHEET_KEYWORDS = (
    "allocation records",
    "subscribers",
    "subscriber cell matrix",
    "multi cell candidates",
    "all cell candidates",
    "imei summary",
    "imei cell matrix",
    "imsi summary",
    "imsi cell matrix",
    "source ip",
    "nat ip",
    "destination ip",
    "destination ports",
    "destination endpoints",
    "uncommon numbers",
    "normalized events",
    "rejected rows",
    "actual event hits",
    "actual location exclusions",
    "event presence",
    "event n-of-m",
    "event strict common",
    "allocation hits",
    "allocation location exclusions",
    "allocation presence",
    "allocation n-of-m",
    "allocation strict",
    "imei event presence",
    "imsi event presence",
)


def _should_preview_sheet(sheet_name: str) -> bool:
    """Return True for large/detail sheets that should be Excel-previewed."""

    name = str(sheet_name or "").lower()

    return any(
        keyword in name
        for keyword in DETAIL_SHEET_KEYWORDS
    )


def _excel_frame_for_sheet(sheet_name: str, dataframe):
    frame = _frame(dataframe)

    if _should_preview_sheet(sheet_name):
        return _excel_preview(frame)

    return frame


def _excel_subtitle_for_sheet(sheet_name: str, dataframe, subtitle: str) -> str:
    frame = _frame(dataframe)

    if (
        _should_preview_sheet(sheet_name)
        and isinstance(frame, pd.DataFrame)
        and len(frame) > TOWER_IPDR_EXCEL_PREVIEW_ROWS
    ):
        return (
            f"{subtitle} Preview only: first "
            f"{TOWER_IPDR_EXCEL_PREVIEW_ROWS:,} rows are shown in Excel; "
            "complete table remains saved in backend CSV."
        )

    return subtitle


from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .excel_security import excel_safe_value
from .report_guidance import append_methodology_sheet


MAX_DATA_ROWS = 1_000_000
TITLE_FILL = PatternFill("solid", fgColor="17365D")
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
SUCCESS_FILL = PatternFill("solid", fgColor="E2F0D9")
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
ERROR_FILL = PatternFill("solid", fgColor="FCE4D6")
HEADER_FONT = Font(bold=True)
THIN = Side(style="thin", color="B7C9DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _safe_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_-]+", "_", str(value).strip())
    return cleaned.strip("_") or "CASE"


def _safe_value(value: Any) -> Any:
    """Compatibility wrapper around the shared Excel security boundary."""
    return excel_safe_value(value)


def _frame(value: Any) -> pd.DataFrame:
    if isinstance(value, pd.DataFrame):
        return value.copy()
    if isinstance(value, list):
        return pd.DataFrame(value)
    if isinstance(value, dict):
        return pd.DataFrame([value])
    return pd.DataFrame()


def _sheet_name(workbook: Workbook, requested: str) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", " ", requested)
    cleaned = re.sub(r"\s+", " ", cleaned).strip() or "Sheet"
    cleaned = cleaned[:31]
    existing = {sheet.title.lower() for sheet in workbook.worksheets}

    if cleaned.lower() not in existing:
        return cleaned

    number = 2
    while True:
        suffix = f" {number}"
        candidate = cleaned[: 31 - len(suffix)] + suffix
        if candidate.lower() not in existing:
            return candidate
        number += 1


def _title(worksheet, text: str, subtitle: str, columns: int) -> int:
    last_column = max(2, columns)
    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=last_column,
    )
    cell = worksheet.cell(1, 1, text)
    cell.fill = TITLE_FILL
    cell.font = Font(color="FFFFFF", bold=True, size=16)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 28

    worksheet.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=last_column,
    )
    sub = worksheet.cell(2, 1, subtitle)
    sub.font = Font(italic=True, color="44546A")
    sub.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[2].height = 28
    return 4


def _widths(worksheet) -> None:
    for cells in worksheet.columns:
        letter = get_column_letter(cells[0].column)
        maximum = 0
        for cell in cells[:5000]:
            if cell.value is not None:
                maximum = max(maximum, len(str(cell.value)))
        worksheet.column_dimensions[letter].width = min(max(maximum + 2, 10), 44)


def _write_page(worksheet, dataframe: pd.DataFrame, *, title: str, subtitle: str) -> None:
    header_row = _title(worksheet, title, subtitle, max(2, len(dataframe.columns)))

    if dataframe.empty:
        cell = worksheet.cell(header_row, 1, "No records found.")
        cell.fill = WARNING_FILL
        cell.font = HEADER_FONT
        worksheet.column_dimensions["A"].width = 34
        worksheet.freeze_panes = f"A{header_row + 1}"
        worksheet.sheet_view.showGridLines = False
        return

    for column_index, column in enumerate(dataframe.columns, start=1):
        cell = worksheet.cell(header_row, column_index, _safe_value(str(column)))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, record in enumerate(
        dataframe.itertuples(index=False, name=None),
        start=header_row + 1,
    ):
        for column_index, value in enumerate(record, start=1):
            cell = worksheet.cell(row_index, column_index, _safe_value(value))
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            if isinstance(cell.value, (datetime, date)):
                cell.number_format = "yyyy-mm-dd hh:mm:ss"

    last_row = header_row + len(dataframe)
    last_column = get_column_letter(len(dataframe.columns))
    worksheet.auto_filter.ref = f"A{header_row}:{last_column}{last_row}"
    worksheet.freeze_panes = f"A{header_row + 1}"
    worksheet.sheet_view.showGridLines = False
    _widths(worksheet)


def _write_table(
    workbook: Workbook,
    base_name: str,
    dataframe: pd.DataFrame,
    *,
    subtitle: str,
) -> list[str]:
    dataframe = _frame(dataframe)

    if dataframe.empty:
        name = _sheet_name(workbook, base_name)
        worksheet = workbook.create_sheet(name)
        _write_page(worksheet, dataframe, title=base_name, subtitle=subtitle)
        return [name]

    pages = max(1, math.ceil(len(dataframe) / MAX_DATA_ROWS))
    names: list[str] = []

    for page_index in range(pages):
        start = page_index * MAX_DATA_ROWS
        end = min(start + MAX_DATA_ROWS, len(dataframe))
        page = dataframe.iloc[start:end].reset_index(drop=True)
        requested = base_name if pages == 1 else f"{base_name} {page_index + 1}"
        name = _sheet_name(workbook, requested)
        worksheet = workbook.create_sheet(name)
        page_subtitle = subtitle

        if pages > 1:
            page_subtitle = (
                f"{subtitle} | Part {page_index + 1}/{pages} | "
                f"Rows {start + 1:,}-{end:,}"
            )

        _write_page(worksheet, page, title=base_name, subtitle=page_subtitle)
        names.append(name)

    return names


def _key_values(worksheet, start_row: int, rows: list[tuple[str, Any]]) -> None:
    for offset, (key, value) in enumerate(rows):
        row = start_row + offset
        key_cell = worksheet.cell(row, 1, key)
        value_cell = worksheet.cell(row, 2, _safe_value(value))
        key_cell.fill = HEADER_FILL
        key_cell.font = HEADER_FONT
        key_cell.border = BORDER
        value_cell.border = BORDER
        value_cell.alignment = Alignment(wrap_text=True, vertical="top")

    worksheet.column_dimensions["A"].width = 38
    worksheet.column_dimensions["B"].width = 68


def _warnings(load_result: dict[str, Any]) -> pd.DataFrame:
    rows: list[dict[str, str]] = []

    for message in load_result.get("warnings", []) or []:
        rows.append({"Level": "WARNING", "Message": str(message)})

    for message in load_result.get("errors", []) or []:
        rows.append({"Level": "ERROR", "Message": str(message)})

    rows.append(
        {
            "Level": "INFO",
            "Message": (
                "Data volumes are repeated across IPDR event rows. Volume sheets use "
                "deduplicated observed allocation-volume records and are not a billing total."
            ),
        }
    )
    rows.append(
        {
            "Level": "INFO",
            "Message": (
                "A First Cell match anchors the allocation to the searched cell. Last Cell "
                "changes require cautious movement/location interpretation."
            ),
        }
    )
    return pd.DataFrame(rows)


def generate_tower_ipdr_excel_report(
    *,
    case: dict[str, Any],
    load_result: dict[str, Any],
    analysis: dict[str, Any],
    partition: dict[str, Any] | None,
    output_dir: str | Path,
    saved: dict[str, Any] | None = None,
) -> Path:
    """Generate one consolidated Tower IPDR/NAT workbook."""

    case_id = str(case.get("case_id", "")).strip() or "CASE"
    case_name = str(case.get("case_name", "")).strip()
    output = Path(output_dir).expanduser().resolve()
    output.mkdir(parents=True, exist_ok=True)
    timestamp = utc_now().strftime("%Y%m%d_%H%M%S_%f")
    report_path = output / (
        f"Tower_IPDR_Dump_Analysis_{_safe_filename(case_id)}_{timestamp}.xlsx"
    )

    workbook = Workbook()
    workbook.remove(workbook.active)

    overview = workbook.create_sheet(_sheet_name(workbook, "1. Executive Summary"))
    start = _title(
        overview,
        "Tower IPDR/NAT Dump Analysis",
        f"Case: {case_id}" + (f" | {case_name}" if case_name else ""),
        6,
    )

    metadata = load_result.get("metadata", {}) or {}
    multi_cell = _frame(analysis.get("subscriber_multi_cell_candidates"))
    all_cell = _frame(analysis.get("subscriber_all_cell_candidates"))
    event_candidates = _frame(
        partition.get("event_n_of_m_candidates") if isinstance(partition, dict) else None
    )
    allocation_candidates = _frame(
        partition.get("allocation_n_of_m_candidates") if isinstance(partition, dict) else None
    )

    _key_values(
        overview,
        start,
        [
            ("Case ID", case_id),
            ("Case Name", case_name),
            ("Source Section", "Tower IPDR Dump Analysis"),
            ("Currently Supported Parser", "Jio CELL ID_IPDRNAT"),
            ("Generated At", utc_now_iso()),
            ("Input Folder", metadata.get("input_folder", "")),
            ("Files Found", metadata.get("files_found", 0)),
            ("Files Loaded", metadata.get("files_loaded", 0)),
            ("Files Failed", metadata.get("files_failed", 0)),
            ("Normalized IPDR Events", metadata.get("records", 0)),
            ("Searched Cell IDs", metadata.get("unique_cells", 0)),
            ("Unique Subscribers", metadata.get("unique_subscribers", 0)),
            ("Multi-cell Subscribers", len(multi_cell)),
            ("All-cell Subscribers", len(all_cell)),
            (
                "CCTV Partitions",
                partition.get("total_partitions", 0) if isinstance(partition, dict) else 0,
            ),
            ("Actual-event N-of-M Candidates", len(event_candidates)),
            ("Allocation-overlap N-of-M Candidates", len(allocation_candidates)),
            (
                "Actual Event Rule",
                partition.get("actual_event_rule", "") if isinstance(partition, dict) else "window_start <= event_time <= window_end",
            ),
            (
                "Allocation Overlap Rule",
                partition.get("allocation_overlap_rule", "") if isinstance(partition, dict) else "allocation_start <= window_end AND allocation_end >= window_start",
            ),
            (
                "Backend Run Directory",
                saved.get("run_directory", "") if isinstance(saved, dict) else "",
            ),
            (
                "Forensic Volume Rule",
                "Repeated row-level volume is not summed directly; observed allocation-volume records are deduplicated.",
            ),
        ],
    )
    overview.freeze_panes = "A4"
    overview.sheet_view.showGridLines = False

    normalized_events_preview = _excel_preview(
        analysis.get("normalized_events"),
        max_rows=TOWER_IPDR_EXCEL_PREVIEW_ROWS,
    )

    tables = [
        ("2. Source Files", load_result.get("file_summary"), "Loaded Jio CELL ID_IPDRNAT files and diagnostics."),
        ("3. Core Summary", analysis.get("summary"), "Core multi-cell Tower IPDR metrics."),
        ("4. Searched Cells", analysis.get("cell_summary"), "Cell-wise event and candidate distribution."),
        ("5. Allocation Records", analysis.get("allocation_records"), "Deduplicated observed allocation-volume records; not a billing total."),
        ("6. Subscribers", analysis.get("subscriber_summary"), "Subscriber-wise IPDR intelligence."),
        ("7. Subscriber Cell Matrix", analysis.get("subscriber_cell_presence"), "Dynamic N-of-M searched-cell presence matrix."),
        ("8. Multi Cell Candidates", analysis.get("subscriber_multi_cell_candidates"), "Subscribers present in at least two searched cells."),
        ("9. All Cell Candidates", analysis.get("subscriber_all_cell_candidates"), "Subscribers present in every loaded searched cell."),
        ("10. IMEI Summary", analysis.get("imei_summary"), "Device-wise event and cell summary."),
        ("11. IMEI Cell Matrix", analysis.get("imei_cell_presence"), "IMEI presence across searched cells."),
        ("12. IMSI Summary", analysis.get("imsi_summary"), "IMSI-wise event and cell summary."),
        ("13. IMSI Cell Matrix", analysis.get("imsi_cell_presence"), "IMSI presence across searched cells."),
        ("14. Source IP", analysis.get("source_ip_summary"), "Source IPv4/IPv6 intelligence."),
        ("15. NAT IP", analysis.get("translated_ip_summary"), "Translated/NAT IPv4/IPv6 intelligence."),
        ("16. Destination IP", analysis.get("destination_ip_summary"), "Destination IP frequency and subscriber spread."),
        ("17. Destination Ports", analysis.get("destination_port_summary"), "Destination port frequency and subscriber spread."),
        ("18. Destination Endpoints", analysis.get("destination_endpoint_summary"), "Destination IP:port endpoint analysis."),
        ("19. APN", analysis.get("apn_summary"), "Access Point Name distribution."),
        ("20. Roaming", analysis.get("roaming_summary"), "Home/roaming distribution."),
        ("21. Cell Movement", analysis.get("cell_movement_summary"), "First-to-last cell transitions; interpret cautiously."),
        ("22. Hourly Activity", analysis.get("hourly_activity"), "Event activity by date and hour."),
        ("23. Data Quality", analysis.get("data_quality"), "Validation flags without altering raw evidence."),
        ("24. Uncommon Priority", analysis.get("uncommon_priority_summary"), "Priority counts for uncommon subscriber leads."),
        ("25. Uncommon Numbers", analysis.get("uncommon_numbers"), "Window-only or rare subscriber presence ranked for investigation."),
        ("26. Normalized Events Preview", normalized_events_preview, "Preview only: first 5,000 normalized events. Complete normalized evidence remains saved in backend CSV."),
        ("27. Rejected Rows", analysis.get("rejected_rows"), "Malformed/non-data rows quarantined with physical source-line provenance."),
    ]

    partition_tables = [
        ("28. Partition Windows", "partition_windows", "User-entered date/time, resolved CGI scope and automatic ±10-minute windows."),
        ("29. Partition Summary", "partition_summary", "Time-and-location scoped actual-event and allocation-overlap counts."),
        ("30. Partition Status", "partition_status", "Valid, time-only and rejected sighting configurations."),
        ("31. Actual Event Hits", "actual_event_hits", "Events matching both the date-time partition and resolved searched cell."),
        ("32. Actual Location Exclusions", "actual_time_only_excluded_by_location", "Time-matching events excluded because searched cell did not match."),
        ("33. Event Presence", "event_subscriber_presence", "Subscriber presence across valid actual-event partitions."),
        ("34. Event N-of-M", "event_n_of_m_candidates", "Actual-event candidates present in the configured minimum valid partitions."),
        ("35. Event Strict Common", "event_strict_common_candidates", "Actual-event candidates present in all valid partitions."),
        ("36. Allocation Hits", "allocation_overlap_hits", "Allocation records matching both overlap time and searched cell."),
        ("37. Allocation Location Exclusions", "allocation_time_only_excluded_by_location", "Allocation overlaps excluded because searched cell did not match."),
        ("38. Allocation Presence", "allocation_subscriber_presence", "Subscriber presence across valid allocation-overlap partitions."),
        ("39. Allocation N-of-M", "allocation_n_of_m_candidates", "Allocation candidates present in the configured minimum valid partitions."),
        ("40. Allocation Strict", "allocation_strict_common_candidates", "Allocation candidates present in all valid partitions."),
        ("41. IMEI Event Presence", "imei_event_presence", "IMEI continuity across valid actual-event partitions."),
        ("42. IMSI Event Presence", "imsi_event_presence", "IMSI continuity across valid actual-event partitions."),
    ]

    for name, dataframe, subtitle in tables:
        _write_table(
            workbook,
            name,
            _excel_frame_for_sheet(name, dataframe),
            subtitle=_excel_subtitle_for_sheet(name, dataframe, subtitle),
        )

    for name, key, subtitle in partition_tables:
        dataframe = (
            partition.get(key)
            if isinstance(partition, dict)
            else pd.DataFrame()
        )
        _write_table(
            workbook,
            name,
            _excel_frame_for_sheet(name, dataframe),
            subtitle=_excel_subtitle_for_sheet(name, dataframe, subtitle),
        )

    status = pd.DataFrame(
        [
            {
                "Stage": "Tower IPDR Loading",
                "Status": "COMPLETED",
                "Details": f"{metadata.get('records', 0):,} normalized events from {metadata.get('files_loaded', 0)} file(s)",
            },
            {
                "Stage": "Multi-cell Analysis",
                "Status": "COMPLETED",
                "Details": f"{analysis.get('total_cells', 0)} searched cell(s)",
            },
            {
                "Stage": "Date-Time Partitioning",
                "Status": "COMPLETED" if isinstance(partition, dict) else "NOT REQUESTED",
                "Details": f"{partition.get('total_partitions', 0)} partition(s)" if isinstance(partition, dict) else "",
            },
            {
                "Stage": "Consolidated Excel",
                "Status": "COMPLETED",
                "Details": str(report_path),
            },
        ]
    )
    _write_table(workbook, "43. Analysis Status", status, subtitle="Execution status for this report.")
    warning_names = _write_table(
        workbook,
        "44. Warnings",
        _warnings(load_result),
        subtitle="Loader warnings and mandatory interpretation notes.",
    )

    for sheet_name in warning_names:
        worksheet = workbook[sheet_name]
        for row in range(5, worksheet.max_row + 1):
            level = str(worksheet.cell(row, 1).value or "").upper()
            worksheet.cell(row, 1).fill = (
                ERROR_FILL if level == "ERROR" else WARNING_FILL if level == "WARNING" else SUCCESS_FILL
            )

    workbook.properties.title = f"Tower IPDR Dump Analysis - {case_id}"
    workbook.properties.subject = "Multi-cell Jio Tower IPDR/NAT forensic analysis"
    workbook.properties.creator = "Telecom Forensics Analysis Suite"
    append_methodology_sheet(workbook, "Tower IPDR / NAT Analysis")
    workbook.save(report_path)
    return report_path
# TOWER_IPDR_COMPLETE_REPORT_V1

COMPLETE_REPORT_SHEET_ORDER = (
    (
        "executive_summary",
        "1. Executive Summary",
        (
            "Answer-first overview of the complete "
            "Tower IPDR investigation."
        ),
    ),
    (
        "data_quality",
        "2. Data Quality",
        (
            "Evidence-quality checks, missing fields, "
            "duplicates and staging coverage."
        ),
    ),
    (
        "spot_cell_summary",
        "3. Spot & Cell Summary",
        (
            "Spot-wise and searched-cell-wise event, "
            "subscriber and source-file coverage."
        ),
    ),
    (
        "priority_review_queue",
        "4. Priority Review Queue",
        (
            "Deduplicated and category-balanced leads "
            "for investigator review."
        ),
    ),
    (
        "rare_presence",
        "5. Rare Presence",
        (
            "Subscribers with limited or unusual "
            "presence requiring contextual verification."
        ),
    ),
    (
        "multi_spot_intelligence",
        "6. Multi-Spot Intelligence",
        (
            "Cross-Spot presence, Spot-exclusive presence "
            "and repeated cells across Spots."
        ),
    ),
    (
        "subscriber_activity",
        "7. Subscriber Activity",
        (
            "Compact subscriber-level event, Spot, cell, "
            "IMEI and IMSI activity summary."
        ),
    ),
    (
        "device_sim_alerts",
        "8. Device & SIM Alerts",
        (
            "Shared or changing IMEI/IMSI indicators. "
            "Verify against CDR, SDR and CAF records."
        ),
    ),
    (
        "hourly_activity",
        "9. Hourly Activity",
        (
            "Actual event coverage grouped by date and hour."
        ),
    ),
    (
        "source_file_summary",
        "10. Source File Summary",
        (
            "Relative evidence provenance only. Absolute "
            "workstation paths are intentionally excluded."
        ),
    ),
    (
        "analysis_status",
        "11. Analysis Status",
        (
            "Execution stages, cache status, record counts "
            "and report-generation timings."
        ),
    ),
    (
        "methodology_limits",
        "12. Methodology & Limits",
        (
            "Interpretation rules, analytical limitations "
            "and recommended verification steps."
        ),
    ),
)


_COMPLETE_REPORT_ROW_LIMITS = {
    "priority_review_queue": 500,
    "rare_presence": 500,
    "multi_spot_intelligence": 1000,
    "subscriber_activity": 1000,
    "device_sim_alerts": 500,
    "source_file_summary": 2000,
}


def _public_source_file_summary(
    dataframe: pd.DataFrame,
) -> pd.DataFrame:
    """Return report-safe relative evidence provenance.

    Absolute local paths are useful internally, but must not
    appear in an investigator-facing workbook.
    """

    frame = _frame(
        dataframe
    )

    if frame.empty:
        return frame

    if (
        "source_relative_path"
        not in frame.columns
    ):
        if "source_file" in frame.columns:
            frame.insert(
                0,
                "source_relative_path",
                frame[
                    "source_file"
                ].map(
                    lambda value: (
                        Path(
                            str(value)
                        ).name
                        if str(
                            value
                            or ""
                        ).strip()
                        else ""
                    )
                ),
            )
        elif "file_name" in frame.columns:
            frame.insert(
                0,
                "source_relative_path",
                frame[
                    "file_name"
                ].fillna(
                    ""
                ).astype(
                    str
                ),
            )

    private_columns = [
        column
        for column in (
            "source_path",
            "source_file",
            "database_path",
            "report_folder",
            "input_folder",
        )
        if column in frame.columns
    ]

    if private_columns:
        frame = frame.drop(
            columns=private_columns
        )

    preferred_columns = [
        "spot_id",
        "spot_name",
        "spot_folder",
        "source_relative_path",
        "file_name",
        "sha256",
        "status",
        "rows_loaded",
        "searched_cell_id",
        "event_time_min",
        "event_time_max",
        "unique_subscribers",
        "warnings",
        "errors",
        "loaded_at",
    ]

    ordered = [
        column
        for column in preferred_columns
        if column in frame.columns
    ]

    remaining = [
        column
        for column in frame.columns
        if column not in ordered
    ]

    return frame[
        ordered
        + remaining
    ].copy()


def _complete_report_frame(
    key: str,
    value: Any,
) -> pd.DataFrame:
    """Normalize and safely limit one compact report table."""

    frame = _frame(
        value
    )

    if key == "source_file_summary":
        frame = _public_source_file_summary(
            frame
        )

    row_limit = _COMPLETE_REPORT_ROW_LIMITS.get(
        key
    )

    if (
        row_limit is not None
        and len(frame) > row_limit
    ):
        frame = frame.head(
            row_limit
        ).copy()

    return frame


def generate_tower_ipdr_complete_excel_report(
    *,
    case: dict[str, Any],
    report_path: str | Path,
    tables: dict[str, Any],
    generated_at: str = "",
) -> Path:
    """Generate the compact complete Tower IPDR workbook.

    Heavy evidence remains in the DuckDB staging backend.
    The workbook contains investigation summaries and
    controlled lead tables only.
    """

    case_id = str(
        case.get(
            "case_id",
            "",
        )
    ).strip() or "CASE"

    case_name = str(
        case.get(
            "case_name",
            "",
        )
    ).strip()

    output_path = Path(
        report_path
    ).expanduser().resolve()

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook = Workbook()
    workbook.remove(
        workbook.active
    )

    subtitle_suffix = (
        f"Case: {case_id}"
        + (
            f" | {case_name}"
            if case_name
            else ""
        )
        + (
            f" | Generated: {generated_at}"
            if generated_at
            else ""
        )
    )

    for (
        key,
        sheet_name,
        description,
    ) in COMPLETE_REPORT_SHEET_ORDER:
        frame = _complete_report_frame(
            key,
            tables.get(
                key
            ),
        )

        worksheet_name = _sheet_name(
            workbook,
            sheet_name,
        )

        worksheet = workbook.create_sheet(
            worksheet_name
        )

        _write_page(
            worksheet,
            frame,
            title=sheet_name,
            subtitle=(
                f"{subtitle_suffix} | "
                f"{description}"
            ),
        )

        worksheet.row_dimensions[2].height = (
            56
            if len(frame.columns) <= 3
            else 36
        )

    workbook.properties.title = (
        "Tower IPDR Complete Analysis "
        f"- {case_id}"
    )
    workbook.properties.subject = (
        "Compact Spot-aware Tower IPDR/NAT "
        "forensic analysis"
    )
    workbook.properties.creator = (
        "Telecom Forensics Analysis Suite"
    )

    workbook.save(
        output_path
    )

    return output_path
