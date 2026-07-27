"""Compact investigator-facing IMEI device report.

The renderer consumes an already-computed unified IMEI investigation bundle.
It does not rerun analysis functions and does not expose developer diagnostics.
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from modules.core.time_utils import utc_now_iso

from .excel_security import excel_safe_value
from .report_paths import (
    get_imei_common_report_path,
    get_imei_device_report_path,
)


IMEI_DEVICE_SHEETS = (
    "1. Device Overview",
    "2. Associated Identities",
    "3. Cross-Source Timeline",
    "4. CDR Evidence",
    "5. IPDR Evidence",
    "6. GPRS Evidence",
    "7. Device Locations",
    "8. Review Indicators",
    "9. Data Quality & Guide",
)

MAX_REPORT_ROWS = 100_000

TITLE_FILL = PatternFill(
    "solid",
    fgColor="17365D",
)

HEADER_FILL = PatternFill(
    "solid",
    fgColor="D9EAF7",
)

SECTION_FILL = PatternFill(
    "solid",
    fgColor="E2F0D9",
)

NOTE_FILL = PatternFill(
    "solid",
    fgColor="FFF2CC",
)

THIN_SIDE = Side(
    style="thin",
    color="B7C9DB",
)

TABLE_BORDER = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=THIN_SIDE,
    bottom=THIN_SIDE,
)

OVERVIEW_COLUMNS = [
    "Section",
    "Item",
    "Value",
    "Evidence Unit",
    "Note",
]

IDENTITY_COLUMNS = [
    "Evidence Source",
    "Identity Type",
    "Identity Value",
    "Related Identity",
    "First Seen",
    "Last Seen",
    "Matched Count",
]

CROSS_TIMELINE_COLUMNS = [
    "Evidence Source",
    "Evidence Type",
    "Start Time",
    "End Time",
    "Target / Subscriber",
    "IMSI",
    "Contact / Endpoint",
    "IP Address",
    "Cell ID",
    "Source File",
    "Source Row Number",
    "Source Detail",
]

CDR_EVIDENCE_COLUMNS = [
    "Date-Time",
    "Target Number",
    "Call Type",
    "Other Party",
    "Contact Category",
    "Duration (Sec)",
    "IMSI",
    "First Cell ID",
    "Last Cell ID",
    "Source File",
    "Source Row Number",
    "Raw IMEI",
    "Normalized IMEI",
]

IPDR_EVIDENCE_COLUMNS = [
    "Event Time",
    "Allocation End",
    "Subscriber / User ID",
    "IMSI",
    "Source IP",
    "Destination IP",
    "Destination Port",
    "Protocol",
    "Cell ID",
    "Source File",
    "Source Row Number",
    "Query Identifier",
    "Normalized IMEI",
    "Match Basis",
    "Match Relation",
]
GPRS_EVIDENCE_COLUMNS = [
    "Session Start",
    "Session End",
    "Duration (Sec)",
    "Subscriber Number",
    "Identifier Type",
    "IMSI",
    "IPv4 Address",
    "IPv6 Address",
    "Total Volume",
    "Technology",
    "Cell ID",
    "Source File",
    "Source Row Number",
    "Raw IMEI",
    "Normalized IMEI",
]

LOCATION_COLUMNS = [
    "Evidence Source",
    "Cell ID",
    "Latitude",
    "Longitude",
    "Linked Identities",
    "Identity Count",
    "Source Files",
    "Spot Names",
    "Matched Count",
    "First Seen",
    "Last Seen",
    "Total Volume",
]

REVIEW_COLUMNS = [
    "Evidence Source",
    "Indicator",
    "Observation",
    "Caution",
]

QUALITY_COLUMNS = [
    "Evidence Source",
    "Check",
    "Count",
    "Meaning",
]

TEXT_IDENTIFIER_COLUMNS = {
    "Requested IMEI / IMEISV",
    "Identity Value",
    "Related Identity",
    "Target / Subscriber",
    "Target Number",
    "Subscriber / User ID",
    "Subscriber Number",
    "IMSI",
    "Other Party",
    "Contact / Endpoint",
    "IP Address",
    "Source IP",
    "Destination IP",
    "Destination Port",
    "IPv4 Address",
    "IPv6 Address",
    "Cell ID",
    "First Cell ID",
    "Last Cell ID",
    "Raw IMEI",
    "Normalized IMEI",
}


def _frame(
    value: Any,
) -> pd.DataFrame:
    if isinstance(
        value,
        pd.DataFrame,
    ):
        return value.copy()

    if isinstance(
        value,
        list,
    ):
        return pd.DataFrame(
            value
        )

    if isinstance(
        value,
        dict,
    ):
        return pd.DataFrame(
            [
                value
            ]
        )

    return pd.DataFrame()


def _column(
    dataframe: pd.DataFrame,
    name: str,
    default: Any = "",
) -> pd.Series:
    if name in dataframe.columns:
        return dataframe[
            name
        ].copy()

    return pd.Series(
        default,
        index=dataframe.index,
        dtype="object",
    )


def _project(
    value: Any,
    columns: list[str],
) -> pd.DataFrame:
    source = _frame(
        value
    )

    result = pd.DataFrame(
        index=source.index
    )

    for column in columns:
        result[
            column
        ] = _column(
            source,
            column,
        )

    return result.reset_index(
        drop=True
    )


def _safe_scalar(
    value: Any,
) -> Any:
    if value is None:
        return None

    try:
        missing = pd.isna(
            value
        )

        if isinstance(
            missing,
            bool,
        ) and missing:
            return None
    except (
        TypeError,
        ValueError,
    ):
        pass

    if isinstance(
        value,
        pd.Timestamp,
    ):
        return value.to_pydatetime()

    if isinstance(
        value,
        pd.Timedelta,
    ):
        return str(
            value
        )

    if (
        hasattr(
            value,
            "item",
        )
        and not isinstance(
            value,
            (
                str,
                bytes,
                date,
                datetime,
            ),
        )
    ):
        try:
            value = value.item()
        except (
            TypeError,
            ValueError,
        ):
            pass

    return value


def _is_identifier_column(
    column: str,
) -> bool:
    """Return True when Excel must preserve the value as exact text."""

    if column in TEXT_IDENTIFIER_COLUMNS:
        return True

    upper = (
        str(
            column
        )
        .upper()
        .replace(
            "_",
            " ",
        )
        .strip()
    )

    if upper.endswith(
        " COUNT"
    ):
        return False

    if (
        "QUERY IDENTIFIER" in upper
        or "DEVICE FAMILY" in upper
        or "SHA-256" in upper
        or "SHA256" in upper
        or upper.endswith(
            " HASH"
        )
        or "IMEI" in upper
        or "IMEISV" in upper
        or "IMSI" in upper
        or "MSISDN" in upper
        or "CELL ID" in upper
        or "IP ADDRESS" in upper
        or "TARGET NUMBER" in upper
        or "CONTACT NUMBER" in upper
        or "OTHER PARTY" in upper
        or "SUBSCRIBER / USER ID" in upper
    ):
        return True

    return False


def _is_identifier_value_cell(
    dataframe: pd.DataFrame,
    *,
    row_position: int,
    column_name: str,
) -> bool:
    """Detect identifiers stored inside a generic Value column."""

    if _is_identifier_column(
        column_name
    ):
        return True

    if (
        column_name == "Value"
        and "Item" in dataframe.columns
        and 0 <= row_position < len(
            dataframe
        )
    ):
        item_name = str(
            dataframe.iloc[
                row_position
            ][
                "Item"
            ]
        )

        return _is_identifier_column(
            item_name
        )

    return False


def _write_title(
    worksheet,
    *,
    title: str,
    subtitle: str,
    columns: int,
) -> int:
    last_column = max(
        2,
        columns,
    )

    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=last_column,
    )

    title_cell = worksheet.cell(
        row=1,
        column=1,
        value=title,
    )

    title_cell.fill = TITLE_FILL
    title_cell.font = Font(
        color="FFFFFF",
        bold=True,
        size=16,
    )
    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet.row_dimensions[
        1
    ].height = 28

    worksheet.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=last_column,
    )

    subtitle_cell = worksheet.cell(
        row=2,
        column=1,
        value=excel_safe_value(
            subtitle
        ),
    )

    subtitle_cell.font = Font(
        italic=True,
        color="44546A",
    )

    subtitle_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    worksheet.row_dimensions[
        2
    ].height = 28

    return 4


def _set_widths(
    worksheet,
) -> None:
    for column_index in range(
        1,
        worksheet.max_column + 1,
    ):
        maximum = 0

        for row_index in range(
            1,
            min(
                worksheet.max_row,
                500,
            )
            + 1,
        ):
            value = worksheet.cell(
                row=row_index,
                column=column_index,
            ).value

            if value is None:
                continue

            maximum = max(
                maximum,
                len(
                    str(
                        value
                    )
                ),
            )

        worksheet.column_dimensions[
            get_column_letter(
                column_index
            )
        ].width = min(
            max(
                maximum + 2,
                11,
            ),
            44,
        )


def _write_page(
    worksheet,
    dataframe: pd.DataFrame,
    *,
    title: str,
    subtitle: str,
) -> None:
    dataframe = dataframe.copy()

    original_rows = len(
        dataframe
    )

    if original_rows > MAX_REPORT_ROWS:
        dataframe = (
            dataframe.head(
                MAX_REPORT_ROWS
            )
            .copy()
            .reset_index(
                drop=True
            )
        )

        subtitle = (
            f"{subtitle} | Showing first "
            f"{MAX_REPORT_ROWS:,} of "
            f"{original_rows:,} rows. "
            "Complete evidence remains in the case data."
        )

    header_row = _write_title(
        worksheet,
        title=title,
        subtitle=subtitle,
        columns=len(
            dataframe.columns
        ),
    )

    for column_index, column in enumerate(
        dataframe.columns,
        start=1,
    ):
        cell = worksheet.cell(
            row=header_row,
            column=column_index,
            value=excel_safe_value(
                str(
                    column
                )
            ),
        )

        cell.fill = HEADER_FILL
        cell.font = Font(
            bold=True
        )
        cell.border = TABLE_BORDER
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    if dataframe.empty:
        worksheet.cell(
            row=header_row + 1,
            column=1,
            value="No records available for this section.",
        ).fill = NOTE_FILL

    else:
        for row_position, values in enumerate(
            dataframe.itertuples(
                index=False,
                name=None,
            )
        ):
            row_index = (
                header_row
                + 1
                + row_position
            )

            for column_index, value in enumerate(
                values,
                start=1,
            ):
                column_name = str(
                    dataframe.columns[
                        column_index - 1
                    ]
                )

                safe_value = _safe_scalar(
                    value
                )

                force_text = _is_identifier_value_cell(
                    dataframe,
                    row_position=row_position,
                    column_name=column_name,
                )

                if force_text:
                    if safe_value is None:
                        output_value = ""

                    else:
                        # Apply formula-injection protection first, then
                        # keep the final OpenPyXL value as exact text.
                        output_value = str(
                            excel_safe_value(
                                str(
                                    safe_value
                                )
                            )
                        )

                else:
                    output_value = excel_safe_value(
                        safe_value
                    )

                cell = worksheet.cell(
                    row=row_index,
                    column=column_index,
                    value=output_value,
                )

                cell.border = TABLE_BORDER
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

                if force_text:
                    cell.number_format = "@"

                elif isinstance(
                    cell.value,
                    (
                        datetime,
                        date,
                    ),
                ):
                    cell.number_format = (
                        "yyyy-mm-dd hh:mm:ss"
                    )

        last_row = (
            header_row
            + len(
                dataframe
            )
        )

        last_column = get_column_letter(
            len(
                dataframe.columns
            )
        )

        worksheet.auto_filter.ref = (
            f"A{header_row}:"
            f"{last_column}{last_row}"
        )

    worksheet.freeze_panes = (
        f"A{header_row + 1}"
    )

    worksheet.sheet_view.showGridLines = False
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0

    _set_widths(
        worksheet
    )


def _safe_source_message(
    status: Any,
    message: Any,
) -> str:
    normalized_status = str(
        status or ""
    ).strip().upper()

    if normalized_status == "ERROR":
        return (
            "Source analysis was unavailable. "
            "See the technical log for details."
        )

    return str(
        message or ""
    ).strip()


def _build_overview(
    *,
    case: dict[str, Any],
    analysis: dict[str, Any],
) -> pd.DataFrame:
    case_id = (
        str(
            case.get(
                "case_id",
                "",
            )
        ).strip()
        or "CASE"
    )

    case_name = str(
        case.get(
            "case_name",
            "",
        )
    ).strip()

    requested_imei = str(
        analysis.get(
            "requested_imei",
            "",
        )
    ).strip()

    overall_status = str(
        analysis.get(
            "overall_status",
            "",
        )
    ).strip().upper()

    rows = [
        {
            "Section": "CASE",
            "Item": "Case ID",
            "Value": case_id,
            "Evidence Unit": "",
            "Note": "",
        },
        {
            "Section": "CASE",
            "Item": "Case Name",
            "Value": case_name,
            "Evidence Unit": "",
            "Note": "",
        },
        {
            "Section": "DEVICE",
            "Item": "Requested Device Query Identifier",
            "Value": requested_imei,
            "Evidence Unit": "",
            "Note": (
                'The report-query identifier is matched exactly. BASE14, IMEI15 and IMEISV16 remain separate.'
            ),
        },
        {
            "Section": "ANALYSIS",
            "Item": "Overall Status",
            "Value": overall_status,
            "Evidence Unit": "",
            "Note": str(
                analysis.get(
                    "message",
                    "",
                )
            ),
        },
        {
            "Section": "ANALYSIS",
            "Item": "Generated At",
            "Value": utc_now_iso(),
            "Evidence Unit": "",
            "Note": "",
        },
    ]

    source_summary = _project(
        analysis.get(
            "source_summary"
        ),
        [
            "Evidence Source",
            "Status",
            "Evidence Unit",
            "Matched Count",
            "Message",
        ],
    )

    for record in source_summary.to_dict(
        orient="records"
    ):
        rows.append(
            {
                "Section": "SOURCE",
                "Item": record.get(
                    "Evidence Source",
                    "",
                ),
                "Value": record.get(
                    "Matched Count",
                    0,
                ),
                "Evidence Unit": record.get(
                    "Evidence Unit",
                    "",
                ),
                "Note": _safe_source_message(
                    record.get(
                        "Status",
                        "",
                    ),
                    record.get(
                        "Message",
                        "",
                    ),
                ),
            }
        )

    rows.append(
        {
            "Section": "GUIDANCE",
            "Item": "Source Count Rule",
            "Value": (
                "Counts remain separate"
            ),
            "Evidence Unit": "",
            "Note": (
                "CDR records, IPDR records and GPRS "
                "sessions are different evidence types. "
                "They are not combined into one total."
            ),
        }
    )

    return pd.DataFrame(
        rows,
        columns=OVERVIEW_COLUMNS,
    )


def _source_timeline(
    analysis: dict[str, Any],
    source_key: str,
) -> pd.DataFrame:
    source = analysis.get(
        source_key
    )

    if not isinstance(
        source,
        dict,
    ):
        return pd.DataFrame()

    return _frame(
        source.get(
            "timeline"
        )
    )


def _build_locations(
    analysis: dict[str, Any],
) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []

    cdr = analysis.get(
        "cdr"
    )

    if isinstance(
        cdr,
        dict,
    ):
        towers = _frame(
            cdr.get(
                "towers"
            )
        )

        if not towers.empty:
            frames.append(
                pd.DataFrame(
                    {
                        "Evidence Source": "CDR",
                        "Cell ID": _column(
                            towers,
                            "Cell ID",
                        ),
                        "Latitude": "",
                        "Longitude": "",
                        "Linked Identities": _column(
                            towers,
                            "Linked Targets",
                        ),
                        "Identity Count": _column(
                            towers,
                            "Target Count",
                            0,
                        ),
                        "Source Files": "",
                        "Spot Names": "",
                        "Matched Count": _column(
                            towers,
                            "Total Events",
                            0,
                        ),
                        "First Seen": _column(
                            towers,
                            "First Seen",
                        ),
                        "Last Seen": _column(
                            towers,
                            "Last Seen",
                        ),
                        "Total Volume": "",
                    }
                )
            )

    ipdr = analysis.get(
        "ipdr"
    )

    if isinstance(
        ipdr,
        dict,
    ):
        cells = _frame(
            ipdr.get(
                "cells"
            )
        )

        if not cells.empty:
            frames.append(
                pd.DataFrame(
                    {
                        "Evidence Source": "IPDR",
                        "Cell ID": _column(
                            cells,
                            "Cell ID",
                        ),
                        "Latitude": "",
                        "Longitude": "",
                        "Linked Identities": _column(
                            cells,
                            "Subscribers / User IDs",
                        ),
                        "Identity Count": _column(
                            cells,
                            "Subscriber Count",
                            0,
                        ),
                        "Source Files": _column(
                            cells,
                            "Source Files",
                        ),
                        "Spot Names": "",
                        "Matched Count": _column(
                            cells,
                            "Total Records",
                            0,
                        ),
                        "First Seen": _column(
                            cells,
                            "First Seen",
                        ),
                        "Last Seen": _column(
                            cells,
                            "Last Seen",
                        ),
                        "Total Volume": "",
                    }
                )
            )

    gprs = analysis.get(
        "gprs"
    )

    if isinstance(
        gprs,
        dict,
    ):
        cells = _frame(
            gprs.get(
                "cells"
            )
        )

        if not cells.empty:
            frames.append(
                pd.DataFrame(
                    {
                        "Evidence Source": "GPRS",
                        "Cell ID": _column(
                            cells,
                            "Cell ID",
                        ),
                        "Latitude": _column(
                            cells,
                            "Latitude",
                        ),
                        "Longitude": _column(
                            cells,
                            "Longitude",
                        ),
                        "Linked Identities": _column(
                            cells,
                            "Subscribers",
                        ),
                        "Identity Count": _column(
                            cells,
                            "Subscriber Count",
                            0,
                        ),
                        "Source Files": _column(
                            cells,
                            "Source Files",
                        ),
                        "Spot Names": _column(
                            cells,
                            "Spot Names",
                        ),
                        "Matched Count": _column(
                            cells,
                            "Total Sessions",
                            0,
                        ),
                        "First Seen": _column(
                            cells,
                            "First Seen",
                        ),
                        "Last Seen": _column(
                            cells,
                            "Last Seen",
                        ),
                        "Total Volume": _column(
                            cells,
                            "Total Volume",
                            0,
                        ),
                    }
                )
            )

    if not frames:
        return pd.DataFrame(
            columns=LOCATION_COLUMNS
        )

    return (
        pd.concat(
            frames,
            ignore_index=True,
            sort=False,
        )[
            LOCATION_COLUMNS
        ]
        .sort_values(
            [
                "Evidence Source",
                "Matched Count",
                "Cell ID",
            ],
            ascending=[
                True,
                False,
                True,
            ],
            kind="stable",
        )
        .reset_index(
            drop=True
        )
    )


def _build_quality_and_guide(
    analysis: dict[str, Any],
) -> pd.DataFrame:
    quality = _project(
        analysis.get(
            "data_quality"
        ),
        QUALITY_COLUMNS,
    )

    guide = pd.DataFrame(
        [
            {
                "Evidence Source": "GUIDE",
                "Check": "Exact identifier matching",
                "Count": "",
                "Meaning": (
                    'Dedicated BASE14 report queries, 15-digit IMEI and 16-digit IMEISV remain distinct and are not silently converted.'
                ),
            },
            {
                "Evidence Source": "GUIDE",
                "Check": "Source-specific counts",
                "Count": "",
                "Meaning": (
                    "CDR records, IPDR records and GPRS "
                    "sessions must be interpreted separately."
                ),
            },
            {
                "Evidence Source": "GUIDE",
                "Check": "Identity association",
                "Count": "",
                "Meaning": (
                    "Association with a number or IMSI does "
                    "not by itself prove ownership."
                ),
            },
            {
                "Evidence Source": "GUIDE",
                "Check": "Location interpretation",
                "Count": "",
                "Meaning": (
                    "Tower evidence indicates network presence "
                    "and must be verified with time, coverage "
                    "and source-record context."
                ),
            },
            {
                "Evidence Source": "GUIDE",
                "Check": "Review indicators",
                "Count": "",
                "Meaning": (
                    "Indicators identify records requiring "
                    "verification; they are not automatic "
                    "suspect labels."
                ),
            },
            {
                "Evidence Source": "GUIDE",
                "Check": "Evidence verification",
                "Count": "",
                "Meaning": (
                    "Use Source File and Source Row Number "
                    "to verify material findings against "
                    "the original evidence."
                ),
            },
        ],
        columns=QUALITY_COLUMNS,
    )

    return pd.concat(
        [
            quality,
            guide,
        ],
        ignore_index=True,
        sort=False,
    )[
        QUALITY_COLUMNS
    ]


def _has_reportable_evidence(
    analysis: dict[str, Any],
) -> bool:
    source_summary = _frame(
        analysis.get(
            "source_summary"
        )
    )

    if (
        not source_summary.empty
        and "Matched Count" in source_summary.columns
    ):
        count = pd.to_numeric(
            source_summary[
                "Matched Count"
            ],
            errors="coerce",
        ).fillna(
            0
        )

        if count.gt(
            0
        ).any():
            return True

    for source_key in (
        "cdr",
        "ipdr",
        "gprs",
    ):
        if not _source_timeline(
            analysis,
            source_key,
        ).empty:
            return True

    return False


ACQUISITION_MANIFEST_REPORT_COLUMNS = [
    "Relative Path",
    "SHA-256",
    "Acquisition Content Role",
    "Analysis Content Role",
    "Duplicate Reference",
    "Format",
    "Operator",
    "Source Type",
    "Query Identifier",
    "Query Identifier Type",
    "Inspection Status",
    "Records Declared",
    "Records Normalized",
    "Rejected Lines",
    "Message",
]


def _manifest_file_reference(
    value: Any,
) -> str:
    """Return a shareable file reference without workstation paths."""

    text = str(
        value or ""
    ).strip()

    if not text:
        return ""

    return Path(
        text
    ).name


def _build_acquisition_manifest(
    value: Any,
) -> pd.DataFrame:
    """Build the compact investigator-facing acquisition manifest."""

    manifest = _frame(
        value
    )

    if manifest.empty:
        return pd.DataFrame(
            columns=ACQUISITION_MANIFEST_REPORT_COLUMNS
        )

    work = manifest.copy(
        deep=True
    )

    acquisition_duplicate = (
        work[
            "Duplicate Of"
        ]
        .astype(
            "string"
        )
        .fillna(
            ""
        )
        .str.strip()
        if "Duplicate Of" in work.columns
        else pd.Series(
            "",
            index=work.index,
            dtype="string",
        )
    )

    analysis_duplicate = (
        work[
            "Analysis Duplicate Of"
        ]
        .astype(
            "string"
        )
        .fillna(
            ""
        )
        .str.strip()
        if "Analysis Duplicate Of" in work.columns
        else pd.Series(
            "",
            index=work.index,
            dtype="string",
        )
    )

    duplicate_reference = analysis_duplicate.where(
        analysis_duplicate.ne(
            ""
        ),
        acquisition_duplicate,
    )

    work[
        "Duplicate Reference"
    ] = duplicate_reference.map(
        _manifest_file_reference
    )

    return _project(
        work,
        ACQUISITION_MANIFEST_REPORT_COLUMNS,
    )

def generate_imei_device_report(
    *,
    case: dict[str, Any] | None,
    analysis: dict[str, Any],
    output_dir: str | Path | None = None,
) -> Path | None:
    """Generate one compact IMEI device investigation workbook.

    Invalid, no-input and not-found investigations do not create an empty
    workbook. The caller receives ``None`` and can show the analysis message.
    """

    if not isinstance(
        analysis,
        dict,
    ):
        return None

    status = str(
        analysis.get(
            "overall_status",
            "",
        )
    ).strip().upper()

    requested_imei = str(
        analysis.get(
            "requested_imei",
            "",
        )
    ).strip()

    manifest = _build_acquisition_manifest(
        analysis.get(
            "acquisition_manifest"
        )
    )

    has_reportable_evidence = _has_reportable_evidence(
        analysis
    )

    if not requested_imei:
        return None

    if (
        status in {
            "FOUND",
            "PARTIAL",
        }
        and not has_reportable_evidence
    ):
        return None

    if (
        status == "EMPTY_NO_DATA"
        and manifest.empty
    ):
        return None

    if status not in {
        "FOUND",
        "PARTIAL",
        "EMPTY_NO_DATA",
    }:
        return None

    case = (
        dict(
            case
        )
        if isinstance(
            case,
            dict,
        )
        else {}
    )

    case_id = str(
        case.get(
            "case_id",
            "",
        )
    ).strip()

    report_path = get_imei_device_report_path(
        requested_imei,
        case_id=case_id,
        output_dir=output_dir,
    )

    workbook = Workbook()

    workbook.remove(
        workbook.active
    )

    subtitle = (
        f"Device Query Identifier: {requested_imei}"
    )

    sheets = [
        (
            IMEI_DEVICE_SHEETS[0],
            _build_overview(
                case=case,
                analysis=analysis,
            ),
            (
                "Case and source-specific device "
                "investigation summary."
            ),
        ),
        (
            IMEI_DEVICE_SHEETS[1],
            _project(
                analysis.get(
                    "associated_identities"
                ),
                IDENTITY_COLUMNS,
            ),
            (
                'Numbers and IMSIs associated with the requested device-query scope.'
            ),
        ),
        (
            IMEI_DEVICE_SHEETS[2],
            _project(
                analysis.get(
                    "cross_source_timeline"
                ),
                CROSS_TIMELINE_COLUMNS,
            ),
            (
                "Chronological view across evidence sources. "
                "Source-specific details remain separate."
            ),
        ),
        (
            IMEI_DEVICE_SHEETS[3],
            _project(
                _source_timeline(
                    analysis,
                    "cdr",
                ),
                CDR_EVIDENCE_COLUMNS,
            ),
            (
                "Matched call and SMS evidence from CDR data."
            ),
        ),
        (
            IMEI_DEVICE_SHEETS[4],
            _project(
                _source_timeline(
                    analysis,
                    "ipdr",
                ),
                IPDR_EVIDENCE_COLUMNS,
            ),
            (
                "Matched internet and allocation evidence "
                "from IPDR data."
            ),
        ),
        (
            IMEI_DEVICE_SHEETS[5],
            _project(
                _source_timeline(
                    analysis,
                    "gprs",
                ),
                GPRS_EVIDENCE_COLUMNS,
            ),
            (
                "Matched data-session evidence from "
                "normalized GPRS records."
            ),
        ),
        (
            IMEI_DEVICE_SHEETS[6],
            _build_locations(
                analysis
            ),
            (
                'Cell IDs observed within the requested device-query scope. Interpret with source and time context.'
            ),
        ),
        (
            IMEI_DEVICE_SHEETS[7],
            _project(
                analysis.get(
                    "review_indicators"
                ),
                REVIEW_COLUMNS,
            ),
            (
                "Evidence conditions requiring investigator "
                "verification. These are not conclusions."
            ),
        ),
        (
            IMEI_DEVICE_SHEETS[8],
            _build_quality_and_guide(
                analysis
            ),
            (
                "Data limitations, quality checks and "
                "interpretation guidance."
            ),
        ),
    ]

    if not manifest.empty:
        sheets.append(
            (
                "10. Acquisition Manifest",
                manifest,
                (
                    "Physical acquisition paths, report status, "
                    "SHA-256 and analytical content role."
                ),
            )
        )

    for sheet_name, dataframe, guidance in sheets:
        worksheet = workbook.create_sheet(
            sheet_name
        )

        _write_page(
            worksheet,
            dataframe,
            title=sheet_name,
            subtitle=(
                f"{subtitle} | {guidance}"
            ),
        )

    workbook.save(
        report_path
    )

    return report_path


def generate_imei_common_report(
    *,
    case: dict[str, Any] | None,
    analysis: dict[str, Any],
    output_dir: str | Path | None = None,
) -> Path | None:
    """Generate one cross-device IMEI CDR workbook."""

    if not isinstance(
        analysis,
        dict,
    ):
        return None

    if str(
        analysis.get(
            "status",
            "",
        )
    ).strip().upper() != "FOUND":
        return None

    if int(
        analysis.get(
            "device_count",
            0,
        )
        or 0
    ) < 2:
        return None

    case_value = (
        dict(
            case
        )
        if isinstance(
            case,
            dict,
        )
        else {}
    )

    case_id = str(
        case_value.get(
            "case_id",
            "",
        )
    ).strip()

    report_path = get_imei_common_report_path(
        case_id,
        output_dir=output_dir,
    )

    workbook = Workbook()

    workbook.remove(
        workbook.active
    )

    sheets = [
        (
            "1. Device Overview",
            _frame(
                analysis.get(
                    "device_overview"
                )
            ),
            (
                "Per-device acquisition and normalized "
                "evidence summary."
            ),
        ),
        (
            "2. Common Targets",
            _frame(
                analysis.get(
                    "common_targets"
                )
            ),
            (
                "Target numbers appearing with two or "
                "more device families."
            ),
        ),
        (
            "3. Common IMSIs",
            _frame(
                analysis.get(
                    "common_imsis"
                )
            ),
            (
                "SIM identities appearing with two or "
                "more device families."
            ),
        ),
        (
            "4. Common Contacts",
            _frame(
                analysis.get(
                    "common_contacts"
                )
            ),
            (
                "Human mobile contacts shared across "
                "distinct device families."
            ),
        ),
        (
            "5. Shared Service IDs",
            _frame(
                analysis.get(
                    "shared_service_identifiers"
                )
            ),
            (
                "Sender IDs, short codes and other non-human "
                "identifiers shared across device families."
            ),
        ),
        (
            "6. Common Towers",
            _frame(
                analysis.get(
                    "common_towers"
                )
            ),
            (
                "Canonically valid Cell IDs shared across "
                "distinct device families."
            ),
        ),
        (
            "7. Cross Device Timeline",
            _frame(
                analysis.get(
                    "cross_device_timeline"
                )
            ),
            (
                "Chronological CDR evidence retaining "
                "query identifier and source provenance."
            ),
        ),
        (
            "8. Acquisition Manifest",
            _build_acquisition_manifest(
                analysis.get(
                    "acquisition_manifest"
                )
            ),
            (
                "Every physical evidence path and SHA-256 "
                "content role."
            ),
        ),
        (
            "9. Review Indicators",
            _frame(
                analysis.get(
                    "review_indicators"
                )
            ),
            (
                "Shared conditions requiring investigator "
                "verification; these are not conclusions."
            ),
        ),
        (
            "10. Data Quality",
            _frame(
                analysis.get(
                    "data_quality"
                )
            ),
            (
                "Data limitations and content-deduplication "
                "checks."
            ),
        ),
    ]

    subtitle = (
        f"Case: {case_id or 'CASE'} | "
        f"{analysis.get('message', '')}"
    )

    for sheet_name, dataframe, guidance in sheets:
        worksheet = workbook.create_sheet(
            sheet_name
        )

        _write_page(
            worksheet,
            dataframe,
            title=sheet_name,
            subtitle=(
                f"{subtitle} | {guidance}"
            ),
        )

    workbook.save(
        report_path
    )

    return report_path
