"""Generate a detailed analyzed Excel workbook for one normalized CDR."""

from __future__ import annotations

from pathlib import Path
from typing import Any
import traceback

import numpy as np
import pandas as pd
from openpyxl import Workbook

from .excel_styles import (
    finish_sheet,
    set_sensible_widths,
    style_data_area,
    style_metadata_block,
    style_table_header,
)
from .excel_security import excel_safe_value
from .report_guidance import append_methodology_sheet
from .report_paths import get_single_report_path
from modules.enrichment.cgi_address_enrichment import enrich_dataframe_with_cgi_address


CANONICAL_COLUMNS = {
    "a_party": pd.NA,
    "b_party": pd.NA,
    "call_type": "unknown",
    "call_direction": "UNKNOWN",
    "call_date": pd.NA,
    "call_time": pd.NA,
    "call_duration": 0,
    "imei": pd.NA,
    "imsi": pd.NA,
    "first_cell_id": pd.NA,
    "last_cell_id": pd.NA,
    "first_location": pd.NA,
    "last_location": pd.NA,
    "roaming_circle": pd.NA,
    "service_type": pd.NA,
    "source_file": pd.NA,
}


def _first_existing(df: pd.DataFrame, candidates: list[str], default: Any = pd.NA) -> pd.Series:
    for column in candidates:
        if column in df.columns:
            return df[column]
    return pd.Series(default, index=df.index)


def _prepare_dataframe(df: pd.DataFrame, target: str) -> pd.DataFrame:
    data = df.copy()

    # CGI_LEGACY_EXPORT_ENRICHMENT
    try:
        data = enrich_dataframe_with_cgi_address(
            data,
            cell_id_column="first_cell_id",
            prefix="tower_",
        )

        if "tower_address" in data.columns:
            data["Address1"] = data["tower_address"]

        if "tower_latitude" in data.columns:
            data["Latitude"] = data["tower_latitude"]

        if "tower_longitude" in data.columns:
            data["Longitude"] = data["tower_longitude"]

    except Exception:
        pass


    for column, default_value in CANONICAL_COLUMNS.items():
        if column not in data.columns:
            data[column] = default_value

    if "datetime" not in data.columns:
        data["datetime"] = pd.to_datetime(
            data["call_date"].astype("string").fillna("")
            + " "
            + data["call_time"].astype("string").fillna(""),
            errors="coerce",
            dayfirst=True,
        )
    else:
        data["datetime"] = pd.to_datetime(data["datetime"], errors="coerce", dayfirst=True)

    data["call_duration"] = pd.to_numeric(data["call_duration"], errors="coerce").fillna(0).clip(lower=0).astype(int)
    data["call_type"] = data["call_type"].astype("string").fillna("unknown").str.lower().str.strip()
    data["call_direction"] = data["call_direction"].astype("string").fillna("UNKNOWN").str.upper().str.strip()
    data["target_number"] = str(target)
    data["other_party"] = _first_existing(data, ["opposite_party", "b_party"])

    # Optional enrichment columns. These stay blank until CGI/SDR/TAC databases are connected.
    data["contact_name"] = _first_existing(data, ["contact_name", "name", "subscriber_name"], "")
    data["contact_address"] = _first_existing(data, ["contact_address", "sdr_address", "address"], "")
    data["level_code"] = _first_existing(data, ["level_code", "operator_circle", "network"], "Missing")
    data["tower_address"] = _first_existing(data, ["tower_address", "first_location", "location"], "")
    data["last_tower_address"] = _first_existing(data, ["last_tower_address", "last_location"], "")
    data["latitude"] = _first_existing(data, ["latitude", "lat"], "")
    data["longitude"] = _first_existing(data, ["longitude", "lon", "lng"], "")
    data["handset_make"] = _first_existing(data, ["handset_make", "make"], "")
    data["handset_model"] = _first_existing(data, ["handset_model", "model"], "")

    valid_datetime = data["datetime"].notna()
    data["report_date"] = data["call_date"].astype("string")
    data["report_time"] = data["call_time"].astype("string")
    data.loc[valid_datetime, "report_date"] = data.loc[valid_datetime, "datetime"].dt.strftime("%b %d, %Y")
    data.loc[valid_datetime, "report_time"] = data.loc[valid_datetime, "datetime"].dt.strftime("%H:%M:%S")

    return data.sort_values("datetime", na_position="last", kind="stable").reset_index(drop=True)


def _metadata(target: str, data: pd.DataFrame, metadata: dict[str, Any] | None) -> dict[str, Any]:
    result = dict(metadata or {})
    attrs = getattr(data, "attrs", {}) or {}
    stored_metadata = attrs.get("metadata", {}) if isinstance(attrs.get("metadata", {}), dict) else {}
    for key, value in stored_metadata.items():
        result.setdefault(key, value)

    valid_dates = data["datetime"].dropna()
    result.setdefault("case_name", "")
    result.setdefault("target", target)
    result.setdefault("subscriber_name", "")
    result.setdefault("address", "")
    result.setdefault("from_date", valid_dates.min().strftime("%b %d, %Y") if not valid_dates.empty else "All Period")
    result.setdefault("to_date", valid_dates.max().strftime("%b %d, %Y") if not valid_dates.empty else "All Period")
    return result


def _metadata_rows(meta: dict[str, Any], report_name: str) -> list[tuple[str, Any]]:
    return [
        ("Case", meta.get("case_name", "")),
        ("Report", report_name),
        ("Report For", meta.get("target", "")),
        ("Name", meta.get("subscriber_name", "")),
        ("Address", meta.get("address", "")),
        ("From Date", meta.get("from_date", "All Period")),
        ("To Date", meta.get("to_date", "All Period")),
    ]


def _voice_out_mask(data: pd.DataFrame) -> pd.Series:
    return data["call_type"].eq("outgoing") | ((data["call_direction"].eq("OUTGOING")) & ~data["call_type"].str.contains("sms", na=False))


def _voice_in_mask(data: pd.DataFrame) -> pd.Series:
    return data["call_type"].eq("incoming") | ((data["call_direction"].eq("INCOMING")) & ~data["call_type"].str.contains("sms", na=False))


def _sms_out_mask(data: pd.DataFrame) -> pd.Series:
    return data["call_type"].isin(["smsout", "outgoing_sms"]) | (data["call_type"].str.contains("sms", na=False) & data["call_direction"].eq("OUTGOING"))


def _sms_in_mask(data: pd.DataFrame) -> pd.Series:
    return data["call_type"].isin(["smsin", "incoming_sms"]) | (data["call_type"].str.contains("sms", na=False) & data["call_direction"].eq("INCOMING"))


def _call_type_display(data: pd.DataFrame) -> pd.Series:
    conditions = [_voice_out_mask(data), _voice_in_mask(data), _sms_out_mask(data), _sms_in_mask(data)]
    choices = ["Outgoing Call", "Incoming Call", "Outgoing SMS", "Incoming SMS"]
    return pd.Series(np.select(conditions, choices, default=data["call_type"].astype(str)), index=data.index)


def _sub_call_type(data: pd.DataFrame) -> pd.Series:
    raw = _first_existing(data, ["raw_call_type", "service_type"], "")
    fallback = pd.Series(np.select(
        [_voice_out_mask(data), _voice_in_mask(data), _sms_out_mask(data), _sms_in_mask(data)],
        ["Voice out", "Voice in", "P2P OUT", "A2P IN"],
        default="",
    ), index=data.index)
    raw_text = raw.astype("string").fillna("").str.strip()
    return raw_text.where(raw_text.ne(""), fallback)


def _contact_summary(data: pd.DataFrame) -> pd.DataFrame:
    work = data[data["other_party"].notna() & data["other_party"].astype("string").str.strip().ne("")].copy()
    if work.empty:
        return pd.DataFrame(columns=[
            "Other Party", "Level Code", "First Call Time", "Last Call Time", "Total Calls", "Total Duration",
            "Out Count", "Out Duration", "IN Count", "In Duration", "Out SMS Count", "In SMS Count",
            "Avg. Call Duration", "Name", "Address",
        ])

    work["voice_out"] = _voice_out_mask(work)
    work["voice_in"] = _voice_in_mask(work)
    work["sms_out"] = _sms_out_mask(work)
    work["sms_in"] = _sms_in_mask(work)
    work["out_duration"] = work["call_duration"].where(work["voice_out"], 0)
    work["in_duration"] = work["call_duration"].where(work["voice_in"], 0)
    work["voice_duration"] = work["call_duration"].where(work["voice_out"] | work["voice_in"], 0)
    work["voice_event"] = (work["voice_out"] | work["voice_in"]).astype(int)

    grouped = work.groupby("other_party", dropna=False)
    result = grouped.agg(
        **{
            "Level Code": ("level_code", lambda x: x.dropna().astype(str).iloc[0] if not x.dropna().empty else "Missing"),
            "First Call Time": ("datetime", "min"),
            "Last Call Time": ("datetime", "max"),
            "Total Calls": ("other_party", "size"),
            "Total Duration": ("voice_duration", "sum"),
            "Out Count": ("voice_out", "sum"),
            "Out Duration": ("out_duration", "sum"),
            "IN Count": ("voice_in", "sum"),
            "In Duration": ("in_duration", "sum"),
            "Out SMS Count": ("sms_out", "sum"),
            "In SMS Count": ("sms_in", "sum"),
            "Voice Events": ("voice_event", "sum"),
            "Name": ("contact_name", lambda x: x.dropna().astype(str).iloc[0] if not x.dropna().empty else ""),
            "Address": ("contact_address", lambda x: x.dropna().astype(str).iloc[0] if not x.dropna().empty else ""),
        }
    ).reset_index().rename(columns={"other_party": "Other Party"})

    average_duration = (
        result["Total Duration"]
        .div(result["Voice Events"].replace(0, np.nan))
        .fillna(0)
        .round()
        .astype(int)
    )
    result["Avg. Call Duration"] = average_duration
    result = result.drop(columns=["Voice Events"])
    return result.sort_values(["Total Calls", "Total Duration"], ascending=False).reset_index(drop=True)


def _cell_summary(data: pd.DataFrame) -> pd.DataFrame:
    work = data[data["first_cell_id"].notna() & data["first_cell_id"].astype("string").str.strip().ne("")]
    if work.empty:
        return pd.DataFrame(columns=["Cell ID", "Total Calls", "Address"])
    result = work.groupby("first_cell_id", dropna=False).agg(
        **{
            "Total Calls": ("first_cell_id", "size"),
            "Address": ("tower_address", lambda x: x.dropna().astype(str).iloc[0] if not x.dropna().empty else ""),
        }
    ).reset_index().rename(columns={"first_cell_id": "Cell ID"})
    return result.sort_values("Total Calls", ascending=False).reset_index(drop=True)


def _imei_summary(data: pd.DataFrame) -> pd.DataFrame:
    work = data[data["imei"].notna() & data["imei"].astype("string").str.strip().ne("")].copy()
    if work.empty:
        return pd.DataFrame(columns=[
            "IMEI", "Handset Make", "Handset Model", "Total Calls", "Total Duration",
            "Out Count", "Out Duration", "In Count", "In Duration", "SMS Count",
        ])
    work["voice_out"] = _voice_out_mask(work)
    work["voice_in"] = _voice_in_mask(work)
    work["sms"] = _sms_out_mask(work) | _sms_in_mask(work)
    work["out_duration"] = work["call_duration"].where(work["voice_out"], 0)
    work["in_duration"] = work["call_duration"].where(work["voice_in"], 0)
    work["voice_duration"] = work["call_duration"].where(work["voice_out"] | work["voice_in"], 0)
    work["voice_event"] = (work["voice_out"] | work["voice_in"]).astype(int)
    result = work.groupby("imei", dropna=False).agg(
        **{
            "Handset Make": ("handset_make", lambda x: x.dropna().astype(str).iloc[0] if not x.dropna().empty else ""),
            "Handset Model": ("handset_model", lambda x: x.dropna().astype(str).iloc[0] if not x.dropna().empty else ""),
            "Total Calls": ("voice_event", "sum"),
            "Total Duration": ("voice_duration", "sum"),
            "Out Count": ("voice_out", "sum"),
            "Out Duration": ("out_duration", "sum"),
            "In Count": ("voice_in", "sum"),
            "In Duration": ("in_duration", "sum"),
            "SMS Count": ("sms", "sum"),
        }
    ).reset_index().rename(columns={"imei": "IMEI"})
    return result.sort_values(["Total Calls", "SMS Count"], ascending=False).reset_index(drop=True)


def _movements(data: pd.DataFrame) -> pd.DataFrame:
    result = pd.DataFrame({
        "Mobile Number": data["target_number"],
        "Other Party": data["other_party"],
        "Name": data["contact_name"],
        "Address": data["contact_address"],
        "Level Code": data["level_code"],
        "Call Type": _call_type_display(data),
        "Sub Call Type": _sub_call_type(data),
        "IMEI": data["imei"],
        "Date": data["report_date"],
        "Time": data["report_time"],
        "Duration": data["call_duration"],
        "Roaming Circle": data["roaming_circle"],
        "Cell ID": data["first_cell_id"],
        "Address1": data.get("tower_address", data.get("Address1", "")),
        "Latitude": data.get("tower_latitude", data.get("Latitude", "")),
        "Longitude": data.get("tower_longitude", data.get("Longitude", "")),
    })
    return result


def _fclc(data: pd.DataFrame) -> pd.DataFrame:
    valid = data[data["datetime"].notna()].copy()
    if valid.empty:
        return pd.DataFrame(columns=["Date", "Time", "Description", "Call Type", "Other Party", "Level Code", "Duration", "Cell ID", "Address"])
    valid["day"] = valid["datetime"].dt.date
    first_rows = valid.loc[valid.groupby("day")["datetime"].idxmin()].copy()
    first_rows["Description"] = "first call"
    last_rows = valid.loc[valid.groupby("day")["datetime"].idxmax()].copy()
    last_rows["Description"] = "last call"

    # Loader provenance is stored in DataFrame.attrs and may contain another
    # DataFrame (the rejected-row ledger). Pandas compares attrs during concat;
    # comparing embedded DataFrames raises "truth value is ambiguous". These
    # temporary report slices do not need inherited attrs, so clear them first.
    first_rows.attrs = {}
    last_rows.attrs = {}
    combined = pd.concat(
        [first_rows, last_rows],
        ignore_index=True,
    ).sort_values(["datetime", "Description"], kind="stable")
    return pd.DataFrame({
        "Date": combined["datetime"].dt.strftime("%b %d, %Y"),
        "Time": combined["datetime"].dt.strftime("%H:%M:%S"),
        "Description": combined["Description"],
        "Call Type": _call_type_display(combined),
        "Other Party": combined["other_party"],
        "Level Code": combined["level_code"],
        "Duration": combined["call_duration"],
        "Cell ID": combined["first_cell_id"],
        "Address": combined["tower_address"],
    })


def _fclc_summary(fclc: pd.DataFrame) -> pd.DataFrame:
    if fclc.empty:
        return pd.DataFrame(columns=["Cell ID", "Total Calls", "Address"])
    work = fclc[fclc["Cell ID"].notna() & fclc["Cell ID"].astype("string").str.strip().ne("")]
    if work.empty:
        return pd.DataFrame(columns=["Cell ID", "Total Calls", "Address"])
    result = work.groupby("Cell ID", dropna=False).agg(
        **{
            "Total Calls": ("Cell ID", "size"),
            "Address": ("Address", lambda x: x.dropna().astype(str).iloc[0] if not x.dropna().empty else ""),
        }
    ).reset_index()
    return result.sort_values("Total Calls", ascending=False).reset_index(drop=True)


def _fclc_op(data: pd.DataFrame) -> pd.DataFrame:
    work = data[data["other_party"].notna() & data["other_party"].astype("string").str.strip().ne("")]
    if work.empty:
        return pd.DataFrame(columns=["Other Party", "Level Code", "Total Calls", "First Call Time", "Last Call Time"])
    result = work.groupby("other_party", dropna=False).agg(
        **{
            "Level Code": ("level_code", lambda x: x.dropna().astype(str).iloc[0] if not x.dropna().empty else "Missing"),
            "Total Calls": ("other_party", "size"),
            "First Call Time": ("datetime", "min"),
            "Last Call Time": ("datetime", "max"),
        }
    ).reset_index().rename(columns={"other_party": "Other Party"})
    return result.sort_values("Total Calls", ascending=False).reset_index(drop=True)


def _moving_calls(data: pd.DataFrame) -> pd.DataFrame:
    work = data[
        data["last_cell_id"].notna()
        & data["first_cell_id"].notna()
        & data["first_cell_id"].astype("string").ne(data["last_cell_id"].astype("string"))
        & (_voice_out_mask(data) | _voice_in_mask(data))
    ].copy()
    return pd.DataFrame({
        "Mobile Number": work["target_number"],
        "Other Party": work["other_party"],
        "Call Type": _call_type_display(work),
        "IMEI": work["imei"],
        "Date": work["report_date"],
        "Time": work["report_time"],
        "Duration": work["call_duration"],
        "Cell ID": work["first_cell_id"],
        "Address": work["tower_address"],
        "Last Cell ID": work["last_cell_id"],
        "End Address": work["last_tower_address"],
    })


def _roaming_summary(data: pd.DataFrame) -> pd.DataFrame:
    circle = data["roaming_circle"].astype("string").fillna("Home Circle").replace("", "Home Circle")
    provider = _first_existing(data, ["provider", "operator"], "Unknown").astype("string").fillna("Unknown").replace("", "Unknown")
    work = pd.DataFrame({"Circle": circle, "Provider": provider})
    return work.groupby(["Circle", "Provider"], dropna=False).size().reset_index(name="Total Calls").sort_values("Total Calls", ascending=False).reset_index(drop=True)


def _direction_sheet(data: pd.DataFrame, mask: pd.Series) -> pd.DataFrame:
    work = data[mask].copy()
    return pd.DataFrame({
        "Sub Call Type": _sub_call_type(work),
        "Other Party": work["other_party"],
        "Name": work["contact_name"],
        "SDR Address": work["contact_address"],
        "Level Code": work["level_code"],
        "Date": work["report_date"],
        "Time": work["report_time"],
        "Duration": work["call_duration"],
        "Cell ID": work["first_cell_id"],
        "Address": work["tower_address"],
        "End Cell ID": work["last_cell_id"],
        "End Address": work["last_tower_address"],
    })


def _extract_table(data: pd.DataFrame, bundle: dict[str, Any]) -> pd.DataFrame:
    voice_out = int(_voice_out_mask(data).sum())
    voice_in = int(_voice_in_mask(data).sum())
    sms_out = int(_sms_out_mask(data).sum())
    sms_in = int(_sms_in_mask(data).sum())
    rows: list[tuple[str, Any]] = [
        ("Total Records", len(data)),
        ("Unique Contacts", data["other_party"].replace("", pd.NA).dropna().nunique()),
        ("Outgoing Calls", voice_out),
        ("Incoming Calls", voice_in),
        ("Outgoing SMS", sms_out),
        ("Incoming SMS", sms_in),
        ("Total Call Duration (Sec)", int(data.loc[_voice_out_mask(data) | _voice_in_mask(data), "call_duration"].sum())),
        ("Unique IMEIs", data["imei"].replace("", pd.NA).dropna().nunique()),
        ("Unique IMSIs", data["imsi"].replace("", pd.NA).dropna().nunique()),
        ("Unique Cell IDs", data["first_cell_id"].replace("", pd.NA).dropna().nunique()),
    ]

    contacts = _contact_summary(data).head(10)
    for _, contact in contacts.iterrows():
        rows.append(("High-Frequency Contact", f"{contact['Other Party']} => Total Events = {contact['Total Calls']}, Duration = {contact['Total Duration']} sec"))

    errors = bundle.get("errors", {}) if isinstance(bundle, dict) else {}
    if errors:
        for name, message in errors.items():
            rows.append(("Analysis Warning", f"{name}: {message}"))

    return pd.DataFrame(rows, columns=["Header", "Details"])


def _write_dataframe_sheet(wb: Workbook, sheet_name: str, report_name: str, meta: dict[str, Any], frame: pd.DataFrame) -> None:
    ws = wb.create_sheet(title=sheet_name)
    headers = list(frame.columns)
    max_column = max(1, len(headers))
    header_row = style_metadata_block(ws, _metadata_rows(meta, report_name), max_column)

    if not headers:
        headers = ["Result"]
        frame = pd.DataFrame(columns=headers)

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=excel_safe_value(str(header)))
    style_table_header(ws, header_row, len(headers))

    for row_idx, row in enumerate(frame.itertuples(index=False, name=None), start=header_row + 1):
        for col_idx, value in enumerate(row, start=1):
            value = _excel_safe_scalar(value)
            ws.cell(row=row_idx, column=col_idx, value=value)

    last_row = header_row + len(frame)
    style_data_area(ws, header_row + 1, last_row, len(headers))
    set_sensible_widths(ws, headers)
    finish_sheet(ws, header_row, last_row, len(headers))

    # Date-time columns should remain readable.
    for col_idx, header in enumerate(headers, start=1):
        if "time" in str(header).lower() or str(header).lower() in {"date", "first call time", "last call time"}:
            for row_idx in range(header_row + 1, last_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if hasattr(cell.value, "year"):
                    cell.number_format = "mmm dd, yyyy hh:mm:ss"



MODULE_RESULT_SHEETS = [
    ("15. CDR Summary", ["cdr_summary"]),
    ("16. Top Human Contacts", ["top_contacts"]),
    ("17. Contact Category Summary", ["contact_category_summary"]),
    ("18. Service Sender IDs", ["top_service_sender_ids"]),
    ("19. Short Codes", ["top_short_codes"]),
    ("20. Contact Ranking", ["contact_ranking"]),
    ("21. Call Type Summary", ["incoming_outgoing"]),
    ("22. Other Unknown Call Types", ["other_call_type_summary"]),
    ("23. Social Network", ["social_network"]),
    ("24. Location Overview", ["analyze_location"]),
    ("25. Frequent Towers", ["frequent_locations"]),
    ("26. Tower Movement", ["tower_movement"]),
    ("27. Tower Transition", ["tower_transition"]),
    ("28. Movement Pattern", ["movement_pattern"]),
    ("29. Tower Intelligence", ["tower_intelligence"]),
    ("30. Home Tower", ["home_tower"]),
    ("31. Work Tower", ["work_tower"]),
    ("32. Missing CGI Lookup", ["missing_cgi_lookup"]),
    ("33. IMEI Module Summary", ["imei_summary"]),
    ("34. IMEI Intelligence", ["imei_intelligence"]),
    ("35. SIM Changes", ["sim_change"]),
    ("36. Activity Summary", ["activity_summary", "analyze_activity"]),
    ("37. Hourly Activity", ["hourly_activity"]),
    ("38. Daily Activity", ["daily_activity"]),
    ("39. Weekly Activity", ["weekly_activity"]),
    ("40. Monthly Activity", ["monthly_activity"]),
    ("41. Behavioral Observations", ["behavioral_intelligence"]),
    ("42. Review Indicators", ["suspicious_activity"]),
    ("43. Top Contact Details", ["top_contact_details"]),
]


def _excel_safe_scalar(value: Any) -> Any:
    """Compatibility wrapper around the shared Excel security boundary."""
    return excel_safe_value(value)


def _result_as_dataframe(value: Any) -> pd.DataFrame:
    """Convert scalar, tabular and nested analysis results to one table."""
    if isinstance(value, pd.DataFrame):
        return value.copy()
    if isinstance(value, pd.Series):
        return value.reset_index()
    if value is None:
        return pd.DataFrame(columns=["Result"])

    if isinstance(value, dict):
        sections: list[pd.DataFrame] = []
        scalar_rows: list[dict[str, Any]] = []
        for section, item in value.items():
            if isinstance(item, pd.DataFrame):
                frame = item.copy()
                frame.insert(0, "Section", str(section))
                sections.append(frame)
            elif isinstance(item, pd.Series):
                frame = item.reset_index()
                frame.insert(0, "Section", str(section))
                sections.append(frame)
            elif isinstance(item, list) and item and all(isinstance(row, dict) for row in item):
                frame = pd.DataFrame(item)
                frame.insert(0, "Section", str(section))
                sections.append(frame)
            else:
                scalar_rows.append({"Section": str(section), "Metric": str(section), "Value": _excel_safe_scalar(item)})
        if scalar_rows:
            sections.insert(0, pd.DataFrame(scalar_rows))
        return pd.concat(sections, ignore_index=True, sort=False) if sections else pd.DataFrame(columns=["Result"])

    if isinstance(value, list) and value and all(isinstance(row, dict) for row in value):
        return pd.DataFrame(value)
    if isinstance(value, (list, tuple, set)):
        return pd.DataFrame({"Result": [_excel_safe_scalar(item) for item in value]})
    return pd.DataFrame({"Result": [_excel_safe_scalar(value)]})

def _bundle_results(bundle: dict[str, Any]) -> dict[str, Any]:
    """Nested aur direct dono bundle formats support karta hai."""
    if not isinstance(bundle, dict):
        return {}
    results = bundle.get("results")
    return results if isinstance(results, dict) else bundle


def _append_module_result_sheets(
    workbook: Workbook,
    metadata: dict[str, Any],
    bundle: dict[str, Any],
) -> None:
    """Console analysis ke cached outputs ko additional Excel sheets mein likhta hai."""
    results = _bundle_results(bundle)

    for sheet_name, possible_keys in MODULE_RESULT_SHEETS:
        result = None
        for key in possible_keys:
            if key in results:
                result = results[key]
                break
        _write_dataframe_sheet(
            workbook,
            sheet_name,
            sheet_name.split(". ", 1)[-1],
            metadata,
            _result_as_dataframe(result),
        )

    status = bundle.get("status") if isinstance(bundle, dict) else None
    status_frame = status.copy() if isinstance(status, pd.DataFrame) else pd.DataFrame(
        columns=["Group", "Result Key", "Module", "Function", "Status", "Duration (sec)", "Error"]
    )
    _write_dataframe_sheet(
        workbook,
        "44. Analysis Status",
        "Analysis Function Status",
        metadata,
        status_frame,
    )

    errors = bundle.get("errors", {}) if isinstance(bundle, dict) else {}
    error_frame = (
        pd.DataFrame(
            [(name, message) for name, message in errors.items()],
            columns=["Module / Function", "Error"],
        )
        if isinstance(errors, dict)
        else pd.DataFrame(columns=["Module / Function", "Error"])
    )
    _write_dataframe_sheet(
        workbook,
        "45. Analysis Errors",
        "Analysis Errors",
        metadata,
        error_frame,
    )

def generate_single_cdr_report(
    df: pd.DataFrame,
    target: str,
    metadata: dict[str, Any] | None = None,
    analysis_bundle: dict[str, Any] | None = None,
    output_dir: str | Path | None = None,
) -> str | None:
    """
    Generate the sample-style workbook plus complete module-output sheets.

    The function returns the absolute output path. A failure is reported and
    None is returned, so the parent CLI does not crash.
    """
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        print("[-] Excel report generation aborted: DataFrame is empty or invalid.")
        return None

    try:
        data = _prepare_dataframe(df, str(target))
        meta = _metadata(str(target), data, metadata)
        if isinstance(analysis_bundle, dict):
            bundle = analysis_bundle
        else:
            from .analysis_bundle import build_single_analysis_bundle
            bundle = build_single_analysis_bundle(data, target=str(target))

        cc_summary = _contact_summary(data)
        cell_summary = _cell_summary(data)
        imei_summary = _imei_summary(data)
        movements = _movements(data)
        fclc = _fclc(data)
        fclc_summary = _fclc_summary(fclc)
        fclc_op = _fclc_op(data)
        moving_calls = _moving_calls(data)
        roaming_summary = _roaming_summary(data)

        workbook = Workbook()
        workbook.remove(workbook.active)

        _write_dataframe_sheet(workbook, "1. Extract", "Extract", meta, _extract_table(data, bundle))
        _write_dataframe_sheet(workbook, "2. CC Summary", "CC Summary", meta, cc_summary)
        _write_dataframe_sheet(workbook, "3. Cell ID Summary", "Cell ID Summary", meta, cell_summary)
        _write_dataframe_sheet(workbook, "4. IMEI Summary", "IMEI Summary", meta, imei_summary)
        _write_dataframe_sheet(workbook, "5. Movements", "Movements", meta, movements)
        _write_dataframe_sheet(workbook, "6. FCLC", "FCLC", meta, fclc)
        _write_dataframe_sheet(workbook, "7. FCLC Summary", "FCLC Summary", meta, fclc_summary)
        _write_dataframe_sheet(workbook, "8. FCLC OP", "FCLC OP", meta, fclc_op)
        _write_dataframe_sheet(workbook, "9. Moving Calls", "Moving Calls", meta, moving_calls)
        _write_dataframe_sheet(workbook, "10. Roaming Summary", "Roaming Summary", meta, roaming_summary)
        _write_dataframe_sheet(workbook, "11. Outgoing Calls", "Outgoing Calls", meta, _direction_sheet(data, _voice_out_mask(data)))
        _write_dataframe_sheet(workbook, "12. Incoming Calls", "Incoming Calls", meta, _direction_sheet(data, _voice_in_mask(data)))
        _write_dataframe_sheet(workbook, "13. Outgoing SMS", "Outgoing SMS", meta, _direction_sheet(data, _sms_out_mask(data)))
        _write_dataframe_sheet(workbook, "14. Incoming SMS", "Incoming SMS", meta, _direction_sheet(data, _sms_in_mask(data)))

        # Cached console-analysis results are rendered here. No analysis module
        # is executed again during Excel generation.
        _append_module_result_sheets(workbook, meta, bundle)
        rejected_rows = df.attrs.get("rejected_rows", pd.DataFrame())
        _write_dataframe_sheet(
            workbook,
            "46. Rejected Rows",
            "Rejected / Quarantined Source Rows",
            meta,
            rejected_rows if isinstance(rejected_rows, pd.DataFrame) else pd.DataFrame(),
        )

        path = get_single_report_path(target, output_dir)
        append_methodology_sheet(workbook, "Single CDR Analysis")
        workbook.save(path)
        print(f"[+] Single CDR Excel report generated: {path}")
        return str(path)

    except Exception as error:
        print("[-] Single CDR Excel report generation failed.")
        print(f"    Error Type : {type(error).__name__}")
        print(f"    Message    : {error}")
        print(traceback.format_exc(limit=4).rstrip())
        return None
