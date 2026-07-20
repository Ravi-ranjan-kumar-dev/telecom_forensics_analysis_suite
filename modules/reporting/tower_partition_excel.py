"""Consolidated Excel report for Tower Dump date-time partition analysis.

This module renders an already-computed partition result. It does not execute
the analytical functions again.
"""

from __future__ import annotations

from modules.core.time_utils import utc_now, utc_now_iso

import math
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .excel_security import excel_safe_value
from .report_guidance import append_methodology_sheet


EXCEL_MAX_ROWS = 1_048_576
MAX_DATA_ROWS_PER_SHEET = 1_000_000

TITLE_FILL = PatternFill("solid", fgColor="17365D")
SECTION_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
SUBHEADER_FILL = PatternFill("solid", fgColor="E2F0D9")
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
ERROR_FILL = PatternFill("solid", fgColor="FCE4D6")
SUCCESS_FILL = PatternFill("solid", fgColor="E2F0D9")
WHITE_FONT = Font(color="FFFFFF", bold=True)
HEADER_FONT = Font(color="000000", bold=True)
TITLE_FONT = Font(color="FFFFFF", bold=True, size=16)
THIN_SIDE = Side(style="thin", color="B7C9DB")
THIN_BORDER = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=THIN_SIDE,
    bottom=THIN_SIDE,
)


def _safe_scalar(value: Any) -> Any:
    """Compatibility wrapper around the shared Excel security boundary."""
    return excel_safe_value(value)


def _safe_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_-]+", "_", str(value).strip())
    cleaned = cleaned.strip("_")
    return cleaned or "CASE"


def _natural_partition_key(value: str) -> tuple[int, str]:
    text = str(value)
    match = re.fullmatch(r"[PS](\d+)", text.upper())

    if match:
        return int(match.group(1)), text

    return 10**9, text


def _unique_sheet_name(
    workbook: Workbook,
    requested: str,
) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", " ", str(requested))
    cleaned = re.sub(r"\s+", " ", cleaned).strip() or "Sheet"
    cleaned = cleaned[:31]

    existing = {sheet.title.lower() for sheet in workbook.worksheets}

    if cleaned.lower() not in existing:
        return cleaned

    suffix = 2

    while True:
        tail = f" {suffix}"
        candidate = cleaned[: 31 - len(tail)] + tail

        if candidate.lower() not in existing:
            return candidate

        suffix += 1


def _apply_title(
    worksheet,
    title: str,
    subtitle: str = "",
    *,
    column_count: int = 8,
) -> int:
    last_column = max(2, column_count)
    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=last_column,
    )
    title_cell = worksheet.cell(1, 1, title)
    title_cell.fill = TITLE_FILL
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )
    worksheet.row_dimensions[1].height = 28

    row = 2

    if subtitle:
        worksheet.merge_cells(
            start_row=2,
            start_column=1,
            end_row=2,
            end_column=last_column,
        )
        subtitle_cell = worksheet.cell(2, 1, subtitle)
        subtitle_cell.font = Font(italic=True, color="44546A")
        subtitle_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        worksheet.row_dimensions[2].height = 24
        row = 4
    else:
        row = 3

    return row


def _set_reasonable_widths(
    worksheet,
    *,
    min_width: int = 10,
    max_width: int = 42,
) -> None:
    for cells in worksheet.columns:
        first = cells[0]
        column_letter = get_column_letter(first.column)
        maximum = 0

        for cell in cells[:5000]:
            value = cell.value

            if value is None:
                continue

            maximum = max(maximum, len(str(value)))

        worksheet.column_dimensions[column_letter].width = min(
            max(maximum + 2, min_width),
            max_width,
        )


def _style_table_header(worksheet, row: int, columns: int) -> None:
    for column in range(1, columns + 1):
        cell = worksheet.cell(row, column)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    worksheet.row_dimensions[row].height = 28


def _write_key_values(
    worksheet,
    start_row: int,
    rows: Iterable[tuple[str, Any]],
) -> int:
    current_row = start_row

    for key, value in rows:
        key_cell = worksheet.cell(current_row, 1, str(key))
        value_cell = worksheet.cell(
            current_row,
            2,
            _safe_scalar(value),
        )

        key_cell.fill = HEADER_FILL
        key_cell.font = HEADER_FONT
        key_cell.border = THIN_BORDER
        value_cell.border = THIN_BORDER
        value_cell.alignment = Alignment(
            vertical="top",
            wrap_text=True,
        )
        current_row += 1

    worksheet.column_dimensions["A"].width = 30
    worksheet.column_dimensions["B"].width = 55
    return current_row


def _normalize_dataframe(
    value: Any,
    columns: list[str] | None = None,
) -> pd.DataFrame:
    if isinstance(value, pd.DataFrame):
        dataframe = value.copy()
    elif isinstance(value, list):
        dataframe = pd.DataFrame(value)
    elif isinstance(value, dict):
        dataframe = pd.DataFrame([value])
    else:
        dataframe = pd.DataFrame()

    if columns is not None:
        for column in columns:
            if column not in dataframe.columns:
                dataframe[column] = ""

        dataframe = dataframe[columns]

    return dataframe



# VISITOR_REPORT_DATAFRAME_HELPER

def _visitor_report_dataframe(
    value: Any,
) -> pd.DataFrame:
    """Prepare simple investigator-facing visitor columns."""

    dataframe = _normalize_dataframe(value)

    rename_map = {
        "partition_id": "Partition",
        "partition_location": "Location",
        "partition_window_start": "Start Date-Time",
        "partition_window_end": "End Date-Time",
        "partition_cgi_group_id": "CGI Group",
        "scope_mode": "Scope Mode",
        "scope_confidence": "Scope Confidence",
        "location_confirmed": "Location Confirmed",
        "scope_basis": "Scope Basis",
        "resolved_cell_count": "Resolved Cell Count",
        "resolved_cells": "Resolved Cells",
        "loaded_cell_count": "Loaded Cell Count",
        "subscriber_number": "Mobile Number",
        "sdr_found": "SDR Found",
        "sdr_subscriber_name": "Subscriber Name",
        "sdr_father_name": "Father / Husband Name",
        "sdr_address": "Full Address",
        "sdr_operator": "SDR Operator",
        "sdr_circle": "Circle",
        "sdr_activation_date": "Activation Date",
        "sdr_caf_number": "CAF Number",
        "visitor_type": "Visitor Type",
        "current_seen_count": "Current Events",
        "baseline_seen_count": "Baseline Events",
        "cells_seen": "Cells Seen",
        "imei_count": "IMEI Count",
        "imsi_count": "IMSI Count",
        "first_seen": "First Seen",
        "last_seen": "Last Seen",
        "rarity_score": "Rarity Score",
        "priority": "Priority",
        "confidence": "Confidence",
        "multi_cell_relevant": "Multi-Cell",
        "why_important": "Why It Matters",
        "next_verification": "Suggested Verification",
        "rank_reason": "Ranking Reason",
        "investigation_hint": "Investigation Hint",
    }

    dataframe = dataframe.rename(
        columns={
            source: target
            for source, target in rename_map.items()
            if source in dataframe.columns
        }
    )

    preferred_columns = [
        "Partition",
        "Location",
        "Start Date-Time",
        "End Date-Time",
        "CGI Group",
        "Scope Mode",
        "Scope Confidence",
        "Location Confirmed",
        "Scope Basis",
        "Resolved Cell Count",
        "Resolved Cells",
        "Loaded Cell Count",
        "Mobile Number",
        "SDR Found",
        "Subscriber Name",
        "Father / Husband Name",
        "Full Address",
        "SDR Operator",
        "Circle",
        "Activation Date",
        "CAF Number",
        "Visitor Type",
        "Current Events",
        "Baseline Events",
        "Cells Seen",
        "IMEI Count",
        "IMSI Count",
        "First Seen",
        "Last Seen",
        "Rarity Score",
        "Priority",
        "Confidence",
        "Multi-Cell",
        "Why It Matters",
        "Suggested Verification",
    ]

    existing = [
        column
        for column in preferred_columns
        if column in dataframe.columns
    ]

    # INVESTIGATOR_ONLY_VISITOR_COLUMNS
    # Internal engine fields are intentionally hidden from the
    # user-facing report to keep the output simple and consistent.
    return dataframe[
        existing
    ].copy()



# VISITOR_SDR_BATCH_ENRICHMENT

VISITOR_SDR_COLUMNS = [
    "sdr_found",
    "sdr_subscriber_name",
    "sdr_father_name",
    "sdr_address",
    "sdr_operator",
    "sdr_circle",
    "sdr_activation_date",
    "sdr_caf_number",
]


def _build_visitor_sdr_lookup(
    visitor_dataframe: Any,
) -> pd.DataFrame:
    """Run one canonical batch SDR lookup for all unique visitors."""

    from modules.enrichment.sdr_subscriber_enrichment import (
        lookup_sdr_subscribers,
        normalize_mobile_number,
    )

    dataframe = _normalize_dataframe(
        visitor_dataframe
    )

    output_columns = [
        "_sdr_lookup_mobile",
        *VISITOR_SDR_COLUMNS,
    ]

    if (
        dataframe.empty
        or "subscriber_number" not in dataframe.columns
    ):
        return pd.DataFrame(
            columns=output_columns
        )

    normalized_numbers = (
        dataframe["subscriber_number"]
        .map(normalize_mobile_number)
    )

    numbers = sorted(
        {
            str(number).strip()
            for number in normalized_numbers
            if str(number).strip()
        }
    )

    if not numbers:
        return pd.DataFrame(
            columns=output_columns
        )

    lookup = lookup_sdr_subscribers(
        numbers
    )

    if lookup is None or lookup.empty:
        return pd.DataFrame(
            columns=output_columns
        )

    lookup = lookup.rename(
        columns={
            "lookup_mobile": "_sdr_lookup_mobile",
            "sdr_found": "sdr_found",
            "subscriber_name": "sdr_subscriber_name",
            "father_name": "sdr_father_name",
            "subscriber_address": "sdr_address",
            "operator": "sdr_operator",
            "circle": "sdr_circle",
            "activation_date": "sdr_activation_date",
            "caf_number": "sdr_caf_number",
        }
    )

    for column in output_columns:
        if column not in lookup.columns:
            lookup[column] = (
                "No"
                if column == "sdr_found"
                else ""
            )

    lookup["_sdr_lookup_mobile"] = (
        lookup["_sdr_lookup_mobile"]
        .map(normalize_mobile_number)
    )

    lookup = (
        lookup[
            output_columns
        ]
        .drop_duplicates(
            subset=["_sdr_lookup_mobile"],
            keep="first",
        )
        .reset_index(drop=True)
    )

    lookup["sdr_found"] = (
        lookup["sdr_found"]
        .fillna("No")
        .astype(str)
        .str.strip()
        .replace(
            {
                "": "No",
                "nan": "No",
                "None": "No",
            }
        )
    )

    for column in VISITOR_SDR_COLUMNS:
        if column == "sdr_found":
            continue

        lookup[column] = (
            lookup[column]
            .fillna("")
        )

    return lookup


def _enrich_visitor_dataframe_with_sdr(
    visitor_dataframe: Any,
    sdr_lookup: pd.DataFrame,
) -> pd.DataFrame:
    """Merge a previously computed SDR lookup into one visitor table."""

    from modules.enrichment.sdr_subscriber_enrichment import (
        normalize_mobile_number,
    )

    dataframe = _normalize_dataframe(
        visitor_dataframe
    )

    if (
        dataframe.empty
        or "subscriber_number" not in dataframe.columns
    ):
        for column in VISITOR_SDR_COLUMNS:
            if column not in dataframe.columns:
                dataframe[column] = (
                    "No"
                    if column == "sdr_found"
                    else ""
                )

        return dataframe

    output = dataframe.drop(
        columns=VISITOR_SDR_COLUMNS,
        errors="ignore",
    ).copy()

    output["_sdr_lookup_mobile"] = (
        output["subscriber_number"]
        .map(normalize_mobile_number)
    )

    if (
        isinstance(sdr_lookup, pd.DataFrame)
        and not sdr_lookup.empty
    ):
        output = output.merge(
            sdr_lookup,
            on="_sdr_lookup_mobile",
            how="left",
            validate="many_to_one",
        )
    else:
        for column in VISITOR_SDR_COLUMNS:
            output[column] = (
                "No"
                if column == "sdr_found"
                else ""
            )

    output["sdr_found"] = (
        output["sdr_found"]
        .fillna("No")
        .astype(str)
        .str.strip()
        .replace(
            {
                "": "No",
                "nan": "No",
                "None": "No",
            }
        )
    )

    for column in VISITOR_SDR_COLUMNS:
        if column == "sdr_found":
            continue

        if column not in output.columns:
            output[column] = ""

        output[column] = (
            output[column]
            .fillna("")
        )

    return output.drop(
        columns=["_sdr_lookup_mobile"],
        errors="ignore",
    )


def _visitor_sdr_summary_counts(
    visitor_dataframe: Any,
    sdr_lookup: pd.DataFrame,
) -> dict[str, int]:
    """Return unique visitor SDR coverage counts."""

    from modules.enrichment.sdr_subscriber_enrichment import (
        normalize_mobile_number,
    )

    dataframe = _normalize_dataframe(
        visitor_dataframe
    )

    if (
        dataframe.empty
        or "subscriber_number" not in dataframe.columns
    ):
        return {
            "unique_visitors": 0,
            "sdr_found": 0,
            "sdr_not_found": 0,
        }

    unique_numbers = {
        normalized
        for normalized in (
            dataframe["subscriber_number"]
            .map(normalize_mobile_number)
        )
        if normalized
    }

    found_numbers: set[str] = set()

    if (
        isinstance(sdr_lookup, pd.DataFrame)
        and not sdr_lookup.empty
        and "_sdr_lookup_mobile" in sdr_lookup.columns
    ):
        found_mask = (
            sdr_lookup.get(
                "sdr_found",
                pd.Series(
                    "No",
                    index=sdr_lookup.index,
                ),
            )
            .fillna("No")
            .astype(str)
            .str.strip()
            .str.upper()
            .eq("YES")
        )

        found_numbers = {
            str(value).strip()
            for value in sdr_lookup.loc[
                found_mask,
                "_sdr_lookup_mobile",
            ]
            if str(value).strip()
        }

    found_count = len(
        unique_numbers.intersection(
            found_numbers
        )
    )

    return {
        "unique_visitors": len(unique_numbers),
        "sdr_found": found_count,
        "sdr_not_found": max(
            len(unique_numbers) - found_count,
            0,
        ),
    }



def _write_dataframe_rows(
    worksheet,
    dataframe: pd.DataFrame,
    *,
    title: str,
    subtitle: str,
) -> None:
    column_count = max(2, len(dataframe.columns))
    header_row = _apply_title(
        worksheet,
        title,
        subtitle,
        column_count=column_count,
    )

    if dataframe.empty:
        worksheet.cell(header_row, 1, "No records found.")
        worksheet.cell(header_row, 1).fill = WARNING_FILL
        worksheet.cell(header_row, 1).font = HEADER_FONT
        worksheet.column_dimensions["A"].width = 36
        worksheet.freeze_panes = f"A{header_row + 1}"
        return

    headers = [str(column) for column in dataframe.columns]

    for column_number, header in enumerate(headers, start=1):
        worksheet.cell(header_row, column_number, _safe_scalar(header))

    _style_table_header(
        worksheet,
        header_row,
        len(headers),
    )

    for row_number, record in enumerate(
        dataframe.itertuples(index=False, name=None),
        start=header_row + 1,
    ):
        for column_number, value in enumerate(record, start=1):
            cell = worksheet.cell(
                row_number,
                column_number,
                _safe_scalar(value),
            )
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=False,
            )

            if isinstance(cell.value, (datetime, date)):
                cell.number_format = "yyyy-mm-dd hh:mm:ss"

    final_row = header_row + len(dataframe)
    final_column = get_column_letter(len(headers))
    worksheet.auto_filter.ref = (
        f"A{header_row}:{final_column}{final_row}"
    )
    worksheet.freeze_panes = f"A{header_row + 1}"
    worksheet.sheet_view.showGridLines = False
    _set_reasonable_widths(worksheet)


def _write_dataframe_sheets(
    workbook: Workbook,
    base_title: str,
    dataframe: pd.DataFrame,
    *,
    subtitle: str,
) -> list[str]:
    """Write a DataFrame, splitting across sheets near Excel's row limit."""

    dataframe = _normalize_dataframe(dataframe)

    if dataframe.empty:
        sheet_name = _unique_sheet_name(workbook, base_title)
        worksheet = workbook.create_sheet(sheet_name)
        _write_dataframe_rows(
            worksheet,
            dataframe,
            title=base_title,
            subtitle=subtitle,
        )
        return [sheet_name]

    page_count = max(
        1,
        math.ceil(len(dataframe) / MAX_DATA_ROWS_PER_SHEET),
    )
    names: list[str] = []

    for page_index in range(page_count):
        start = page_index * MAX_DATA_ROWS_PER_SHEET
        stop = min(
            start + MAX_DATA_ROWS_PER_SHEET,
            len(dataframe),
        )
        page = dataframe.iloc[start:stop].reset_index(drop=True)

        requested = (
            base_title
            if page_count == 1
            else f"{base_title} {page_index + 1}"
        )
        sheet_name = _unique_sheet_name(workbook, requested)
        worksheet = workbook.create_sheet(sheet_name)

        page_subtitle = subtitle

        if page_count > 1:
            page_subtitle = (
                f"{subtitle} | Part {page_index + 1} of {page_count} | "
                f"Rows {start + 1:,} to {stop:,}"
            )

        _write_dataframe_rows(
            worksheet,
            page,
            title=base_title,
            subtitle=page_subtitle,
        )
        names.append(sheet_name)

    return names


def _partition_columns(dataframe: pd.DataFrame) -> list[str]:
    return sorted(
        [
            str(column)
            for column in dataframe.columns
            if re.fullmatch(r"[PS]\d+", str(column), re.IGNORECASE)
        ],
        key=_natural_partition_key,
    )


def _candidate_matrix(
    subscriber_presence: pd.DataFrame,
) -> pd.DataFrame:
    dataframe = _normalize_dataframe(subscriber_presence)

    if dataframe.empty:
        return dataframe

    identity_columns = [
        column
        for column in (
            "subscriber_number",
            "match_count",
            "total_sightings",
            "match_ratio",
            "total_events",
            "operators",
            "first_seen",
            "last_seen",
        )
        if column in dataframe.columns
    ]
    partition_columns = _partition_columns(dataframe)
    return dataframe[identity_columns + partition_columns].copy()


def _status_dataframe(
    *,
    result: dict[str, Any],
    saved: dict[str, Any] | None,
    report_path: Path,
) -> pd.DataFrame:
    rows = [
        {
            "Stage": "Tower Dump Loading",
            "Status": "Completed",
            "Details": (
                f"{int(result.get('total_input_records', 0)):,} "
                "normalized records"
            ),
        },
        {
            "Stage": "Logical Partitioning",
            "Status": "Completed",
            "Details": (
                f"{int(result.get('total_sightings', 0))} "
                "dynamic partition window(s)"
            ),
        },
        {
            "Stage": "N-of-M Candidate Extraction",
            "Status": "Completed",
            "Details": (
                f"{len(_normalize_dataframe(result.get('n_of_m_candidates'))):,} "
                "candidate(s) in 2 or more partitions"
            ),
        },
        {
            "Stage": "Strict Common Extraction",
            "Status": "Completed",
            "Details": (
                f"{len(_normalize_dataframe(result.get('strict_common_candidates'))):,} "
                "candidate(s) in all partitions"
            ),
        },
        {
            "Stage": "Visitor Intelligence",
            "Status": "Completed",
            "Details": (
                f"{len(_normalize_dataframe(result.get('partition_visitor_intelligence'))):,} "
                "partition visitor classification row(s)"
            ),
        },
        {
            "Stage": "Internal Backend Tables",
            "Status": "Completed",
            "Details": (
                str(saved.get("run_directory", ""))
                if isinstance(saved, dict)
                else ""
            ),
        },
        {
            "Stage": "Consolidated Excel Report",
            "Status": "Completed",
            "Details": str(report_path),
        },
    ]
    return pd.DataFrame(rows)


def _warnings_dataframe(result: dict[str, Any]) -> pd.DataFrame:
    warnings = result.get("warnings", [])
    errors = result.get("errors", [])

    rows: list[dict[str, Any]] = []

    for item in warnings if isinstance(warnings, list) else [warnings]:
        if str(item).strip():
            rows.append(
                {
                    "Level": "WARNING",
                    "Message": str(item),
                }
            )

    for item in errors if isinstance(errors, list) else [errors]:
        if str(item).strip():
            rows.append(
                {
                    "Level": "ERROR",
                    "Message": str(item),
                }
            )

    if not rows:
        rows.append(
            {
                "Level": "INFO",
                "Message": "No load or partition warnings were reported.",
            }
        )

    return pd.DataFrame(rows)


def generate_tower_partition_excel_report(
    result: dict[str, Any],
    *,
    case: dict[str, Any],
    sightings: list[dict[str, Any]],
    output_dir: str | Path,
    input_folder: str | Path = "",
    saved: dict[str, Any] | None = None,
) -> Path:
    """Generate one user-facing workbook from all partition outputs."""

    if not isinstance(result, dict):
        raise TypeError("Partition result dictionary required hai.")

    output_path = Path(output_dir).expanduser().resolve()
    output_path.mkdir(parents=True, exist_ok=True)

    case_id = str(case.get("case_id", "")).strip() or "CASE"
    case_name = str(case.get("case_name", "")).strip()
    timestamp = utc_now().strftime("%Y%m%d_%H%M%S_%f")
    filename = (
        f"Tower_Dump_Partition_Analysis_"
        f"{_safe_filename(case_id)}_{timestamp}.xlsx"
    )
    report_path = output_path / filename

    workbook = Workbook()
    workbook.remove(workbook.active)

    partition_summary = _normalize_dataframe(
        result.get("partition_summary")
    )
    partition_status = _normalize_dataframe(
        result.get("partition_status")
    )
    rejected_rows = _normalize_dataframe(
        result.get("rejected_rows")
    )
    subscriber_presence = _normalize_dataframe(
        result.get("subscriber_presence")
    )
    n_of_m_candidates = _normalize_dataframe(
        result.get("n_of_m_candidates")
    )
    strict_common = _normalize_dataframe(
        result.get("strict_common_candidates")
    )
    imei_presence = _normalize_dataframe(
        result.get("imei_presence")
    )
    imsi_presence = _normalize_dataframe(
        result.get("imsi_presence")
    )
    matrix = _candidate_matrix(subscriber_presence)

    # ONE_BATCH_VISITOR_SDR_LOOKUP
    raw_visitor_intelligence = _normalize_dataframe(
        result.get("partition_visitor_intelligence")
    )

    visitor_sdr_lookup = _build_visitor_sdr_lookup(
        raw_visitor_intelligence
    )

    visitor_sdr_counts = _visitor_sdr_summary_counts(
        raw_visitor_intelligence,
        visitor_sdr_lookup,
    )

    visitor_intelligence = _visitor_report_dataframe(
        _enrich_visitor_dataframe_with_sdr(
            raw_visitor_intelligence,
            visitor_sdr_lookup,
        )
    )

    new_visitors = _visitor_report_dataframe(
        _enrich_visitor_dataframe_with_sdr(
            result.get("new_visitors"),
            visitor_sdr_lookup,
        )
    )

    rare_visitors = _visitor_report_dataframe(
        _enrich_visitor_dataframe_with_sdr(
            result.get("rare_visitors"),
            visitor_sdr_lookup,
        )
    )

    repeat_relevant = _visitor_report_dataframe(
        _enrich_visitor_dataframe_with_sdr(
            result.get("repeat_relevant_visitors"),
            visitor_sdr_lookup,
        )
    )

    regular_local = _visitor_report_dataframe(
        _enrich_visitor_dataframe_with_sdr(
            result.get("regular_local_presence"),
            visitor_sdr_lookup,
        )
    )

    multi_cell_relevant = _visitor_report_dataframe(
        _enrich_visitor_dataframe_with_sdr(
            result.get("multi_cell_relevant"),
            visitor_sdr_lookup,
        )
    )

    visitor_priority_leads = _visitor_report_dataframe(
        _enrich_visitor_dataframe_with_sdr(
            result.get("partition_priority_leads"),
            visitor_sdr_lookup,
        )
    )

    sighting_frame = _normalize_dataframe(sightings)

    # PARTITION_SCOPE_SUMMARY_COUNTS
    scope_modes = (
        partition_summary.get(
            "scope_mode",
            pd.Series(dtype="object"),
        )
        .fillna("")
        .astype(str)
        .str.strip()
        .str.upper()
    )

    scope_confidence = (
        partition_summary.get(
            "scope_confidence",
            pd.Series(dtype="object"),
        )
        .fillna("")
        .astype(str)
        .str.strip()
        .str.upper()
    )

    location_confirmed_values = (
        partition_summary.get(
            "location_confirmed",
            pd.Series(dtype="object"),
        )
        .fillna("")
        .astype(str)
        .str.strip()
        .str.upper()
    )

    auto_active_partitions = int(
        scope_modes.eq(
            "AUTO_ACTIVE_CELLS"
        ).sum()
    )

    location_scoped_partitions = int(
        scope_modes.eq(
            "LOCATION_SCOPED"
        ).sum()
    )

    high_scope_confidence = int(
        scope_confidence.eq("HIGH").sum()
    )

    medium_scope_confidence = int(
        scope_confidence.eq("MEDIUM").sum()
    )

    low_scope_confidence = int(
        scope_confidence.eq("LOW").sum()
    )

    confirmed_location_partitions = int(
        location_confirmed_values.eq("YES").sum()
    )

    # GENERIC_PARTITION_WINDOW_DISPLAY
    if not sighting_frame.empty:
        sighting_frame = sighting_frame.rename(
            columns={
                "sighting_id": "Partition",
                "location_name": "Location",
                "window_start": "Start Date-Time",
                "window_end": "End Date-Time",
                "cgi_group_id": "CGI Group",
                "source_types": "Source Types",
            }
        )

        preferred = [
            "Partition",
            "Location",
            "Start Date-Time",
            "End Date-Time",
            "CGI Group",
            "Source Types",
        ]

        legacy_hidden = {
            "cctv_timestamp",
            "minutes_before",
            "minutes_after",
        }

        existing = [
            column
            for column in preferred
            if column in sighting_frame.columns
        ]

        remaining = [
            column
            for column in sighting_frame.columns
            if column not in existing
            and column not in legacy_hidden
        ]

        sighting_frame = sighting_frame[
            existing + remaining
        ]

    # 1. Executive Summary
    overview = workbook.create_sheet(
        _unique_sheet_name(workbook, "1. Executive Summary")
    )
    start_row = _apply_title(
        overview,
        "Tower Dump Date-Time Partition Analysis",
        (
            f"Case: {case_id}"
            + (f" | {case_name}" if case_name else "")
        ),
        column_count=6,
    )

    load_metadata = result.get("load_metadata", {})
    operators = result.get("operators", [])
    cell_ids = result.get("cell_ids", [])

    _write_key_values(
        overview,
        start_row,
        [
            ("Case ID", case_id),
            ("Case Name", case_name),
            ("Analysis Type", "Pair-Based Date-Time Tower Dump Partitioning"),
            ("Generated At", utc_now_iso()),
            ("Input Folder", str(input_folder)),
            ("Total Input Records", int(result.get("total_input_records", 0))),
            ("Configured Sightings", int(result.get("total_configured_sightings", len(sightings)))),
            ("Valid Partitions", int(result.get("total_sightings", 0))),
            ("Auto Active-Cell Partitions", auto_active_partitions),
            ("Configured Location-Scoped Partitions", location_scoped_partitions),
            ("Location-Confirmed Partitions", confirmed_location_partitions),
            ("High Scope Confidence", high_scope_confidence),
            ("Medium Scope Confidence", medium_scope_confidence),
            ("Low Scope Confidence", low_scope_confidence),
            (
                "Scope Interpretation",
                (
                    "Location Confirmed = YES means a configured CGI group "
                    "was matched. AUTO_ACTIVE_CELLS means cells were inferred "
                    "from activity inside the selected Date-Time Part."
                ),
            ),
            ("Rejected Source Rows", len(rejected_rows)),
            ("Range Rule", "Start Date-Time <= Event Time < End Date-Time"),
            ("Operators", ", ".join(map(str, operators)) if operators else ""),
            ("Loaded Searched CGI/Cells", len(cell_ids) if isinstance(cell_ids, list) else 0),
            ("Files Found", load_metadata.get("files_found", "")),
            ("Files Loaded", load_metadata.get("files_loaded", "")),
            ("Files Failed", load_metadata.get("files_failed", "")),
            ("Unique Subscribers", len(subscriber_presence)),
            ("Candidates in 2+ Partitions", len(n_of_m_candidates)),
            ("Candidates in All Partitions", len(strict_common)),
            ("Unique IMEI Entries", len(imei_presence)),
            ("Unique IMSI Entries", len(imsi_presence)),
            ("Visitor Classification Rows", len(visitor_intelligence)),
            (
                "Unique Visitor Mobile Numbers",
                visitor_sdr_counts["unique_visitors"],
            ),
            (
                "Visitor SDR Profiles Found",
                visitor_sdr_counts["sdr_found"],
            ),
            (
                "Visitor SDR Profiles Not Found",
                visitor_sdr_counts["sdr_not_found"],
            ),
            (
                "Visitor SDR Lookup Method",
                (
                    "Single batch lookup across unique visitor numbers; "
                    "large DuckDB SDR table first, primary SDR table fallback."
                ),
            ),
            ("New Visitors", len(new_visitors)),
            ("Rare Visitors", len(rare_visitors)),
            ("Repeat Relevant Visitors", len(repeat_relevant)),
            ("Regular / Local Presence", len(regular_local)),
            ("Multi-Cell Relevant Visitors", len(multi_cell_relevant)),
            ("Priority Visitor Leads", len(visitor_priority_leads)),
            (
                "Backend Run Directory",
                saved.get("run_directory", "")
                if isinstance(saved, dict)
                else "",
            ),
        ],
    )

    overview.sheet_view.showGridLines = False
    overview.freeze_panes = "A4"

    # 2-11. All analytical outputs in the same workbook.
    _write_dataframe_sheets(
        workbook,
        "2. Partition Windows",
        sighting_frame,
        subtitle="User-entered exact Start and End Date-Time pairs using half-open ranges.",
    )
    _write_dataframe_sheets(
        workbook,
        "3. Partition Summary",
        partition_summary,
        subtitle="Record and identifier counts for every dynamic partition.",
    )
    _write_dataframe_sheets(
        workbook,
        "4. Subscriber Presence",
        subscriber_presence,
        subtitle="All subscriber identities and their partition-wise presence.",
    )
    _write_dataframe_sheets(
        workbook,
        "5. N-of-M Candidates",
        n_of_m_candidates,
        subtitle="Subscribers present in two or more configured partitions.",
    )
    _write_dataframe_sheets(
        workbook,
        "6. Strict Common",
        strict_common,
        subtitle="Subscribers present in every configured partition.",
    )
    _write_dataframe_sheets(
        workbook,
        "7. IMEI Continuity",
        imei_presence,
        subtitle="Device continuity across partitions using IMEI.",
    )
    _write_dataframe_sheets(
        workbook,
        "8. IMSI Continuity",
        imsi_presence,
        subtitle="SIM/subscriber identity continuity across partitions using IMSI.",
    )
    _write_dataframe_sheets(
        workbook,
        "9. Candidate Matrix",
        matrix,
        subtitle="Dynamic P1/P2/... or S1/S2/... presence matrix.",
    )

    # PARTITION_VISITOR_INTELLIGENCE_SHEETS
    _write_dataframe_sheets(
        workbook,
        "10. Visitor Intelligence",
        visitor_intelligence,
        subtitle=(
            "All selected-period visitors classified against the same CGI "
            "scope outside each partition time range."
        ),
    )
    _write_dataframe_sheets(
        workbook,
        "11. New Visitors",
        new_visitors,
        subtitle=(
            "Numbers present in the selected partition but absent from "
            "the same CGI-scope baseline."
        ),
    )
    _write_dataframe_sheets(
        workbook,
        "12. Rare Visitors",
        rare_visitors,
        subtitle=(
            "Numbers with only one or two baseline events in the same CGI scope."
        ),
    )
    _write_dataframe_sheets(
        workbook,
        "13. Repeat Relevant",
        repeat_relevant,
        subtitle=(
            "Numbers with low baseline presence and repeated activity "
            "inside the selected partition."
        ),
    )
    _write_dataframe_sheets(
        workbook,
        "14. Regular Local",
        regular_local,
        subtitle=(
            "Numbers with established baseline presence; generally not new visitors."
        ),
    )
    _write_dataframe_sheets(
        workbook,
        "15. Multi-Cell Relevant",
        multi_cell_relevant,
        subtitle=(
            "Selected-period visitors observed across multiple relevant cells."
        ),
    )
    _write_dataframe_sheets(
        workbook,
        "16. Visitor Priority Leads",
        visitor_priority_leads,
        subtitle=(
            "Visitor leads ranked by rarity, current activity, cell presence "
            "and device/SIM consistency."
        ),
    )

    status_placeholder = output_path / filename
    status_frame = _status_dataframe(
        result=result,
        saved=saved,
        report_path=status_placeholder,
    )
    status_names = _write_dataframe_sheets(
        workbook,
        "17. Analysis Status",
        status_frame,
        subtitle="Execution status for the consolidated partition workflow.",
    )

    warning_frame = _warnings_dataframe(result)
    warning_names = _write_dataframe_sheets(
        workbook,
        "18. Warnings",
        warning_frame,
        subtitle="Load, normalization and partition warnings.",
    )
    _write_dataframe_sheets(
        workbook,
        "19. Partition Status",
        partition_status,
        subtitle="Validation, source-type and CGI scope status for every configured partition.",
    )
    _write_dataframe_sheets(
        workbook,
        "20. Rejected Rows",
        rejected_rows,
        subtitle="Malformed or non-data source rows quarantined with physical line provenance.",
    )

    # Highlight status/warning levels.
    for name in status_names:
        worksheet = workbook[name]
        for row in range(5, worksheet.max_row + 1):
            status = str(worksheet.cell(row, 2).value or "").upper()
            if status == "COMPLETED":
                worksheet.cell(row, 2).fill = SUCCESS_FILL

    for name in warning_names:
        worksheet = workbook[name]
        for row in range(5, worksheet.max_row + 1):
            level = str(worksheet.cell(row, 1).value or "").upper()

            if level == "ERROR":
                worksheet.cell(row, 1).fill = ERROR_FILL
            elif level == "WARNING":
                worksheet.cell(row, 1).fill = WARNING_FILL
            else:
                worksheet.cell(row, 1).fill = SUCCESS_FILL

    workbook.properties.title = (
        f"Tower Dump Partition Analysis - {case_id}"
    )
    workbook.properties.subject = (
        "Date-time pair driven Tower Dump partition and visitor intelligence analysis"
    )
    workbook.properties.creator = "Telecom Forensics Analysis Suite"
    workbook.properties.description = (
        "Consolidated user-facing report. Internal CSV files remain backend data."
    )

    append_methodology_sheet(workbook, "Date-Time Tower Dump Partition Analysis")

    workbook.save(report_path)
    return report_path
