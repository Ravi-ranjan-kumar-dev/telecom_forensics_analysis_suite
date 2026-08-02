"""Reusable investigator-facing Excel report viewer."""

from __future__ import annotations

import re
from itertools import islice
from pathlib import Path
from typing import Any, Final

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from PySide6.QtCore import QSignalBlocker, Qt, QUrl
from PySide6.QtGui import QColor, QCloseEvent, QDesktopServices
from PySide6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QComboBox,
    QDialog,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from modules.reporting.report_section_reader import (
    ReportSection,
    discover_report_sections,
    read_report_section_rows,
)

__all__ = [
    "RelatedRecordsDialog",
    "ReportViewerDialog",
    "canonical_phone_number",
    "detect_identifier_columns",
    "detect_number_columns",
    "prepare_related_records",
]


_NUMBER_HEADER_PATTERN: Final[re.Pattern[str]] = re.compile(
    r"(?:^|\b)(?:contact|other party|common number|b[ -]?party|a[ -]?party|"
    r"source target|destination target|target(?: number)?|msisdn|"
    r"subscriber(?: number)?|mobile(?: number)?|phone(?: number)?)(?:\b|$)",
    re.IGNORECASE,
)
_CELL_HEADER_PATTERN: Final[re.Pattern[str]] = re.compile(
    r"(?:^|\b)(?:cell(?: site)? id|cgi|towers?)(?:\b|$)",
    re.IGNORECASE,
)
_IMEI_HEADER_PATTERN: Final[re.Pattern[str]] = re.compile(
    r"(?:^|\b)(?:imei|old imei|new imei|device id)(?:\b|$)", re.IGNORECASE
)
_IMSI_HEADER_PATTERN: Final[re.Pattern[str]] = re.compile(
    r"(?:^|\b)(?:imsi|old imsi|new imsi|sim id)(?:\b|$)", re.IGNORECASE
)
_NON_IDENTIFIER_HEADER_PATTERN: Final[re.Pattern[str]] = re.compile(
    r"(?:status|address|town|district|latitude|longitude|count|events?|"
    r"records?|duration|name|valid|invalid|found|unique|total)",
    re.IGNORECASE,
)
_MAX_VISIBLE_ROWS: Final[int] = 500
_HEADER_SCAN_ROWS: Final[int] = 25
_IDENTIFIER_LABELS: Final[dict[str, str]] = {
    "phone": "Phone Number",
    "cell_id": "Cell ID",
    "imei": "IMEI",
    "imsi": "IMSI",
}
_MINIMUM_IDENTIFIER_DIGITS: Final[dict[str, int]] = {
    "phone": 8,
    "cell_id": 4,
    "imei": 14,
    "imsi": 14,
}

_RELATED_RECORD_COLUMNS: Final[tuple[tuple[str, str], ...]] = (
    ("a_party", "Target Number"),
    ("target_number", "Target Number"),
    ("b_party", "Other Party"),
    ("call_date", "Date"),
    ("call_time", "Time"),
    ("call_type", "Event Type"),
    ("call_duration", "Duration (Seconds)"),
    ("first_cell_id", "First Cell ID"),
    ("first_bts_location", "First Tower Location"),
    ("last_cell_id", "Last Cell ID"),
    ("last_bts_location", "Last Tower Location"),
    ("imei", "IMEI"),
    ("imsi", "IMSI"),
    ("roaming_network", "Roaming Network"),
    ("service_type", "Service Type"),
)


class RelatedRecordsDialog(QDialog):
    """Display verified source Call/SMS records for one telecom identifier."""

    def __init__(
        self,
        identifier: str,
        records: Any,
        parent: QWidget | None = None,
        identifier_label: str = "Number",
    ) -> None:
        super().__init__(parent)
        self.setWindowTitle(
            f"Related {identifier_label} Records - {identifier}"
        )
        self.resize(1280, 760)
        self.setModal(True)

        display_records, duplicates_hidden = prepare_related_records(records)

        title = QLabel(f"Related records for {identifier_label}: {identifier}")
        title.setObjectName("moduleTitle")
        duplicate_note = (
            f" Exact duplicates hidden: {duplicates_hidden:,}."
            if duplicates_hidden
            else ""
        )
        metadata = getattr(records, "attrs", {})
        result_limited = bool(metadata.get("result_limited", False))
        limit_note = ""
        if result_limited:
            try:
                result_limit = max(
                    1,
                    int(metadata.get("result_limit", len(records))),
                )
            except (TypeError, ValueError):
                result_limit = len(records)

            minimum_matches = max(
                result_limit + 1,
                len(records) + 1,
            )
            limit_note = (
                " Query limit reached: at least "
                f"{minimum_matches:,} matching source records were found; "
                f"the first {len(records):,} source records were loaded."
            )

        status = QLabel(
            f"Verified records shown: {len(display_records):,}."
            f"{duplicate_note}{limit_note} "
            "Source evidence was read without modification."
        )
        status.setWordWrap(True)

        table = QTableWidget()
        table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        table.setAlternatingRowColors(True)
        table.verticalHeader().setVisible(False)
        columns = list(map(str, display_records.columns))
        table.setColumnCount(len(columns))
        table.setHorizontalHeaderLabels(columns)
        table.setRowCount(len(display_records))
        for row_index, row in enumerate(
            display_records.itertuples(index=False, name=None)
        ):
            for column_index, value in enumerate(row):
                table.setItem(
                    row_index,
                    column_index,
                    QTableWidgetItem(_clean_cell(value)),
                )
        table.setSortingEnabled(True)
        table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Interactive
        )
        table.horizontalHeader().setDefaultSectionSize(150)

        close_button = QPushButton("Close")
        close_button.setObjectName("primaryButton")
        close_button.clicked.connect(self.accept)
        controls = QHBoxLayout()
        controls.addStretch(1)
        controls.addWidget(close_button)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 22, 24, 22)
        layout.setSpacing(12)
        layout.addWidget(title)
        layout.addWidget(status)
        layout.addWidget(table, stretch=1)
        layout.addLayout(controls)

        self._table = table
        self._status = status


def prepare_related_records(records: Any) -> tuple[Any, int]:
    """Build a clean display copy and count exact duplicate source rows."""

    source = records.copy()
    duplicate_mask = source.duplicated(keep="first")
    duplicates_hidden = int(duplicate_mask.sum())
    unique_records = source.loc[~duplicate_mask].copy()

    selected: list[str] = []
    labels: list[str] = []
    used_labels: set[str] = set()
    for source_name, display_label in _RELATED_RECORD_COLUMNS:
        if source_name not in unique_records.columns:
            continue
        if display_label in used_labels:
            continue
        selected.append(source_name)
        labels.append(display_label)
        used_labels.add(display_label)

    if selected:
        display_records = unique_records.loc[:, selected].copy()
        display_records.columns = labels
    else:
        visible_columns = [
            column
            for column in unique_records.columns
            if not str(column).startswith("_")
        ]
        display_records = unique_records.loc[:, visible_columns].copy()

    return display_records.reset_index(drop=True), duplicates_hidden


def _clean_cell(value: object) -> str:
    """Return a safe display value without changing workbook data."""

    if value is None:
        return ""
    return str(value).strip()


def _canonical_digits(value: object) -> str:
    """Return digits from a telecom identifier without altering its source."""

    return re.sub(r"\D", "", _clean_cell(value))


def detect_number_columns(headers: list[object] | tuple[object, ...]) -> tuple[int, ...]:
    """Return indexes of columns that represent telephone numbers."""

    return tuple(
        index
        for index, header in enumerate(headers)
        if _NUMBER_HEADER_PATTERN.search(_clean_cell(header))
    )


def detect_identifier_columns(
    headers: list[object] | tuple[object, ...],
) -> dict[int, str]:
    """Map workbook columns to safe telecom identifier types."""

    patterns = (
        ("imei", _IMEI_HEADER_PATTERN),
        ("imsi", _IMSI_HEADER_PATTERN),
        ("cell_id", _CELL_HEADER_PATTERN),
        ("phone", _NUMBER_HEADER_PATTERN),
    )
    detected: dict[int, str] = {}
    for index, header in enumerate(headers):
        text = _clean_cell(header)
        if _NON_IDENTIFIER_HEADER_PATTERN.search(text):
            continue
        for identifier_type, pattern in patterns:
            if pattern.search(text):
                detected[index] = identifier_type
                break
    return detected


def canonical_phone_number(value: object) -> str:
    """Normalize an Indian phone value for safe report-summary matching."""

    digits = _canonical_digits(value)
    if len(digits) == 12 and digits.startswith("91"):
        return digits[2:]
    if len(digits) == 11 and digits.startswith("0"):
        return digits[1:]
    return digits


def _canonical_identifier(value: object, identifier_type: str) -> str:
    """Normalize an identifier for matching and verified source queries."""

    if identifier_type == "phone":
        return canonical_phone_number(value)
    return _canonical_digits(value)


def _header_index(rows: list[tuple[object, ...]]) -> int:
    """Find the most likely table header within the report title area."""

    if not rows:
        return 0

    def score(row: tuple[object, ...]) -> tuple[int, int]:
        populated = sum(bool(_clean_cell(value)) for value in row)
        number_labels = len(detect_identifier_columns(row))
        return number_labels, populated

    return max(range(len(rows)), key=lambda index: score(rows[index]))


class ReportViewerDialog(QDialog):
    """Display an Excel workbook without modifying the evidence file."""

    def __init__(
        self,
        report_path: str | Path,
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self._report_path = Path(report_path).expanduser().resolve(strict=False)
        self._workbook = None
        self._number_columns: tuple[int, ...] = ()
        self._identifier_columns: dict[int, str] = {}
        self._verified_source_link = None
        self._active_sheet_name = ""
        self._sections_by_sheet: dict[str, tuple[ReportSection, ...]] = {}
        self._active_section: ReportSection | None = None
        self._legacy_header_row: int | None = None
        self._legacy_headers: tuple[str, ...] = ()
        self._page_index = 0
        self._loaded_headers: tuple[str, ...] = ()
        self._loaded_rows: tuple[tuple[object, ...], ...] = ()
        self._section_total_records = 0

        self.setWindowTitle("Investigation Report Viewer")
        self.resize(1200, 760)
        self.setModal(True)

        self._file_label = QLabel(self._report_path.name)
        self._file_label.setObjectName("moduleTitle")

        self._sheet_selector = QComboBox()
        self._sheet_selector.setAccessibleName("Report sheet")
        self._sheet_selector.currentTextChanged.connect(self._show_sheet)

        self._section_label = QLabel("Section:")
        self._section_label.setVisible(False)
        self._section_selector = QComboBox()
        self._section_selector.setAccessibleName("Report section")
        self._section_selector.setVisible(False)
        self._section_selector.currentIndexChanged.connect(self._show_section)

        self._status = QLabel()
        self._status.setObjectName("cardText")
        self._status.setWordWrap(True)

        self._search_input = QLineEdit()
        self._search_input.setAccessibleName("Search loaded report records")
        self._search_input.setPlaceholderText("Search all columns on this page")
        self._search_input.setToolTip(
            "Search is case-insensitive and checks only records loaded on the "
            "current page."
        )
        self._search_input.setClearButtonEnabled(True)
        self._search_input.setEnabled(False)
        self._search_input.textChanged.connect(self._apply_search_filter)

        self._record_count_label = QLabel("Visible records: 0 of 0 loaded.")
        self._record_count_label.setAccessibleName("Visible record count")

        self._previous_page_button = QPushButton("Previous Page")
        self._previous_page_button.setObjectName("secondaryButton")
        self._previous_page_button.setAccessibleName("Previous report page")
        self._previous_page_button.clicked.connect(
            lambda _checked=False: self._change_page(-1)
        )
        self._page_label = QLabel("Page 0 of 0")
        self._page_label.setAccessibleName("Report page position")
        self._next_page_button = QPushButton("Next Page")
        self._next_page_button.setObjectName("secondaryButton")
        self._next_page_button.setAccessibleName("Next report page")
        self._next_page_button.clicked.connect(
            lambda _checked=False: self._change_page(1)
        )
        self._page_controls = QWidget()
        self._page_controls.setVisible(False)
        page_controls = QHBoxLayout(self._page_controls)
        page_controls.setContentsMargins(0, 0, 0, 0)
        page_controls.addStretch(1)
        page_controls.addWidget(self._previous_page_button)
        page_controls.addWidget(self._page_label)
        page_controls.addWidget(self._next_page_button)
        page_controls.addStretch(1)

        self._table = QTableWidget()
        self._table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._table.setSelectionBehavior(
            QAbstractItemView.SelectionBehavior.SelectRows
        )
        self._table.setAlternatingRowColors(True)
        self._table.setSortingEnabled(True)
        self._table.verticalHeader().setVisible(False)
        self._table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.ResizeToContents
        )
        self._table.cellDoubleClicked.connect(self._show_identifier_records)

        external_button = QPushButton("Open Externally")
        external_button.setObjectName("secondaryButton")
        external_button.clicked.connect(self._open_externally)

        close_button = QPushButton("Close")
        close_button.setObjectName("primaryButton")
        close_button.clicked.connect(self.accept)

        controls = QHBoxLayout()
        controls.addWidget(QLabel("Sheet:"))
        controls.addWidget(self._sheet_selector, stretch=1)
        controls.addWidget(external_button)
        controls.addWidget(close_button)

        section_controls = QHBoxLayout()
        section_controls.addWidget(self._section_label)
        section_controls.addWidget(self._section_selector, stretch=1)

        search_controls = QHBoxLayout()
        search_controls.addWidget(QLabel("Search:"))
        search_controls.addWidget(self._search_input, stretch=1)
        search_controls.addWidget(self._record_count_label)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 22, 24, 22)
        layout.setSpacing(12)
        layout.addWidget(self._file_label)
        layout.addLayout(controls)
        layout.addLayout(section_controls)
        layout.addLayout(search_controls)
        layout.addWidget(self._page_controls)
        layout.addWidget(self._status)
        layout.addWidget(self._table, stretch=1)

        self._load_workbook()

    @property
    def number_columns(self) -> tuple[int, ...]:
        """Return number columns detected in the visible sheet."""

        return self._number_columns

    def _load_workbook(self) -> None:
        """Open the workbook in read-only mode and list its sheets."""

        try:
            self._workbook = load_workbook(
                self._report_path,
                read_only=True,
                data_only=True,
            )
        except Exception as error:
            self._status.setText(
                "This workbook could not be read. "
                f"{type(error).__name__}: {error}"
            )
            self._search_input.setEnabled(False)
            self._page_controls.setVisible(False)
            self._table.setEnabled(False)
            return

        blocker = QSignalBlocker(self._sheet_selector)
        self._sheet_selector.addItems(self._workbook.sheetnames)
        del blocker
        if self._workbook.sheetnames:
            self._show_sheet(self._workbook.sheetnames[0])

    def _show_sheet(self, sheet_name: str) -> None:
        """Render a bounded number of rows from the selected sheet."""

        if self._workbook is None or not sheet_name:
            return

        self._active_sheet_name = sheet_name
        worksheet: Worksheet = self._workbook[sheet_name]
        sections = self._sections_by_sheet.get(sheet_name)
        if sections is None:
            sections = discover_report_sections(worksheet)
            self._sections_by_sheet[sheet_name] = sections

        self._set_section_choices(sections)
        if sections:
            self._render_section(sections[0])
            return

        self._render_legacy_sheet(worksheet)

    def _set_section_choices(
        self,
        sections: tuple[ReportSection, ...],
    ) -> None:
        """Show only canonical table sections for the active sheet."""

        blocker = QSignalBlocker(self._section_selector)
        self._section_selector.clear()
        for section in sections:
            self._section_selector.addItem(section.title, section)
        del blocker

        has_sections = bool(sections)
        self._section_label.setVisible(has_sections)
        self._section_selector.setVisible(has_sections)

    def _show_section(self, index: int) -> None:
        """Render the selected canonical section without rescanning the sheet."""

        if (
            self._workbook is None
            or not self._active_sheet_name
            or index < 0
        ):
            return

        section = self._section_selector.itemData(index)
        if not isinstance(section, ReportSection):
            return

        self._render_section(section)

    def _render_section(
        self,
        section: ReportSection,
    ) -> None:
        """Render one bounded report section and its investigator context."""

        self._active_section = section
        self._legacy_header_row = None
        self._legacy_headers = ()
        self._page_index = 0
        self._load_active_page()

    def _render_legacy_sheet(
        self,
        worksheet: Worksheet,
    ) -> None:
        """Preserve bounded viewing for older reports without sections."""

        rows = worksheet.iter_rows(values_only=True)
        scanned_rows = list(islice(rows, _HEADER_SCAN_ROWS))
        header_index = _header_index(scanned_rows)
        header_row = scanned_rows[header_index] if scanned_rows else ()
        headers = tuple(
            _clean_cell(value) or f"Column {index + 1}"
            for index, value in enumerate(header_row)
        )

        worksheet_rows = worksheet.max_row or 0
        self._active_section = None
        self._legacy_header_row = header_index + 1
        self._legacy_headers = headers
        self._section_total_records = max(
            worksheet_rows - self._legacy_header_row,
            0,
        )
        self._page_index = 0
        self._load_active_page()

    def _load_active_page(self) -> None:
        """Load one bounded page for the active section or legacy sheet."""

        if self._workbook is None or not self._active_sheet_name:
            return

        worksheet: Worksheet = self._workbook[self._active_sheet_name]
        offset = self._page_index * _MAX_VISIBLE_ROWS
        if self._active_section is not None:
            section = self._active_section
            visible_rows = read_report_section_rows(
                worksheet,
                section,
                offset=offset,
                limit=_MAX_VISIBLE_ROWS,
            )
            self._render_table(
                section.headers,
                visible_rows,
                total_records=section.record_count,
            )

            empty_note = (
                " No records are available for this section."
                if section.is_empty
                else ""
            )
            guidance_note = (
                f" Review guidance: {section.guidance}"
                if section.guidance
                else ""
            )
            identifier_note = (
                " Highlighted identifiers open verified source records."
                if self._identifier_columns
                else ""
            )
            self._status.setText(
                f"Sheet: {self._active_sheet_name} | Section: {section.title} | "
                f"Records: {section.record_count:,}."
                f"{self._page_range_note(len(visible_rows))}"
                f"{empty_note}{guidance_note}{identifier_note}"
            )
        else:
            header_row = self._legacy_header_row
            if header_row is None:
                return

            total_rows = self._section_total_records
            visible_rows: tuple[tuple[object, ...], ...] = ()
            if offset < total_rows and self._legacy_headers:
                first_row = header_row + 1 + offset
                last_row = min(
                    header_row + total_rows,
                    first_row + _MAX_VISIBLE_ROWS - 1,
                )
                visible_rows = tuple(
                    tuple(row)
                    for row in worksheet.iter_rows(
                        min_row=first_row,
                        max_row=last_row,
                        min_col=1,
                        max_col=len(self._legacy_headers),
                        values_only=True,
                    )
                )

            self._render_table(
                self._legacy_headers,
                visible_rows,
                total_records=total_rows,
            )
            identifier_note = (
                " Highlighted identifiers open verified source records."
                if self._identifier_columns
                else ""
            )
            self._status.setText(
                f"Sheet: {self._active_sheet_name} | Records: {total_rows:,}."
                f"{self._page_range_note(len(visible_rows))}{identifier_note}"
            )

        self._update_page_controls()

    def _page_range_note(self, visible_count: int) -> str:
        """Describe the active page range when a table spans multiple pages."""

        if self._section_total_records <= _MAX_VISIBLE_ROWS or visible_count < 1:
            return ""

        first_record = self._page_index * _MAX_VISIBLE_ROWS + 1
        last_record = first_record + visible_count - 1
        return (
            f" Showing records {first_record:,}-{last_record:,} "
            f"of {self._section_total_records:,}."
        )

    def _update_page_controls(self) -> None:
        """Show valid navigation actions for the active record page."""

        page_count = (
            self._section_total_records + _MAX_VISIBLE_ROWS - 1
        ) // _MAX_VISIBLE_ROWS
        has_multiple_pages = page_count > 1
        self._page_controls.setVisible(has_multiple_pages)
        self._page_label.setText(
            f"Page {self._page_index + 1} of {page_count}"
            if page_count
            else "Page 0 of 0"
        )
        self._previous_page_button.setEnabled(
            has_multiple_pages and self._page_index > 0
        )
        self._next_page_button.setEnabled(
            has_multiple_pages and self._page_index + 1 < page_count
        )

    def _change_page(self, delta: int) -> None:
        """Move to one valid adjacent record page."""

        page_count = (
            self._section_total_records + _MAX_VISIBLE_ROWS - 1
        ) // _MAX_VISIBLE_ROWS
        next_index = self._page_index + delta
        if next_index < 0 or next_index >= page_count:
            return

        self._page_index = next_index
        self._load_active_page()

    def _render_table(
        self,
        headers: list[str] | tuple[str, ...],
        visible_rows: (
            list[tuple[object, ...]]
            | tuple[tuple[object, ...], ...]
        ),
        *,
        total_records: int,
    ) -> None:
        """Store loaded rows and render them with safe local filtering."""

        self._loaded_headers = tuple(headers)
        self._loaded_rows = tuple(tuple(row) for row in visible_rows)
        self._section_total_records = max(
            total_records,
            len(self._loaded_rows),
        )

        blocker = QSignalBlocker(self._search_input)
        self._search_input.clear()
        del blocker
        self._search_input.setEnabled(bool(self._loaded_rows))

        self._identifier_columns = detect_identifier_columns(
            self._loaded_headers
        )
        self._number_columns = tuple(
            index
            for index, identifier_type in self._identifier_columns.items()
            if identifier_type == "phone"
        )

        self._apply_search_filter("")

    def _apply_search_filter(self, query: str) -> None:
        """Show loaded rows containing a case-insensitive literal query."""

        normalized_query = query.strip().casefold()
        if normalized_query:
            filtered_rows = tuple(
                row
                for row in self._loaded_rows
                if any(
                    normalized_query in _clean_cell(value).casefold()
                    for value in row
                )
            )
        else:
            filtered_rows = self._loaded_rows

        self._populate_table(filtered_rows)

        loaded_count = len(self._loaded_rows)
        count_text = (
            f"Visible records: {len(filtered_rows):,} "
            f"of {loaded_count:,} loaded."
        )
        if self._section_total_records > loaded_count:
            count_text += (
                f" Section total: {self._section_total_records:,}."
            )
        self._record_count_label.setText(count_text)

    def _populate_table(
        self,
        visible_rows: tuple[tuple[object, ...], ...],
    ) -> None:
        """Populate filtered rows while preserving sorting and drill-down."""

        self._table.setSortingEnabled(False)
        self._table.clear()
        self._table.setColumnCount(len(self._loaded_headers))
        self._table.setHorizontalHeaderLabels(self._loaded_headers)
        self._table.setRowCount(len(visible_rows))

        for row_index, row in enumerate(visible_rows):
            for column_index in range(len(self._loaded_headers)):
                value = row[column_index] if column_index < len(row) else None
                item = QTableWidgetItem(_clean_cell(value))
                if column_index in self._identifier_columns:
                    label = _IDENTIFIER_LABELS[
                        self._identifier_columns[column_index]
                    ]
                    item.setForeground(QColor("#285A1F"))
                    item.setToolTip(f"Double-click to view related {label} records")
                self._table.setItem(row_index, column_index, item)

        self._table.setSortingEnabled(True)

    def _show_number_summary(self, row: int, column: int) -> None:
        """Backward-compatible phone-number drill-down entry point."""

        self._show_identifier_records(row, column)

    def _show_identifier_records(self, row: int, column: int) -> None:
        """Show verified records for a phone, Cell ID, IMEI or IMSI."""

        identifier_type = self._identifier_columns.get(column)
        if identifier_type is None:
            return

        label = _IDENTIFIER_LABELS[identifier_type]
        selected = self._table.item(row, column)
        identifier = selected.text().strip() if selected is not None else ""
        canonical_value = _canonical_identifier(identifier, identifier_type)
        if len(canonical_value) < _MINIMUM_IDENTIFIER_DIGITS[identifier_type]:
            QMessageBox.information(
                self,
                f"Invalid {label}",
                f"The selected value is not a valid {label}.",
            )
            return

        matches = 0
        for row_index in range(self._table.rowCount()):
            item = self._table.item(row_index, column)
            if item is not None:
                candidate = _canonical_identifier(
                    item.text(),
                    identifier_type,
                )
                if candidate == canonical_value:
                    matches += 1

        try:
            from modules.reporting import cdr_report_source
        except ImportError:
            QMessageBox.warning(
                self,
                "Related Records Unavailable",
                "The verified source reader is not available in this installation.",
            )
            return

        QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
        try:
            if self._verified_source_link is None:
                self._verified_source_link = cdr_report_source.load_verified_source_link(
                    self._report_path
                )
            if self._verified_source_link is None:
                QMessageBox.information(
                    self,
                    f"{label} Summary",
                    (
                        f"Selected {label}: {identifier}\n"
                        f"Matching rows in the visible report sheet: {matches}\n\n"
                        "This older report has no verified source-data link. "
                        "Only the workbook summary can be shown safely."
                    ),
                )
                return

            records = cdr_report_source.query_related_records(
                self._verified_source_link,
                canonical_value,
                identifier_type=identifier_type,
            )
        except cdr_report_source.SourceLinkError as error:
            QMessageBox.warning(
                self,
                "Source Verification Failed",
                (
                    "Complete records were not opened because the report-to-source "
                    f"verification failed.\n\n{error}"
                ),
            )
            return
        except Exception as error:
            QMessageBox.warning(
                self,
                "Related Records",
                f"Related records could not be read.\n{type(error).__name__}: {error}",
            )
            return
        finally:
            QApplication.restoreOverrideCursor()

        if records.empty:
            QMessageBox.information(
                self,
                "Related Records",
                (
                    f"Selected {label}: {identifier}\n"
                    f"Matching rows in the visible report sheet: {matches}\n\n"
                    "No matching Call/SMS records were found in the verified source."
                ),
            )
            return

        RelatedRecordsDialog(
            identifier,
            records,
            self,
            identifier_label=label,
        ).exec()

    def _open_externally(self) -> None:
        """Open the report with the operating-system application."""

        if not QDesktopServices.openUrl(QUrl.fromLocalFile(str(self._report_path))):
            QMessageBox.warning(
                self,
                "Open Report",
                "The operating system could not open the report.",
            )

    def closeEvent(self, event: QCloseEvent) -> None:  # noqa: N802
        """Release the read-only workbook handle when the dialog closes."""

        if self._workbook is not None:
            self._workbook.close()
        super().closeEvent(event)
